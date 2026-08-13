# ==========================================================
# MonAgric - MVP unificado (estable)
# Home + Riego (sector validado) + Cosechas (validado) + SQLite
# Con manejo de errores (no se cierra: muestra dialogo y log)
# ==========================================================

from __future__ import annotations

import csv
import hashlib
import json
import math
import os
import sqlite3
import traceback
from pathlib import Path
from datetime import datetime, date, time, timedelta
from urllib import parse, request, error

from kivy.config import Config
from kivy.utils import platform as _kivy_platform

# Solo en escritorio: el botón del medio/derecho del mouse simula multitouch y
# deja puntos rojos en pantalla.
# En Android NO hay que tocar el proveedor de input: SDL sintetiza eventos de
# mouse a partir del táctil, así que sumar el proveedor "mouse" hace que CADA
# toque se despache DOS veces (menús que se abren y cierran solos, teclado que
# no aparece, scroll errático, registros duplicados y crash al reabrir un menú).
_ES_MOVIL = _kivy_platform in ("android", "ios")
if not _ES_MOVIL:
    # En escritorio el botón del medio/derecho simula multitouch (puntos rojos).
    Config.set("input", "mouse", "mouse,disable_multitouch")

from kivy.lang import Builder
from kivy.factory import Factory
from kivy.clock import Clock
from kivy.metrics import dp
from kivy.properties import StringProperty, BooleanProperty
from kivy.utils import get_color_from_hex, platform
from kivy.core.window import Window
from kivy.uix.image import Image
from kivy.uix.progressbar import ProgressBar
from kivy.uix.scrollview import ScrollView
from kivy.uix.behaviors import ButtonBehavior
from kivy.uix.widget import Widget
from kivy.graphics import Color, Line
from kivymd.app import MDApp
from kivymd.uix.screen import MDScreen
from kivymd.uix.card import MDCard
from kivy.uix.screenmanager import ScreenManager
from kivymd.uix.snackbar import Snackbar
from kivymd.uix.list import OneLineListItem, TwoLineAvatarIconListItem, IconLeftWidget, IconRightWidget
from kivymd.uix.menu import MDDropdownMenu
from kivymd.uix.dialog import MDDialog
from kivymd.uix.button import MDFlatButton, MDIconButton, MDRaisedButton
from kivymd.uix.boxlayout import MDBoxLayout
from kivymd.uix.label import MDLabel, MDIcon
from kivymd.uix.textfield import MDTextField
from kivymd.uix.dropdownitem import MDDropDownItem
from kivymd.uix.selectioncontrol import MDCheckbox
# kivymd.uix.pickers (MDDatePicker) es de los imports más caros y solo hace
# falta al abrir un calendario: se importa perezosamente en _date_picker().
# (kivymd.uix.fitimage no se usaba: se quitó.)


def _date_picker(**kwargs):
    from kivymd.uix.pickers import MDDatePicker
    return MDDatePicker(**kwargs)


# ==========================================================
# CONFIG
# ==========================================================

APP_NAME = "MonAgric"

def _resolve_data_dir() -> Path:
    # En Android los datos deben vivir en el almacenamiento interno de la app:
    # el directorio de trabajo se reextrae en cada actualizacion y se pierde.
    if platform == "android":
        from android.storage import app_storage_path
        base = Path(app_storage_path())
    else:
        base = Path(__file__).resolve().parent
    base.mkdir(parents=True, exist_ok=True)
    return base

DATA_DIR = _resolve_data_dir()

def _migrar_archivos_monagro() -> None:
    # La app se llamaba MonAgro: si existen datos con el nombre viejo y no hay
    # archivo nuevo, se renombran para no perder nada al actualizar.
    renombres = {
        "monagro.sqlite3": "monagric.sqlite3",
        "monagro_error.log": "monagric_error.log",
    }
    for viejo, nuevo in renombres.items():
        p_viejo, p_nuevo = DATA_DIR / viejo, DATA_DIR / nuevo
        if p_viejo.exists() and not p_nuevo.exists():
            try:
                p_viejo.rename(p_nuevo)
            except OSError:
                pass

_migrar_archivos_monagro()
DB_PATH = DATA_DIR / "monagric.sqlite3"
LOG_PATH = DATA_DIR / "monagric_error.log"
CSV_DIR = DATA_DIR / "csv"
SESSION_PATH = DATA_DIR / "stock_session.json"
STOCK_CULTIVOS_PATH = DATA_DIR / "stock_cultivos.json"
CHACRAS_PATH = DATA_DIR / "stock_chacras.json"
SAGE_GREEN  = get_color_from_hex("#5C7A5E")
WARM_AMBER  = get_color_from_hex("#D4860B")
CARD_BG     = get_color_from_hex("#F7F9F5")
SURFACE_BG  = get_color_from_hex("#EEF2EC")
CARD_BORDER = get_color_from_hex("#B8CDB9")
BANNER_PATH = "img/fondo.jpg"
DEFAULT_API_BASE_URL = "http://127.0.0.1:8000"
DEFAULT_PREVIEW_DEVICE = "desktop"
PREVIEW_SIZES = {
    "phone": (360, 800),        # Android estandar (aprox.)
    "phone_large": (412, 915),  # Android grande
    "tablet": (800, 1280),      # Tablet vertical
    "desktop": (560, 900),      # Ventana inicial comoda en PC
}

SECTORES = list("ABCDEFGHIJK")
BANCAL_MIN, BANCAL_MAX = 1, 15
BANCALES = [str(i) for i in range(BANCAL_MIN, BANCAL_MAX + 1)]

SIN_ASIGNAR = "S/A"

CULTIVOS = [
    SIN_ASIGNAR,
    "Acelga", "Ajo", "Ajo de verdeo", "Albahaca", "Apio", "Berenjena", "Brocoli",
    "Cebolla", "Cherry", "Choclos", "Coliflor", "Espinaca", "Habas", "Hakusai",
    "Kale", "Khol Rabi", "Lechuga", "Mix de hojas", "Morron", "Nabo Hakurei",
    "Perejil", "Puerro", "Rabanito", "Remolacha", "Repollo bco", "Repollo col",
    "Rucula", "Tomate", "Zanahoria", "Zapallito", "Zapallo", "Zucchini"
]

LOGROS_RIEGO = [
    "Capacidad de campo",
    "Riego aceptable",
    "Riego insuficiente",
]

# P´ onderación de calidad de riego para la barra del home (0..1)
PESOS_LOGRO_RIEGO = {
    "Capacidad de campo": 1.0,
    "Riego aceptable": 0.75,
    "Riego insuficiente": 0.25,
}
RIEGO_VERDE = 0.75   # promedio "aceptable o mejor"
RIEGO_AMBAR = 0.50
LOGRO_RIEGO_ABREV = {
    "Capacidad de campo": "CC",
    "Riego aceptable": "OK",
    "Riego insuficiente": "INSUF",
}

FILTRO_TODOS = "Todos"
DEFAULT_CHACRA = "Chacra 1"

TIPOS_SIEMBRA = [
    "Siembra directa",
    "Siembra almácigo",
    "Trasplante",
    "Esqueje",
]

UNIDADES_SUPERFICIE = ["bancales", "m2", "ha"]
DEFAULT_BANCAL_M2 = 15.0
DEFAULT_LARGO_BANCAL_M = 20.0
DEFAULT_PASILLO_M = 0.5

TIPOS_RIEGO = ["Aspersión", "Goteo", "Surco", "Superficie"]

# Las mismas actividades que usa la planilla de horas del proyecto, para que los
# registros cargados desde el celular y desde la PC sean comparables.
ACTIVIDADES_TRABAJO = ["Planificación", "Siembra", "Trasplante", "Manejo productivo",
                       "Cosecha y acondicionado", "Administración", "Comercialización",
                       "Comunicación", "Mantenimiento"]
ROLES_INTEGRANTE = ["Administrador", "Encargado", "Operario", "Otro"]

# ============================================================================
# SANIDAD — monitoreo fitosanitario
# Base de síntomas, signos y enfermedades tomada de "Apuntes de Patología
# Vegetal" (Rivera y Wright, 2020, Cátedra de Fitopatología FAUBA): cap. 3
# (síntomas y signos), cuadros 6.8/7.2/8.3/9.2 y cap. 13 (patologías).
# ============================================================================

# Síntomas más comunes (una modificación en el aspecto o función del órgano)
SINTOMAS_FITO = [
    "Ninguno",
    "Podredumbre (húmeda o seca)",
    "Cancro",
    "Antracnosis",
    "Mancha foliar",
    "Tizón (secado rápido)",
    "Escaldadura",
    "Ausencia de órganos / granos",
    "Enanismo",
    "Agalla / tumor",
    "Torsión y ampollado",
    "Escoba de brujas (brotes)",
    "Edema (intumescencia)",
    "Sarna",
    "Filodia / virescencia",
    "Clorosis (amarillez)",
    "Mosaico",
    "Pigmentación / enrojecimiento",
    "Marchitez",
    "Marchitamiento vascular",
    "Menor producción",
    "Otro",
]

# Signos más comunes (evidencia directa del agente causal)
SIGNOS_FITO = [
    "Ninguno",
    "Moho gris",
    "Moho blanco pulverulento (oídio)",
    "Mildiu (pelusa en el envés)",
    "Pústulas (roya)",
    "Soros negros (carbón)",
    "Micelio / esclerocios",
    "Fumagina (moho negro)",
    "Exudado bacteriano",
    "Insectos visibles",
    "Huevos / larvas",
    "Telarañas (ácaros)",
    "Galerías / minas",
    "Planta parásita (cuscuta)",
    "Otro",
]

# Bioinsumos disponibles para tratamiento (producto -> dosis y objetivo).
# Lista base pedida por el productor + cobre/azufre (usados como preventivos
# en producción orgánica según cap. 12 del libro).
PRODUCTOS_SANIDAD = {
    "MML (leche + melaza)":   {"dosis": "5 % (500 cc/10 L)",     "objetivo": "Preventivo, oídio"},
    "Té de compost":          {"dosis": "1:5 en agua",           "objetivo": "Antagonistas, hongos"},
    "Té de cola de caballo":  {"dosis": "5 % del extracto",      "objetivo": "Fungicida (oídio/mildiu/botritis)"},
    "InsectBio":              {"dosis": "3-5 cc/L",              "objetivo": "Insectos chupadores"},
    "Trichodermas":           {"dosis": "1-2 g/L o al sustrato", "objetivo": "Hongos de suelo"},
    "Té de ortiga":           {"dosis": "1:10 en agua",          "objetivo": "Fortificante, pulgones"},
    "Tierra de diatomeas":    {"dosis": "20-30 g/L o espolvoreo","objetivo": "Insectos y ácaros"},
    "Bacillus thuringiensis": {"dosis": "1-2 g/L",               "objetivo": "Orugas y larvas"},
    "Cobre (caldo bordelés)": {"dosis": "máx 6 kg/ha/año",       "objetivo": "Bacterias y hongos"},
    "Azufre":                 {"dosis": "3-5 g/L",               "objetivo": "Oídio y ácaros"},
    "Otro":                   {"dosis": "",                      "objetivo": ""},
}
PRODUCTOS_SANIDAD_LISTA = list(PRODUCTOS_SANIDAD.keys())

# Frecuencias de aplicación ofrecidas (días entre aplicaciones)
FRECUENCIAS_APLICACION = [7, 10, 15, 21, 30]

# Base de diagnóstico: cada entrada asocia síntomas y signos con una posible
# enfermedad o plaga y los bioinsumos recomendados. El motor de diagnóstico
# puntúa cada entrada por coincidencia (los signos pesan más porque revelan
# al agente). Tags de síntomas/signos deben coincidir con SINTOMAS_FITO/SIGNOS_FITO.
ENFERMEDADES_DB = [
    {"nombre": "Podredumbre gris (Botrytis)", "tipo": "Hongo",
     "sintomas": ["Podredumbre (húmeda o seca)", "Mancha foliar", "Tizón (secado rápido)"],
     "signos": ["Moho gris"],
     "productos": ["Té de cola de caballo", "Té de compost", "Trichodermas"]},
    {"nombre": "Mal de los almácigos (damping-off)", "tipo": "Hongo",
     "sintomas": ["Podredumbre (húmeda o seca)", "Marchitez"],
     "signos": ["Micelio / esclerocios", "Moho gris"],
     "productos": ["Trichodermas", "Té de compost"]},
    {"nombre": "Oídio", "tipo": "Hongo",
     "sintomas": ["Clorosis (amarillez)", "Menor producción"],
     "signos": ["Moho blanco pulverulento (oídio)"],
     "productos": ["Té de cola de caballo", "Azufre", "MML (leche + melaza)"]},
    {"nombre": "Mildiu", "tipo": "Hongo/Oomycete",
     "sintomas": ["Mancha foliar", "Clorosis (amarillez)"],
     "signos": ["Mildiu (pelusa en el envés)"],
     "productos": ["Té de cola de caballo", "Cobre (caldo bordelés)", "Té de compost"]},
    {"nombre": "Roya", "tipo": "Hongo",
     "sintomas": ["Clorosis (amarillez)", "Menor producción", "Pigmentación / enrojecimiento"],
     "signos": ["Pústulas (roya)"],
     "productos": ["Té de cola de caballo", "Azufre"]},
    {"nombre": "Carbón", "tipo": "Hongo",
     "sintomas": ["Ausencia de órganos / granos"],
     "signos": ["Soros negros (carbón)"],
     "productos": ["Té de compost"]},
    {"nombre": "Antracnosis", "tipo": "Hongo",
     "sintomas": ["Antracnosis", "Mancha foliar", "Podredumbre (húmeda o seca)"],
     "signos": ["Micelio / esclerocios"],
     "productos": ["Té de cola de caballo", "Cobre (caldo bordelés)"]},
    {"nombre": "Torque / lepra (Taphrina)", "tipo": "Hongo",
     "sintomas": ["Torsión y ampollado", "Pigmentación / enrojecimiento"],
     "signos": ["Micelio / esclerocios"],
     "productos": ["Cobre (caldo bordelés)", "Té de cola de caballo"]},
    {"nombre": "Fumagina", "tipo": "Hongo",
     "sintomas": ["Clorosis (amarillez)", "Menor producción"],
     "signos": ["Fumagina (moho negro)"],
     "productos": ["InsectBio", "Té de ortiga"]},
    {"nombre": "Fusariosis / marchitamiento vascular", "tipo": "Hongo",
     "sintomas": ["Marchitamiento vascular", "Marchitez", "Clorosis (amarillez)"],
     "signos": ["Micelio / esclerocios"],
     "productos": ["Trichodermas", "Té de compost"]},
    {"nombre": "Mancha bacteriana / podredumbre negra", "tipo": "Bacteria",
     "sintomas": ["Mancha foliar", "Clorosis (amarillez)", "Tizón (secado rápido)", "Podredumbre (húmeda o seca)"],
     "signos": ["Exudado bacteriano"],
     "productos": ["Cobre (caldo bordelés)", "Té de compost"]},
    {"nombre": "Cancrosis bacteriana", "tipo": "Bacteria",
     "sintomas": ["Cancro", "Mancha foliar", "Sarna"],
     "signos": ["Exudado bacteriano"],
     "productos": ["Cobre (caldo bordelés)"]},
    {"nombre": "Marchitez bacteriana (Ralstonia)", "tipo": "Bacteria",
     "sintomas": ["Marchitamiento vascular", "Marchitez"],
     "signos": ["Exudado bacteriano"],
     "productos": ["Té de compost"]},
    {"nombre": "Agalla de corona (Agrobacterium)", "tipo": "Bacteria",
     "sintomas": ["Agalla / tumor"],
     "signos": [],
     "productos": ["Trichodermas"]},
    {"nombre": "Sarna común (Streptomyces)", "tipo": "Bacteria",
     "sintomas": ["Sarna"],
     "signos": [],
     "productos": ["Té de compost"]},
    {"nombre": "Virosis / mosaico", "tipo": "Virus",
     "sintomas": ["Mosaico", "Clorosis (amarillez)", "Torsión y ampollado", "Enanismo", "Filodia / virescencia"],
     "signos": [],
     "productos": ["InsectBio", "Té de ortiga"]},
    {"nombre": "Fitoplasma (escoba de brujas / amarillez)", "tipo": "Mollicute",
     "sintomas": ["Escoba de brujas (brotes)", "Enanismo", "Clorosis (amarillez)", "Filodia / virescencia"],
     "signos": [],
     "productos": ["InsectBio"]},
    {"nombre": "Pulgones / áfidos", "tipo": "Plaga",
     "sintomas": ["Torsión y ampollado", "Clorosis (amarillez)", "Menor producción"],
     "signos": ["Insectos visibles", "Fumagina (moho negro)"],
     "productos": ["InsectBio", "Té de ortiga", "Tierra de diatomeas"]},
    {"nombre": "Mosca blanca", "tipo": "Plaga",
     "sintomas": ["Clorosis (amarillez)", "Menor producción"],
     "signos": ["Insectos visibles", "Fumagina (moho negro)"],
     "productos": ["InsectBio", "Tierra de diatomeas"]},
    {"nombre": "Trips", "tipo": "Plaga",
     "sintomas": ["Pigmentación / enrojecimiento", "Mancha foliar", "Menor producción"],
     "signos": ["Insectos visibles"],
     "productos": ["InsectBio", "Tierra de diatomeas"]},
    {"nombre": "Ácaros / arañuela", "tipo": "Plaga",
     "sintomas": ["Clorosis (amarillez)", "Pigmentación / enrojecimiento"],
     "signos": ["Telarañas (ácaros)", "Insectos visibles"],
     "productos": ["Azufre", "Tierra de diatomeas", "Té de cola de caballo"]},
    {"nombre": "Orugas / isocas (lepidópteros)", "tipo": "Plaga",
     "sintomas": ["Ausencia de órganos / granos", "Menor producción"],
     "signos": ["Huevos / larvas", "Insectos visibles"],
     "productos": ["Bacillus thuringiensis", "InsectBio"]},
    {"nombre": "Minadores de hoja", "tipo": "Plaga",
     "sintomas": ["Mancha foliar"],
     "signos": ["Galerías / minas", "Huevos / larvas"],
     "productos": ["InsectBio", "Bacillus thuringiensis"]},
    {"nombre": "Nematodes", "tipo": "Plaga",
     "sintomas": ["Agalla / tumor", "Enanismo", "Marchitez"],
     "signos": [],
     "productos": ["Tierra de diatomeas", "Té de compost"]},
    {"nombre": "Cuscuta (planta parásita)", "tipo": "Planta",
     "sintomas": ["Marchitez", "Enanismo", "Menor producción"],
     "signos": ["Planta parásita (cuscuta)"],
     "productos": []},
    {"nombre": "Trastorno no parasitario (nutricional / ambiental)", "tipo": "Abiótico",
     "sintomas": ["Edema (intumescencia)", "Clorosis (amarillez)", "Escaldadura",
                  "Pigmentación / enrojecimiento", "Marchitez"],
     "signos": [],
     "productos": ["Té de compost"]},
]

SEVERIDADES_SANIDAD = [0, 25, 50, 75, 100]


def diagnosticar_sanidad(sintomas, signos, limite: int = 3):
    """Devuelve [(nombre, tipo, score, productos)] ordenado por probabilidad.

    Los signos revelan al agente causal, así que pesan más que los síntomas.
    'Ninguno' y 'Otro' no aportan al diagnóstico."""
    ignora = {"", "Ninguno", "Otro"}
    if isinstance(sintomas, str):
        sintomas = [sintomas]
    if isinstance(signos, str):
        signos = [signos]
    sint = {s for s in sintomas if s not in ignora}
    sig = {s for s in signos if s not in ignora}
    if not sint and not sig:
        return []
    resultados = []
    for e in ENFERMEDADES_DB:
        score = len(sint & set(e["sintomas"])) + len(sig & set(e["signos"])) * 3
        if score > 0:
            resultados.append((e["nombre"], e["tipo"], score, e.get("productos", [])))
    resultados.sort(key=lambda r: r[2], reverse=True)
    return resultados[:limite]

# Semáforo del tablero: real / esperado a la fecha
SEMAFORO_VERDE = 0.90
SEMAFORO_AMBAR = 0.60

TIPO_SIEMBRA_DIRECTA = "Directa"
TIPO_SIEMBRA_ALMACIGO = "Almácigo"
DEFAULT_ANCHO_BANCAL_M = 0.8

# Tipo de cosecha (forma de la curva en el tiempo):
#  - concentrada: cosecha en pocos días (lechuga, brócoli, zapallo).
#  - escalonada: varias recolecciones a lo largo de semanas (tomate, berenjena).
#  - continua: se cosecha progresivamente hasta el fin de la temporada
#    (bienales de hoja/raíz: acelga, kale, remolacha, zanahoria, perejil).
COSECHA_CONCENTRADA = "concentrada"
COSECHA_ESCALONADA = "escalonada"
COSECHA_CONTINUA = "continua"
TIPOS_COSECHA = [COSECHA_CONCENTRADA, COSECHA_ESCALONADA, COSECHA_CONTINUA]

# Perfil por cultivo: (tipo de siembra, dias de almacigo a trasplante,
# dias de trasplante/siembra a inicio de cosecha, ventana de cosecha en dias,
# rinde kg/m2, distancia entre plantas cm, lineas por bancal, tipo de cosecha).
# Fuente: "Planificación 26-27 - Información de cultivos" (datos de Martín);
# la ventana surge de "días en cosecha" (promedio máx/mín). Los cultivos
# 'continua' llevan una ventana amplia pero se recortan a la fecha de fin de
# temporada. Ajuste: Ajo rinde 18->1.8 (posible error de tipeo). Editable.
# Valores de la planilla "Información de cultivos" (referencias/): los días de
# almácigo, de cama-a-cosecha y de ventana de cosecha son el promedio de la
# columna máximo/mínimo; la ventana tiene un piso de 5 días. Rinde 0 en la
# planilla (Ajo de verdeo, Khol Rabi) se reemplaza por un estimado razonable.
PERFIL_CULTIVO_DEFAULTS = {
    "Acelga":        (TIPO_SIEMBRA_ALMACIGO, 35,  52, 140, 5.5, 40, 3, COSECHA_CONTINUA),
    "Ajo":           (TIPO_SIEMBRA_DIRECTA,   0, 270,   5, 2.4, 15, 4, COSECHA_CONCENTRADA),
    "Ajo de verdeo": (TIPO_SIEMBRA_ALMACIGO, 45,  60,   8, 2.5, 20, 4, COSECHA_ESCALONADA),
    "Albahaca":      (TIPO_SIEMBRA_ALMACIGO, 40,  75,  75, 3.0, 25, 4, COSECHA_ESCALONADA),
    "Apio":          (TIPO_SIEMBRA_ALMACIGO, 53, 105,  11, 4.0, 30, 4, COSECHA_CONCENTRADA),
    "Berenjena":     (TIPO_SIEMBRA_ALMACIGO, 26,  85,  45, 3.0, 40, 2, COSECHA_ESCALONADA),
    "Brocoli":       (TIPO_SIEMBRA_ALMACIGO, 35,  90,   8, 4.0, 50, 3, COSECHA_CONCENTRADA),
    "Cebolla":       (TIPO_SIEMBRA_ALMACIGO, 75, 130,  16, 4.0, 10, 4, COSECHA_CONCENTRADA),
    "Cherry":        (TIPO_SIEMBRA_ALMACIGO, 55,  95,  65, 4.0, 45, 2, COSECHA_ESCALONADA),
    "Choclos":       (TIPO_SIEMBRA_ALMACIGO, 35, 110,  16, 2.5, 25, 2, COSECHA_CONCENTRADA),
    "Coliflor":      (TIPO_SIEMBRA_ALMACIGO, 35, 105,   8, 3.0, 50, 3, COSECHA_CONCENTRADA),
    "Espinaca":      (TIPO_SIEMBRA_ALMACIGO, 26,  65,   8, 3.0, 20, 5, COSECHA_CONCENTRADA),
    "Habas":         (TIPO_SIEMBRA_DIRECTA,   0,  85,  18, 2.5, 15, 1, COSECHA_ESCALONADA),
    "Hakusai":       (TIPO_SIEMBRA_ALMACIGO, 38, 105,  16, 6.0, 40, 3, COSECHA_CONCENTRADA),
    "Kale":          (TIPO_SIEMBRA_ALMACIGO, 35,  50, 150, 5.0, 40, 3, COSECHA_CONTINUA),
    "Khol Rabi":     (TIPO_SIEMBRA_ALMACIGO, 35,  70,  11, 3.0, 20, 4, COSECHA_CONCENTRADA),
    "Lechuga":       (TIPO_SIEMBRA_ALMACIGO, 38,  58,   8, 4.8, 30, 4, COSECHA_CONCENTRADA),
    "Mix de hojas":  (TIPO_SIEMBRA_ALMACIGO, 38,  52,   8, 3.0,  5, 5, COSECHA_CONCENTRADA),
    "Morron":        (TIPO_SIEMBRA_ALMACIGO, 26,  70,  40, 5.0, 45, 2, COSECHA_ESCALONADA),
    "Nabo Hakurei":  (TIPO_SIEMBRA_ALMACIGO, 26,  55,  11, 4.5, 15, 5, COSECHA_CONCENTRADA),
    "Perejil":       (TIPO_SIEMBRA_ALMACIGO, 38,  70,  50, 4.5, 10, 5, COSECHA_CONTINUA),
    "Puerro":        (TIPO_SIEMBRA_ALMACIGO, 75, 100,  61, 5.0, 15, 4, COSECHA_CONTINUA),
    "Rabanito":      (TIPO_SIEMBRA_DIRECTA,   0,  50,   6, 3.0,  5, 5, COSECHA_CONCENTRADA),
    "Remolacha":     (TIPO_SIEMBRA_ALMACIGO, 38,  65, 126, 5.0, 15, 5, COSECHA_CONTINUA),
    "Repollo bco":   (TIPO_SIEMBRA_ALMACIGO, 35, 105,  16, 5.0, 45, 3, COSECHA_CONCENTRADA),
    "Repollo col":   (TIPO_SIEMBRA_ALMACIGO, 35, 105,  16, 5.0, 45, 3, COSECHA_CONCENTRADA),
    "Rucula":        (TIPO_SIEMBRA_DIRECTA,   0,  40,   8, 5.0,  5, 5, COSECHA_CONCENTRADA),
    "Tomate":        (TIPO_SIEMBRA_ALMACIGO, 55,  95,  65, 4.0, 45, 2, COSECHA_ESCALONADA),
    "Zanahoria":     (TIPO_SIEMBRA_DIRECTA,   0,  93, 126, 6.0, 10, 5, COSECHA_CONTINUA),
    "Zapallito":     (TIPO_SIEMBRA_ALMACIGO, 35,  33,  38, 4.5, 50, 1, COSECHA_ESCALONADA),
    "Zapallo":       (TIPO_SIEMBRA_ALMACIGO, 35, 130,  16, 3.0, 80, 1, COSECHA_CONCENTRADA),
    "Zucchini":      (TIPO_SIEMBRA_ALMACIGO, 35,  33,  38, 4.5, 50, 1, COSECHA_ESCALONADA),
}


# Iconos ilustrativos para el selector de cultivos (Material Design Icons);
# los cultivos sin icono propio usan "sprout".
CULTIVO_ICONO_DEFAULT = "sprout"
CULTIVO_ICONOS = {
    "Acelga": "leaf",
    "Ajo": "seed",
    "Ajo de verdeo": "sprout",
    "Albahaca": "leaf",
    "Berenjena": "fruit-grapes",
    "Brocoli": "tree",
    "Cherry": "fruit-cherries",
    "Choclos": "corn",
    "Coliflor": "flower",
    "Espinaca": "leaf",
    "Habas": "peanut",
    "Kale": "leaf",
    "Lechuga": "sprout",
    "Mix de hojas": "leaf-maple",
    "Morron": "chili-mild",
    "Nabo Hakurei": "carrot",
    "Perejil": "grass",
    "Rabanito": "carrot",
    "Remolacha": "carrot",
    "Repollo bco": "flower",
    "Repollo col": "flower",
    "Rucula": "leaf",
    "Tomate": "food-apple",
    "Zanahoria": "carrot",
    "Zapallito": "pumpkin",
    "Zapallo": "pumpkin",
    "Zucchini": "pumpkin",
}


def dias_a_cosecha_total(tipo: str, dias_almacigo: int, dias_trasplante_cosecha: int) -> int:
    """Dias desde la siembra (en almacigo o directa) hasta el inicio de cosecha."""
    if norm_text(tipo) == TIPO_SIEMBRA_ALMACIGO:
        return int(dias_almacigo) + int(dias_trasplante_cosecha)
    return int(dias_trasplante_cosecha)


def calcular_plantas(superficie_m2: float, distancia_cm: float, lineas: int, ancho_bancal_m: float) -> int:
    """Numero de plantas segun marco de plantacion y superficie."""
    if superficie_m2 <= 0 or distancia_cm <= 0 or lineas <= 0 or ancho_bancal_m <= 0:
        return 0
    metros_lineales_de_bancal = superficie_m2 / ancho_bancal_m
    return int(round(metros_lineales_de_bancal * lineas / (distancia_cm / 100.0)))


def sugerir_nombre_temporada(hoy: date | None = None) -> str:
    d = hoy or date.today()
    if d.month >= 7:
        return f"{d.year}-{(d.year + 1) % 100:02d}"
    return f"{d.year - 1}-{d.year % 100:02d}"


def superficie_a_m2(cantidad: float, unidad: str, bancal_m2: float) -> float:
    unidad = norm_text(unidad).lower()
    if unidad == "bancales":
        return round(cantidad * bancal_m2, 2)
    if unidad == "ha":
        return round(cantidad * 10000.0, 2)
    return round(cantidad, 2)


def api_healthcheck(base_url: str, timeout: float = 2.0) -> bool:
    base = norm_text(base_url).rstrip("/")
    if not base:
        return False
    url = f"{base}/health"
    req = request.Request(url=url, method="GET", headers={"Accept": "application/json"})
    try:
        with request.urlopen(req, timeout=timeout) as res:
            return 200 <= int(getattr(res, "status", 0)) < 300
    except Exception:
        return False

# ==========================================================
# HELPERS / VALIDACIONES
# ==========================================================

def _instalar_filtro_mouse_sintetico():
    """En Android/iOS, SDL2 inyecta por cada toque real un evento de mouse
    sintético (device == "mouse"). Eso hacía que CADA toque se despachara dos
    veces: menús que se abrían y cerraban solos, teclado que no aparecía, scroll
    que solo andaba por los márgenes y registros/tareas/horas duplicados.

    Confirmado en la tablet: un toque físico generaba un SDL2MotionEvent (real)
    y un MouseMotionEvent (dev='mouse', duplicado). Interceptamos los tres
    eventos de la ventana y descartamos los que vienen del mouse sintético."""
    if not _ES_MOVIL:
        return
    try:
        from kivy.core.window import Window
    except Exception:
        return

    def _es_sintetico(touch):
        return getattr(touch, "device", None) == "mouse"

    # Un observador bindeado que devuelve True frena el despacho al resto de la
    # ventana (a los widgets): así el evento de mouse sintético se descarta y solo
    # queda el toque real. Devolver None deja seguir el flujo normal.
    def _filtro_down(_win, touch):
        return True if _es_sintetico(touch) else None

    def _filtro_move(_win, touch):
        return True if _es_sintetico(touch) else None

    def _filtro_up(_win, touch):
        return True if _es_sintetico(touch) else None

    Window.bind(on_touch_down=_filtro_down,
                on_touch_move=_filtro_move,
                on_touch_up=_filtro_up)


class SectionCard(MDCard):
    """Tarjeta de sección.

    Es una MDCard normal: varias tarjetas del tablero también son botones
    (se les hace bind(on_release=...) para abrir su detalle), así que NO hay
    que interceptar on_touch_down. Se intentó hacerlo para arreglar el scroll,
    pero el scroll fallaba por otra causa (el toque duplicado del mouse
    sintético, ver _instalar_filtro_mouse_sintetico) y el override rompía el
    click de las tarjetas.
    """


def abrir_menu(menu):
    """Abre un MDDropdownMenu de forma segura.

    Si el menú ya está abierto, KivyMD intenta agregarlo de nuevo a la ventana y
    lanza WidgetException, que tumbaba la app. Un toque repetido (o un evento
    duplicado) no debe romper nada."""
    if menu is None:
        return
    if getattr(menu, "parent", None) is not None:
        return  # ya está abierto
    try:
        menu.open()
    except Exception as e:
        log_exception("No se pudo abrir el menú", e)


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")

def norm_text(s: str) -> str:
    return (s or "").strip()

def validate_fecha(fecha: str) -> str:
    fecha = norm_text(fecha)
    # Normaliza a ISO con ceros (acepta 2026-8-1 y devuelve 2026-08-01)
    return datetime.strptime(fecha, "%Y-%m-%d").date().isoformat()

def validate_optional_fecha(fecha: str) -> str:
    fecha = norm_text(fecha)
    if not fecha:
        return ""
    return validate_fecha(fecha)

def parse_hhmm(s: str) -> time:
    return datetime.strptime(norm_text(s), "%H:%M").time()

def validate_sector(sector: str) -> str:
    sector = norm_text(sector).upper()
    if sector in SECTORES:
        return sector
    # Sectores definidos por el usuario en la temporada activa
    try:
        if sector in {norm_text(s).upper() for s in sectores_de_temporada_activa()}:
            return sector
    except Exception:
        pass
    raise ValueError("Sector inválido (definí los sectores en la temporada).")

def validate_bancal(bancal_str: str) -> int:
    try:
        b = int(norm_text(bancal_str))
    except Exception:
        raise ValueError("Bancal debe ser un numero entero.")
    if not (BANCAL_MIN <= b <= BANCAL_MAX):
        raise ValueError("Bancal fuera de rango (1..15).")
    return b

def validate_kg(kg_str: str) -> float:
    s = norm_text(kg_str).replace(",", ".")
    try:
        kg = float(s)
    except Exception:
        raise ValueError("Kg debe ser un numero (ej: 12.5).")
    if kg <= 0:
        raise ValueError("Kg debe ser > 0.")
    return round(kg, 3)

def validate_positive_float(value: str, field_name: str) -> float:
    s = norm_text(value).replace(",", ".")
    try:
        number = float(s)
    except Exception:
        raise ValueError(f"{field_name} debe ser un numero (ej: 12.5).")
    if number <= 0:
        raise ValueError(f"{field_name} debe ser > 0.")
    return round(number, 3)

def validate_cultivo(cultivo: str) -> str:
    cultivo = norm_text(cultivo)
    if cultivo == SIN_ASIGNAR or cultivo in CULTIVOS:
        return cultivo
    # Cultivos creados por el usuario
    try:
        if cultivo in get_cultivos():
            return cultivo
    except Exception:
        pass
    raise ValueError("Cultivo no valido (use la lista).")

def validate_chacra_nombre(chacra: str) -> str:
    chacra = norm_text(chacra)
    if not chacra:
        raise ValueError("El nombre de chacra es obligatorio.")
    if len(chacra) > 40:
        raise ValueError("El nombre de chacra no puede superar 40 caracteres.")
    if chacra.lower() == "agregar chacra":
        raise ValueError("'Agregar Chacra' es una opcion reservada.")
    return chacra

def validate_peso_unit(peso_str: str) -> float:
    s = norm_text(peso_str).replace(",", ".")
    try:
        peso = float(s)
    except Exception:
        raise ValueError("Peso por unidad debe ser un numero (ej: 0.35).")
    if peso <= 0:
        raise ValueError("Peso por unidad debe ser > 0.")
    return round(peso, 3)

def validate_unidades(unidades_str: str) -> int:
    s = norm_text(unidades_str)
    try:
        unidades = int(s)
    except Exception:
        raise ValueError("Cantidad de unidades debe ser un numero entero (ej: 20).")
    if unidades <= 0:
        raise ValueError("Cantidad de unidades debe ser > 0.")
    return unidades

def validate_tipo_siembra(tipo: str) -> str:
    tipo = norm_text(tipo)
    if tipo not in TIPOS_SIEMBRA:
        raise ValueError("Tipo no válido. Opciones: " + ", ".join(TIPOS_SIEMBRA))
    return tipo

def validate_generacion(gen_str: str) -> int:
    s = norm_text(gen_str)
    try:
        gen = int(s)
    except Exception:
        raise ValueError("Generación debe ser un número entero (ej: 1).")
    if gen < 1:
        raise ValueError("Generación debe ser >= 1.")
    return gen

def validate_bandejas(ban_str: str) -> int:
    s = norm_text(ban_str)
    try:
        ban = int(s)
    except Exception:
        raise ValueError("Cantidad de bandejas debe ser un número entero (ej: 2).")
    if ban < 1:
        raise ValueError("Cantidad de bandejas debe ser >= 1.")
    return ban

def validate_horas(horas_str: str) -> int:
    s = norm_text(horas_str)
    try:
        horas = int(s)
    except Exception:
        raise ValueError("Horas de riego debe ser un número entero (ej: 2).")
    if horas <= 0:
        raise ValueError("Horas de riego debe ser > 0.")
    return horas

def validate_logro(logro: str) -> str:
    logro = norm_text(logro)
    if logro not in LOGROS_RIEGO:
        raise ValueError("Logro de riego no válido.")
    return logro

def is_sin_asignar(cultivo: str) -> bool:
    return norm_text(cultivo) == SIN_ASIGNAR

def calc_horas(inicio: time, fin: time) -> float:
    dt0 = datetime.combine(date.today(), inicio)
    dt1 = datetime.combine(date.today(), fin)
    if dt1 < dt0:
        dt1 = dt1 + timedelta(days=1)
    return round((dt1 - dt0).total_seconds() / 3600, 2)

# ==========================================================
# LOG / ERRORES
# ==========================================================

def log_exception(prefix: str, exc: BaseException):
    tb = traceback.format_exc()
    msg = f"\n[{now_iso()}] {prefix}\n{tb}\n"
    try:
        LOG_PATH.write_text((LOG_PATH.read_text(encoding="utf-8") if LOG_PATH.exists() else "") + msg, encoding="utf-8")
    except Exception:
        # si falla escribir log, al menos lo tiramos por consola
        print(msg)

def show_error_dialog(title: str, text: str):
    app = MDApp.get_running_app()
    if not app:
        return
    # Cerramos dialogo previo si existia
    if getattr(app, "_err_dialog", None):
        try:
            app._err_dialog.dismiss()
        except Exception:
            pass
    app._err_dialog = MDDialog(
        title=title,
        text=text,
        md_bg_color=SAGE_GREEN,
        buttons=[MDFlatButton(
            text="OK",
            text_color=(1, 1, 1, 1),
            on_release=lambda *_: app._err_dialog.dismiss()
        )],
    )
    app._err_dialog.open()

# ==========================================================
# DB
# ==========================================================

def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON;")
    return conn

def append_csv(path: Path, headers: list[str], row: list):
    CSV_DIR.mkdir(exist_ok=True)
    new_file = not path.exists()
    with path.open("a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if new_file:
            writer.writerow(headers)
        writer.writerow(row)

def init_db():
    with get_conn() as conn:
        # ---- RIEGO
        conn.execute("""
        CREATE TABLE IF NOT EXISTS riego (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha TEXT NOT NULL,
            hora_inicio TEXT NOT NULL,
            hora_fin TEXT NOT NULL,
            horas_riego REAL NOT NULL,
            logro_riego TEXT NOT NULL,
            operador TEXT NOT NULL,
            created_at TEXT NOT NULL
        );
        """)
        # Migracion: agregar sector si falta
        cols = [r[1] for r in conn.execute("PRAGMA table_info(riego);").fetchall()]
        if "sector" not in cols:
            conn.execute("ALTER TABLE riego ADD COLUMN sector TEXT NOT NULL DEFAULT 'A';")
        if "logro_riego" not in cols:
            conn.execute("ALTER TABLE riego ADD COLUMN logro_riego TEXT NOT NULL DEFAULT 'Riego aceptable';")

        # ---- COSECHAS
        conn.execute("""
        CREATE TABLE IF NOT EXISTS cosechas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha TEXT NOT NULL,
            cultivo TEXT NOT NULL,
            kg REAL NOT NULL,
            sector TEXT NOT NULL,
            bancal INTEGER NOT NULL,
            created_at TEXT NOT NULL
        );
        """)

        # ---- STOCK
        conn.execute("""
        CREATE TABLE IF NOT EXISTS stock (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha TEXT NOT NULL,
            chacra TEXT NOT NULL DEFAULT 'Chacra 1',
            sector TEXT NOT NULL,
            bancal INTEGER NOT NULL,
            cultivo TEXT NOT NULL,
            kg_disponible REAL NOT NULL,
            peso_unitario REAL NOT NULL,
            unidades REAL NOT NULL,
            created_at TEXT NOT NULL
        );
        """)
        stock_cols = [r[1] for r in conn.execute("PRAGMA table_info(stock);").fetchall()]
        if "chacra" not in stock_cols:
            conn.execute("ALTER TABLE stock ADD COLUMN chacra TEXT NOT NULL DEFAULT 'Chacra 1';")

        # ---- OBJETIVOS
        conn.execute("""
        CREATE TABLE IF NOT EXISTS objetivos (
            cultivo TEXT PRIMARY KEY,
            superficie_m2 REAL NOT NULL,
            cosecha_esperada_kg REAL NOT NULL,
            temporada_inicio TEXT NOT NULL DEFAULT '',
            temporada_fin TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );
        """)
        objetivos_cols = [r[1] for r in conn.execute("PRAGMA table_info(objetivos);").fetchall()]
        if "temporada_inicio" not in objetivos_cols:
            conn.execute("ALTER TABLE objetivos ADD COLUMN temporada_inicio TEXT NOT NULL DEFAULT '';")
        if "temporada_fin" not in objetivos_cols:
            conn.execute("ALTER TABLE objetivos ADD COLUMN temporada_fin TEXT NOT NULL DEFAULT '';")

        # ---- SIEMBRAS Y TRASPLANTES
        conn.execute("""
        CREATE TABLE IF NOT EXISTS siembras (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha TEXT NOT NULL,
            cultivo TEXT NOT NULL,
            variedad TEXT NOT NULL DEFAULT '',
            tipo TEXT NOT NULL,
            generacion INTEGER NOT NULL DEFAULT 1,
            bandejas INTEGER NOT NULL DEFAULT 1,
            tipo_bandeja INTEGER NOT NULL DEFAULT 72,
            observaciones TEXT NOT NULL DEFAULT '',
            operador TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL
        );
        """)
        siembras_cols = [r[1] for r in conn.execute("PRAGMA table_info(siembras);").fetchall()]
        if "tipo_bandeja" not in siembras_cols:
            conn.execute("ALTER TABLE siembras ADD COLUMN tipo_bandeja INTEGER NOT NULL DEFAULT 72;")

        # ---- TAREAS
        conn.execute("""
        CREATE TABLE IF NOT EXISTS tareas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tarea TEXT NOT NULL,
            fecha TEXT NOT NULL,
            importancia TEXT NOT NULL DEFAULT 'Media',
            n_personas INTEGER NOT NULL DEFAULT 1,
            realizada INTEGER NOT NULL DEFAULT 0,
            fecha_realizada TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL
        );
        """)

        conn.execute("""
        CREATE TABLE IF NOT EXISTS app_config (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        );
        """)

        # Migracion: cada registro operativo queda enlazado a su temporada
        for tabla in ("cosechas", "riego", "siembras", "tareas", "stock"):
            cols = [r[1] for r in conn.execute(f"PRAGMA table_info({tabla});").fetchall()]
            if "temporada_id" not in cols:
                conn.execute(f"ALTER TABLE {tabla} ADD COLUMN temporada_id INTEGER;")

        # ---- CULTIVOS CREADOS POR EL USUARIO
        conn.execute("""
        CREATE TABLE IF NOT EXISTS cultivos_extra (
            nombre TEXT PRIMARY KEY,
            created_at TEXT NOT NULL
        );
        """)

        # ---- SECTORES DE RIEGO POR TEMPORADA
        conn.execute("""
        CREATE TABLE IF NOT EXISTS sectores_riego (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            temporada_id INTEGER NOT NULL,
            sector TEXT NOT NULL,
            bancales INTEGER NOT NULL DEFAULT 1,
            tipo_riego TEXT NOT NULL DEFAULT '',
            UNIQUE(temporada_id, sector)
        );
        """)

        # ---- INTEGRANTES DEL PROYECTO Y HORAS DE TRABAJO
        conn.execute("""
        CREATE TABLE IF NOT EXISTS integrantes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            direccion TEXT NOT NULL DEFAULT '',
            telefono TEXT NOT NULL DEFAULT '',
            rol TEXT NOT NULL DEFAULT '',
            valor_hora REAL NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL
        );
        """)
        conn.execute("""
        CREATE TABLE IF NOT EXISTS horas_trabajo (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha TEXT NOT NULL,
            integrante TEXT NOT NULL,
            horas REAL NOT NULL,
            actividades TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL,
            temporada_id INTEGER
        );
        """)

        # ---- TRASPLANTES (un registro = un bancal trasplantado)
        conn.execute("""
        CREATE TABLE IF NOT EXISTS trasplantes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha TEXT NOT NULL,
            integrante TEXT NOT NULL DEFAULT '',
            cultivo TEXT NOT NULL,
            generacion INTEGER NOT NULL DEFAULT 1,
            sector TEXT NOT NULL DEFAULT '',
            bancal INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL,
            temporada_id INTEGER
        );
        """)

        # ---- FECHAS DE SIEMBRA PLANIFICADAS (una por generación)
        conn.execute("""
        CREATE TABLE IF NOT EXISTS plan_fechas_siembra (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            temporada_id INTEGER NOT NULL,
            cultivo TEXT NOT NULL,
            fecha TEXT NOT NULL
        );
        """)

        # Siembra directa: se registra tambien sector y bancal
        siembras_cols2 = [r[1] for r in conn.execute("PRAGMA table_info(siembras);").fetchall()]
        if "sector" not in siembras_cols2:
            conn.execute("ALTER TABLE siembras ADD COLUMN sector TEXT NOT NULL DEFAULT '';")
        if "bancal" not in siembras_cols2:
            conn.execute("ALTER TABLE siembras ADD COLUMN bancal INTEGER NOT NULL DEFAULT 0;")

        # ---- SANIDAD ----
        # Planilla única de TODAS las aplicaciones (rutina/preventivas y de
        # tratamiento por enfermedad). Es la base para descargar y analizar.
        conn.execute("""
        CREATE TABLE IF NOT EXISTS sanidad_aplicaciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha TEXT NOT NULL,
            producto TEXT NOT NULL DEFAULT '',
            dosis TEXT NOT NULL DEFAULT '',
            cultivo TEXT NOT NULL DEFAULT '',
            sector TEXT NOT NULL DEFAULT '',
            bancal INTEGER NOT NULL DEFAULT 0,
            tipo TEXT NOT NULL DEFAULT 'rutina',
            tratamiento_id INTEGER,
            diagnostico TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL,
            temporada_id INTEGER
        );
        """)
        # Detecciones de enfermedad o plaga durante el monitoreo
        conn.execute("""
        CREATE TABLE IF NOT EXISTS sanidad_detecciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha TEXT NOT NULL,
            cultivo TEXT NOT NULL DEFAULT '',
            sector TEXT NOT NULL DEFAULT '',
            bancal INTEGER NOT NULL DEFAULT 0,
            sintoma TEXT NOT NULL DEFAULT '',
            signo TEXT NOT NULL DEFAULT '',
            diagnostico TEXT NOT NULL DEFAULT '',
            severidad INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL,
            temporada_id INTEGER
        );
        """)
        # Tratamientos planificados a partir de una detección
        conn.execute("""
        CREATE TABLE IF NOT EXISTS sanidad_tratamientos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            deteccion_id INTEGER,
            cultivo TEXT NOT NULL DEFAULT '',
            sector TEXT NOT NULL DEFAULT '',
            bancal INTEGER NOT NULL DEFAULT 0,
            producto TEXT NOT NULL DEFAULT '',
            dosis TEXT NOT NULL DEFAULT '',
            n_aplicaciones INTEGER NOT NULL DEFAULT 1,
            frecuencia_dias INTEGER NOT NULL DEFAULT 7,
            fecha_inicio TEXT NOT NULL,
            diagnostico TEXT NOT NULL DEFAULT '',
            activo INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL,
            temporada_id INTEGER
        );
        """)
        # Calendario de aplicaciones de cada tratamiento (con casilla realizada)
        conn.execute("""
        CREATE TABLE IF NOT EXISTS sanidad_plan_aplicaciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tratamiento_id INTEGER NOT NULL,
            fecha_programada TEXT NOT NULL,
            realizada INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL
        );
        """)

        # ---- TEMPORADAS Y PLAN
        conn.execute("""
        CREATE TABLE IF NOT EXISTS temporadas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            fecha_inicio TEXT NOT NULL,
            fecha_fin TEXT NOT NULL DEFAULT '',
            activa INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL
        );
        """)
        conn.execute("""
        CREATE TABLE IF NOT EXISTS plan_temporada (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            temporada_id INTEGER NOT NULL,
            cultivo TEXT NOT NULL,
            superficie_m2 REAL NOT NULL,
            cosecha_esperada_kg REAL NOT NULL,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            UNIQUE(temporada_id, cultivo)
        );
        """)

        # ---- PERFIL DE CULTIVO (curva de cosecha y marco de plantacion, editable)
        conn.execute("""
        CREATE TABLE IF NOT EXISTS cultivo_perfil (
            cultivo TEXT PRIMARY KEY,
            dias_a_cosecha INTEGER NOT NULL,
            ventana_cosecha_dias INTEGER NOT NULL,
            rinde_ref_kg_m2 REAL NOT NULL DEFAULT 0
        );
        """)
        perfil_cols = [r[1] for r in conn.execute("PRAGMA table_info(cultivo_perfil);").fetchall()]
        for col, ddl in (
            ("tipo_siembra", "TEXT NOT NULL DEFAULT 'Directa'"),
            ("dias_almacigo", "INTEGER NOT NULL DEFAULT 0"),
            ("dias_trasplante_cosecha", "INTEGER NOT NULL DEFAULT 0"),
            ("distancia_cm", "REAL NOT NULL DEFAULT 25"),
            ("lineas_bancal", "INTEGER NOT NULL DEFAULT 3"),
            ("tipo_cosecha", "TEXT NOT NULL DEFAULT 'escalonada'"),
        ):
            if col not in perfil_cols:
                conn.execute(f"ALTER TABLE cultivo_perfil ADD COLUMN {col} {ddl};")

        for cultivo, (tipo, alm, tc, ventana, rinde, dist, lineas, tcosecha) in PERFIL_CULTIVO_DEFAULTS.items():
            conn.execute(
                """INSERT OR IGNORE INTO cultivo_perfil
                   (cultivo, dias_a_cosecha, ventana_cosecha_dias, rinde_ref_kg_m2,
                    tipo_siembra, dias_almacigo, dias_trasplante_cosecha, distancia_cm, lineas_bancal, tipo_cosecha)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (cultivo, dias_a_cosecha_total(tipo, alm, tc), ventana, rinde, tipo, alm, tc, dist, lineas, tcosecha),
            )

        # Migracion unica: reemplaza los defaults estimados por los datos de la
        # planilla de planificacion 26-27 (solo la primera vez).
        ya_migrado = conn.execute(
            "SELECT value FROM app_config WHERE key = 'perfil_defaults_v2'"
        ).fetchone()
        if not ya_migrado:
            for cultivo, (tipo, alm, tc, ventana, rinde, dist, lineas, _tc) in PERFIL_CULTIVO_DEFAULTS.items():
                conn.execute(
                    """UPDATE cultivo_perfil SET
                           tipo_siembra = ?, dias_almacigo = ?, dias_trasplante_cosecha = ?,
                           dias_a_cosecha = ?, ventana_cosecha_dias = ?, rinde_ref_kg_m2 = ?,
                           distancia_cm = ?, lineas_bancal = ?
                       WHERE cultivo = ?""",
                    (tipo, alm, tc, dias_a_cosecha_total(tipo, alm, tc), ventana, rinde, dist, lineas, cultivo),
                )
            conn.execute(
                "INSERT OR REPLACE INTO app_config (key, value) VALUES ('perfil_defaults_v2', '1')"
            )

        # Migracion v3: ventanas de cosecha del PDF + tipo de cosecha
        # (concentrada/escalonada/continua). Solo pisa lo que el usuario no editó.
        ya_v3 = conn.execute(
            "SELECT value FROM app_config WHERE key = 'perfil_defaults_v3'"
        ).fetchone()
        if not ya_v3:
            for cultivo, (_ts, _alm, _tc, ventana, _rinde, _dist, _lin, tcosecha) in PERFIL_CULTIVO_DEFAULTS.items():
                conn.execute(
                    "UPDATE cultivo_perfil SET ventana_cosecha_dias = ?, tipo_cosecha = ? WHERE cultivo = ?",
                    (ventana, tcosecha, cultivo),
                )
            conn.execute(
                "INSERT OR REPLACE INTO app_config (key, value) VALUES ('perfil_defaults_v3', '1')"
            )

        # Migracion v4: valores oficiales de la planilla "Informacion de cultivos"
        # (referencias/). Reescribe TODOS los campos del perfil de una sola vez,
        # para que las bases ya existentes (PC y celular) queden con los defaults
        # de la planilla. A partir de aca el usuario puede editar cada perfil y su
        # cambio se respeta. Tambien se suman cultivos nuevos (Apio, Cebolla,
        # Hakusai, Khol Rabi, Puerro) via el INSERT OR IGNORE de mas arriba.
        ya_v4 = conn.execute(
            "SELECT value FROM app_config WHERE key = 'perfil_defaults_v4'"
        ).fetchone()
        if not ya_v4:
            for cultivo, (tipo, alm, tc, ventana, rinde, dist, lineas, tcosecha) in PERFIL_CULTIVO_DEFAULTS.items():
                conn.execute(
                    """UPDATE cultivo_perfil SET
                           tipo_siembra = ?, dias_almacigo = ?, dias_trasplante_cosecha = ?,
                           dias_a_cosecha = ?, ventana_cosecha_dias = ?, rinde_ref_kg_m2 = ?,
                           distancia_cm = ?, lineas_bancal = ?, tipo_cosecha = ?
                       WHERE cultivo = ?""",
                    (tipo, alm, tc, dias_a_cosecha_total(tipo, alm, tc), ventana, rinde,
                     dist, lineas, tcosecha, cultivo),
                )
            conn.execute(
                "INSERT OR REPLACE INTO app_config (key, value) VALUES ('perfil_defaults_v4', '1')"
            )

        # Marco de plantacion elegido por cultivo dentro del plan
        plan_cols = [r[1] for r in conn.execute("PRAGMA table_info(plan_temporada);").fetchall()]
        for col, ddl in (
            ("tipo_siembra", "TEXT NOT NULL DEFAULT ''"),
            ("distancia_cm", "REAL NOT NULL DEFAULT 0"),
            ("lineas", "INTEGER NOT NULL DEFAULT 0"),
            ("plantas", "INTEGER NOT NULL DEFAULT 0"),
        ):
            if col not in plan_cols:
                conn.execute(f"ALTER TABLE plan_temporada ADD COLUMN {col} {ddl};")
        conn.commit()

    global _cultivos_cache
    _cultivos_cache = None

_cultivos_cache = None

def get_cultivos() -> list:
    """Catalogo completo: cultivos base + los creados por el usuario."""
    global _cultivos_cache
    if _cultivos_cache is None:
        extras = []
        try:
            with get_conn() as conn:
                extras = [r[0] for r in conn.execute(
                    "SELECT nombre FROM cultivos_extra ORDER BY nombre").fetchall()]
        except Exception:
            extras = []
        base = [c for c in CULTIVOS if c != SIN_ASIGNAR]
        _cultivos_cache = sorted(set(base + extras), key=str.lower)
    return _cultivos_cache

def add_cultivo_extra(nombre: str) -> str:
    global _cultivos_cache
    nombre = norm_text(nombre)
    if not nombre:
        raise ValueError("El nombre del cultivo es obligatorio.")
    if len(nombre) > 30:
        raise ValueError("El nombre no puede superar 30 caracteres.")
    existentes = {c.lower() for c in get_cultivos()} | {SIN_ASIGNAR.lower()}
    if nombre.lower() in existentes:
        raise ValueError("Ese cultivo ya existe.")
    with get_conn() as conn:
        conn.execute(
            "INSERT INTO cultivos_extra (nombre, created_at) VALUES (?, ?)",
            (nombre, now_iso()),
        )
        conn.commit()
    _cultivos_cache = None
    return nombre

def _temporada_activa_id():
    try:
        t = get_temporada_activa()
        return t["id"] if t else None
    except Exception:
        return None

def insert_riego(fecha, horas, operador, sector, logro=""):
    created = now_iso()
    with get_conn() as conn:
        conn.execute(
            """INSERT INTO riego
            (fecha, hora_inicio, hora_fin, horas_riego, logro_riego, operador, sector, created_at, temporada_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (fecha, "", "", horas, logro, operador, sector, created, _temporada_activa_id())
        )
        conn.commit()
    append_csv(
        CSV_DIR / "riego.csv",
        ["fecha", "horas_riego", "logro_riego", "operador", "sector", "created_at"],
        [fecha, horas, logro, operador, sector, created],
    )

def list_riego_by_fecha(fecha):
    with get_conn() as conn:
        cur = conn.execute(
            """SELECT horas_riego, sector, operador
               FROM riego
               WHERE fecha = ?
               ORDER BY created_at""",
            (fecha,)
        )
        return cur.fetchall()

def list_riego_all(limit: int | None = None):
    with get_conn() as conn:
        sql = """SELECT fecha, horas_riego, sector, operador
                 FROM riego
                 ORDER BY fecha DESC, created_at DESC"""
        if limit:
            sql += " LIMIT ?"
            return conn.execute(sql, (limit,)).fetchall()
        return conn.execute(sql).fetchall()

def list_riego_between(fecha_desde: str, fecha_hasta: str):
    with get_conn() as conn:
        cur = conn.execute(
            """SELECT fecha, sector, horas_riego, operador
               FROM riego
               WHERE fecha >= ? AND fecha <= ?
               ORDER BY fecha ASC, sector ASC, created_at ASC""",
            (fecha_desde, fecha_hasta)
        )
        return cur.fetchall()


def insert_siembra(fecha, cultivo, variedad, tipo, generacion, bandejas, tipo_bandeja,
                   observaciones, operador, sector="", bancal=0):
    created = now_iso()
    with get_conn() as conn:
        conn.execute(
            """INSERT INTO siembras
            (fecha, cultivo, variedad, tipo, generacion, bandejas, tipo_bandeja, observaciones,
             operador, sector, bancal, created_at, temporada_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (fecha, cultivo, variedad, tipo, generacion, bandejas, tipo_bandeja, observaciones,
             operador, norm_text(sector), bancal, created, _temporada_activa_id())
        )
        conn.commit()
    append_csv(
        CSV_DIR / "siembras.csv",
        ["fecha", "cultivo", "variedad", "tipo", "generacion", "bandejas", "tipo_bandeja",
         "observaciones", "operador", "sector", "bancal", "created_at"],
        [fecha, cultivo, variedad, tipo, generacion, bandejas, tipo_bandeja, observaciones,
         operador, sector, bancal, created],
    )

def list_siembras_by_fecha(fecha):
    with get_conn() as conn:
        return conn.execute(
            """SELECT cultivo, variedad, tipo, generacion, bandejas, observaciones, operador
               FROM siembras WHERE fecha = ? ORDER BY created_at""",
            (fecha,)
        ).fetchall()

def list_siembras_all(limit: int | None = None):
    with get_conn() as conn:
        sql = """SELECT fecha, cultivo, variedad, tipo, generacion, bandejas, observaciones, operador
                 FROM siembras ORDER BY fecha DESC, created_at DESC"""
        if limit:
            sql += " LIMIT ?"
            return conn.execute(sql, (limit,)).fetchall()
        return conn.execute(sql).fetchall()

def list_siembras_between(fecha_desde: str, fecha_hasta: str):
    with get_conn() as conn:
        return conn.execute(
            """SELECT fecha, cultivo, variedad, tipo, generacion, bandejas, observaciones, operador
               FROM siembras WHERE fecha >= ? AND fecha <= ? ORDER BY fecha ASC, created_at ASC""",
            (fecha_desde, fecha_hasta)
        ).fetchall()


class DataRepository:
    """Contrato de acceso a datos para desacoplar UI de almacenamiento."""

    def insert_riego(self, fecha, horas, operador, sector, logro=""):
        raise NotImplementedError

    def list_riego_by_fecha(self, fecha):
        raise NotImplementedError

    def list_riego_all(self, limit: int | None = None):
        raise NotImplementedError

    def list_riego_between(self, fecha_desde: str, fecha_hasta: str):
        raise NotImplementedError

    def list_riego_last_48h(self):
        raise NotImplementedError

    def list_last_riego_by_sector(self):
        raise NotImplementedError

    def insert_cosecha(self, fecha, cultivo, kg, sector, bancal):
        raise NotImplementedError

    def list_cosechas_by_fecha(self, fecha):
        raise NotImplementedError

    def list_cosechas_all(self, limit: int | None = None):
        raise NotImplementedError

    def list_cosechas_between(self, fecha_desde: str, fecha_hasta: str):
        raise NotImplementedError

    def upsert_objetivo(
        self,
        cultivo: str,
        superficie_m2: float,
        cosecha_esperada_kg: float,
    ):
        raise NotImplementedError

    def get_objetivo(self, cultivo: str):
        raise NotImplementedError

    def set_temporada(self, inicio: str, fin: str):
        raise NotImplementedError

    def get_temporada(self):
        raise NotImplementedError

    def list_objetivos(self):
        raise NotImplementedError

    def delete_objetivo(self, cultivo: str):
        raise NotImplementedError

    def clear_objetivos(self):
        raise NotImplementedError

    def insert_stock(self, fecha, chacra, sector, bancal, cultivo, kg_disponible, peso_unitario, unidades):
        raise NotImplementedError

    def list_stock_by_fecha(self, fecha, chacra: str | None = None):
        raise NotImplementedError

    def list_stock_all(self, limit: int | None = None, chacra: str | None = None):
        raise NotImplementedError

    def insert_siembra(self, fecha, cultivo, variedad, tipo, generacion, bandejas, tipo_bandeja,
                       observaciones, operador, sector="", bancal=0):
        raise NotImplementedError

    def list_siembras_by_fecha(self, fecha):
        raise NotImplementedError

    def list_siembras_all(self, limit: int | None = None):
        raise NotImplementedError

    def list_siembras_between(self, fecha_desde: str, fecha_hasta: str):
        raise NotImplementedError

    def insert_tarea(self, tarea: str, fecha: str, importancia: str, n_personas: int):
        raise NotImplementedError

    def list_tareas_pendientes(self):
        raise NotImplementedError

    def marcar_tarea_realizada(self, tarea_id: int):
        raise NotImplementedError

    def list_tareas_realizadas_between(self, fecha_desde: str, fecha_hasta: str):
        raise NotImplementedError

    # ---- Edicion / borrado (v1: solo modo local)
    def list_cosechas_con_id(self, fecha: str | None = None, limit: int | None = None):
        raise NotImplementedError

    def update_cosecha(self, cosecha_id: int, fecha: str, kg: float):
        raise NotImplementedError

    def delete_cosecha(self, cosecha_id: int):
        raise NotImplementedError

    def list_riego_con_id(self, fecha: str | None = None, limit: int | None = None):
        raise NotImplementedError

    def update_riego(self, riego_id: int, fecha: str, horas: float, operador: str, logro: str = ""):
        raise NotImplementedError

    def delete_riego(self, riego_id: int):
        raise NotImplementedError

    def list_siembras_con_id(self, fecha: str | None = None, limit: int | None = None):
        raise NotImplementedError

    def update_siembra(self, siembra_id: int, fecha: str, variedad: str, generacion: int,
                       bandejas: int, tipo_bandeja: int, observaciones: str, operador: str):
        raise NotImplementedError

    def delete_siembra(self, siembra_id: int):
        raise NotImplementedError


class LocalDataRepository(DataRepository):
    """Implementacion local actual (SQLite + CSV)."""

    def insert_riego(self, fecha, horas, operador, sector, logro=""):
        insert_riego(fecha, horas, operador, sector, logro)

    def list_riego_by_fecha(self, fecha):
        return list_riego_by_fecha(fecha)

    def list_riego_all(self, limit: int | None = None):
        return list_riego_all(limit=limit)

    def list_riego_between(self, fecha_desde: str, fecha_hasta: str):
        return list_riego_between(fecha_desde, fecha_hasta)

    def list_riego_last_48h(self):
        return list_riego_last_48h()

    def list_last_riego_by_sector(self):
        return list_last_riego_by_sector()

    def insert_cosecha(self, fecha, cultivo, kg, sector, bancal):
        insert_cosecha(fecha, cultivo, kg, sector, bancal)

    def list_cosechas_by_fecha(self, fecha):
        return list_cosechas_by_fecha(fecha)

    def list_cosechas_all(self, limit: int | None = None):
        return list_cosechas_all(limit=limit)

    def list_cosechas_between(self, fecha_desde: str, fecha_hasta: str):
        return list_cosechas_between(fecha_desde, fecha_hasta)

    def upsert_objetivo(
        self,
        cultivo: str,
        superficie_m2: float,
        cosecha_esperada_kg: float,
    ):
        upsert_objetivo(cultivo, superficie_m2, cosecha_esperada_kg)

    def get_objetivo(self, cultivo: str):
        return get_objetivo(cultivo)

    def set_temporada(self, inicio: str, fin: str):
        set_temporada_config(inicio, fin)

    def get_temporada(self):
        return get_temporada_config()

    def list_objetivos(self):
        return list_objetivos()

    def delete_objetivo(self, cultivo: str):
        delete_objetivo(cultivo)

    def clear_objetivos(self):
        clear_objetivos()

    def insert_stock(self, fecha, chacra, sector, bancal, cultivo, kg_disponible, peso_unitario, unidades):
        insert_stock(fecha, chacra, sector, bancal, cultivo, kg_disponible, peso_unitario, unidades)

    def list_stock_by_fecha(self, fecha, chacra: str | None = None):
        return list_stock_by_fecha(fecha, chacra=chacra)

    def list_stock_all(self, limit: int | None = None, chacra: str | None = None):
        return list_stock_all(limit=limit, chacra=chacra)

    def insert_siembra(self, fecha, cultivo, variedad, tipo, generacion, bandejas, tipo_bandeja,
                       observaciones, operador, sector="", bancal=0):
        insert_siembra(fecha, cultivo, variedad, tipo, generacion, bandejas, tipo_bandeja,
                       observaciones, operador, sector, bancal)

    def list_siembras_by_fecha(self, fecha):
        return list_siembras_by_fecha(fecha)

    def list_siembras_all(self, limit: int | None = None):
        return list_siembras_all(limit=limit)

    def list_siembras_between(self, fecha_desde: str, fecha_hasta: str):
        return list_siembras_between(fecha_desde, fecha_hasta)

    def insert_tarea(self, tarea: str, fecha: str, importancia: str, n_personas: int):
        insert_tarea(tarea, fecha, importancia, n_personas)

    def list_tareas_pendientes(self):
        return list_tareas_pendientes()

    def marcar_tarea_realizada(self, tarea_id: int):
        marcar_tarea_realizada(tarea_id)

    def list_tareas_realizadas_between(self, fecha_desde: str, fecha_hasta: str):
        return list_tareas_realizadas_between(fecha_desde, fecha_hasta)

    def list_cosechas_con_id(self, fecha: str | None = None, limit: int | None = None):
        return list_cosechas_con_id(fecha=fecha, limit=limit)

    def update_cosecha(self, cosecha_id: int, fecha: str, kg: float):
        update_cosecha(cosecha_id, fecha, kg)

    def delete_cosecha(self, cosecha_id: int):
        delete_cosecha(cosecha_id)

    def list_riego_con_id(self, fecha: str | None = None, limit: int | None = None):
        return list_riego_con_id(fecha=fecha, limit=limit)

    def update_riego(self, riego_id: int, fecha: str, horas: float, operador: str, logro: str = ""):
        update_riego(riego_id, fecha, horas, operador, logro)

    def delete_riego(self, riego_id: int):
        delete_riego(riego_id)

    def list_siembras_con_id(self, fecha: str | None = None, limit: int | None = None):
        return list_siembras_con_id(fecha=fecha, limit=limit)

    def update_siembra(self, siembra_id: int, fecha: str, variedad: str, generacion: int,
                       bandejas: int, tipo_bandeja: int, observaciones: str, operador: str):
        update_siembra(siembra_id, fecha, variedad, generacion, bandejas, tipo_bandeja, observaciones, operador)

    def delete_siembra(self, siembra_id: int):
        delete_siembra(siembra_id)


class RemoteApiRepository(DataRepository):
    """Repositorio remoto HTTP para datos compartidos entre dispositivos."""

    def __init__(self, base_url: str, timeout: float = 10.0):
        self.base_url = (base_url or DEFAULT_API_BASE_URL).strip().rstrip("/")
        self.timeout = timeout

    def _url(self, path: str) -> str:
        return f"{self.base_url}{path}"

    def _request_json(self, method: str, path: str, params: dict | None = None, payload: dict | None = None):
        url = self._url(path)
        if params:
            clean = {k: v for k, v in params.items() if v is not None}
            qs = parse.urlencode(clean)
            if qs:
                url = f"{url}?{qs}"

        data = None
        headers = {"Accept": "application/json"}
        if payload is not None:
            data = json.dumps(payload).encode("utf-8")
            headers["Content-Type"] = "application/json"

        req = request.Request(url=url, data=data, method=method, headers=headers)
        try:
            with request.urlopen(req, timeout=self.timeout) as res:
                raw = res.read().decode("utf-8")
                if not raw:
                    return None
                return json.loads(raw)
        except error.HTTPError as e:
            body = ""
            try:
                body = e.read().decode("utf-8")
            except Exception:
                pass
            raise RuntimeError(f"HTTP {e.code} {path}: {body or e.reason}") from e
        except error.URLError as e:
            raise RuntimeError(f"No se pudo conectar a API: {e.reason}") from e

    def insert_riego(self, fecha, horas, operador, sector, logro=""):
        self._request_json(
            "POST",
            "/riego",
            payload={
                "fecha": fecha,
                "horas_riego": horas,
                "operador": operador,
                "sector": sector,
                "logro_riego": logro,
            },
        )

    def list_riego_by_fecha(self, fecha):
        rows = self._request_json("GET", "/riego", params={"fecha": fecha}) or []
        return [(r["horas_riego"], r["sector"], r["operador"]) for r in rows]

    def list_riego_all(self, limit: int | None = None):
        rows = self._request_json("GET", "/riego/all", params={"limit": limit}) or []
        return [(r["fecha"], r["horas_riego"], r["sector"], r["operador"]) for r in rows]

    def list_riego_between(self, fecha_desde: str, fecha_hasta: str):
        rows = self._request_json(
            "GET",
            "/riego/range",
            params={"fecha_desde": fecha_desde, "fecha_hasta": fecha_hasta},
        ) or []
        return [(r["fecha"], r["sector"], r["horas_riego"], r["operador"]) for r in rows]

    def list_riego_last_48h(self):
        rows = self._request_json("GET", "/riego/last-48h") or []
        return [(r["fecha"], r["sector"], r["horas_riego"], r["operador"]) for r in rows]

    def list_last_riego_by_sector(self):
        data = self._request_json("GET", "/riego/last-by-sector") or {}
        if not isinstance(data, dict):
            return {}
        return data

    def insert_cosecha(self, fecha, cultivo, kg, sector, bancal):
        self._request_json(
            "POST",
            "/cosechas",
            payload={
                "fecha": fecha,
                "cultivo": cultivo,
                "kg": kg,
                "sector": sector,
                "bancal": bancal,
            },
        )

    def list_cosechas_by_fecha(self, fecha):
        rows = self._request_json("GET", "/cosechas", params={"fecha": fecha}) or []
        return [(r["cultivo"], r["kg"], r["sector"], r["bancal"]) for r in rows]

    def list_cosechas_all(self, limit: int | None = None):
        rows = self._request_json("GET", "/cosechas/all", params={"limit": limit}) or []
        return [(r["fecha"], r["cultivo"], r["kg"], r["sector"], r["bancal"]) for r in rows]

    def list_cosechas_between(self, fecha_desde: str, fecha_hasta: str):
        rows = self.list_cosechas_all(limit=5000)
        return [r for r in rows if fecha_desde <= r[0] <= fecha_hasta]

    def upsert_objetivo(
        self,
        cultivo: str,
        superficie_m2: float,
        cosecha_esperada_kg: float,
    ):
        # Fallback local hasta exponer endpoint remoto de objetivos.
        upsert_objetivo(cultivo, superficie_m2, cosecha_esperada_kg)

    def get_objetivo(self, cultivo: str):
        return get_objetivo(cultivo)

    def set_temporada(self, inicio: str, fin: str):
        set_temporada_config(inicio, fin)

    def get_temporada(self):
        return get_temporada_config()

    def list_objetivos(self):
        return list_objetivos()

    def delete_objetivo(self, cultivo: str):
        delete_objetivo(cultivo)

    def clear_objetivos(self):
        clear_objetivos()

    def insert_stock(self, fecha, chacra, sector, bancal, cultivo, kg_disponible, peso_unitario, unidades):
        payload = {
            "fecha": fecha,
            "chacra": chacra,
            "sector": sector,
            "bancal": bancal,
            "cultivo": cultivo,
            "kg_disponible": kg_disponible,
            "peso_unitario": peso_unitario,
            "unidades": unidades,
        }
        try:
            self._request_json("POST", "/stock", payload=payload)
        except Exception:
            # Compatibilidad con backend viejo sin campo chacra
            payload.pop("chacra", None)
            self._request_json("POST", "/stock", payload=payload)

    def list_stock_by_fecha(self, fecha, chacra: str | None = None):
        params = {"fecha": fecha}
        if chacra:
            params["chacra"] = chacra
        try:
            rows = self._request_json("GET", "/stock", params=params) or []
        except Exception:
            if chacra:
                rows = self._request_json("GET", "/stock", params={"fecha": fecha}) or []
            else:
                raise
        return [(r["sector"], r["bancal"], r["cultivo"], r["kg_disponible"], r["peso_unitario"], r["unidades"]) for r in rows]

    def insert_siembra(self, fecha, cultivo, variedad, tipo, generacion, bandejas, tipo_bandeja,
                       observaciones, operador, sector="", bancal=0):
        self._request_json("POST", "/siembras", payload={
            "fecha": fecha, "cultivo": cultivo, "variedad": variedad,
            "tipo": tipo, "generacion": generacion, "bandejas": bandejas,
            "tipo_bandeja": tipo_bandeja,
            "observaciones": observaciones, "operador": operador,
            "sector": sector, "bancal": bancal,
        })

    def list_siembras_by_fecha(self, fecha):
        rows = self._request_json("GET", "/siembras", params={"fecha": fecha}) or []
        return [(r["cultivo"], r["variedad"], r["tipo"], r["generacion"], r["bandejas"], r["observaciones"], r["operador"]) for r in rows]

    def list_siembras_all(self, limit: int | None = None):
        rows = self._request_json("GET", "/siembras/all", params={"limit": limit}) or []
        return [(r["fecha"], r["cultivo"], r["variedad"], r["tipo"], r["generacion"], r["bandejas"], r["observaciones"], r["operador"]) for r in rows]

    def list_siembras_between(self, fecha_desde: str, fecha_hasta: str):
        rows = self._request_json("GET", "/siembras/range", params={"fecha_desde": fecha_desde, "fecha_hasta": fecha_hasta}) or []
        return [(r["fecha"], r["cultivo"], r["variedad"], r["tipo"], r["generacion"], r["bandejas"], r["observaciones"], r["operador"]) for r in rows]

    def list_stock_all(self, limit: int | None = None, chacra: str | None = None):
        params = {"limit": limit}
        if chacra:
            params["chacra"] = chacra
        try:
            rows = self._request_json("GET", "/stock/all", params=params) or []
        except Exception:
            if chacra:
                rows = self._request_json("GET", "/stock/all", params={"limit": limit}) or []
            else:
                raise
        return [
            (r["fecha"], r["sector"], r["bancal"], r["cultivo"], r["kg_disponible"], r["peso_unitario"], r["unidades"])
            for r in rows
        ]

    def insert_tarea(self, tarea: str, fecha: str, importancia: str, n_personas: int):
        insert_tarea(tarea, fecha, importancia, n_personas)

    def list_tareas_pendientes(self):
        return list_tareas_pendientes()

    def marcar_tarea_realizada(self, tarea_id: int):
        marcar_tarea_realizada(tarea_id)

    def list_tareas_realizadas_between(self, fecha_desde: str, fecha_hasta: str):
        return list_tareas_realizadas_between(fecha_desde, fecha_hasta)

    # Edicion/borrado remoto: sin endpoints en el backend todavia.
    # Se rechaza para no divergir los datos entre servidor y telefono.
    def list_cosechas_con_id(self, fecha: str | None = None, limit: int | None = None):
        raise RuntimeError("Editar/eliminar requiere modo de datos LOCAL (el servidor aún no lo soporta).")

    def update_cosecha(self, cosecha_id: int, fecha: str, kg: float):
        raise RuntimeError("Editar/eliminar requiere modo de datos LOCAL (el servidor aún no lo soporta).")

    def delete_cosecha(self, cosecha_id: int):
        raise RuntimeError("Editar/eliminar requiere modo de datos LOCAL (el servidor aún no lo soporta).")

    def list_riego_con_id(self, fecha: str | None = None, limit: int | None = None):
        raise RuntimeError("Editar/eliminar requiere modo de datos LOCAL (el servidor aún no lo soporta).")

    def update_riego(self, riego_id: int, fecha: str, horas: float, operador: str, logro: str = ""):
        raise RuntimeError("Editar/eliminar requiere modo de datos LOCAL (el servidor aún no lo soporta).")

    def delete_riego(self, riego_id: int):
        raise RuntimeError("Editar/eliminar requiere modo de datos LOCAL (el servidor aún no lo soporta).")

    def list_siembras_con_id(self, fecha: str | None = None, limit: int | None = None):
        raise RuntimeError("Editar/eliminar requiere modo de datos LOCAL (el servidor aún no lo soporta).")

    def update_siembra(self, siembra_id: int, fecha: str, variedad: str, generacion: int,
                       bandejas: int, tipo_bandeja: int, observaciones: str, operador: str):
        raise RuntimeError("Editar/eliminar requiere modo de datos LOCAL (el servidor aún no lo soporta).")

    def delete_siembra(self, siembra_id: int):
        raise RuntimeError("Editar/eliminar requiere modo de datos LOCAL (el servidor aún no lo soporta).")


class FailoverRepository(DataRepository):
    """Usa remoto y si falla pasa a local para no interrumpir operación."""

    def __init__(self, primary: DataRepository, fallback: DataRepository):
        self.primary = primary
        self.fallback = fallback

    def _call(self, method_name: str, *args, **kwargs):
        method = getattr(self.primary, method_name)
        try:
            return method(*args, **kwargs)
        except Exception:
            return getattr(self.fallback, method_name)(*args, **kwargs)

    def insert_riego(self, fecha, horas, operador, sector, logro=""):
        return self._call("insert_riego", fecha, horas, operador, sector, logro)

    def list_riego_by_fecha(self, fecha):
        return self._call("list_riego_by_fecha", fecha)

    def list_riego_all(self, limit: int | None = None):
        return self._call("list_riego_all", limit=limit)

    def list_riego_between(self, fecha_desde: str, fecha_hasta: str):
        return self._call("list_riego_between", fecha_desde, fecha_hasta)

    def list_riego_last_48h(self):
        return self._call("list_riego_last_48h")

    def list_last_riego_by_sector(self):
        return self._call("list_last_riego_by_sector")

    def insert_cosecha(self, fecha, cultivo, kg, sector, bancal):
        return self._call("insert_cosecha", fecha, cultivo, kg, sector, bancal)

    def list_cosechas_by_fecha(self, fecha):
        return self._call("list_cosechas_by_fecha", fecha)

    def list_cosechas_all(self, limit: int | None = None):
        return self._call("list_cosechas_all", limit=limit)

    def list_cosechas_between(self, fecha_desde: str, fecha_hasta: str):
        return self._call("list_cosechas_between", fecha_desde, fecha_hasta)

    def upsert_objetivo(
        self,
        cultivo: str,
        superficie_m2: float,
        cosecha_esperada_kg: float,
    ):
        return self._call("upsert_objetivo", cultivo, superficie_m2, cosecha_esperada_kg)

    def get_objetivo(self, cultivo: str):
        return self._call("get_objetivo", cultivo)

    def set_temporada(self, inicio: str, fin: str):
        return self._call("set_temporada", inicio, fin)

    def get_temporada(self):
        return self._call("get_temporada")

    def list_objetivos(self):
        return self._call("list_objetivos")

    def delete_objetivo(self, cultivo: str):
        return self._call("delete_objetivo", cultivo)

    def clear_objetivos(self):
        return self._call("clear_objetivos")

    def insert_stock(self, fecha, chacra, sector, bancal, cultivo, kg_disponible, peso_unitario, unidades):
        return self._call("insert_stock", fecha, chacra, sector, bancal, cultivo, kg_disponible, peso_unitario, unidades)

    def list_stock_by_fecha(self, fecha, chacra: str | None = None):
        return self._call("list_stock_by_fecha", fecha, chacra=chacra)

    def list_stock_all(self, limit: int | None = None, chacra: str | None = None):
        return self._call("list_stock_all", limit=limit, chacra=chacra)

    def insert_siembra(self, fecha, cultivo, variedad, tipo, generacion, bandejas, tipo_bandeja,
                       observaciones, operador, sector="", bancal=0):
        return self._call("insert_siembra", fecha, cultivo, variedad, tipo, generacion, bandejas,
                          tipo_bandeja, observaciones, operador, sector, bancal)

    def list_siembras_by_fecha(self, fecha):
        return self._call("list_siembras_by_fecha", fecha)

    def list_siembras_all(self, limit: int | None = None):
        return self._call("list_siembras_all", limit=limit)

    def list_siembras_between(self, fecha_desde: str, fecha_hasta: str):
        return self._call("list_siembras_between", fecha_desde, fecha_hasta)

    def insert_tarea(self, tarea: str, fecha: str, importancia: str, n_personas: int):
        return self._call("insert_tarea", tarea, fecha, importancia, n_personas)

    def list_tareas_pendientes(self):
        return self._call("list_tareas_pendientes")

    def marcar_tarea_realizada(self, tarea_id: int):
        return self._call("marcar_tarea_realizada", tarea_id)

    def list_tareas_realizadas_between(self, fecha_desde: str, fecha_hasta: str):
        return self._call("list_tareas_realizadas_between", fecha_desde, fecha_hasta)

    # Edicion/borrado: sin failover al local para evitar divergencia de datos.
    def list_cosechas_con_id(self, fecha: str | None = None, limit: int | None = None):
        return self.primary.list_cosechas_con_id(fecha=fecha, limit=limit)

    def update_cosecha(self, cosecha_id: int, fecha: str, kg: float):
        return self.primary.update_cosecha(cosecha_id, fecha, kg)

    def delete_cosecha(self, cosecha_id: int):
        return self.primary.delete_cosecha(cosecha_id)

    def list_riego_con_id(self, fecha: str | None = None, limit: int | None = None):
        return self.primary.list_riego_con_id(fecha=fecha, limit=limit)

    def update_riego(self, riego_id: int, fecha: str, horas: float, operador: str, logro: str = ""):
        return self.primary.update_riego(riego_id, fecha, horas, operador, logro)

    def delete_riego(self, riego_id: int):
        return self.primary.delete_riego(riego_id)

    def list_siembras_con_id(self, fecha: str | None = None, limit: int | None = None):
        return self.primary.list_siembras_con_id(fecha=fecha, limit=limit)

    def update_siembra(self, siembra_id: int, fecha: str, variedad: str, generacion: int,
                       bandejas: int, tipo_bandeja: int, observaciones: str, operador: str):
        return self.primary.update_siembra(siembra_id, fecha, variedad, generacion,
                                           bandejas, tipo_bandeja, observaciones, operador)

    def delete_siembra(self, siembra_id: int):
        return self.primary.delete_siembra(siembra_id)


def create_repository_from_env(backend: str | None = None, api_base_url: str | None = None) -> DataRepository:
    backend = norm_text(backend if backend is not None else os.getenv("MONAGRIC_DATA_BACKEND", "local")).lower()
    api_base_url = norm_text(
        api_base_url if api_base_url is not None else os.getenv("MONAGRIC_API_BASE_URL", DEFAULT_API_BASE_URL)
    )
    local_repo = LocalDataRepository()

    if backend == "local":
        return local_repo
    if backend == "remote":
        return RemoteApiRepository(api_base_url)
    if backend == "auto":
        return FailoverRepository(RemoteApiRepository(api_base_url), local_repo)
    raise ValueError("MONAGRIC_DATA_BACKEND invalido. Use: local, remote o auto.")


def list_riego_last_48h():
    start_dt = datetime.now() - timedelta(hours=48)
    with get_conn() as conn:
        cur = conn.execute(
            """SELECT fecha, sector, horas_riego, operador
               FROM riego
               WHERE created_at >= ?
               ORDER BY created_at DESC, sector""",
            (start_dt.isoformat(timespec="seconds"),)
        )
        return cur.fetchall()

def list_last_riego_by_sector():
    with get_conn() as conn:
        cur = conn.execute(
            """SELECT sector, MAX(created_at)
               FROM riego
               GROUP BY sector"""
        )
        return {row[0]: row[1] for row in cur.fetchall()}

_IMPORTANCIA_ORDER = {"Alta": 0, "Media": 1, "Baja": 2}

def insert_tarea(tarea: str, fecha: str, importancia: str, n_personas: int):
    created = now_iso()
    with get_conn() as conn:
        conn.execute(
            """INSERT INTO tareas (tarea, fecha, importancia, n_personas, realizada, fecha_realizada, created_at, temporada_id)
               VALUES (?, ?, ?, ?, 0, '', ?, ?)""",
            (tarea, fecha, importancia, n_personas, created, _temporada_activa_id()),
        )
        conn.commit()

def list_tareas_pendientes():
    tid = _temporada_activa_id()
    with get_conn() as conn:
        sql = """SELECT id, tarea, fecha, importancia, n_personas
                 FROM tareas
                 WHERE realizada = 0"""
        params = []
        if tid is not None:
            sql += " AND temporada_id = ?"
            params.append(tid)
        sql += " ORDER BY fecha ASC, created_at ASC"
        rows = conn.execute(sql, tuple(params)).fetchall()
    orden = _IMPORTANCIA_ORDER
    return sorted(rows, key=lambda r: (r[2], orden.get(r[3], 99)))

def marcar_tarea_realizada(tarea_id: int):
    with get_conn() as conn:
        conn.execute(
            "UPDATE tareas SET realizada = 1, fecha_realizada = ? WHERE id = ?",
            (now_iso(), tarea_id),
        )
        conn.commit()

def list_tareas_realizadas_between(fecha_desde: str, fecha_hasta: str):
    with get_conn() as conn:
        return conn.execute(
            """SELECT id, tarea, fecha, importancia, n_personas, fecha_realizada
               FROM tareas
               WHERE realizada = 1 AND fecha >= ? AND fecha <= ?
               ORDER BY fecha_realizada DESC""",
            (fecha_desde, fecha_hasta),
        ).fetchall()


def insert_cosecha(fecha, cultivo, kg, sector, bancal):
    created = now_iso()
    with get_conn() as conn:
        conn.execute(
            """INSERT INTO cosechas
            (fecha, cultivo, kg, sector, bancal, created_at, temporada_id)
            VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (fecha, cultivo, kg, sector, bancal, created, _temporada_activa_id())
        )
        conn.commit()
    append_csv(
        CSV_DIR / "cosechas.csv",
        ["fecha", "cultivo", "kg", "sector", "bancal", "created_at"],
        [fecha, cultivo, kg, sector, bancal, created],
    )

def list_cosechas_by_fecha(fecha):
    with get_conn() as conn:
        cur = conn.execute(
            """SELECT cultivo, kg, sector, bancal
               FROM cosechas
               WHERE fecha = ?
               ORDER BY cultivo""",
            (fecha,)
        )
        return cur.fetchall()

def list_cosechas_all(limit: int | None = None):
    with get_conn() as conn:
        sql = """SELECT fecha, cultivo, kg, sector, bancal
                 FROM cosechas
                 ORDER BY fecha DESC, cultivo"""
        if limit:
            sql += " LIMIT ?"
            return conn.execute(sql, (limit,)).fetchall()
        return conn.execute(sql).fetchall()

def list_cosechas_between(fecha_desde: str, fecha_hasta: str):
    with get_conn() as conn:
        cur = conn.execute(
            """SELECT fecha, cultivo, kg, sector, bancal
               FROM cosechas
               WHERE fecha >= ? AND fecha <= ?
               ORDER BY fecha ASC, cultivo ASC, sector ASC, bancal ASC""",
            (fecha_desde, fecha_hasta),
        )
        return cur.fetchall()

def set_temporada_config(inicio: str, fin: str):
    inicio = validate_fecha(inicio)
    fin = validate_optional_fecha(fin)
    if fin and inicio > fin:
        raise ValueError("La fecha de inicio de temporada no puede ser mayor a la fecha fin.")
    with get_conn() as conn:
        conn.execute(
            """
            INSERT INTO app_config (key, value) VALUES (?, ?)
            ON CONFLICT(key) DO UPDATE SET value = excluded.value
            """,
            ("temporada_inicio", inicio),
        )
        conn.execute(
            """
            INSERT INTO app_config (key, value) VALUES (?, ?)
            ON CONFLICT(key) DO UPDATE SET value = excluded.value
            """,
            ("temporada_fin", fin),
        )
        conn.commit()

def get_temporada_config():
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT key, value FROM app_config WHERE key IN ('temporada_inicio', 'temporada_fin')"
        ).fetchall()
    data = {k: v for k, v in rows}
    return {
        "inicio": norm_text(data.get("temporada_inicio", "")),
        "fin": norm_text(data.get("temporada_fin", "")),
    }

def upsert_objetivo(
    cultivo: str,
    superficie_m2: float,
    cosecha_esperada_kg: float,
):
    cultivo = validate_cultivo(cultivo)
    now = now_iso()
    with get_conn() as conn:
        conn.execute(
            """
            INSERT INTO objetivos (cultivo, superficie_m2, cosecha_esperada_kg, temporada_inicio, temporada_fin, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(cultivo) DO UPDATE SET
                superficie_m2 = excluded.superficie_m2,
                cosecha_esperada_kg = excluded.cosecha_esperada_kg,
                temporada_inicio = excluded.temporada_inicio,
                temporada_fin = excluded.temporada_fin,
                updated_at = excluded.updated_at
            """,
            (cultivo, superficie_m2, cosecha_esperada_kg, "", "", now, now),
        )
        conn.commit()

def get_objetivo(cultivo: str):
    cultivo = validate_cultivo(cultivo)
    with get_conn() as conn:
        cur = conn.execute(
            """SELECT cultivo, superficie_m2, cosecha_esperada_kg, temporada_inicio, temporada_fin
               FROM objetivos
               WHERE cultivo = ?""",
            (cultivo,),
        )
        row = cur.fetchone()
        return row if row else None

def list_objetivos():
    with get_conn() as conn:
        cur = conn.execute(
            """SELECT cultivo, superficie_m2, cosecha_esperada_kg, temporada_inicio, temporada_fin
               FROM objetivos
               ORDER BY cultivo ASC"""
        )
        return cur.fetchall()

def delete_objetivo(cultivo: str):
    cultivo = validate_cultivo(cultivo)
    with get_conn() as conn:
        conn.execute("DELETE FROM objetivos WHERE cultivo = ?", (cultivo,))
        conn.commit()

def clear_objetivos():
    with get_conn() as conn:
        conn.execute("DELETE FROM objetivos")
        conn.commit()

def insert_stock(fecha, chacra, sector, bancal, cultivo, kg_disponible, peso_unitario, unidades):
    created = now_iso()
    chacra = validate_chacra_nombre(chacra)
    with get_conn() as conn:
        conn.execute(
            """INSERT INTO stock
            (fecha, chacra, sector, bancal, cultivo, kg_disponible, peso_unitario, unidades, created_at, temporada_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (fecha, chacra, sector, bancal, cultivo, kg_disponible, peso_unitario, unidades, created, _temporada_activa_id())
        )
        conn.commit()
    append_csv(
        CSV_DIR / "stock.csv",
        ["fecha", "chacra", "sector", "bancal", "cultivo", "kg_disponible", "peso_unitario", "unidades", "created_at"],
        [fecha, chacra, sector, bancal, cultivo, kg_disponible, peso_unitario, unidades, created],
    )

def list_stock_by_fecha(fecha, chacra: str | None = None):
    with get_conn() as conn:
        if chacra:
            cur = conn.execute(
                """SELECT sector, bancal, cultivo, kg_disponible, peso_unitario, unidades
                   FROM stock
                   WHERE fecha = ? AND chacra = ?
                   ORDER BY sector, bancal""",
                (fecha, chacra)
            )
        else:
            cur = conn.execute(
                """SELECT sector, bancal, cultivo, kg_disponible, peso_unitario, unidades
                   FROM stock
                   WHERE fecha = ?
                   ORDER BY sector, bancal""",
                (fecha,)
            )
        return cur.fetchall()

def list_stock_all(limit: int | None = None, chacra: str | None = None):
    tid = _temporada_activa_id()
    with get_conn() as conn:
        sql = """SELECT fecha, sector, bancal, cultivo, kg_disponible, peso_unitario, unidades
                 FROM stock"""
        condiciones = []
        params = []
        if chacra:
            condiciones.append("chacra = ?")
            params.append(chacra)
        if tid is not None:
            condiciones.append("temporada_id = ?")
            params.append(tid)
        if condiciones:
            sql += " WHERE " + " AND ".join(condiciones)
        sql += " ORDER BY fecha DESC, sector, bancal"
        if limit:
            sql += " LIMIT ?"
            params.append(limit)
        return conn.execute(sql, tuple(params)).fetchall()

# ==========================================================
# EDICION / BORRADO DE REGISTROS (solo modo local)
# ==========================================================

def list_cosechas_con_id(fecha: str | None = None, limit: int | None = None):
    tid = _temporada_activa_id()
    with get_conn() as conn:
        if fecha:
            sql = """SELECT id, fecha, cultivo, kg, sector, bancal FROM cosechas
                     WHERE fecha = ? ORDER BY cultivo"""
            return conn.execute(sql, (fecha,)).fetchall()
        params = []
        sql = "SELECT id, fecha, cultivo, kg, sector, bancal FROM cosechas"
        if tid is not None:
            sql += " WHERE temporada_id = ?"
            params.append(tid)
        sql += " ORDER BY fecha DESC, cultivo"
        if limit:
            sql += " LIMIT ?"
            params.append(limit)
        return conn.execute(sql, tuple(params)).fetchall()

def update_cosecha(cosecha_id: int, fecha: str, kg: float):
    with get_conn() as conn:
        conn.execute("UPDATE cosechas SET fecha = ?, kg = ? WHERE id = ?", (fecha, kg, cosecha_id))
        conn.commit()

def delete_cosecha(cosecha_id: int):
    with get_conn() as conn:
        conn.execute("DELETE FROM cosechas WHERE id = ?", (cosecha_id,))
        conn.commit()

def list_riego_con_id(fecha: str | None = None, limit: int | None = None):
    tid = _temporada_activa_id()
    with get_conn() as conn:
        if fecha:
            sql = """SELECT id, fecha, horas_riego, sector, operador, logro_riego FROM riego
                     WHERE fecha = ? ORDER BY created_at"""
            return conn.execute(sql, (fecha,)).fetchall()
        params = []
        sql = "SELECT id, fecha, horas_riego, sector, operador, logro_riego FROM riego"
        if tid is not None:
            sql += " WHERE temporada_id = ?"
            params.append(tid)
        sql += " ORDER BY fecha DESC, created_at DESC"
        if limit:
            sql += " LIMIT ?"
            params.append(limit)
        return conn.execute(sql, tuple(params)).fetchall()

def update_riego(riego_id: int, fecha: str, horas: float, operador: str, logro: str = ""):
    with get_conn() as conn:
        conn.execute(
            "UPDATE riego SET fecha = ?, horas_riego = ?, operador = ?, logro_riego = ? WHERE id = ?",
            (fecha, horas, operador, logro, riego_id),
        )
        conn.commit()

def calidad_riego_ultimos_dias(dias: int = 7):
    """Puntaje ponderado de calidad de riego de los últimos N días.

    Devuelve (puntaje 0..1 o None si no hay riegos valorados, total_riegos,
    cantidad_sin_logro)."""
    desde = (date.today() - timedelta(days=dias)).isoformat()
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT logro_riego FROM riego WHERE fecha >= ?", (desde,)
        ).fetchall()
    total = len(rows)
    valorados = [PESOS_LOGRO_RIEGO[norm_text(r[0])] for r in rows
                 if norm_text(r[0]) in PESOS_LOGRO_RIEGO]
    sin_logro = total - len(valorados)
    puntaje = (sum(valorados) / len(valorados)) if valorados else None
    return puntaje, total, sin_logro

def delete_riego(riego_id: int):
    with get_conn() as conn:
        conn.execute("DELETE FROM riego WHERE id = ?", (riego_id,))
        conn.commit()

_SIEMBRA_COLS = ("id, fecha, cultivo, variedad, tipo, generacion, bandejas, tipo_bandeja, "
                 "observaciones, operador, sector, bancal")

def list_siembras_con_id(fecha: str | None = None, limit: int | None = None):
    tid = _temporada_activa_id()
    with get_conn() as conn:
        if fecha:
            sql = f"""SELECT {_SIEMBRA_COLS} FROM siembras
                      WHERE fecha = ? ORDER BY created_at"""
            return conn.execute(sql, (fecha,)).fetchall()
        params = []
        sql = f"SELECT {_SIEMBRA_COLS} FROM siembras"
        if tid is not None:
            sql += " WHERE temporada_id = ?"
            params.append(tid)
        sql += " ORDER BY fecha DESC, created_at DESC"
        if limit:
            sql += " LIMIT ?"
            params.append(limit)
        return conn.execute(sql, tuple(params)).fetchall()

def update_siembra(siembra_id: int, fecha: str, variedad: str, generacion: int,
                   bandejas: int, tipo_bandeja: int, observaciones: str, operador: str):
    with get_conn() as conn:
        conn.execute(
            """UPDATE siembras SET fecha = ?, variedad = ?, generacion = ?, bandejas = ?,
                                   tipo_bandeja = ?, observaciones = ?, operador = ?
               WHERE id = ?""",
            (fecha, variedad, generacion, bandejas, tipo_bandeja, observaciones, operador, siembra_id),
        )
        conn.commit()

def delete_siembra(siembra_id: int):
    with get_conn() as conn:
        conn.execute("DELETE FROM siembras WHERE id = ?", (siembra_id,))
        conn.commit()

def semillas_por_cultivo(temporada_id: int) -> dict:
    """Semillas sembradas por cultivo en la temporada (bandejas x tipo de bandeja)."""
    with get_conn() as conn:
        rows = conn.execute(
            """SELECT cultivo, SUM(bandejas * tipo_bandeja) FROM siembras
               WHERE temporada_id = ? GROUP BY cultivo""",
            (temporada_id,),
        ).fetchall()
    return {r[0]: int(r[1] or 0) for r in rows}

# ==========================================================
# CONFIG DE USUARIO / TEMPORADAS / PLAN / PERFIL DE CULTIVO
# ==========================================================

def set_config_value(key: str, value: str):
    with get_conn() as conn:
        conn.execute(
            """INSERT INTO app_config (key, value) VALUES (?, ?)
               ON CONFLICT(key) DO UPDATE SET value = excluded.value""",
            (key, value),
        )
        conn.commit()

def get_config_value(key: str, default: str = "") -> str:
    with get_conn() as conn:
        row = conn.execute("SELECT value FROM app_config WHERE key = ?", (key,)).fetchone()
    return norm_text(row[0]) if row else default

def set_perfil_usuario(productor: str, chacra: str, largo_bancal_m: float,
                       ancho_bancal_m: float, pasillo_m: float, n_bancales: int):
    set_config_value("productor_nombre", norm_text(productor))
    set_config_value("chacra_nombre", norm_text(chacra))
    set_config_value("largo_bancal_m", str(largo_bancal_m))
    set_config_value("ancho_bancal_m", str(ancho_bancal_m))
    set_config_value("pasillo_m", str(pasillo_m))
    set_config_value("n_bancales", str(n_bancales))
    # El resto de la app trabaja con la superficie del bancal en m2
    set_config_value("bancal_m2", str(round(largo_bancal_m * ancho_bancal_m, 2)))

def _config_float(key: str, default: float) -> float:
    try:
        return float(get_config_value(key, str(default)))
    except Exception:
        return default

def get_perfil_usuario() -> dict:
    bancal_m2 = _config_float("bancal_m2", DEFAULT_BANCAL_M2)
    ancho = _config_float("ancho_bancal_m", DEFAULT_ANCHO_BANCAL_M)
    largo = _config_float("largo_bancal_m", round(bancal_m2 / ancho, 2) if ancho > 0 else DEFAULT_LARGO_BANCAL_M)
    pasillo = _config_float("pasillo_m", DEFAULT_PASILLO_M)
    try:
        n_bancales = int(float(get_config_value("n_bancales", "0")))
    except Exception:
        n_bancales = 0
    superficie_neta = round(largo * ancho * n_bancales, 1)
    superficie_ocupada = round(largo * (ancho + pasillo) * n_bancales, 1)
    return {
        "productor": get_config_value("productor_nombre"),
        "chacra": get_config_value("chacra_nombre"),
        "bancal_m2": bancal_m2,
        "ancho_bancal_m": ancho,
        "largo_bancal_m": largo,
        "pasillo_m": pasillo,
        "n_bancales": n_bancales,
        "superficie_neta_m2": superficie_neta,
        "superficie_ocupada_m2": superficie_ocupada,
    }

def create_temporada(nombre: str, fecha_inicio: str, fecha_fin: str = "") -> int:
    nombre = norm_text(nombre)
    if not nombre:
        raise ValueError("El nombre de temporada es obligatorio (ej: 2026-27).")
    fecha_inicio = validate_fecha(fecha_inicio)
    fecha_fin = validate_optional_fecha(fecha_fin)
    if fecha_fin and fecha_inicio > fecha_fin:
        raise ValueError("El inicio de temporada no puede ser mayor al fin.")
    with get_conn() as conn:
        conn.execute("UPDATE temporadas SET activa = 0")
        cur = conn.execute(
            "INSERT INTO temporadas (nombre, fecha_inicio, fecha_fin, activa, created_at) VALUES (?, ?, ?, 1, ?)",
            (nombre, fecha_inicio, fecha_fin, now_iso()),
        )
        conn.commit()
        nueva_id = int(cur.lastrowid)
    # Adopta registros huerfanos cuya fecha caiga dentro de la nueva temporada
    try:
        migrar_registros_a_temporadas()
    except Exception:
        pass
    return nueva_id

def get_temporada_activa():
    with get_conn() as conn:
        row = conn.execute(
            "SELECT id, nombre, fecha_inicio, fecha_fin FROM temporadas WHERE activa = 1 ORDER BY id DESC LIMIT 1"
        ).fetchone()
    if not row:
        return None
    return {"id": row[0], "nombre": row[1], "inicio": row[2], "fin": row[3]}

def list_temporadas():
    with get_conn() as conn:
        rows = conn.execute(
            """SELECT id, nombre, fecha_inicio, fecha_fin, activa FROM temporadas
               ORDER BY fecha_inicio DESC, id DESC"""
        ).fetchall()
    return [
        {"id": r[0], "nombre": r[1], "inicio": r[2], "fin": r[3], "activa": bool(r[4])}
        for r in rows
    ]

def update_temporada(temporada_id: int, nombre: str, fecha_inicio: str, fecha_fin: str = ""):
    nombre = norm_text(nombre)
    if not nombre:
        raise ValueError("El nombre de temporada es obligatorio.")
    fecha_inicio = validate_fecha(fecha_inicio)
    fecha_fin = validate_optional_fecha(fecha_fin)
    if fecha_fin and fecha_inicio > fecha_fin:
        raise ValueError("El inicio de temporada no puede ser mayor al fin.")
    with get_conn() as conn:
        conn.execute(
            "UPDATE temporadas SET nombre = ?, fecha_inicio = ?, fecha_fin = ? WHERE id = ?",
            (nombre, fecha_inicio, fecha_fin, temporada_id),
        )
        conn.commit()
    try:
        migrar_registros_a_temporadas()
    except Exception:
        pass

def activar_temporada(temporada_id: int):
    with get_conn() as conn:
        conn.execute("UPDATE temporadas SET activa = 0")
        conn.execute("UPDATE temporadas SET activa = 1 WHERE id = ?", (temporada_id,))
        conn.commit()

def delete_temporada(temporada_id: int):
    """Elimina la temporada, su plan y TODOS los registros enlazados a ella."""
    with get_conn() as conn:
        for tabla in ("cosechas", "riego", "siembras", "tareas", "stock", "horas_trabajo",
                      "trasplantes", "sanidad_aplicaciones", "sanidad_detecciones"):
            conn.execute(f"DELETE FROM {tabla} WHERE temporada_id = ?", (temporada_id,))
        conn.execute("""DELETE FROM sanidad_plan_aplicaciones WHERE tratamiento_id IN
                        (SELECT id FROM sanidad_tratamientos WHERE temporada_id = ?)""",
                     (temporada_id,))
        conn.execute("DELETE FROM sanidad_tratamientos WHERE temporada_id = ?", (temporada_id,))
        conn.execute("DELETE FROM sectores_riego WHERE temporada_id = ?", (temporada_id,))
        conn.execute("DELETE FROM plan_fechas_siembra WHERE temporada_id = ?", (temporada_id,))
        conn.execute("DELETE FROM plan_temporada WHERE temporada_id = ?", (temporada_id,))
        conn.execute("DELETE FROM temporadas WHERE id = ?", (temporada_id,))
        conn.commit()

def save_sectores_riego(temporada_id: int, filas: list):
    """Reemplaza la configuración de sectores de riego de la temporada."""
    with get_conn() as conn:
        conn.execute("DELETE FROM sectores_riego WHERE temporada_id = ?", (temporada_id,))
        for f in filas:
            conn.execute(
                "INSERT INTO sectores_riego (temporada_id, sector, bancales, tipo_riego) VALUES (?, ?, ?, ?)",
                (temporada_id, norm_text(f["sector"]), int(f["bancales"]), norm_text(f["tipo"])),
            )
        conn.commit()

def list_sectores_riego(temporada_id: int):
    with get_conn() as conn:
        return conn.execute(
            """SELECT sector, bancales, tipo_riego FROM sectores_riego
               WHERE temporada_id = ? ORDER BY sector""",
            (temporada_id,),
        ).fetchall()

def sectores_de_temporada_activa() -> list:
    t = get_temporada_activa()
    if not t:
        return []
    return [r[0] for r in list_sectores_riego(t["id"])]

# ---- Integrantes del proyecto

def insert_integrante(nombre: str, direccion: str, telefono: str, rol: str, valor_hora: float) -> int:
    nombre = norm_text(nombre)
    if not nombre:
        raise ValueError("El nombre del integrante es obligatorio.")
    with get_conn() as conn:
        cur = conn.execute(
            """INSERT INTO integrantes (nombre, direccion, telefono, rol, valor_hora, created_at)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (nombre, norm_text(direccion), norm_text(telefono), norm_text(rol), valor_hora, now_iso()),
        )
        conn.commit()
        return int(cur.lastrowid)

def update_integrante(integrante_id: int, nombre: str, direccion: str, telefono: str,
                      rol: str, valor_hora: float):
    nombre = norm_text(nombre)
    if not nombre:
        raise ValueError("El nombre del integrante es obligatorio.")
    with get_conn() as conn:
        conn.execute(
            """UPDATE integrantes SET nombre = ?, direccion = ?, telefono = ?, rol = ?, valor_hora = ?
               WHERE id = ?""",
            (nombre, norm_text(direccion), norm_text(telefono), norm_text(rol), valor_hora, integrante_id),
        )
        conn.commit()

def delete_integrante(integrante_id: int):
    with get_conn() as conn:
        conn.execute("DELETE FROM integrantes WHERE id = ?", (integrante_id,))
        conn.commit()

def list_integrantes():
    with get_conn() as conn:
        rows = conn.execute(
            """SELECT id, nombre, direccion, telefono, rol, valor_hora
               FROM integrantes ORDER BY nombre"""
        ).fetchall()
    return [{"id": r[0], "nombre": r[1], "direccion": r[2], "telefono": r[3],
             "rol": r[4], "valor_hora": float(r[5])} for r in rows]

# ---- Horas de trabajo

def insert_horas_trabajo(fecha: str, integrante: str, horas: float, actividades: list):
    with get_conn() as conn:
        conn.execute(
            """INSERT INTO horas_trabajo (fecha, integrante, horas, actividades, created_at, temporada_id)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (fecha, norm_text(integrante), horas, ", ".join(actividades), now_iso(), _temporada_activa_id()),
        )
        conn.commit()

def list_horas_con_id(limit: int | None = None):
    tid = _temporada_activa_id()
    with get_conn() as conn:
        sql = "SELECT id, fecha, integrante, horas, actividades FROM horas_trabajo"
        params = []
        if tid is not None:
            sql += " WHERE temporada_id = ?"
            params.append(tid)
        sql += " ORDER BY fecha DESC, created_at DESC"
        if limit:
            sql += " LIMIT ?"
            params.append(limit)
        return conn.execute(sql, tuple(params)).fetchall()

def update_horas_trabajo(registro_id: int, fecha: str, horas: float, actividades: list):
    with get_conn() as conn:
        conn.execute(
            "UPDATE horas_trabajo SET fecha = ?, horas = ?, actividades = ? WHERE id = ?",
            (fecha, horas, ", ".join(actividades), registro_id),
        )
        conn.commit()

def delete_horas_trabajo(registro_id: int):
    with get_conn() as conn:
        conn.execute("DELETE FROM horas_trabajo WHERE id = ?", (registro_id,))
        conn.commit()

def resumen_horas_trabajo():
    """Totales de la temporada activa: por integrante (horas, monto $) y por
    actividad (horas, monto $, repartiendo las horas entre las actividades
    marcadas en cada registro)."""
    tarifas = {i["nombre"]: i["valor_hora"] for i in list_integrantes()}
    por_persona = {}
    por_actividad = {}
    for _rid, _fecha, integrante, horas, actividades in list_horas_con_id():
        horas = float(horas)
        tarifa = tarifas.get(integrante, 0.0)
        p = por_persona.setdefault(integrante, {"horas": 0.0, "monto": 0.0})
        p["horas"] += horas
        p["monto"] += horas * tarifa
        lista = [a for a in (norm_text(x) for x in str(actividades).split(",")) if a]
        if lista:
            cuota = horas / len(lista)
            for act in lista:
                a = por_actividad.setdefault(act, {"horas": 0.0, "monto": 0.0})
                a["horas"] += cuota
                a["monto"] += cuota * tarifa
    return por_persona, por_actividad

# ---- Trasplantes

def insert_trasplante(fecha: str, integrante: str, cultivo: str, generacion: int,
                      sector: str, bancal: int):
    with get_conn() as conn:
        conn.execute(
            """INSERT INTO trasplantes (fecha, integrante, cultivo, generacion, sector, bancal,
                                        created_at, temporada_id)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (fecha, norm_text(integrante), cultivo, generacion, norm_text(sector), bancal,
             now_iso(), _temporada_activa_id()),
        )
        conn.commit()

def list_trasplantes_con_id(limit: int | None = None):
    tid = _temporada_activa_id()
    with get_conn() as conn:
        sql = "SELECT id, fecha, integrante, cultivo, generacion, sector, bancal FROM trasplantes"
        params = []
        if tid is not None:
            sql += " WHERE temporada_id = ?"
            params.append(tid)
        sql += " ORDER BY fecha DESC, created_at DESC"
        if limit:
            sql += " LIMIT ?"
            params.append(limit)
        return conn.execute(sql, tuple(params)).fetchall()

def update_trasplante(registro_id: int, fecha: str, generacion: int, sector: str, bancal: int):
    with get_conn() as conn:
        conn.execute(
            "UPDATE trasplantes SET fecha = ?, generacion = ?, sector = ?, bancal = ? WHERE id = ?",
            (fecha, generacion, norm_text(sector), bancal, registro_id),
        )
        conn.commit()

def delete_trasplante(registro_id: int):
    with get_conn() as conn:
        conn.execute("DELETE FROM trasplantes WHERE id = ?", (registro_id,))
        conn.commit()

def ventana_trasplante_cultivo(cultivo: str, temporada: dict, primera_siembra: str | None = None):
    """(inicio, fin) esperados de la campaña de trasplante del cultivo: desde
    que el primer almácigo está listo hasta la última fecha que todavía llega
    a cosecharse dentro de la temporada."""
    perfil = get_perfil_cultivo(cultivo)
    base = primera_siembra or temporada.get("inicio")
    try:
        d0 = date.fromisoformat(str(base))
    except Exception:
        return None, None
    ini = d0 + timedelta(days=max(0, perfil["dias_almacigo"]))
    fin_str = norm_text(temporada.get("fin", ""))
    try:
        fin_temp = date.fromisoformat(fin_str) if fin_str else d0 + timedelta(days=330)
    except Exception:
        fin_temp = d0 + timedelta(days=330)
    fin = fin_temp - timedelta(days=max(0, perfil["dias_trasplante_cosecha"]))
    if fin < ini + timedelta(days=14):
        fin = ini + timedelta(days=14)
    return ini, fin

def plantas_por_bancal_cultivo(cultivo: str) -> int:
    """Plantas que entran en un bancal según el marco del perfil y la chacra."""
    perfil = get_perfil_cultivo(cultivo)
    usuario = get_perfil_usuario()
    return calcular_plantas(usuario["bancal_m2"], perfil["distancia_cm"],
                            perfil["lineas_bancal"], usuario["ancho_bancal_m"])

def trasplantes_por_cultivo(hoy: str | None = None):
    """{cultivo: {"registros": n, "plantas": estimadas, "por_mes": {(a,m): plantas}}}."""
    salida = {}
    for _rid, fecha, _integrante, cultivo, _gen, _sec, _b in list_trasplantes_con_id():
        if hoy and str(fecha) > hoy:
            continue
        d = salida.setdefault(cultivo, {"registros": 0, "plantas": 0, "por_mes": {}})
        plantas = plantas_por_bancal_cultivo(cultivo)
        d["registros"] += 1
        d["plantas"] += plantas
        try:
            fd = date.fromisoformat(str(fecha))
            clave = (fd.year, fd.month)
            d["por_mes"][clave] = d["por_mes"].get(clave, 0) + plantas
        except Exception:
            pass
    return salida

# ---- Stock (un registro = kg de un cultivo en un bancal, en una fecha)

def insert_stock_registro(fecha: str, sector: str, bancal: int, cultivo: str, kg: float):
    with get_conn() as conn:
        conn.execute(
            """INSERT INTO stock (fecha, chacra, sector, bancal, cultivo,
                                  kg_disponible, peso_unitario, unidades,
                                  created_at, temporada_id)
               VALUES (?, ?, ?, ?, ?, ?, 0, 0, ?, ?)""",
            (fecha, norm_text(get_perfil_usuario().get("chacra", "")), norm_text(sector),
             int(bancal), norm_text(cultivo), float(kg), now_iso(), _temporada_activa_id()),
        )
        conn.commit()


def list_stock_registros(limit: int | None = None):
    """(id, fecha, sector, bancal, cultivo, kg) de la temporada activa, más nuevos primero."""
    tid = _temporada_activa_id()
    with get_conn() as conn:
        sql = "SELECT id, fecha, sector, bancal, cultivo, kg_disponible FROM stock"
        params = []
        if tid is not None:
            sql += " WHERE temporada_id = ?"
            params.append(tid)
        sql += " ORDER BY fecha DESC, created_at DESC"
        if limit:
            sql += " LIMIT ?"
            params.append(limit)
        return conn.execute(sql, tuple(params)).fetchall()


def update_stock_registro(registro_id: int, fecha: str, sector: str, bancal: int,
                          cultivo: str, kg: float):
    with get_conn() as conn:
        conn.execute(
            """UPDATE stock SET fecha = ?, sector = ?, bancal = ?, cultivo = ?,
                                kg_disponible = ? WHERE id = ?""",
            (fecha, norm_text(sector), int(bancal), norm_text(cultivo), float(kg), registro_id),
        )
        conn.commit()


def delete_stock_registro(registro_id: int):
    with get_conn() as conn:
        conn.execute("DELETE FROM stock WHERE id = ?", (registro_id,))
        conn.commit()


def stock_entre_fechas(desde: str, hasta: str):
    """(fecha, sector, bancal, cultivo, kg) del rango, para exportar."""
    tid = _temporada_activa_id()
    with get_conn() as conn:
        sql = ("SELECT fecha, sector, bancal, cultivo, kg_disponible FROM stock "
               "WHERE fecha >= ? AND fecha <= ?")
        params = [desde, hasta]
        if tid is not None:
            sql += " AND temporada_id = ?"
            params.append(tid)
        sql += " ORDER BY fecha, cultivo, sector, bancal"
        return conn.execute(sql, tuple(params)).fetchall()


def stock_totales_por_cultivo(desde: str, hasta: str):
    """{cultivo: kg} sumando todos los bancales del rango."""
    totales = {}
    for _f, _s, _b, cultivo, kg in stock_entre_fechas(desde, hasta):
        totales[cultivo] = round(totales.get(cultivo, 0.0) + float(kg), 2)
    return totales


def resumen_stock_fecha(fecha: str):
    """{cultivo: kg} de una fecha, sumando el mismo cultivo en varios bancales."""
    return stock_totales_por_cultivo(fecha, fecha)


# ---- Exportar a planilla (.xlsx) y compartir

MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def exportar_xlsx(destino: Path, hoja: str, encabezados: list, filas: list) -> Path:
    """Escribe una planilla .xlsx (abre en LibreOffice Calc, Excel y Google Sheets)."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = (norm_text(hoja) or "Datos")[:31]
    ws.append(list(encabezados))
    for fila in filas:
        ws.append(list(fila))
    # negrita en el encabezado y anchos cómodos
    for celda in ws[1]:
        celda.font = celda.font.copy(bold=True)
    for i, titulo in enumerate(encabezados, start=1):
        largo = max([len(str(titulo))] + [len(str(f[i - 1])) for f in filas] or [0])
        ws.column_dimensions[get_column_letter(i)].width = min(40, max(12, largo + 4))
    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destino)
    return destino


def guardar_en_descargas(archivo: Path, mime: str = MIME_XLSX):
    """Copia el archivo a la carpeta Descargas del celular usando MediaStore.

    Android 10+ bloquea escribir en Descargas a mano (scoped storage), pero sí
    deja insertarlo por MediaStore: queda visible en el explorador de archivos y
    devuelve un content:// que se puede compartir. Fuera de Android no hace nada.
    Devuelve la URI (o None)."""
    if platform != "android":
        return None
    try:
        from jnius import autoclass
        PythonActivity = autoclass("org.kivy.android.PythonActivity")
        ContentValues = autoclass("android.content.ContentValues")
        Downloads = autoclass("android.provider.MediaStore$Downloads")
        resolver = PythonActivity.mActivity.getContentResolver()

        valores = ContentValues()
        valores.put("_display_name", archivo.name)
        valores.put("mime_type", mime)
        valores.put("relative_path", "Download/MonAgric")
        uri = resolver.insert(Downloads.EXTERNAL_CONTENT_URI, valores)
        if uri is None:
            return None
        salida = resolver.openOutputStream(uri)
        salida.write(bytearray(archivo.read_bytes()))
        salida.flush()
        salida.close()
        return uri
    except Exception as e:
        log_exception("No se pudo copiar a Descargas (MediaStore)", e)
        return None


def compartir_archivo(archivo: Path, mime: str = MIME_XLSX, asunto: str = "MonAgric"):
    """Abre el menú de compartir de Android (WhatsApp, mail, Drive...) con el archivo."""
    if platform != "android":
        return False
    try:
        from jnius import autoclass, cast
        PythonActivity = autoclass("org.kivy.android.PythonActivity")
        Intent = autoclass("android.content.Intent")
        String = autoclass("java.lang.String")

        uri = guardar_en_descargas(archivo, mime)
        if uri is None:
            return False
        intent = Intent(Intent.ACTION_SEND)
        intent.setType(mime)
        intent.putExtra(Intent.EXTRA_STREAM, cast("android.os.Parcelable", uri))
        intent.putExtra(Intent.EXTRA_SUBJECT, cast("java.lang.CharSequence", String(asunto)))
        intent.addFlags(Intent.FLAG_GRANT_READ_URI_PERMISSION)
        elegir = Intent.createChooser(intent, cast("java.lang.CharSequence", String("Compartir con...")))
        PythonActivity.mActivity.startActivity(elegir)
        return True
    except Exception as e:
        log_exception("No se pudo compartir el archivo", e)
        return False


# ---- Sanidad (aplicaciones, detecciones, tratamientos)

def insert_aplicacion_sanidad(fecha, producto, dosis, cultivo, sector, bancal,
                              tipo="rutina", tratamiento_id=None, diagnostico=""):
    """Agrega una fila a la planilla única de aplicaciones (rutina o tratamiento)."""
    with get_conn() as conn:
        conn.execute(
            """INSERT INTO sanidad_aplicaciones
               (fecha, producto, dosis, cultivo, sector, bancal, tipo, tratamiento_id,
                diagnostico, created_at, temporada_id)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (fecha, norm_text(producto), norm_text(dosis), norm_text(cultivo),
             norm_text(sector), int(bancal or 0), norm_text(tipo), tratamiento_id,
             norm_text(diagnostico), now_iso(), _temporada_activa_id()),
        )
        conn.commit()


def list_aplicaciones_sanidad(limit=None, solo_hoy=False):
    """Filas (id, fecha, producto, dosis, cultivo, sector, bancal, tipo, diagnostico)."""
    tid = _temporada_activa_id()
    with get_conn() as conn:
        sql = ("SELECT id, fecha, producto, dosis, cultivo, sector, bancal, tipo, diagnostico "
               "FROM sanidad_aplicaciones")
        cond, params = [], []
        if tid is not None:
            cond.append("temporada_id = ?")
            params.append(tid)
        if solo_hoy:
            cond.append("fecha = ?")
            params.append(date.today().isoformat())
        if cond:
            sql += " WHERE " + " AND ".join(cond)
        sql += " ORDER BY fecha DESC, created_at DESC"
        if limit:
            sql += " LIMIT ?"
            params.append(limit)
        return conn.execute(sql, tuple(params)).fetchall()


def delete_aplicacion_sanidad(registro_id):
    with get_conn() as conn:
        conn.execute("DELETE FROM sanidad_aplicaciones WHERE id = ?", (registro_id,))
        conn.commit()


def resumen_aplicaciones_sanidad():
    """Totales para el resumen: por producto, por tipo y total."""
    filas = list_aplicaciones_sanidad()
    por_producto, por_tipo = {}, {"rutina": 0, "tratamiento": 0}
    for _id, _f, producto, _d, _c, _s, _b, tipo, _diag in filas:
        por_producto[producto] = por_producto.get(producto, 0) + 1
        por_tipo[tipo] = por_tipo.get(tipo, 0) + 1
    return {"total": len(filas), "por_producto": por_producto, "por_tipo": por_tipo}


def insert_deteccion(fecha, cultivo, sector, bancal, sintoma, signo, diagnostico, severidad):
    with get_conn() as conn:
        cur = conn.execute(
            """INSERT INTO sanidad_detecciones
               (fecha, cultivo, sector, bancal, sintoma, signo, diagnostico, severidad,
                created_at, temporada_id)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (fecha, norm_text(cultivo), norm_text(sector), int(bancal or 0),
             norm_text(sintoma), norm_text(signo), norm_text(diagnostico),
             int(severidad or 0), now_iso(), _temporada_activa_id()),
        )
        conn.commit()
        return cur.lastrowid


def list_detecciones(limit=None):
    tid = _temporada_activa_id()
    with get_conn() as conn:
        sql = ("SELECT id, fecha, cultivo, sector, bancal, sintoma, signo, diagnostico, severidad "
               "FROM sanidad_detecciones")
        params = []
        if tid is not None:
            sql += " WHERE temporada_id = ?"
            params.append(tid)
        sql += " ORDER BY fecha DESC, created_at DESC"
        if limit:
            sql += " LIMIT ?"
            params.append(limit)
        return conn.execute(sql, tuple(params)).fetchall()


def crear_tratamiento(deteccion_id, cultivo, sector, bancal, producto, dosis,
                      n_aplicaciones, frecuencia_dias, fecha_inicio, diagnostico=""):
    """Crea el tratamiento y genera su calendario de aplicaciones. Devuelve el id."""
    n = max(1, int(n_aplicaciones))
    freq = max(1, int(frecuencia_dias))
    try:
        d0 = date.fromisoformat(str(fecha_inicio))
    except Exception:
        d0 = date.today()
    with get_conn() as conn:
        cur = conn.execute(
            """INSERT INTO sanidad_tratamientos
               (deteccion_id, cultivo, sector, bancal, producto, dosis, n_aplicaciones,
                frecuencia_dias, fecha_inicio, diagnostico, activo, created_at, temporada_id)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?)""",
            (deteccion_id, norm_text(cultivo), norm_text(sector), int(bancal or 0),
             norm_text(producto), norm_text(dosis), n, freq, d0.isoformat(),
             norm_text(diagnostico), now_iso(), _temporada_activa_id()),
        )
        trat_id = cur.lastrowid
        for i in range(n):
            fprog = (d0 + timedelta(days=i * freq)).isoformat()
            conn.execute(
                """INSERT INTO sanidad_plan_aplicaciones
                   (tratamiento_id, fecha_programada, realizada, created_at)
                   VALUES (?, ?, 0, ?)""",
                (trat_id, fprog, now_iso()),
            )
        conn.commit()
        return trat_id


def list_tratamientos_activos():
    """Tratamientos con aplicaciones pendientes en la temporada activa."""
    tid = _temporada_activa_id()
    with get_conn() as conn:
        sql = ("SELECT id, cultivo, sector, bancal, producto, dosis, n_aplicaciones, "
               "frecuencia_dias, fecha_inicio, diagnostico FROM sanidad_tratamientos "
               "WHERE activo = 1")
        params = []
        if tid is not None:
            sql += " AND temporada_id = ?"
            params.append(tid)
        sql += " ORDER BY fecha_inicio ASC, created_at ASC"
        return conn.execute(sql, tuple(params)).fetchall()


def list_plan_aplicaciones(tratamiento_id):
    """(id, fecha_programada, realizada) del calendario de un tratamiento."""
    with get_conn() as conn:
        return conn.execute(
            """SELECT id, fecha_programada, realizada FROM sanidad_plan_aplicaciones
               WHERE tratamiento_id = ? ORDER BY fecha_programada ASC, id ASC""",
            (tratamiento_id,),
        ).fetchall()


def marcar_aplicacion_plan(plan_id):
    """Marca una aplicación del calendario como realizada, la vuelca a la planilla
    general y desactiva el tratamiento cuando ya no quedan pendientes."""
    with get_conn() as conn:
        row = conn.execute(
            "SELECT tratamiento_id, fecha_programada, realizada FROM sanidad_plan_aplicaciones WHERE id = ?",
            (plan_id,),
        ).fetchone()
        if not row:
            return
        trat_id, fprog, realizada = row
        if realizada:
            return
        conn.execute("UPDATE sanidad_plan_aplicaciones SET realizada = 1 WHERE id = ?", (plan_id,))
        t = conn.execute(
            """SELECT cultivo, sector, bancal, producto, dosis, diagnostico, temporada_id
               FROM sanidad_tratamientos WHERE id = ?""",
            (trat_id,),
        ).fetchone()
        if t:
            cultivo, sector, bancal, producto, dosis, diagnostico, temp_id = t
            fecha = norm_text(str(fprog)) or date.today().isoformat()
            conn.execute(
                """INSERT INTO sanidad_aplicaciones
                   (fecha, producto, dosis, cultivo, sector, bancal, tipo, tratamiento_id,
                    diagnostico, created_at, temporada_id)
                   VALUES (?, ?, ?, ?, ?, ?, 'tratamiento', ?, ?, ?, ?)""",
                (fecha, producto, dosis, cultivo, sector, bancal, trat_id,
                 diagnostico, now_iso(), temp_id),
            )
        pendientes = conn.execute(
            "SELECT COUNT(*) FROM sanidad_plan_aplicaciones WHERE tratamiento_id = ? AND realizada = 0",
            (trat_id,),
        ).fetchone()[0]
        if pendientes == 0:
            conn.execute("UPDATE sanidad_tratamientos SET activo = 0 WHERE id = ?", (trat_id,))
        conn.commit()


def delete_tratamiento(tratamiento_id):
    """Cancela un tratamiento y su calendario (no toca la planilla ya realizada)."""
    with get_conn() as conn:
        conn.execute("DELETE FROM sanidad_plan_aplicaciones WHERE tratamiento_id = ?", (tratamiento_id,))
        conn.execute("DELETE FROM sanidad_tratamientos WHERE id = ?", (tratamiento_id,))
        conn.commit()


# ---- PIN de administrador (protege Configuración de Temporada)

def set_admin_pin(pin: str):
    pin = norm_text(pin)
    valor = hashlib.sha256(pin.encode("utf-8")).hexdigest() if pin else ""
    set_config_value("admin_pin_hash", valor)

def admin_pin_definido() -> bool:
    return bool(get_config_value("admin_pin_hash", ""))

def verificar_admin_pin(pin: str) -> bool:
    guardado = get_config_value("admin_pin_hash", "")
    if not guardado:
        return True
    return hashlib.sha256(norm_text(pin).encode("utf-8")).hexdigest() == guardado

def migrar_registros_a_temporadas():
    """Asigna temporada a registros viejos (sin enlace) segun su fecha."""
    with get_conn() as conn:
        temporadas = conn.execute(
            "SELECT id, fecha_inicio, fecha_fin FROM temporadas ORDER BY fecha_inicio"
        ).fetchall()
        for tid, inicio, fin in temporadas:
            for tabla in ("cosechas", "riego", "siembras", "tareas", "stock", "horas_trabajo",
                          "trasplantes", "sanidad_aplicaciones", "sanidad_detecciones"):
                if fin:
                    conn.execute(
                        f"UPDATE {tabla} SET temporada_id = ? WHERE temporada_id IS NULL AND fecha >= ? AND fecha <= ?",
                        (tid, inicio, fin),
                    )
                else:
                    conn.execute(
                        f"UPDATE {tabla} SET temporada_id = ? WHERE temporada_id IS NULL AND fecha >= ?",
                        (tid, inicio),
                    )
        conn.commit()

def list_cosechas_de_temporada(temporada_id: int):
    with get_conn() as conn:
        return conn.execute(
            """SELECT fecha, cultivo, kg FROM cosechas
               WHERE temporada_id = ? ORDER BY fecha ASC""",
            (temporada_id,),
        ).fetchall()

def list_siembras_de_temporada(temporada_id: int):
    with get_conn() as conn:
        return conn.execute(
            """SELECT fecha, cultivo FROM siembras
               WHERE temporada_id = ? ORDER BY fecha ASC""",
            (temporada_id,),
        ).fetchall()

def upsert_plan_item(temporada_id: int, cultivo: str, superficie_m2: float, esperado_kg: float,
                     tipo_siembra: str = "", distancia_cm: float = 0, lineas: int = 0, plantas: int = 0):
    cultivo = validate_cultivo(cultivo)
    if is_sin_asignar(cultivo):
        raise ValueError("Elegí un cultivo real para el plan.")
    now = now_iso()
    with get_conn() as conn:
        conn.execute(
            """INSERT INTO plan_temporada
               (temporada_id, cultivo, superficie_m2, cosecha_esperada_kg,
                tipo_siembra, distancia_cm, lineas, plantas, created_at, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
               ON CONFLICT(temporada_id, cultivo) DO UPDATE SET
                   superficie_m2 = excluded.superficie_m2,
                   cosecha_esperada_kg = excluded.cosecha_esperada_kg,
                   tipo_siembra = excluded.tipo_siembra,
                   distancia_cm = excluded.distancia_cm,
                   lineas = excluded.lineas,
                   plantas = excluded.plantas,
                   updated_at = excluded.updated_at""",
            (temporada_id, cultivo, superficie_m2, esperado_kg,
             norm_text(tipo_siembra), distancia_cm, lineas, plantas, now, now),
        )
        conn.commit()

def delete_plan_item(temporada_id: int, cultivo: str):
    with get_conn() as conn:
        conn.execute(
            "DELETE FROM plan_temporada WHERE temporada_id = ? AND cultivo = ?",
            (temporada_id, cultivo),
        )
        conn.execute(
            "DELETE FROM plan_fechas_siembra WHERE temporada_id = ? AND cultivo = ?",
            (temporada_id, cultivo),
        )
        conn.commit()

def save_fechas_siembra_plan(temporada_id: int, cultivo: str, fechas: list):
    """Reemplaza las fechas de siembra planificadas del cultivo (una por generación)."""
    limpias = sorted({validate_fecha(f) for f in fechas if norm_text(str(f))})
    with get_conn() as conn:
        conn.execute(
            "DELETE FROM plan_fechas_siembra WHERE temporada_id = ? AND cultivo = ?",
            (temporada_id, cultivo),
        )
        for f in limpias:
            conn.execute(
                "INSERT INTO plan_fechas_siembra (temporada_id, cultivo, fecha) VALUES (?, ?, ?)",
                (temporada_id, cultivo, f),
            )
        conn.commit()

def get_fechas_siembra_plan(temporada_id: int, cultivo: str) -> list:
    with get_conn() as conn:
        return [r[0] for r in conn.execute(
            """SELECT fecha FROM plan_fechas_siembra
               WHERE temporada_id = ? AND cultivo = ? ORDER BY fecha""",
            (temporada_id, cultivo),
        ).fetchall()]

def list_plan(temporada_id: int):
    with get_conn() as conn:
        return conn.execute(
            """SELECT cultivo, superficie_m2, cosecha_esperada_kg,
                      tipo_siembra, distancia_cm, lineas, plantas
               FROM plan_temporada
               WHERE temporada_id = ? ORDER BY cultivo""",
            (temporada_id,),
        ).fetchall()

def get_perfil_cultivo(cultivo: str):
    with get_conn() as conn:
        row = conn.execute(
            """SELECT dias_a_cosecha, ventana_cosecha_dias, rinde_ref_kg_m2,
                      tipo_siembra, dias_almacigo, dias_trasplante_cosecha,
                      distancia_cm, lineas_bancal, tipo_cosecha
               FROM cultivo_perfil WHERE cultivo = ?""",
            (cultivo,),
        ).fetchone()
    if not row:
        return {
            "dias_a_cosecha": 60, "ventana": 30, "rinde_ref": 0.0,
            "tipo_siembra": TIPO_SIEMBRA_DIRECTA, "dias_almacigo": 0,
            "dias_trasplante_cosecha": 60, "distancia_cm": 25.0, "lineas_bancal": 3,
            "tipo_cosecha": COSECHA_ESCALONADA,
        }
    tipo_cosecha = norm_text(row[8]) if len(row) > 8 and row[8] else COSECHA_ESCALONADA
    return {
        "dias_a_cosecha": int(row[0]),
        "ventana": int(row[1]),
        "rinde_ref": float(row[2]),
        "tipo_siembra": norm_text(row[3]) or TIPO_SIEMBRA_DIRECTA,
        "dias_almacigo": int(row[4]),
        "dias_trasplante_cosecha": int(row[5]),
        "distancia_cm": float(row[6]),
        "lineas_bancal": int(row[7]),
        "tipo_cosecha": tipo_cosecha if tipo_cosecha in TIPOS_COSECHA else COSECHA_ESCALONADA,
    }

def upsert_perfil_cultivo(cultivo: str, tipo_siembra: str, dias_almacigo: int,
                          dias_trasplante_cosecha: int, ventana: int, rinde_ref: float,
                          distancia_cm: float, lineas_bancal: int,
                          tipo_cosecha: str = COSECHA_ESCALONADA):
    cultivo = validate_cultivo(cultivo)
    tipo_siembra = norm_text(tipo_siembra) or TIPO_SIEMBRA_DIRECTA
    tipo_cosecha = norm_text(tipo_cosecha) if norm_text(tipo_cosecha) in TIPOS_COSECHA else COSECHA_ESCALONADA
    dias_totales = dias_a_cosecha_total(tipo_siembra, dias_almacigo, dias_trasplante_cosecha)
    with get_conn() as conn:
        conn.execute(
            """INSERT INTO cultivo_perfil
               (cultivo, dias_a_cosecha, ventana_cosecha_dias, rinde_ref_kg_m2,
                tipo_siembra, dias_almacigo, dias_trasplante_cosecha, distancia_cm, lineas_bancal, tipo_cosecha)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
               ON CONFLICT(cultivo) DO UPDATE SET
                   dias_a_cosecha = excluded.dias_a_cosecha,
                   ventana_cosecha_dias = excluded.ventana_cosecha_dias,
                   rinde_ref_kg_m2 = excluded.rinde_ref_kg_m2,
                   tipo_siembra = excluded.tipo_siembra,
                   dias_almacigo = excluded.dias_almacigo,
                   dias_trasplante_cosecha = excluded.dias_trasplante_cosecha,
                   distancia_cm = excluded.distancia_cm,
                   lineas_bancal = excluded.lineas_bancal,
                   tipo_cosecha = excluded.tipo_cosecha""",
            (cultivo, dias_totales, ventana, rinde_ref, tipo_siembra,
             dias_almacigo, dias_trasplante_cosecha, distancia_cm, lineas_bancal, tipo_cosecha),
        )
        conn.commit()

def migrar_objetivos_a_temporada():
    """Si hay objetivos viejos y ninguna temporada, crea la temporada activa y copia el plan."""
    if get_temporada_activa():
        return
    objetivos = list_objetivos()
    if not objetivos:
        # Sin objetivos no hay nada que preservar: el setup inicial hace el resto
        return
    legacy = get_temporada_config()
    inicio = legacy.get("inicio") or date.today().isoformat()
    fin = legacy.get("fin", "")
    temporada_id = create_temporada(sugerir_nombre_temporada(), inicio, fin)
    for row in objetivos:
        try:
            upsert_plan_item(temporada_id, row[0], float(row[1]), float(row[2]))
        except Exception:
            continue

# ==========================================================
# LOGICA DE RELOJES (prorrateo por fecha, kg/m2, semaforo)
# ==========================================================


def _fraccion_suave(x: float) -> float:
    """Acumulada en S (sigmoide): la cosecha arranca lento, acelera en el medio
    de la ventana y desacelera al final. Es la forma de una curva de crecimiento
    (área foliar / llenado de fruto). Ventana corta = subida abrupta (concentrada);
    larga = subida gradual (escalonada/continua)."""
    if x <= 0.0:
        return 0.0
    if x >= 1.0:
        return 1.0
    return x - math.sin(2.0 * math.pi * x) / (2.0 * math.pi)


# Fracción del rendimiento ya "presente" en la planta al iniciar la cosecha,
# por tipo. Alta en cosecha entera (lechuga: la planta ya está a tamaño); baja
# en escalonada (tomate: primeros frutos). Da la altura de la rampa de desarrollo.
UMBRAL_DESARROLLO = {
    COSECHA_CONCENTRADA: 0.90,
    COSECHA_ESCALONADA: 0.15,
    COSECHA_CONTINUA: 0.25,
}


def _fin_de_mes(anio: int, mes: int) -> date:
    return date(anio + (mes == 12), (mes % 12) + 1, 1)

def _span_temporada(meses: list):
    """(fecha_inicio, total_dias) del eje de meses."""
    ini = date(meses[0][0], meses[0][1], 1)
    fin = _fin_de_mes(*meses[-1])
    return ini, max(1, (fin - ini).days)

def ventanas_cosecha_cultivo(cultivo: str, temporada: dict, fechas_siembra: list | None = None,
                             inicio_fallback: str | None = None,
                             fecha_primera_cosecha: str | None = None) -> list:
    """Ventanas (ini, fin) de cosecha, UNA POR GENERACIÓN si hay fechas de
    siembra planificadas; recortadas al fin de temporada. Sin fechas usa una
    única ventana desde el inicio de ciclo (comportamiento anterior)."""
    perfil = get_perfil_cultivo(cultivo)
    ventana = max(1, perfil["ventana"])
    bases = []
    for f in (fechas_siembra or []):
        try:
            bases.append(date.fromisoformat(str(f)))
        except Exception:
            continue
    bases.sort()
    if not bases:
        base = inicio_fallback or temporada.get("inicio")
        try:
            bases = [date.fromisoformat(str(base))]
        except Exception:
            return []
    fin_temp = None
    fin_str = norm_text(temporada.get("fin", ""))
    if fin_str:
        try:
            fin_temp = date.fromisoformat(fin_str)
        except Exception:
            pass
    ventanas = []
    for i, b in enumerate(bases):
        ini = b + timedelta(days=max(0, perfil["dias_a_cosecha"]))
        if i == 0 and fecha_primera_cosecha:
            # La primera cosecha real calibra la primera ventana
            try:
                pc = date.fromisoformat(str(fecha_primera_cosecha))
                if pc < ini:
                    ini = pc
            except Exception:
                pass
        fin = ini + timedelta(days=ventana)
        if fin_temp:
            if ini >= fin_temp:
                continue  # esta generación no llega a cosecharse en la temporada
            if fin > fin_temp:
                fin = fin_temp
        if fin <= ini:
            fin = ini + timedelta(days=1)
        ventanas.append((b, ini, fin))
    return ventanas


def resumen_cultivo_a_fecha(cultivo: str, fecha: str, temporada: dict, superficie_m2: float,
                            esperado_kg: float, kg_real: float, fecha_siembra: str | None = None,
                            fecha_primera_cosecha: str | None = None,
                            fechas_plan: list | None = None) -> dict:
    """Calcula el estado del reloj de un cultivo: % a la fecha, kg/m2 y semaforo.

    Con fechas de siembra planificadas (fechas_plan) el esperado se reparte en
    una ventana de cosecha por generación (continuidades e intermitencias);
    sin fechas, una única ventana calibrada por la primera cosecha real y
    recortada al fin de temporada.
    """
    perfil = get_perfil_cultivo(cultivo)
    tipo_cosecha = perfil.get("tipo_cosecha", COSECHA_ESCALONADA)
    inicio_ciclo = (fechas_plan[0] if fechas_plan else None) or fecha_siembra \
        or temporada.get("inicio") or fecha

    ventanas = ventanas_cosecha_cultivo(cultivo, temporada, fechas_plan,
                                        fecha_siembra or temporada.get("inicio"),
                                        fecha_primera_cosecha)
    inicio_ventana = ventanas[0][1] if ventanas else None
    fin_ventana = ventanas[-1][2] if ventanas else None
    try:
        siembra_dt = date.fromisoformat(str(inicio_ciclo))
    except Exception:
        siembra_dt = None
    inicio_teorico = siembra_dt + timedelta(days=max(0, perfil["dias_a_cosecha"])) if siembra_dt else None

    fraccion = 0.0
    if ventanas:
        try:
            f = date.fromisoformat(fecha)
            acumulado = 0.0
            for _siembra_g, ini, fin in ventanas:
                if f >= ini:
                    dur = max(1, (fin - ini).days)
                    acumulado += _fraccion_suave((f - ini).days / dur)
            fraccion = acumulado / len(ventanas)
        except Exception:
            pass

    esperado_a_fecha = esperado_kg * fraccion
    kg_m2_real = round(kg_real / superficie_m2, 2) if superficie_m2 > 0 else 0.0
    kg_m2_obj = round(esperado_kg / superficie_m2, 2) if superficie_m2 > 0 else 0.0

    if fraccion <= 0:
        try:
            dias_para_inicio = (inicio_teorico - date.fromisoformat(fecha)).days if inicio_teorico else 0
        except Exception:
            dias_para_inicio = 0
        if kg_real > 0:
            estado = "verde"
            pct = round((kg_real / esperado_kg) * 100.0, 1) if esperado_kg > 0 else 0.0
        else:
            estado = "pre"
            pct = 0.0
    else:
        dias_para_inicio = 0
        pct = round((kg_real / esperado_a_fecha) * 100.0, 1) if esperado_a_fecha > 0 else 0.0
        pct = min(pct, 999.0)
        ratio = (kg_real / esperado_a_fecha) if esperado_a_fecha > 0 else 0.0
        if ratio >= SEMAFORO_VERDE:
            estado = "verde"
        elif ratio >= SEMAFORO_AMBAR:
            estado = "ambar"
        else:
            estado = "rojo"

    return {
        "cultivo": cultivo,
        "estado": estado,                      # pre | verde | ambar | rojo
        "pct_a_fecha": pct,                    # real vs esperado a la fecha
        "esperado_a_fecha": round(esperado_a_fecha, 1),
        "esperado_total": round(esperado_kg, 1),
        "kg_real": round(kg_real, 1),
        "kg_m2_real": kg_m2_real,
        "kg_m2_obj": kg_m2_obj,
        "fraccion": round(fraccion, 3),
        "dias_para_inicio": max(0, dias_para_inicio),
        "tipo_cosecha": tipo_cosecha,
        "umbral_frac": UMBRAL_DESARROLLO.get(tipo_cosecha, 0.2),
        "siembra": str(inicio_ciclo),
        "ventana_ini": inicio_ventana.isoformat() if inicio_ventana else "",
        "ventana_fin": fin_ventana.isoformat() if fin_ventana else "",
        "ventanas": [(i.isoformat(), f.isoformat()) for _s, i, f in ventanas],
        "gens": [(s.isoformat(), i.isoformat(), f.isoformat()) for s, i, f in ventanas],
    }

SEMAFORO_COLORES = {
    "verde": (0.22, 0.60, 0.28, 1),
    "ambar": (0.83, 0.53, 0.04, 1),
    "rojo": (0.75, 0.18, 0.18, 1),
    "pre": (0.55, 0.58, 0.56, 1),
}


# ==========================================================
# UI KV
# (usamos MDDropDownItem para seleccionar con menú)
# ==========================================================

KV = """
<SectionCard>:
    orientation: "vertical"
    padding: "14dp"
    spacing: "10dp"
    size_hint_x: 1
    size_hint_y: None
    adaptive_height: True
    md_bg_color: app.card_bg
    radius: [16, 16, 16, 16]
    elevation: 1
    shadow_softness: 0
    line_color: app.card_outline
    line_width: 1.2

<MobilePrimaryButton@MDRaisedButton>:
    size_hint_y: None
    height: "48dp"
    font_size: "14sp"
    md_bg_color: app.accent_color
    text_color: (1, 1, 1, 1)

<MobileActionButton@MDRaisedButton>:
    size_hint_y: None
    height: "44dp"
    font_size: "13sp"
    md_bg_color: app.header_color
    text_color: (1, 1, 1, 1)

<HomeNavCard@MDCard>:
    orientation: "vertical"
    padding: ["12dp", "16dp", "12dp", "12dp"]
    spacing: "6dp"
    size_hint_y: None
    height: "120dp"
    md_bg_color: app.card_bg
    radius: [16, 16, 16, 16]
    elevation: 0
    line_color: app.card_outline
    line_width: 1

<HomeWideCard@MDCard>:
    orientation: "horizontal"
    padding: ["20dp", "12dp"]
    spacing: "16dp"
    size_hint_y: None
    height: "72dp"
    md_bg_color: app.card_bg
    radius: [16, 16, 16, 16]
    elevation: 0
    line_color: app.card_outline
    line_width: 1

<HomeDisabledCard@MDCard>:
    orientation: "horizontal"
    padding: ["20dp", "12dp"]
    spacing: "16dp"
    size_hint_y: None
    height: "56dp"
    md_bg_color: (0.91, 0.91, 0.91, 1)
    radius: [12, 12, 12, 12]
    elevation: 0

<DisabledActionButton@MDRaisedButton>:
    size_hint_y: None
    height: "44dp"
    font_size: "10sp"
    md_bg_color: (0.75, 0.75, 0.75, 1)
    text_color: (1, 1, 1, 1)
    disabled: True

# Solo Home se instancia al arrancar (es la pantalla del uso diario). El resto
# —incluido Setup, que solo se ve la primera vez— se crea bajo demanda al
# navegar y se precarga en segundo plano tras el primer frame (ver MonAgricSM).
# Instanciar las ~20 pantallas de golpe agregaba ~15 s al arranque.
MonAgricSM:
    HomeScreen:

<HomeScreen>:
    name: "home"
    canvas.before:
        Color:
            rgba: app.surface_bg
        Rectangle:
            pos: self.pos
            size: self.size

    MDBoxLayout:
        orientation: "vertical"
        spacing: "0dp"

        # --- Hero header ---
        MDBoxLayout:
            orientation: "vertical"
            size_hint_y: None
            height: "148dp"
            md_bg_color: app.header_color
            padding: ["20dp", "14dp", "20dp", "10dp"]
            spacing: "4dp"

            Image:
                source: "img/logo1.png"
                size_hint_y: None
                height: "76dp"
                allow_stretch: True
                keep_ratio: True

            MDLabel:
                text: app.temporada_header_text
                size_hint_y: None
                height: "18dp"
                font_style: "Caption"
                halign: "center"
                theme_text_color: "Custom"
                text_color: (0.95, 0.97, 0.94, 0.95)
                shorten: True
                shorten_from: "right"
                max_lines: 1

            MDLabel:
                text: app.backend_status_text
                size_hint_y: None
                height: "18dp"
                font_style: "Caption"
                halign: "right"
                theme_text_color: "Custom"
                text_color: (0.75, 0.2, 0.2, 1) if (app.backend_mode_name == "remote" and not app.backend_online) else (0.9, 0.95, 0.9, 0.75)
                shorten: True
                shorten_from: "right"
                max_lines: 1

        # --- Grilla de módulos ---
        ScrollView:
            MDBoxLayout:
                orientation: "vertical"
                padding: ["12dp", "16dp", "12dp", "8dp"]
                spacing: "12dp"
                size_hint_y: None
                height: self.minimum_height

                MDGridLayout:
                    cols: 2
                    adaptive_height: True
                    spacing: "12dp"

                    # Siembras
                    HomeNavCard:
                        on_release: root.ir_siembras()
                        MDIcon:
                            icon: "sprout"
                            halign: "center"
                            theme_text_color: "Custom"
                            text_color: app.header_color
                            font_size: "38sp"
                        MDLabel:
                            text: "SIEMBRAS"
                            halign: "center"
                            font_style: "Button"
                            theme_text_color: "Custom"
                            text_color: app.header_color

                    # Trasplantes
                    HomeNavCard:
                        on_release: root.ir_trasplantes()
                        MDIcon:
                            icon: "shovel"
                            halign: "center"
                            theme_text_color: "Custom"
                            text_color: app.header_color
                            font_size: "38sp"
                        MDLabel:
                            text: "TRASPLANTES"
                            halign: "center"
                            font_style: "Button"
                            theme_text_color: "Custom"
                            text_color: app.header_color

                    # Cosechas
                    HomeNavCard:
                        on_release: root.ir_cosechas()
                        MDIcon:
                            icon: "basket"
                            halign: "center"
                            theme_text_color: "Custom"
                            text_color: app.header_color
                            font_size: "38sp"
                        MDLabel:
                            text: "COSECHAS"
                            halign: "center"
                            font_style: "Button"
                            theme_text_color: "Custom"
                            text_color: app.header_color

                    # Horas de trabajo
                    HomeNavCard:
                        on_release: root.ir_horas()
                        MDIcon:
                            icon: "account-clock"
                            halign: "center"
                            theme_text_color: "Custom"
                            text_color: app.accent_color
                            font_size: "38sp"
                        MDLabel:
                            text: "HORAS DE TRABAJO"
                            halign: "center"
                            font_style: "Button"
                            theme_text_color: "Custom"
                            text_color: app.header_color

                    # Riego
                    HomeNavCard:
                        on_release: root.ir_riego()
                        MDIcon:
                            icon: "water"
                            halign: "center"
                            theme_text_color: "Custom"
                            text_color: (0.23, 0.51, 0.82, 1)
                            font_size: "38sp"
                        MDLabel:
                            text: "RIEGO"
                            halign: "center"
                            font_style: "Button"
                            theme_text_color: "Custom"
                            text_color: app.header_color

                    # Sanidad
                    HomeNavCard:
                        on_release: root.ir_sanidad()
                        MDIcon:
                            icon: "leaf"
                            halign: "center"
                            theme_text_color: "Custom"
                            text_color: app.header_color
                            font_size: "38sp"
                        MDLabel:
                            text: "SANIDAD"
                            halign: "center"
                            font_style: "Button"
                            theme_text_color: "Custom"
                            text_color: app.header_color

                    # Stock
                    HomeNavCard:
                        on_release: root.ir_stock()
                        MDIcon:
                            icon: "inbox-multiple"
                            halign: "center"
                            theme_text_color: "Custom"
                            text_color: app.header_color
                            font_size: "38sp"
                        MDLabel:
                            text: "STOCK"
                            halign: "center"
                            font_style: "Button"
                            theme_text_color: "Custom"
                            text_color: app.header_color

                    # Tareas
                    HomeNavCard:
                        on_release: root.ir_tareas()
                        MDIcon:
                            icon: "clipboard-check-outline"
                            halign: "center"
                            theme_text_color: "Custom"
                            text_color: app.accent_color
                            font_size: "38sp"
                        MDLabel:
                            text: "TAREAS"
                            halign: "center"
                            font_style: "Button"
                            theme_text_color: "Custom"
                            text_color: app.header_color

                # --- Tablero de monitoreo: ¿cómo vamos? (se arma desde Python) ---
                MDBoxLayout:
                    id: tablero_box
                    orientation: "vertical"
                    spacing: "8dp"
                    size_hint_y: None
                    height: self.minimum_height

                # Configuración de temporada — al final, protegida con PIN
                HomeWideCard:
                    on_release: root.ir_objetivo()
                    MDIcon:
                        icon: "target"
                        halign: "center"
                        valign: "middle"
                        theme_text_color: "Custom"
                        text_color: app.accent_color
                        font_size: "32sp"
                        size_hint_x: None
                        width: "40dp"
                    MDLabel:
                        text: "CONFIGURACIÓN DE TEMPORADA"
                        font_style: "Button"
                        valign: "middle"
                        theme_text_color: "Custom"
                        text_color: app.header_color
                    Widget:
                    MDIcon:
                        icon: "lock-outline"
                        halign: "right"
                        valign: "middle"
                        theme_text_color: "Custom"
                        text_color: app.header_color
                        font_size: "18sp"
                        size_hint_x: None
                        width: "24dp"

                # Salir + créditos
                MDBoxLayout:
                    orientation: "vertical"
                    size_hint_y: None
                    height: "80dp"
                    spacing: "6dp"
                    padding: ["0dp", "8dp", "0dp", "0dp"]

                    MobilePrimaryButton:
                        text: "SALIR"
                        md_bg_color: (0.29, 0.40, 0.30, 1)
                        on_release: root.confirm_exit()

                    MDLabel:
                        size_hint_y: None
                        height: "16dp"
                        text: "M[a]rtox - MonAgric 1.0 '26"
                        font_style: "Caption"
                        halign: "center"
                        theme_text_color: "Custom"
                        text_color: (0.45, 0.52, 0.46, 1)

<SetupScreen>:
    name: "setup"
    canvas.before:
        Color:
            rgba: app.surface_bg
        Rectangle:
            pos: self.pos
            size: self.size

    MDBoxLayout:
        orientation: "vertical"
        padding: "0dp"
        spacing: "0dp"

        MDBoxLayout:
            orientation: "vertical"
            size_hint_y: None
            height: "64dp"
            md_bg_color: app.header_color
            padding: ["4dp", "0dp", "16dp", "0dp"]

            MDBoxLayout:
                spacing: "4dp"
                size_hint_y: 1

                MDIconButton:
                    icon: "arrow-left"
                    theme_text_color: "Custom"
                    text_color: (1, 1, 1, 1)
                    on_release: root.ir_home()

                MDLabel:
                    text: "Nueva Temporada"
                    font_style: "H6"
                    theme_text_color: "Custom"
                    text_color: (1, 1, 1, 1)

                Widget:

                MDLabel:
                    text: "MonAgric"
                    font_style: "Caption"
                    halign: "right"
                    theme_text_color: "Custom"
                    text_color: (0.9, 0.95, 0.9, 0.75)

        ScrollView:
            MDBoxLayout:
                orientation: "vertical"
                spacing: "12dp"
                padding: ["14dp", "12dp", "14dp", "14dp"]
                size_hint_y: None
                height: self.minimum_height

                MDLabel:
                    text: "Definí el plan de la temporada: qué cultivos vas a producir, cuánta superficie y cuánto esperás cosechar. Con eso se activan los relojes del tablero."
                    font_style: "Caption"
                    theme_text_color: "Secondary"
                    size_hint_y: None
                    height: self.texture_size[1] + dp(6)

                SectionCard:
                    MDLabel:
                        text: "Tu chacra"
                        font_style: "Subtitle1"
                        size_hint_y: None
                        height: "24dp"
                    MDTextField:
                        id: productor
                        hint_text: "Tu nombre"
                    MDTextField:
                        id: chacra
                        hint_text: "Nombre de la chacra"
                    MDBoxLayout:
                        adaptive_height: True
                        spacing: "8dp"
                        MDTextField:
                            id: largo_bancal
                            hint_text: "Largo bancal (m)"
                            input_filter: "float"
                            on_text: root.actualizar_superficie()
                        MDTextField:
                            id: ancho_bancal
                            hint_text: "Ancho bancal (m)"
                            input_filter: "float"
                            on_text: root.actualizar_superficie()
                    MDBoxLayout:
                        adaptive_height: True
                        spacing: "8dp"
                        MDTextField:
                            id: pasillo
                            hint_text: "Pasillo (m)"
                            input_filter: "float"
                            on_text: root.actualizar_superficie()
                        MDTextField:
                            id: n_bancales
                            hint_text: "N° de bancales"
                            input_filter: "int"
                            on_text: root.actualizar_superficie()
                    MDLabel:
                        id: superficie_info
                        text: ""
                        font_style: "Caption"
                        theme_text_color: "Custom"
                        text_color: app.header_color
                        size_hint_y: None
                        height: "32dp"

                SectionCard:
                    MDLabel:
                        text: "Temporada"
                        font_style: "Subtitle1"
                        size_hint_y: None
                        height: "24dp"
                    MDTextField:
                        id: temporada_nombre
                        hint_text: "Nombre (ej: 2026-27)"
                    MDBoxLayout:
                        adaptive_height: True
                        spacing: "4dp"
                        MDTextField:
                            id: fecha
                            hint_text: "Inicio (AAAA-MM-DD)"
                        MDIconButton:
                            icon: "calendar"
                            size_hint: None, None
                            size: "48dp", "48dp"
                            pos_hint: {"center_y": 0.5}
                            on_release: root.abrir_calendario()
                    MDBoxLayout:
                        adaptive_height: True
                        spacing: "4dp"
                        MDTextField:
                            id: fecha_fin
                            hint_text: "Fin (opcional)"
                        MDIconButton:
                            icon: "calendar"
                            size_hint: None, None
                            size: "48dp", "48dp"
                            pos_hint: {"center_y": 0.5}
                            on_release: root.abrir_calendario_fin()

                SectionCard:
                    MDLabel:
                        text: "Sectores de riego"
                        font_style: "Subtitle1"
                        size_hint_y: None
                        height: "24dp"
                    MDLabel:
                        text: "Bancales por sector y tipo de riego (aspersión, goteo, surco, superficie)."
                        font_style: "Caption"
                        theme_text_color: "Secondary"
                        size_hint_y: None
                        height: "18dp"
                    MDBoxLayout:
                        id: sectores_box
                        orientation: "vertical"
                        spacing: "4dp"
                        size_hint_y: None
                        height: self.minimum_height
                    MobileActionButton:
                        text: "+ AGREGAR SECTOR"
                        on_release: root.agregar_sector_riego()

                SectionCard:
                    MDLabel:
                        text: "Plan de cultivos"
                        font_style: "Subtitle1"
                        size_hint_y: None
                        height: "24dp"
                    MobilePrimaryButton:
                        text: "+ AGREGAR CULTIVO"
                        on_release: root.abrir_agregar_cultivo()
                    MDList:
                        id: plan_lista
                        size_hint_y: None
                        height: self.minimum_height

                MobilePrimaryButton:
                    text: "COMENZAR TEMPORADA"
                    md_bg_color: app.accent_color
                    on_release: root.comenzar_temporada()

<TareasScreen>:
    name: "tareas"
    canvas.before:
        Color:
            rgba: app.surface_bg
        Rectangle:
            pos: self.pos
            size: self.size

    MDBoxLayout:
        orientation: "vertical"
        padding: "0dp"
        spacing: "0dp"

        MDBoxLayout:
            orientation: "vertical"
            size_hint_y: None
            height: "64dp"
            md_bg_color: app.header_color
            padding: ["4dp", "0dp", "16dp", "0dp"]

            MDBoxLayout:
                spacing: "4dp"
                size_hint_y: 1

                MDIconButton:
                    icon: "arrow-left"
                    theme_text_color: "Custom"
                    text_color: (1, 1, 1, 1)
                    on_release: root.ir_home()

                MDLabel:
                    text: "TAREAS"
                    font_style: "H6"
                    theme_text_color: "Custom"
                    text_color: (1, 1, 1, 1)

                Widget:

                MDLabel:
                    text: "MonAgric"
                    halign: "right"
                    theme_text_color: "Custom"
                    text_color: (0.9, 0.95, 0.9, 0.75)

        ScrollView:
            MDBoxLayout:
                orientation: "vertical"
                padding: ["14dp", "12dp", "14dp", "14dp"]
                spacing: "12dp"
                size_hint_y: None
                height: self.minimum_height

                # Botón nueva tarea
                MobilePrimaryButton:
                    text: "  + NUEVA TAREA"
                    size_hint_x: 1
                    on_release: root.abrir_dialogo_nueva_tarea()

                # Lista de tareas pendientes
                SectionCard:
                    MDBoxLayout:
                        orientation: "vertical"
                        spacing: "4dp"
                        size_hint_y: None
                        adaptive_height: True

                        MDLabel:
                            text: "Pendientes"
                            font_style: "Subtitle2"
                            theme_text_color: "Custom"
                            text_color: app.header_color
                            size_hint_y: None
                            height: "24dp"

                        MDList:
                            id: lista_tareas
                            size_hint_y: None
                            height: self.minimum_height

                # Botones de acción
                MDGridLayout:
                    cols: 3
                    adaptive_height: True
                    spacing: "8dp"
                    size_hint_y: None
                    height: "52dp"

                    MobileActionButton:
                        text: "RESUMEN"
                        font_size: "11sp"
                        on_release: root.abrir_resumen()

                    MobileActionButton:
                        text: "DESCARGAR"
                        font_size: "11sp"
                        on_release: root.abrir_descarga()

                    DisabledActionButton:
                        text: "SUGERENCIA IA"

<RiegoScreen>:
    name: "riego"
    canvas.before:
        Color:
            rgba: app.surface_bg
        Rectangle:
            pos: self.pos
            size: self.size

    MDBoxLayout:
        orientation: "vertical"
        padding: "0dp"
        spacing: "0dp"

        MDBoxLayout:
            orientation: "vertical"
            size_hint_y: None
            height: "64dp"
            md_bg_color: app.header_color
            padding: ["4dp", "0dp", "16dp", "0dp"]

            MDBoxLayout:
                spacing: "4dp"
                size_hint_y: 1

                MDIconButton:
                    icon: "arrow-left"
                    size_hint_x: None
                    size_hint_y: None
                    size: "48dp", "48dp"
                    theme_text_color: "Custom"
                    text_color: (1, 1, 1, 1)
                    on_release: root.ir_home()

                MDLabel:
                    text: "Riego"
                    font_style: "H6"
                    valign: "middle"
                    theme_text_color: "Custom"
                    text_color: (1, 1, 1, 1)

                Widget:

                MDLabel:
                    text: "MonAgric"
                    font_style: "Caption"
                    halign: "right"
                    valign: "middle"
                    theme_text_color: "Custom"
                    text_color: (0.9, 0.95, 0.9, 0.75)

        ScrollView:
            MDBoxLayout:
                orientation: "vertical"
                spacing: "8dp"
                padding: ["14dp", "12dp", "14dp", "14dp"]
                size_hint_y: None
                height: self.minimum_height

                MDLabel:
                    text: app.backend_status_text
                    size_hint_y: None
                    height: "18dp"
                    font_style: "Caption"
                    halign: "right"
                    theme_text_color: "Custom"
                    text_color: (0.75, 0.2, 0.2, 1) if (app.backend_mode_name == "remote" and not app.backend_online) else (0.4, 0.4, 0.4, 1)
                    shorten: True
                    shorten_from: "right"
                    max_lines: 1

                MDBoxLayout:
                    adaptive_height: True
                    spacing: "4dp"

                    MDTextField:
                        id: fecha
                        hint_text: "AAAA-MM-DD"
                        text: root.hoy()

                    MDIconButton:
                        icon: "calendar"
                        size_hint_x: None
                        size_hint_y: None
                        size: "48dp", "48dp"
                        pos_hint: {"center_y": 0.5}
                        on_release: root.abrir_calendario()

                MDBoxLayout:
                    adaptive_height: True
                    spacing: "12dp"

                    MDLabel:
                        text: "Sector"
                        size_hint_x: None
                        width: "70dp"
                        valign: "middle"

                    MDDropDownItem:
                        id: sector_item
                        text: "A"
                        on_release: root.open_sector_menu()

                MDBoxLayout:
                    size_hint_y: None
                    height: "48dp"
                    spacing: "6dp"

                    MDLabel:
                        text: "Horas de riego"
                        size_hint_x: None
                        width: "110dp"
                        valign: "middle"

                    MDIconButton:
                        icon: "minus-circle-outline"
                        size_hint_x: None
                        size_hint_y: None
                        size: "40dp", "40dp"
                        pos_hint: {"center_y": 0.5}
                        theme_text_color: "Custom"
                        text_color: app.header_color
                        on_release: root.dec_horas()

                    MDLabel:
                        id: horas
                        text: "1"
                        halign: "center"
                        valign: "middle"
                        size_hint_x: None
                        width: "44dp"
                        font_style: "H6"
                        theme_text_color: "Custom"
                        text_color: app.header_color

                    MDIconButton:
                        icon: "plus-circle-outline"
                        size_hint_x: None
                        size_hint_y: None
                        size: "40dp", "40dp"
                        pos_hint: {"center_y": 0.5}
                        theme_text_color: "Custom"
                        text_color: app.header_color
                        on_release: root.inc_horas()

                MDBoxLayout:
                    adaptive_height: True
                    spacing: "12dp"

                    MDLabel:
                        text: "Logro"
                        size_hint_x: None
                        width: "70dp"
                        valign: "middle"

                    MDDropDownItem:
                        id: logro_item
                        text: "Riego aceptable"
                        on_release: root.open_logro_menu()

                MDBoxLayout:
                    adaptive_height: True
                    spacing: "12dp"

                    MDLabel:
                        text: "Filtro sector"
                        size_hint_x: None
                        width: "90dp"
                        valign: "middle"

                    MDDropDownItem:
                        id: filtro_item
                        text: "Todos"
                        on_release: root.open_filtro_menu()

                MDBoxLayout:
                    adaptive_height: True
                    spacing: "12dp"
                    MDLabel:
                        text: "Nombre"
                        size_hint_x: None
                        width: "70dp"
                        valign: "middle"
                    MDDropDownItem:
                        id: operador_item
                        text: "Seleccionar"
                        on_release: root.open_operador_menu()

                MDGridLayout:
                    cols: 2
                    adaptive_height: True
                    spacing: "8dp"
                    row_force_default: True
                    row_default_height: "44dp"

                    MobilePrimaryButton:
                        text: "REGISTRAR"
                        on_release: root.registrar()

                    MobileActionButton:
                        text: "RESUMEN"
                        on_release: root.resumen_ultimas_48h()

                    MobileActionButton:
                        text: "DESCARGAR"
                        on_release: root.abrir_descarga()

                    MobileActionButton:
                        id: ver_btn
                        text: "VER REGISTROS"
                        on_release: root.toggle_ver_todos()

                SectionCard:
                    MDLabel:
                        text: "Prioridad de riego por sector"
                        font_style: "Subtitle1"
                    ScrollView:
                        size_hint_y: None
                        height: "160dp"
                        MDList:
                            id: prioridad_lista

                MDLabel:
                    id: modo_label
                    text: "Mostrando: Hoy"
                    size_hint_y: None
                    height: "20dp"
                    font_style: "Caption"

                SectionCard:
                    MDLabel:
                        text: "Riegos realizados"
                        font_style: "Subtitle1"
                    ScrollView:
                        size_hint_y: None
                        height: "180dp"
                        MDList:
                            id: lista

<CosechasScreen>:
    name: "cosechas"
    canvas.before:
        Color:
            rgba: app.surface_bg
        Rectangle:
            pos: self.pos
            size: self.size

    MDBoxLayout:
        orientation: "vertical"
        padding: "0dp"
        spacing: "0dp"

        MDBoxLayout:
            orientation: "vertical"
            size_hint_y: None
            height: "64dp"
            md_bg_color: app.header_color
            padding: ["4dp", "0dp", "16dp", "0dp"]

            MDBoxLayout:
                spacing: "4dp"
                size_hint_y: 1

                MDIconButton:
                    icon: "arrow-left"
                    size_hint_x: None
                    size_hint_y: None
                    size: "48dp", "48dp"
                    theme_text_color: "Custom"
                    text_color: (1, 1, 1, 1)
                    on_release: root.ir_home()

                MDLabel:
                    text: "Cosechas"
                    font_style: "H6"
                    valign: "middle"
                    theme_text_color: "Custom"
                    text_color: (1, 1, 1, 1)

                Widget:

                MDLabel:
                    text: "MonAgric"
                    font_style: "Caption"
                    halign: "right"
                    valign: "middle"
                    theme_text_color: "Custom"
                    text_color: (0.9, 0.95, 0.9, 0.75)

        ScrollView:
            MDBoxLayout:
                orientation: "vertical"
                spacing: "8dp"
                padding: ["14dp", "12dp", "14dp", "14dp"]
                size_hint_y: None
                height: self.minimum_height

                MDLabel:
                    text: app.backend_status_text
                    size_hint_y: None
                    height: "18dp"
                    font_style: "Caption"
                    halign: "right"
                    theme_text_color: "Custom"
                    text_color: (0.75, 0.2, 0.2, 1) if (app.backend_mode_name == "remote" and not app.backend_online) else (0.4, 0.4, 0.4, 1)
                    shorten: True
                    shorten_from: "right"
                    max_lines: 1

                MDBoxLayout:
                    adaptive_height: True
                    spacing: "4dp"

                    MDTextField:
                        id: fecha
                        hint_text: "AAAA-MM-DD"
                        text: root.hoy()

                    MDIconButton:
                        icon: "calendar"
                        size_hint_x: None
                        size_hint_y: None
                        size: "48dp", "48dp"
                        pos_hint: {"center_y": 0.5}
                        on_release: root.abrir_calendario()

                MDBoxLayout:
                    adaptive_height: True
                    spacing: "12dp"

                    MDLabel:
                        text: "Cultivo"
                        size_hint_x: None
                        width: "70dp"
                        valign: "middle"

                    MDDropDownItem:
                        id: cultivo_item
                        text: "Lechuga"
                        on_release: root.open_cultivo_menu()

                MDTextField:
                    id: kg
                    hint_text: "Cantidad (kg)"
                    text: ""
                    input_filter: "float"
                    input_type: "number"
                    size_hint_y: None
                    height: "48dp"

                MobilePrimaryButton:
                    text: "GUARDAR"
                    on_release: root.guardar()

                SectionCard:
                    MDLabel:
                        id: logro_label
                        text: "Logro: sin objetivo cargado para este cultivo"
                        font_style: "Caption"
                        theme_text_color: "Secondary"

                    ProgressBar:
                        id: logro_bar
                        max: 100
                        value: 0

                MDGridLayout:
                    cols: 2
                    adaptive_height: True
                    spacing: "8dp"
                    row_force_default: True
                    row_default_height: "44dp"

                    MobileActionButton:
                        text: "RESUMEN"
                        on_release: root.resumen_del_dia()

                    MobileActionButton:
                        text: "DESCARGAR"
                        on_release: root.abrir_descarga()

                    MobileActionButton:
                        id: ver_btn
                        text: "VER REGISTROS"
                        on_release: root.toggle_ver_todos()

                MDLabel:
                    id: modo_label
                    text: "Mostrando: Hoy"
                    size_hint_y: None
                    height: "20dp"
                    font_style: "Caption"

                SectionCard:
                    MDLabel:
                        text: "Cosechas"
                        font_style: "Subtitle1"
                    ScrollView:
                        size_hint_y: None
                        height: "180dp"
                        MDList:
                            id: lista

<ObjetivoScreen>:
    name: "objetivo"
    canvas.before:
        Color:
            rgba: app.surface_bg
        Rectangle:
            pos: self.pos
            size: self.size

    MDBoxLayout:
        orientation: "vertical"
        padding: "0dp"
        spacing: "0dp"

        MDBoxLayout:
            orientation: "vertical"
            size_hint_y: None
            height: "64dp"
            md_bg_color: app.header_color
            padding: ["4dp", "0dp", "16dp", "0dp"]

            MDBoxLayout:
                spacing: "4dp"
                size_hint_y: 1

                MDIconButton:
                    icon: "arrow-left"
                    size_hint_x: None
                    size_hint_y: None
                    size: "48dp", "48dp"
                    theme_text_color: "Custom"
                    text_color: (1, 1, 1, 1)
                    on_release: root.ir_home()

                MDLabel:
                    text: "Configuración de Temporada"
                    font_style: "H6"
                    valign: "middle"
                    theme_text_color: "Custom"
                    text_color: (1, 1, 1, 1)

                Widget:

                MDLabel:
                    text: "MonAgric"
                    font_style: "Caption"
                    halign: "right"
                    valign: "middle"
                    theme_text_color: "Custom"
                    text_color: (0.9, 0.95, 0.9, 0.75)

        ScrollView:
            MDBoxLayout:
                orientation: "vertical"
                spacing: "8dp"
                padding: ["14dp", "12dp", "14dp", "14dp"]
                size_hint_y: None
                height: self.minimum_height

                MDLabel:
                    text: app.backend_status_text
                    size_hint_y: None
                    height: "18dp"
                    font_style: "Caption"
                    halign: "right"
                    theme_text_color: "Custom"
                    text_color: (0.75, 0.2, 0.2, 1) if (app.backend_mode_name == "remote" and not app.backend_online) else (0.4, 0.4, 0.4, 1)
                    shorten: True
                    shorten_from: "right"
                    max_lines: 1

                SectionCard:
                    MDLabel:
                        text: "Temporada"
                        font_style: "Subtitle1"
                        size_hint_y: None
                        height: "24dp"

                    MDBoxLayout:
                        adaptive_height: True
                        spacing: "12dp"

                        MDLabel:
                            text: "Ver:"
                            size_hint_x: None
                            width: "36dp"
                            valign: "middle"

                        MDDropDownItem:
                            id: temporada_item
                            text: "Seleccionar"
                            on_release: root.open_temporada_menu()

                    MDLabel:
                        id: temporada_label
                        text: "Temporada: sin definir"
                        font_style: "Caption"
                        theme_text_color: "Secondary"

                    MDGridLayout:
                        cols: 2
                        adaptive_height: True
                        spacing: "8dp"
                        row_force_default: True
                        row_default_height: "44dp"

                        MobileActionButton:
                            text: "EDITAR"
                            on_release: root.abrir_editar_temporada()

                        MobileActionButton:
                            text: "ACTIVAR"
                            on_release: root.activar_temporada_vista()

                        MobilePrimaryButton:
                            text: "NUEVA"
                            on_release: root.ir_setup()

                        MobileActionButton:
                            text: "ELIMINAR"
                            md_bg_color: (0.75, 0.18, 0.18, 1)
                            on_release: root.eliminar_temporada_vista()

                MobilePrimaryButton:
                    text: "+ AGREGAR CULTIVO"
                    on_release: root.abrir_agregar_cultivo()

                MDGridLayout:
                    cols: 2
                    adaptive_height: True
                    spacing: "8dp"
                    row_force_default: True
                    row_default_height: "44dp"

                    MobileActionButton:
                        text: "PERFILES CULTIVO"
                        font_size: "11sp"
                        on_release: root.abrir_perfil_cultivo()

                    MobileActionButton:
                        text: "SECTORES RIEGO"
                        font_size: "11sp"
                        on_release: root.abrir_sectores_riego()

                    MobileActionButton:
                        text: "INTEGRANTES"
                        font_size: "11sp"
                        on_release: root.abrir_integrantes()

                    MobileActionButton:
                        text: "PIN ADMIN"
                        font_size: "11sp"
                        on_release: root.configurar_pin_admin()

                    MobileActionButton:
                        text: "REINICIAR TODO"
                        on_release: root.confirmar_reiniciar_objetivos()

                SectionCard:
                    MDLabel:
                        text: "Plan de cultivos (tocá un cultivo para editarlo)"
                        font_style: "Subtitle1"
                    ScrollView:
                        size_hint_y: None
                        height: "300dp"
                        MDList:
                            id: objetivos_lista

<SiembrasScreen>:
    name: "siembras"
    canvas.before:
        Color:
            rgba: app.surface_bg
        Rectangle:
            pos: self.pos
            size: self.size

    MDBoxLayout:
        orientation: "vertical"
        padding: "0dp"
        spacing: "0dp"

        MDBoxLayout:
            orientation: "vertical"
            size_hint_y: None
            height: "64dp"
            md_bg_color: app.header_color
            padding: ["4dp", "0dp", "16dp", "0dp"]

            MDBoxLayout:
                spacing: "4dp"
                size_hint_y: 1

                MDIconButton:
                    icon: "arrow-left"
                    size_hint_x: None
                    size_hint_y: None
                    size: "48dp", "48dp"
                    theme_text_color: "Custom"
                    text_color: (1, 1, 1, 1)
                    on_release: root.ir_home()

                MDLabel:
                    text: "Siembras y Trasplantes"
                    font_style: "H6"
                    valign: "middle"
                    theme_text_color: "Custom"
                    text_color: (1, 1, 1, 1)

                Widget:

                MDLabel:
                    text: "MonAgric"
                    font_style: "Caption"
                    halign: "right"
                    valign: "middle"
                    theme_text_color: "Custom"
                    text_color: (0.9, 0.95, 0.9, 0.75)

        ScrollView:
            MDBoxLayout:
                orientation: "vertical"
                spacing: "8dp"
                padding: ["14dp", "12dp", "14dp", "14dp"]
                size_hint_y: None
                height: self.minimum_height

                MDLabel:
                    text: app.backend_status_text
                    size_hint_y: None
                    height: "18dp"
                    font_style: "Caption"
                    halign: "right"
                    theme_text_color: "Custom"
                    text_color: (0.75, 0.2, 0.2, 1) if (app.backend_mode_name == "remote" and not app.backend_online) else (0.4, 0.4, 0.4, 1)
                    shorten: True
                    shorten_from: "right"
                    max_lines: 1

                MDBoxLayout:
                    adaptive_height: True
                    spacing: "4dp"

                    MDTextField:
                        id: fecha
                        hint_text: "AAAA-MM-DD"
                        text: root.hoy()

                    MDIconButton:
                        icon: "calendar"
                        size_hint_x: None
                        size_hint_y: None
                        size: "48dp", "48dp"
                        pos_hint: {"center_y": 0.5}
                        on_release: root.abrir_calendario()

                MDBoxLayout:
                    adaptive_height: True
                    spacing: "12dp"

                    MDLabel:
                        text: "Cultivo"
                        size_hint_x: None
                        width: "70dp"
                        valign: "middle"

                    MDDropDownItem:
                        id: cultivo_item
                        text: "Seleccionar"
                        on_release: root.open_cultivo_menu()

                MDLabel:
                    id: plan_info
                    text: ""
                    font_style: "Caption"
                    theme_text_color: "Custom"
                    text_color: app.header_color
                    size_hint_y: None
                    height: "18dp"
                    shorten: True
                    shorten_from: "right"
                    max_lines: 1

                MDTextField:
                    id: variedad
                    hint_text: "Variedad (libre)"
                    text: ""
                    size_hint_y: None
                    height: "48dp"

                MDBoxLayout:
                    adaptive_height: True
                    spacing: "12dp"

                    MDLabel:
                        text: "Tipo"
                        size_hint_x: None
                        width: "70dp"
                        valign: "middle"

                    MDDropDownItem:
                        id: tipo_item
                        text: "Siembra directa"
                        on_release: root.open_tipo_menu()

                MDBoxLayout:
                    adaptive_height: True
                    spacing: "12dp"

                    MDLabel:
                        id: ubicacion_label
                        text: "Sector"
                        size_hint_x: None
                        width: "70dp"
                        valign: "middle"

                    MDDropDownItem:
                        id: sector_item
                        text: "-"
                        on_release: root.open_sector_menu()

                    MDLabel:
                        text: "Bancal"
                        size_hint_x: None
                        width: "60dp"
                        valign: "middle"

                    MDDropDownItem:
                        id: bancal_item
                        text: "-"
                        on_release: root.open_bancal_menu()

                MDBoxLayout:
                    size_hint_y: None
                    height: "48dp"
                    spacing: "6dp"

                    MDLabel:
                        text: "Generación"
                        size_hint_x: None
                        width: "90dp"
                        valign: "middle"

                    MDIconButton:
                        icon: "minus-circle-outline"
                        size_hint_x: None
                        size_hint_y: None
                        size: "40dp", "40dp"
                        pos_hint: {"center_y": 0.5}
                        theme_text_color: "Custom"
                        text_color: app.header_color
                        on_release: root.dec_generacion()

                    MDLabel:
                        id: generacion
                        text: "1"
                        halign: "center"
                        valign: "middle"
                        size_hint_x: None
                        width: "44dp"
                        font_style: "H6"
                        theme_text_color: "Custom"
                        text_color: app.header_color

                    MDIconButton:
                        icon: "plus-circle-outline"
                        size_hint_x: None
                        size_hint_y: None
                        size: "40dp", "40dp"
                        pos_hint: {"center_y": 0.5}
                        theme_text_color: "Custom"
                        text_color: app.header_color
                        on_release: root.inc_generacion()

                MDBoxLayout:
                    size_hint_y: None
                    height: "48dp"
                    spacing: "6dp"

                    MDLabel:
                        text: "Bandejas"
                        size_hint_x: None
                        width: "90dp"
                        valign: "middle"

                    MDIconButton:
                        icon: "minus-circle-outline"
                        size_hint_x: None
                        size_hint_y: None
                        size: "40dp", "40dp"
                        pos_hint: {"center_y": 0.5}
                        theme_text_color: "Custom"
                        text_color: app.header_color
                        on_release: root.dec_bandejas()

                    MDLabel:
                        id: bandejas
                        text: "1"
                        halign: "center"
                        valign: "middle"
                        size_hint_x: None
                        width: "44dp"
                        font_style: "H6"
                        theme_text_color: "Custom"
                        text_color: app.header_color

                    MDIconButton:
                        icon: "plus-circle-outline"
                        size_hint_x: None
                        size_hint_y: None
                        size: "40dp", "40dp"
                        pos_hint: {"center_y": 0.5}
                        theme_text_color: "Custom"
                        text_color: app.header_color
                        on_release: root.inc_bandejas()

                MDBoxLayout:
                    size_hint_y: None
                    height: "48dp"
                    spacing: "8dp"

                    MDLabel:
                        text: "Tipo bandeja"
                        size_hint_x: None
                        width: "90dp"
                        valign: "middle"

                    MDDropDownItem:
                        id: tipo_bandeja_item
                        text: "72"
                        size_hint_x: None
                        width: "70dp"
                        on_release: root.open_tipo_bandeja_menu()

                    MDLabel:
                        id: semillas_label
                        text: "= 72 semillas"
                        valign: "middle"
                        font_style: "Caption"
                        theme_text_color: "Custom"
                        text_color: app.header_color

                MDBoxLayout:
                    adaptive_height: True
                    spacing: "12dp"
                    MDLabel:
                        text: "Nombre"
                        size_hint_x: None
                        width: "70dp"
                        valign: "middle"
                    MDDropDownItem:
                        id: operador_item
                        text: "Seleccionar"
                        on_release: root.open_operador_menu()

                MDTextField:
                    id: observaciones
                    hint_text: "Observaciones (libre)"
                    text: ""
                    multiline: True
                    size_hint_y: None
                    height: "80dp"

                MDGridLayout:
                    cols: 2
                    adaptive_height: True
                    spacing: "8dp"
                    row_force_default: True
                    row_default_height: "44dp"

                    MobilePrimaryButton:
                        text: "GUARDAR"
                        on_release: root.guardar()

                    MobileActionButton:
                        text: "RESUMEN"
                        on_release: root.resumen_del_dia()

                    MobileActionButton:
                        text: "DESCARGAR"
                        on_release: root.abrir_descarga()

                    MobileActionButton:
                        id: ver_btn
                        text: "VER REGISTROS"
                        on_release: root.toggle_ver_todos()

                MDLabel:
                    id: modo_label
                    text: "Mostrando: Hoy"
                    size_hint_y: None
                    height: "20dp"
                    font_style: "Caption"

                SectionCard:
                    MDLabel:
                        text: "Registros"
                        font_style: "Subtitle1"
                    ScrollView:
                        size_hint_y: None
                        height: "180dp"
                        MDList:
                            id: lista

<CosechasLogroScreen>:
    name: "cosechas_logro"
    canvas.before:
        Color:
            rgba: app.surface_bg
        Rectangle:
            pos: self.pos
            size: self.size

    MDBoxLayout:
        orientation: "vertical"
        padding: "0dp"
        spacing: "0dp"

        MDBoxLayout:
            orientation: "vertical"
            size_hint_y: None
            height: "64dp"
            md_bg_color: app.header_color
            padding: ["4dp", "0dp", "16dp", "0dp"]

            MDBoxLayout:
                spacing: "4dp"
                size_hint_y: 1

                MDIconButton:
                    icon: "arrow-left"
                    theme_text_color: "Custom"
                    text_color: (1, 1, 1, 1)
                    on_release: root.ir_home()

                MDLabel:
                    text: "Logro de cosechas"
                    font_style: "H6"
                    theme_text_color: "Custom"
                    text_color: (1, 1, 1, 1)

                Widget:

                MDLabel:
                    text: "MonAgric"
                    font_style: "Caption"
                    halign: "right"
                    theme_text_color: "Custom"
                    text_color: (0.9, 0.95, 0.9, 0.75)

        ScrollView:
            MDBoxLayout:
                orientation: "vertical"
                spacing: "12dp"
                padding: ["14dp", "12dp", "14dp", "14dp"]
                size_hint_y: None
                height: self.minimum_height

                MDBoxLayout:
                    id: logro_box
                    orientation: "vertical"
                    spacing: "8dp"
                    size_hint_y: None
                    height: self.minimum_height

<HorasScreen>:
    name: "horas"
    canvas.before:
        Color:
            rgba: app.surface_bg
        Rectangle:
            pos: self.pos
            size: self.size

    MDBoxLayout:
        orientation: "vertical"
        padding: "0dp"
        spacing: "0dp"

        MDBoxLayout:
            orientation: "vertical"
            size_hint_y: None
            height: "76dp"
            md_bg_color: app.header_color
            padding: ["4dp", "4dp", "16dp", "6dp"]

            MDBoxLayout:
                spacing: "4dp"
                size_hint_y: None
                height: "44dp"

                MDIconButton:
                    icon: "arrow-left"
                    theme_text_color: "Custom"
                    text_color: (1, 1, 1, 1)
                    on_release: root.ir_home()

                MDLabel:
                    text: "Horas de trabajo"
                    font_style: "H6"
                    theme_text_color: "Custom"
                    text_color: (1, 1, 1, 1)

                Widget:

                MDLabel:
                    text: "MonAgric"
                    font_style: "Caption"
                    halign: "right"
                    theme_text_color: "Custom"
                    text_color: (0.9, 0.95, 0.9, 0.75)

            MDLabel:
                text: "Registro por día y actividad"
                font_style: "Caption"
                size_hint_y: None
                height: "18dp"
                padding_x: "48dp"
                theme_text_color: "Custom"
                text_color: (0.92, 0.96, 0.92, 0.85)

        ScrollView:
            MDBoxLayout:
                orientation: "vertical"
                spacing: "10dp"
                padding: ["14dp", "12dp", "14dp", "14dp"]
                size_hint_y: None
                height: self.minimum_height

                MDBoxLayout:
                    adaptive_height: True
                    spacing: "4dp"

                    MDTextField:
                        id: fecha
                        hint_text: "AAAA-MM-DD"
                        text: root.hoy()

                    MDIconButton:
                        icon: "calendar"
                        size_hint: None, None
                        size: "48dp", "48dp"
                        pos_hint: {"center_y": 0.5}
                        on_release: root.abrir_calendario()

                MDBoxLayout:
                    adaptive_height: True
                    spacing: "12dp"

                    MDLabel:
                        text: "Nombre"
                        size_hint_x: None
                        width: "70dp"
                        valign: "middle"

                    MDDropDownItem:
                        id: nombre_item
                        text: "Seleccionar"
                        on_release: root.open_nombre_menu()

                MDBoxLayout:
                    size_hint_y: None
                    height: "48dp"
                    spacing: "12dp"

                    MDLabel:
                        text: "Horas"
                        size_hint_x: None
                        width: "70dp"
                        valign: "middle"

                    MDRaisedButton:
                        id: horas_btn
                        text: "0"
                        md_bg_color: app.header_color
                        text_color: (1, 1, 1, 1)
                        size_hint: None, None
                        size: "110dp", "44dp"
                        pos_hint: {"center_y": 0.5}
                        on_release: root.abrir_teclado_horas()

                SectionCard:
                    MDLabel:
                        text: "Tareas realizadas"
                        font_style: "Subtitle1"
                        size_hint_y: None
                        height: "24dp"
                    MDBoxLayout:
                        id: actividades_box
                        orientation: "vertical"
                        spacing: "0dp"
                        size_hint_y: None
                        height: self.minimum_height

                MobilePrimaryButton:
                    text: "GUARDAR"
                    on_release: root.guardar()

                MDGridLayout:
                    cols: 2
                    adaptive_height: True
                    spacing: "8dp"
                    row_force_default: True
                    row_default_height: "44dp"

                    MobileActionButton:
                        text: "RESUMEN $"
                        on_release: root.abrir_resumen()

                    MobileActionButton:
                        text: "DESCARGAR"
                        on_release: root.abrir_descarga()

                SectionCard:
                    MDLabel:
                        text: "Registros (tocá para editar)"
                        font_style: "Subtitle1"
                    ScrollView:
                        size_hint_y: None
                        height: "200dp"
                        MDList:
                            id: lista

<TrasplantesScreen>:
    name: "trasplantes"
    canvas.before:
        Color:
            rgba: app.surface_bg
        Rectangle:
            pos: self.pos
            size: self.size

    MDBoxLayout:
        orientation: "vertical"
        padding: "0dp"
        spacing: "0dp"

        MDBoxLayout:
            orientation: "vertical"
            size_hint_y: None
            height: "64dp"
            md_bg_color: app.header_color
            padding: ["4dp", "0dp", "16dp", "0dp"]

            MDBoxLayout:
                spacing: "4dp"
                size_hint_y: 1

                MDIconButton:
                    icon: "arrow-left"
                    theme_text_color: "Custom"
                    text_color: (1, 1, 1, 1)
                    on_release: root.ir_home()

                MDLabel:
                    text: "Trasplantes"
                    font_style: "H6"
                    theme_text_color: "Custom"
                    text_color: (1, 1, 1, 1)

                Widget:

                MDLabel:
                    text: "MonAgric"
                    font_style: "Caption"
                    halign: "right"
                    theme_text_color: "Custom"
                    text_color: (0.9, 0.95, 0.9, 0.75)

        ScrollView:
            MDBoxLayout:
                orientation: "vertical"
                spacing: "10dp"
                padding: ["14dp", "12dp", "14dp", "14dp"]
                size_hint_y: None
                height: self.minimum_height

                MDBoxLayout:
                    adaptive_height: True
                    spacing: "4dp"

                    MDTextField:
                        id: fecha
                        hint_text: "AAAA-MM-DD"
                        text: root.hoy()

                    MDIconButton:
                        icon: "calendar"
                        size_hint: None, None
                        size: "48dp", "48dp"
                        pos_hint: {"center_y": 0.5}
                        on_release: root.abrir_calendario()

                MDBoxLayout:
                    adaptive_height: True
                    spacing: "12dp"

                    MDLabel:
                        text: "Nombre"
                        size_hint_x: None
                        width: "70dp"
                        valign: "middle"

                    MDDropDownItem:
                        id: nombre_item
                        text: "Seleccionar"
                        on_release: root.open_nombre_menu()

                MDBoxLayout:
                    adaptive_height: True
                    spacing: "12dp"

                    MDLabel:
                        text: "Cultivo"
                        size_hint_x: None
                        width: "70dp"
                        valign: "middle"

                    MDDropDownItem:
                        id: cultivo_item
                        text: "Seleccionar"
                        on_release: root.open_cultivo_menu()

                MDBoxLayout:
                    size_hint_y: None
                    height: "48dp"
                    spacing: "6dp"

                    MDLabel:
                        text: "Generación"
                        size_hint_x: None
                        width: "90dp"
                        valign: "middle"

                    MDIconButton:
                        icon: "minus-circle-outline"
                        size_hint: None, None
                        size: "40dp", "40dp"
                        pos_hint: {"center_y": 0.5}
                        theme_text_color: "Custom"
                        text_color: app.header_color
                        on_release: root.dec_generacion()

                    MDLabel:
                        id: generacion
                        text: "1"
                        halign: "center"
                        valign: "middle"
                        size_hint_x: None
                        width: "44dp"
                        font_style: "H6"
                        theme_text_color: "Custom"
                        text_color: app.header_color

                    MDIconButton:
                        icon: "plus-circle-outline"
                        size_hint: None, None
                        size: "40dp", "40dp"
                        pos_hint: {"center_y": 0.5}
                        theme_text_color: "Custom"
                        text_color: app.header_color
                        on_release: root.inc_generacion()

                MDBoxLayout:
                    adaptive_height: True
                    spacing: "12dp"

                    MDLabel:
                        text: "Sector"
                        size_hint_x: None
                        width: "70dp"
                        valign: "middle"

                    MDDropDownItem:
                        id: sector_item
                        text: "A"
                        on_release: root.open_sector_menu()

                    MDLabel:
                        text: "Bancal"
                        size_hint_x: None
                        width: "60dp"
                        valign: "middle"

                    MDDropDownItem:
                        id: bancal_item
                        text: "1"
                        on_release: root.open_bancal_menu()

                MobilePrimaryButton:
                    text: "GUARDAR"
                    on_release: root.guardar()

                SectionCard:
                    MDLabel:
                        text: "Registros (tocá para editar)"
                        font_style: "Subtitle1"
                    ScrollView:
                        size_hint_y: None
                        height: "200dp"
                        MDList:
                            id: lista

<TrasplantesLogroScreen>:
    name: "trasplantes_logro"
    canvas.before:
        Color:
            rgba: app.surface_bg
        Rectangle:
            pos: self.pos
            size: self.size

    MDBoxLayout:
        orientation: "vertical"
        padding: "0dp"
        spacing: "0dp"

        MDBoxLayout:
            orientation: "vertical"
            size_hint_y: None
            height: "64dp"
            md_bg_color: app.header_color
            padding: ["4dp", "0dp", "16dp", "0dp"]

            MDBoxLayout:
                spacing: "4dp"
                size_hint_y: 1

                MDIconButton:
                    icon: "arrow-left"
                    theme_text_color: "Custom"
                    text_color: (1, 1, 1, 1)
                    on_release: root.ir_home()

                MDLabel:
                    text: "Logro de trasplantes"
                    font_style: "H6"
                    theme_text_color: "Custom"
                    text_color: (1, 1, 1, 1)

                Widget:

                MDLabel:
                    text: "MonAgric"
                    font_style: "Caption"
                    halign: "right"
                    theme_text_color: "Custom"
                    text_color: (0.9, 0.95, 0.9, 0.75)

        ScrollView:
            MDBoxLayout:
                orientation: "vertical"
                spacing: "12dp"
                padding: ["14dp", "12dp", "14dp", "14dp"]
                size_hint_y: None
                height: self.minimum_height

                MDBoxLayout:
                    id: logro_box
                    orientation: "vertical"
                    spacing: "8dp"
                    size_hint_y: None
                    height: self.minimum_height

<SiembrasLogroScreen>:
    name: "siembras_logro"
    canvas.before:
        Color:
            rgba: app.surface_bg
        Rectangle:
            pos: self.pos
            size: self.size

    MDBoxLayout:
        orientation: "vertical"
        padding: "0dp"
        spacing: "0dp"

        MDBoxLayout:
            orientation: "vertical"
            size_hint_y: None
            height: "64dp"
            md_bg_color: app.header_color
            padding: ["4dp", "0dp", "16dp", "0dp"]

            MDBoxLayout:
                spacing: "4dp"
                size_hint_y: 1

                MDIconButton:
                    icon: "arrow-left"
                    theme_text_color: "Custom"
                    text_color: (1, 1, 1, 1)
                    on_release: root.ir_home()

                MDLabel:
                    text: "Siembras efectivas"
                    font_style: "H6"
                    theme_text_color: "Custom"
                    text_color: (1, 1, 1, 1)

                Widget:

                MDLabel:
                    text: "MonAgric"
                    font_style: "Caption"
                    halign: "right"
                    theme_text_color: "Custom"
                    text_color: (0.9, 0.95, 0.9, 0.75)

        ScrollView:
            MDBoxLayout:
                orientation: "vertical"
                spacing: "12dp"
                padding: ["14dp", "12dp", "14dp", "14dp"]
                size_hint_y: None
                height: self.minimum_height

                MDBoxLayout:
                    id: logro_box
                    orientation: "vertical"
                    spacing: "8dp"
                    size_hint_y: None
                    height: self.minimum_height

<SanidadScreen>:
    name: "sanidad"
    canvas.before:
        Color:
            rgba: app.surface_bg
        Rectangle:
            pos: self.pos
            size: self.size

    MDBoxLayout:
        orientation: "vertical"
        padding: "0dp"
        spacing: "0dp"

        MDBoxLayout:
            orientation: "vertical"
            size_hint_y: None
            height: "64dp"
            md_bg_color: app.header_color
            padding: ["4dp", "0dp", "16dp", "0dp"]

            MDBoxLayout:
                spacing: "4dp"
                size_hint_y: 1

                MDIconButton:
                    icon: "arrow-left"
                    theme_text_color: "Custom"
                    text_color: (1, 1, 1, 1)
                    on_release: root.ir_home()

                MDLabel:
                    text: "Sanidad"
                    font_style: "H6"
                    theme_text_color: "Custom"
                    text_color: (1, 1, 1, 1)

                Widget:

                MDLabel:
                    text: "MonAgric"
                    font_style: "Caption"
                    halign: "right"
                    theme_text_color: "Custom"
                    text_color: (0.9, 0.95, 0.9, 0.75)

        ScrollView:
            MDBoxLayout:
                orientation: "vertical"
                spacing: "12dp"
                padding: ["14dp", "12dp", "14dp", "16dp"]
                size_hint_y: None
                height: self.minimum_height

                # ---- APLICACIONES DE RUTINA ----
                SectionCard:
                    MDLabel:
                        text: "Aplicaciones de rutina"
                        font_style: "Subtitle1"
                        bold: True
                        theme_text_color: "Custom"
                        text_color: app.header_color
                        size_hint_y: None
                        height: "26dp"
                    MDLabel:
                        text: "Registro de aplicaciones preventivas o de mantenimiento (MML, bioles, compost, etc.)."
                        font_style: "Caption"
                        size_hint_y: None
                        height: "34dp"

                    MDBoxLayout:
                        adaptive_height: True
                        spacing: "4dp"
                        MDTextField:
                            id: r_fecha
                            hint_text: "AAAA-MM-DD"
                            text: root.hoy()
                        MDIconButton:
                            icon: "calendar"
                            size_hint: None, None
                            size: "48dp", "48dp"
                            pos_hint: {"center_y": 0.5}
                            on_release: root.abrir_calendario()

                    MDBoxLayout:
                        adaptive_height: True
                        spacing: "12dp"
                        MDLabel:
                            text: "Producto"
                            size_hint_x: None
                            width: "70dp"
                            valign: "middle"
                        MDDropDownItem:
                            id: r_producto_item
                            text: "Seleccionar"
                            on_release: root.open_r_producto_menu()

                    MDTextField:
                        id: r_dosis
                        hint_text: "Dosis (ej. 5%, 3 cc/L)"

                    MDBoxLayout:
                        adaptive_height: True
                        spacing: "12dp"
                        MDLabel:
                            text: "Cultivo"
                            size_hint_x: None
                            width: "70dp"
                            valign: "middle"
                        MDDropDownItem:
                            id: r_cultivo_item
                            text: "Seleccionar"
                            on_release: root.open_r_cultivo_menu()

                    MDBoxLayout:
                        adaptive_height: True
                        spacing: "12dp"
                        MDLabel:
                            text: "Sector"
                            size_hint_x: None
                            width: "70dp"
                            valign: "middle"
                        MDDropDownItem:
                            id: r_sector_item
                            text: "A"
                            on_release: root.open_r_sector_menu()
                        MDLabel:
                            text: "Bancal"
                            size_hint_x: None
                            width: "60dp"
                            valign: "middle"
                        MDDropDownItem:
                            id: r_bancal_item
                            text: "1"
                            on_release: root.open_r_bancal_menu()

                    MobilePrimaryButton:
                        text: "REGISTRAR APLICACIÓN"
                        on_release: root.guardar_rutina()

                # ---- DETECCIÓN DE ENFERMEDAD O PLAGA (se despliega) ----
                MobileActionButton:
                    id: btn_deteccion
                    text: "DETECCIÓN DE ENFERMEDAD O PLAGA"
                    on_release: root.toggle_deteccion()

                MDBoxLayout:
                    id: deteccion_box
                    orientation: "vertical"
                    spacing: "10dp"
                    size_hint_y: None
                    height: self.minimum_height

                # ---- TRATAMIENTOS ACTIVOS (tabla con casillas) ----
                MDBoxLayout:
                    id: tratamientos_box
                    orientation: "vertical"
                    spacing: "10dp"
                    size_hint_y: None
                    height: self.minimum_height

                # ---- PLANILLA GENERAL DE APLICACIONES ----
                SectionCard:
                    MDBoxLayout:
                        adaptive_height: True
                        spacing: "6dp"
                        MDLabel:
                            text: "Aplicaciones registradas"
                            font_style: "Subtitle1"
                            bold: True
                            theme_text_color: "Custom"
                            text_color: app.header_color
                        MDLabel:
                            id: planilla_modo
                            text: ""
                            font_style: "Caption"
                            halign: "right"

                    MDBoxLayout:
                        id: planilla_box
                        orientation: "vertical"
                        spacing: "2dp"
                        size_hint_y: None
                        height: self.minimum_height

                    MDBoxLayout:
                        adaptive_height: True
                        spacing: "8dp"
                        MobileActionButton:
                            text: "RESUMEN"
                            on_release: root.abrir_resumen()
                        MobileActionButton:
                            text: "VER HOY"
                            on_release: root.toggle_ver_hoy()
                        MobileActionButton:
                            text: "DESCARGAR"
                            on_release: root.abrir_descarga()

<StockScreen>:
    name: "stock"
    canvas.before:
        Color:
            rgba: app.surface_bg
        Rectangle:
            pos: self.pos
            size: self.size

    MDBoxLayout:
        orientation: "vertical"

        MDBoxLayout:
            orientation: "vertical"
            size_hint_y: None
            height: "64dp"
            md_bg_color: app.header_color
            padding: ["4dp", "0dp", "16dp", "0dp"]

            MDBoxLayout:
                spacing: "4dp"
                size_hint_y: 1

                MDIconButton:
                    icon: "arrow-left"
                    theme_text_color: "Custom"
                    text_color: (1, 1, 1, 1)
                    on_release: root.ir_home()

                MDLabel:
                    text: "Stock"
                    font_style: "H6"
                    theme_text_color: "Custom"
                    text_color: (1, 1, 1, 1)

                Widget:

                MDLabel:
                    text: "MonAgric"
                    font_style: "Caption"
                    halign: "right"
                    theme_text_color: "Custom"
                    text_color: (0.9, 0.95, 0.9, 0.75)

        ScrollView:
            MDBoxLayout:
                orientation: "vertical"
                spacing: "12dp"
                padding: ["14dp", "12dp", "14dp", "16dp"]
                size_hint_y: None
                height: self.minimum_height

                # ---- Carga rápida ----
                SectionCard:
                    MDLabel:
                        text: "Cargar stock"
                        font_style: "Subtitle1"
                        bold: True
                        theme_text_color: "Custom"
                        text_color: app.header_color
                        size_hint_y: None
                        height: "26dp"

                    MDBoxLayout:
                        adaptive_height: True
                        spacing: "4dp"
                        MDTextField:
                            id: fecha
                            hint_text: "AAAA-MM-DD"
                            text: root.hoy()
                        MDIconButton:
                            icon: "calendar"
                            size_hint: None, None
                            size: "48dp", "48dp"
                            pos_hint: {"center_y": 0.5}
                            on_release: root.abrir_calendario()

                    MDBoxLayout:
                        adaptive_height: True
                        spacing: "12dp"
                        MDLabel:
                            text: "Sector"
                            size_hint_x: None
                            width: "62dp"
                            valign: "middle"
                        MDDropDownItem:
                            id: sector_item
                            text: "A"
                            on_release: root.open_sector_menu()
                        MDLabel:
                            text: "Bancal"
                            size_hint_x: None
                            width: "62dp"
                            valign: "middle"
                        MDDropDownItem:
                            id: bancal_item
                            text: "1"
                            on_release: root.open_bancal_menu()

                    MDBoxLayout:
                        adaptive_height: True
                        spacing: "12dp"
                        MDLabel:
                            text: "Cultivo"
                            size_hint_x: None
                            width: "62dp"
                            valign: "middle"
                        MDDropDownItem:
                            id: cultivo_item
                            text: "Seleccionar"
                            on_release: root.open_cultivo_menu()

                    MDTextField:
                        id: kg
                        hint_text: "Cant. (kg)"
                        input_filter: "float"
                        text: ""

                    MobilePrimaryButton:
                        text: "INGRESAR DATOS"
                        on_release: root.guardar()

                # ---- Lo cargado ----
                SectionCard:
                    MDBoxLayout:
                        adaptive_height: True
                        spacing: "6dp"
                        MDLabel:
                            text: "Registros"
                            font_style: "Subtitle1"
                            bold: True
                            theme_text_color: "Custom"
                            text_color: app.header_color
                        MDLabel:
                            id: modo_label
                            text: ""
                            font_style: "Caption"
                            halign: "right"

                    MDBoxLayout:
                        id: lista_box
                        orientation: "vertical"
                        spacing: "2dp"
                        size_hint_y: None
                        height: self.minimum_height

                    MobileActionButton:
                        id: ver_mas_btn
                        text: "VER MÁS"
                        on_release: root.ver_mas()

                    MDBoxLayout:
                        adaptive_height: True
                        spacing: "8dp"
                        MobileActionButton:
                            text: "RESUMEN"
                            on_release: root.abrir_resumen()
                        MobileActionButton:
                            text: "DESCARGAR"
                            on_release: root.abrir_descarga()

                    MobilePrimaryButton:
                        text: "COMPARTIR STOCK"
                        on_release: root.compartir()
"""


# ==========================================================
# SCREENS
# ==========================================================

class SectorRiegoRow(MDBoxLayout):
    """Fila de la tabla de sectores: letra automática, n° de bancales y tipo de riego."""

    def __init__(self, letra: str, on_eliminar=None, bancales: str = "", tipo: str = "", **kwargs):
        super().__init__(orientation="horizontal", spacing=dp(6),
                         size_hint_y=None, height=dp(48), **kwargs)
        self._tipo = tipo if tipo in TIPOS_RIEGO else TIPOS_RIEGO[0]
        self.sector_label = MDLabel(text=letra, bold=True, valign="middle",
                                    size_hint_x=None, width=dp(26),
                                    theme_text_color="Custom", text_color=SAGE_GREEN)
        self.bancales_input = MDTextField(hint_text="Bancales", input_filter="int",
                                          text=bancales, size_hint_x=None, width=dp(84))
        self.tipo_btn = MDFlatButton(text=self._tipo, theme_text_color="Custom",
                                     text_color=SAGE_GREEN, pos_hint={"center_y": 0.5})
        self._menu = MDDropdownMenu(
            caller=self.tipo_btn, width_mult=3,
            items=[{"text": t, "on_release": (lambda x=t: self._sel_tipo(x))}
                   for t in TIPOS_RIEGO],
        )
        self.tipo_btn.bind(on_release=lambda *_: abrir_menu(self._menu))
        borrar = MDIconButton(icon="delete-outline", theme_text_color="Custom",
                              text_color=(0.7, 0.25, 0.25, 1),
                              pos_hint={"center_y": 0.5})
        if on_eliminar:
            borrar.bind(on_release=lambda *_: on_eliminar(self))
        self.add_widget(self.sector_label)
        self.add_widget(self.bancales_input)
        self.add_widget(self.tipo_btn)
        self.add_widget(Widget())
        self.add_widget(borrar)

    def _sel_tipo(self, tipo: str):
        self._tipo = tipo
        self.tipo_btn.text = tipo
        self._menu.dismiss()

    def set_letra(self, letra: str):
        self.sector_label.text = letra

    def get_datos(self):
        """None si la fila está vacía; dict validado si tiene datos."""
        texto = norm_text(self.bancales_input.text)
        if not texto:
            return None
        bancales = int(texto)
        if bancales < 1:
            raise ValueError(f"Sector {self.sector_label.text}: bancales debe ser >= 1.")
        return {"sector": self.sector_label.text, "bancales": bancales, "tipo": self._tipo}


class CultivoCelda(ButtonBehavior, MDBoxLayout):
    """Celda de la cuadrícula del selector: ícono + nombre, tocable."""

    def __init__(self, cultivo: str, on_select, **kwargs):
        super().__init__(orientation="vertical", size_hint_y=None, height=dp(84),
                         spacing=dp(2), **kwargs)
        self.add_widget(MDIcon(
            icon=CULTIVO_ICONOS.get(cultivo, CULTIVO_ICONO_DEFAULT),
            halign="center", theme_text_color="Custom", text_color=SAGE_GREEN,
            font_size="34sp", size_hint_y=None, height=dp(44),
        ))
        nombre = MDLabel(text=cultivo, halign="center", font_style="Caption",
                         size_hint_y=None, height=dp(34))
        nombre.font_size = "11sp"
        self.add_widget(nombre)
        self.bind(on_release=lambda *_: on_select(cultivo))


class BaseScreen(MDScreen):
    """Utilidades comunes a pantallas (manejo de errores/diálogos)."""

    @property
    def repo(self) -> DataRepository:
        app = MDApp.get_running_app()
        if app and getattr(app, "repo", None):
            return app.repo
        raise RuntimeError("Repositorio de datos no inicializado.")

    def refresh_backend_status(self):
        app = MDApp.get_running_app()
        if app and hasattr(app, "refresh_backend_status"):
            app.refresh_backend_status()

    def ensure_write_allowed(self):
        app = MDApp.get_running_app()
        if not app:
            return
        # En modo remoto exigimos API online para no mezclar datos en local.
        if getattr(app, "backend_mode_name", "local") == "remote":
            app.refresh_backend_status()
            if not getattr(app, "backend_online", False):
                raise ConnectionError("Servidor remoto sin conexion. Guardado bloqueado para no mezclar datos.")

    def _abrir_calendario_para(self, tf):
        try:
            try:
                current = date.fromisoformat(norm_text(tf.text))
            except Exception:
                current = date.today()
            picker = _date_picker(year=current.year, month=current.month, day=current.day)
            picker.bind(on_save=lambda inst, val, _r: setattr(tf, "text", val.strftime("%Y-%m-%d")))
            picker.open()
        except Exception as e:
            self.safe_error("Error al abrir calendario", e, "BaseScreen._abrir_calendario_para()")

    def abrir_calendario(self):
        self._abrir_calendario_para(self.ids.fecha)

    def _make_fecha_row(self, tf):
        row = MDBoxLayout(size_hint_y=None, height=dp(48), spacing=dp(4))
        row.add_widget(tf)
        btn = MDIconButton(icon="calendar", size_hint=(None, None), size=(dp(48), dp(48)))
        btn.bind(on_release=lambda *_: self._abrir_calendario_para(tf))
        row.add_widget(btn)
        return row

    # ---- Nombre del operador: lista desplegable de los integrantes cargados
    _operador_value = ""

    def preparar_menu_operador(self):
        """Arma el desplegable con los integrantes de Configuración de Temporada."""
        item = self.ids.get("operador_item")
        if item is None:
            return
        try:
            nombres = [i["nombre"] for i in list_integrantes()]
        except Exception:
            nombres = []
        self._operador_menu = MDDropdownMenu(
            caller=item, width_mult=5,
            items=[{"text": n, "on_release": (lambda x=n: self._set_operador(x))}
                   for n in nombres],
        )
        if self._operador_value not in nombres:
            self._operador_value = nombres[0] if nombres else ""
        item.text = self._operador_value or "Seleccionar"
        if self._operador_value:
            try:
                item.set_item(self._operador_value)
            except Exception:
                pass

    def _set_operador(self, nombre):
        self._operador_value = nombre
        item = self.ids.get("operador_item")
        if item is not None:
            item.text = nombre
            try:
                item.set_item(nombre)
            except Exception:
                pass
        menu = getattr(self, "_operador_menu", None)
        if menu:
            menu.dismiss()

    def open_operador_menu(self):
        if not list_integrantes():
            self.safe_snackbar("Cargá los nombres en Configuración de Temporada → INTEGRANTES.")
            return
        abrir_menu(getattr(self, "_operador_menu", None))

    def safe_snackbar(self, text: str):
        try:
            Snackbar(text=text).open()
        except Exception:
            # En caso rarísimo de error UI, no derribar la app
            print(text)

    def modo_local(self) -> bool:
        app = MDApp.get_running_app()
        return getattr(app, "backend_mode_name", "local") == "local"

    def abrir_acciones_registro(self, descripcion: str, on_editar=None, on_eliminar=None):
        """Diálogo Editar/Eliminar para un registro tocado en una lista."""
        dialog = None

        def _cerrar(*_):
            try:
                dialog.dismiss()
            except Exception:
                pass

        botones = []
        if on_editar:
            botones.append(MDFlatButton(
                text="EDITAR", text_color=(1, 1, 1, 1),
                on_release=lambda *_: (_cerrar(), on_editar()),
            ))
        if on_eliminar:
            botones.append(MDFlatButton(
                text="ELIMINAR", text_color=(1, 1, 1, 1),
                on_release=lambda *_: (_cerrar(), self.confirmar_eliminar(descripcion, on_eliminar)),
            ))
        botones.append(MDFlatButton(text="CANCELAR", text_color=(1, 1, 1, 1), on_release=_cerrar))

        dialog = MDDialog(
            title="Registro",
            text=descripcion,
            md_bg_color=SAGE_GREEN,
            buttons=botones,
        )
        dialog.open()

    def confirmar_eliminar(self, descripcion: str, on_ok):
        dialog = None

        def _ok(*_):
            try:
                dialog.dismiss()
            except Exception:
                pass
            on_ok()

        dialog = MDDialog(
            title="Eliminar registro",
            text=f"{descripcion}\n\nEsta acción no se puede deshacer.",
            md_bg_color=SAGE_GREEN,
            buttons=[
                MDFlatButton(text="CANCELAR", text_color=(1, 1, 1, 1),
                             on_release=lambda *_: dialog.dismiss()),
                MDFlatButton(text="ELIMINAR", text_color=(1, 1, 1, 1), on_release=_ok),
            ],
        )
        dialog.open()

    def abrir_teclado_numerico(self, titulo: str, on_ok, inicial: str = "",
                               permitir_coma: bool = True, oculto: bool = False):
        """Teclado numérico propio: 0-9, coma decimal, borrar y OK.

        `oculto` muestra puntos en lugar del valor (para el PIN)."""
        estado = {"valor": norm_text(inicial)}

        display = MDLabel(
            text="", halign="center", font_style="H4", bold=True,
            theme_text_color="Custom", text_color=SAGE_GREEN,
            size_hint_y=None, height=dp(52),
        )

        def _refrescar():
            v = estado["valor"]
            display.text = ("•" * len(v)) if (oculto and v) else (v or "0")

        def _tocar(tecla):
            v = estado["valor"]
            if tecla == "borrar":
                estado["valor"] = v[:-1]
            elif tecla == ",":
                if permitir_coma and "," not in v:
                    estado["valor"] = (v or "0") + ","
            else:
                if len(v) < 8:
                    estado["valor"] = v + tecla
            _refrescar()

        _refrescar()
        teclas = [["1", "2", "3"], ["4", "5", "6"], ["7", "8", "9"],
                  [("," if permitir_coma else ""), "0", "borrar"]]
        grilla = MDBoxLayout(orientation="vertical", spacing=dp(6),
                             size_hint_y=None, height=dp(4 * 52 + 18))
        for fila_teclas in teclas:
            fila = MDBoxLayout(orientation="horizontal", spacing=dp(6),
                               size_hint_y=None, height=dp(52))
            for t in fila_teclas:
                if not t:
                    fila.add_widget(Widget())
                    continue
                btn = MDRaisedButton(
                    text=("⌫" if t == "borrar" else t),
                    md_bg_color=(WARM_AMBER if t == "borrar" else SAGE_GREEN),
                    text_color=(1, 1, 1, 1), font_size="20sp",
                    size_hint=(1, None), height=dp(52),
                )
                btn.bind(on_release=lambda _w, x=t: _tocar(x))
                fila.add_widget(btn)
            grilla.add_widget(fila)

        content = MDBoxLayout(orientation="vertical", spacing=dp(8),
                              padding=[dp(8), dp(8), dp(8), 0],
                              size_hint_y=None, height=dp(52 + 4 * 52 + 34))
        content.add_widget(display)
        content.add_widget(grilla)
        dialog = None

        def _ok(*_):
            try:
                dialog.dismiss()
            except Exception:
                pass
            on_ok(estado["valor"])

        dialog = MDDialog(
            title=titulo,
            type="custom",
            content_cls=content,
            md_bg_color=CARD_BG,
            buttons=[
                MDFlatButton(text="CANCELAR", text_color=SAGE_GREEN,
                             on_release=lambda *_: dialog.dismiss()),
                MDFlatButton(text="OK", text_color=SAGE_GREEN, on_release=_ok),
            ],
        )
        dialog.open()

    def abrir_selector_cultivo(self, on_select, titulo: str = "Elegir cultivo",
                               incluir_crear: bool = True):
        """Cuadrícula de cultivos con ícono, buscador y creación de cultivos propios."""
        from kivymd.uix.gridlayout import MDGridLayout as _MDGrid

        buscador = MDTextField(hint_text="Buscar cultivo...", size_hint_y=None, height=dp(44))
        grid = _MDGrid(cols=3, spacing=dp(4), size_hint_y=None, padding=[0, dp(4), 0, 0])
        grid.bind(minimum_height=grid.setter("height"))
        scroll = ScrollView(size_hint=(1, None), height=dp(300))
        scroll.add_widget(grid)
        dialog = None

        def _elegir(cultivo):
            try:
                dialog.dismiss()
            except Exception:
                pass
            on_select(cultivo)

        def _llenar(filtro: str = ""):
            grid.clear_widgets()
            f = norm_text(filtro).lower()
            for c in get_cultivos():
                if f and f not in c.lower():
                    continue
                grid.add_widget(CultivoCelda(c, _elegir))

        buscador.bind(text=lambda _w, v: _llenar(v))
        _llenar()

        alto = dp(44) + dp(300) + (dp(52) if incluir_crear else 0) + dp(20)
        content = MDBoxLayout(orientation="vertical", spacing=dp(8),
                              padding=[dp(4), dp(8), dp(4), 0],
                              size_hint_y=None, height=alto)
        content.add_widget(buscador)
        content.add_widget(scroll)
        if incluir_crear:
            crear_btn = MDRaisedButton(
                text="+ NUEVO CULTIVO",
                md_bg_color=WARM_AMBER, text_color=(1, 1, 1, 1),
                size_hint=(1, None), height=dp(44),
            )
            crear_btn.bind(on_release=lambda *_: self.abrir_crear_cultivo(_elegir))
            content.add_widget(crear_btn)

        dialog = MDDialog(
            title=titulo,
            type="custom",
            content_cls=content,
            md_bg_color=CARD_BG,
            buttons=[MDFlatButton(text="CANCELAR", text_color=SAGE_GREEN,
                                  on_release=lambda *_: dialog.dismiss())],
        )
        dialog.open()

    def abrir_crear_cultivo(self, on_created=None):
        """Alta de un cultivo propio con los datos básicos de su perfil."""
        estado = {"tipo": TIPO_SIEMBRA_DIRECTA}
        nombre_input = MDTextField(hint_text="Nombre del cultivo", size_hint_y=None, height=dp(48))
        tipo_btn = MDRaisedButton(
            text=f"Siembra: {estado['tipo']}",
            md_bg_color=SAGE_GREEN, text_color=(1, 1, 1, 1),
            size_hint=(1, None), height=dp(40),
        )
        dias_input = MDTextField(hint_text="Siembra a cosecha (días)", input_filter="int",
                                 text="60", size_hint_y=None, height=dp(48))
        ventana_input = MDTextField(hint_text="Ventana (días)", input_filter="int",
                                    text="30", size_hint_y=None, height=dp(48))
        rinde_input = MDTextField(hint_text="Rinde (kg/m²)", input_filter="float",
                                  size_hint_y=None, height=dp(48))
        ayuda = MDLabel(
            text="Distancia y líneas quedan con valores genéricos: ajustalos en PERFILES CULTIVO.",
            font_style="Caption", theme_text_color="Secondary",
            size_hint_y=None, height=dp(32),
        )

        def _toggle_tipo(*_):
            estado["tipo"] = (
                TIPO_SIEMBRA_ALMACIGO if estado["tipo"] == TIPO_SIEMBRA_DIRECTA
                else TIPO_SIEMBRA_DIRECTA
            )
            tipo_btn.text = f"Siembra: {estado['tipo']}"

        tipo_btn.bind(on_release=_toggle_tipo)

        fila = MDBoxLayout(orientation="horizontal", spacing=dp(8),
                           size_hint_y=None, height=dp(52))
        fila.add_widget(dias_input)
        fila.add_widget(ventana_input)

        content = MDBoxLayout(orientation="vertical", spacing=dp(10),
                              padding=[dp(8), dp(16), dp(8), 0],
                              size_hint_y=None, height=dp(288))
        content.add_widget(nombre_input)
        content.add_widget(tipo_btn)
        content.add_widget(fila)
        content.add_widget(rinde_input)
        content.add_widget(ayuda)
        dialog = None

        def _guardar(*_):
            try:
                nombre = add_cultivo_extra(nombre_input.text)
                dias = int(norm_text(dias_input.text) or 60)
                ventana = int(norm_text(ventana_input.text) or 30)
                if dias < 1 or ventana < 1:
                    raise ValueError("Días y ventana deben ser >= 1.")
                rinde = float(norm_text(rinde_input.text).replace(",", ".") or 0)
                upsert_perfil_cultivo(nombre, estado["tipo"], 0, dias, ventana, rinde, 25.0, 3)
                if dialog:
                    dialog.dismiss()
                self.safe_snackbar(f"Cultivo creado: {nombre}")
                if on_created:
                    on_created(nombre)
            except Exception as e:
                self.safe_error("Error al crear Cultivo", e, "BaseScreen.abrir_crear_cultivo()")

        dialog = MDDialog(
            title="Nuevo cultivo",
            type="custom",
            content_cls=content,
            md_bg_color=CARD_BG,
            buttons=[
                MDFlatButton(text="CANCELAR", text_color=SAGE_GREEN,
                             on_release=lambda *_: dialog.dismiss()),
                MDFlatButton(text="CREAR", text_color=SAGE_GREEN, on_release=_guardar),
            ],
        )
        dialog.open()

    def resumenes_cosecha_temporada(self, temporada: dict, hoy: str) -> list:
        """Reloj de cosecha de cada cultivo del plan (lista de dicts de resumen)."""
        try:
            plan = list_plan(temporada["id"])
        except Exception:
            plan = []
        kg_por_cultivo = {}
        primera_cosecha = {}
        primera_siembra = {}
        try:
            if self.modo_local():
                cosechas = list_cosechas_de_temporada(temporada["id"])
            else:
                cosechas = [(f, c, kg) for f, c, kg, _s, _b in
                            self.repo.list_cosechas_between(temporada["inicio"], hoy)]
            for f, cultivo, kg in cosechas:
                if str(f) > hoy:
                    continue  # registros con fecha futura no cuentan "a hoy"
                kg_por_cultivo[cultivo] = kg_por_cultivo.get(cultivo, 0.0) + float(kg)
                primera_cosecha.setdefault(cultivo, f)
        except Exception:
            pass
        try:
            if self.modo_local():
                siembras = list_siembras_de_temporada(temporada["id"])
            else:
                siembras = [(f, c) for f, c, *_resto in
                            self.repo.list_siembras_between(temporada["inicio"], hoy)]
            for f, cultivo in siembras:
                if str(f) > hoy:
                    continue
                primera_siembra.setdefault(cultivo, f)
        except Exception:
            pass

        resumenes = []
        for cultivo, superficie_m2, esperado_kg, *_marco in plan:
            try:
                try:
                    fechas_plan = get_fechas_siembra_plan(temporada["id"], cultivo)
                except Exception:
                    fechas_plan = []
                resumenes.append(resumen_cultivo_a_fecha(
                    cultivo, hoy, temporada, float(superficie_m2), float(esperado_kg),
                    kg_por_cultivo.get(cultivo, 0.0), primera_siembra.get(cultivo),
                    primera_cosecha.get(cultivo), fechas_plan,
                ))
            except Exception:
                continue
        return resumenes

    def abrir_dialogo_plan_cultivo(self, bancal_m2: float, on_confirm,
                                   titulo: str = "Agregar cultivo al plan",
                                   cultivo_fijo: str | None = None,
                                   cantidad_inicial: str = "",
                                   unidad_inicial: str = "bancales",
                                   esperado_inicial: str = "",
                                   tipo_inicial: str = "",
                                   distancia_inicial: str = "",
                                   lineas_inicial: str = "",
                                   ancho_bancal_m: float = DEFAULT_ANCHO_BANCAL_M,
                                   fechas_iniciales: list | None = None):
        """Diálogo compartido para cargar/editar un cultivo del plan.

        Al elegir cultivo se precargan del perfil: kg esperados, tipo de
        siembra, distancia entre plantas y líneas por bancal (todo editable).
        El número de plantas se calcula en vivo con el marco de plantación.
        Alturas fijas para evitar superposiciones en MDDialog.
        on_confirm(cultivo, m2, esperado, cantidad_txt, unidad, tipo, dist_cm, lineas, plantas).
        """
        estado = {"cultivo": cultivo_fijo or "", "unidad": unidad_inicial,
                  "tipo": tipo_inicial or TIPO_SIEMBRA_DIRECTA}

        if cultivo_fijo:
            cultivo_widget = MDLabel(
                text=cultivo_fijo, bold=True, halign="center",
                theme_text_color="Custom", text_color=SAGE_GREEN,
                size_hint_y=None, height=dp(32),
            )
        else:
            cultivo_widget = MDRaisedButton(
                text="SELECCIONAR CULTIVO",
                md_bg_color=SAGE_GREEN, text_color=(1, 1, 1, 1),
                size_hint=(1, None), height=dp(44),
            )
        cantidad_input = MDTextField(
            hint_text="Cantidad", input_filter="float", text=cantidad_inicial,
            size_hint_y=None, height=dp(48),
        )
        unidad_btn = MDRaisedButton(
            text=estado["unidad"],
            md_bg_color=WARM_AMBER, text_color=(1, 1, 1, 1),
            size_hint=(None, None), size=(dp(110), dp(44)),
            pos_hint={"center_y": 0.5},
        )
        esperado_input = MDTextField(
            hint_text="Cosecha esperada (kg)", input_filter="float", text=esperado_inicial,
            size_hint_y=None, height=dp(48),
        )
        sugerencia_label = MDLabel(
            text="", font_style="Caption", theme_text_color="Secondary",
            size_hint_y=None, height=dp(18),
        )
        tipo_btn = MDRaisedButton(
            text=f"Siembra: {estado['tipo']}",
            md_bg_color=SAGE_GREEN, text_color=(1, 1, 1, 1),
            size_hint=(1, None), height=dp(40),
        )
        distancia_input = MDTextField(
            hint_text="Dist. plantas (cm)", input_filter="float", text=distancia_inicial,
            size_hint_y=None, height=dp(48),
        )
        lineas_input = MDTextField(
            hint_text="Líneas/bancal", input_filter="int", text=lineas_inicial,
            size_hint_y=None, height=dp(48),
        )
        plantas_label = MDLabel(
            text="", font_style="Caption", bold=True,
            theme_text_color="Custom", text_color=SAGE_GREEN,
            size_hint_y=None, height=dp(20),
        )

        unidad_menu = MDDropdownMenu(
            caller=unidad_btn,
            items=[{"text": u, "on_release": (lambda x=u: _sel_unidad(x))}
                   for u in UNIDADES_SUPERFICIE],
            width_mult=3,
        )
        if not cultivo_fijo:
            cultivo_widget.bind(
                on_release=lambda *_: self.abrir_selector_cultivo(_sel_cultivo))

        def _superficie_m2() -> float:
            try:
                cantidad = float(norm_text(cantidad_input.text).replace(",", "."))
            except Exception:
                return 0.0
            if cantidad <= 0:
                return 0.0
            return superficie_a_m2(cantidad, estado["unidad"], bancal_m2)

        def _plantas() -> int:
            try:
                dist = float(norm_text(distancia_input.text).replace(",", ".") or 0)
            except Exception:
                dist = 0.0
            try:
                lineas = int(norm_text(lineas_input.text) or 0)
            except Exception:
                lineas = 0
            return calcular_plantas(_superficie_m2(), dist, lineas, ancho_bancal_m)

        def _recalcular_plantas(*_):
            n = _plantas()
            plantas_label.text = f"≈ {n} plantas" if n > 0 else ""

        def _sugerir(*_):
            if not estado["cultivo"]:
                return
            m2 = _superficie_m2()
            perfil = get_perfil_cultivo(estado["cultivo"])
            if m2 > 0 and perfil["rinde_ref"] > 0:
                esperado_input.text = str(round(m2 * perfil["rinde_ref"], 1))
                sugerencia_label.text = f"{m2:g} m² × {perfil['rinde_ref']:g} kg/m² (referencia editable)"
            else:
                sugerencia_label.text = f"{m2:g} m²" if m2 > 0 else ""
            _recalcular_plantas()

        def _sel_cultivo(c):
            estado["cultivo"] = c
            cultivo_widget.text = c
            perfil = get_perfil_cultivo(c)
            estado["tipo"] = perfil["tipo_siembra"]
            tipo_btn.text = f"Siembra: {estado['tipo']}"
            distancia_input.text = f"{perfil['distancia_cm']:g}"
            lineas_input.text = str(perfil["lineas_bancal"])
            _sugerir()

        def _sel_unidad(u):
            estado["unidad"] = u
            unidad_btn.text = u
            unidad_menu.dismiss()
            _sugerir()

        def _toggle_tipo(*_):
            estado["tipo"] = (
                TIPO_SIEMBRA_ALMACIGO if estado["tipo"] == TIPO_SIEMBRA_DIRECTA else TIPO_SIEMBRA_DIRECTA
            )
            tipo_btn.text = f"Siembra: {estado['tipo']}"

        unidad_btn.bind(on_release=lambda *_: abrir_menu(unidad_menu))
        tipo_btn.bind(on_release=_toggle_tipo)
        cantidad_input.bind(text=_sugerir)
        distancia_input.bind(text=_recalcular_plantas)
        lineas_input.bind(text=_recalcular_plantas)

        fila_superficie = MDBoxLayout(
            orientation="horizontal", spacing=dp(8),
            size_hint_y=None, height=dp(52),
        )
        fila_superficie.add_widget(cantidad_input)
        fila_superficie.add_widget(unidad_btn)

        fila_marco = MDBoxLayout(
            orientation="horizontal", spacing=dp(8),
            size_hint_y=None, height=dp(52),
        )
        fila_marco.add_widget(distancia_input)
        fila_marco.add_widget(lineas_input)

        # --- Fechas de siembra planificadas (una por generación) ---
        fechas_titulo = MDLabel(text="Fechas de siembra (por generación)",
                                font_style="Subtitle2", bold=True,
                                theme_text_color="Custom", text_color=SAGE_GREEN,
                                size_hint_y=None, height=dp(24))
        fechas_box = MDBoxLayout(orientation="vertical", spacing=dp(4), size_hint_y=None)
        fechas_box.bind(minimum_height=fechas_box.setter("height"))
        filas_fechas = []

        def _renumerar_fechas():
            for i, (_row, _tf, lbl) in enumerate(filas_fechas):
                lbl.text = f"Gen {i + 1}"

        def _quitar_fecha(entrada):
            if entrada in filas_fechas:
                filas_fechas.remove(entrada)
            if entrada[0] in fechas_box.children:
                fechas_box.remove_widget(entrada[0])
            _renumerar_fechas()

        def _agregar_fecha(valor: str = ""):
            row = MDBoxLayout(orientation="horizontal", spacing=dp(4),
                              size_hint_y=None, height=dp(48))
            lbl = MDLabel(text="Gen ?", bold=True, valign="middle",
                          size_hint_x=None, width=dp(48),
                          theme_text_color="Custom", text_color=SAGE_GREEN)
            tf = MDTextField(hint_text="AAAA-MM-DD", text=valor,
                             size_hint_y=None, height=dp(48))
            cal = MDIconButton(icon="calendar", pos_hint={"center_y": 0.5},
                               theme_text_color="Custom", text_color=SAGE_GREEN)
            cal.bind(on_release=lambda *_, t=tf: self._abrir_calendario_para(t))
            borrar = MDIconButton(icon="delete-outline", pos_hint={"center_y": 0.5},
                                  theme_text_color="Custom", text_color=(0.7, 0.25, 0.25, 1))
            row.add_widget(lbl)
            row.add_widget(tf)
            row.add_widget(cal)
            row.add_widget(borrar)
            entrada = (row, tf, lbl)
            borrar.bind(on_release=lambda *_: _quitar_fecha(entrada))
            filas_fechas.append(entrada)
            fechas_box.add_widget(row)
            _renumerar_fechas()

        for f in (fechas_iniciales or []):
            _agregar_fecha(str(f))

        agregar_fecha_btn = MDRaisedButton(
            text="+ FECHA DE SIEMBRA",
            md_bg_color=WARM_AMBER, text_color=(1, 1, 1, 1),
            size_hint=(1, None), height=dp(40),
        )
        agregar_fecha_btn.bind(on_release=lambda *_: _agregar_fecha())

        # Contenido con scroll: la tabla de fechas puede crecer
        inner = MDBoxLayout(orientation="vertical", spacing=dp(10),
                            padding=[0, dp(4), 0, dp(4)], size_hint_y=None)
        inner.bind(minimum_height=inner.setter("height"))
        inner.add_widget(cultivo_widget)
        inner.add_widget(fila_superficie)
        inner.add_widget(esperado_input)
        inner.add_widget(sugerencia_label)
        inner.add_widget(tipo_btn)
        inner.add_widget(fila_marco)
        inner.add_widget(plantas_label)
        inner.add_widget(fechas_titulo)
        inner.add_widget(fechas_box)
        inner.add_widget(agregar_fecha_btn)

        scroll = ScrollView(size_hint=(1, None), height=dp(420))
        scroll.add_widget(inner)
        content = MDBoxLayout(orientation="vertical",
                              padding=[dp(8), dp(8), dp(8), 0],
                              size_hint_y=None, height=dp(430))
        content.add_widget(scroll)
        dialog = None
        _recalcular_plantas()

        def _confirmar(*_):
            try:
                cultivo = validate_cultivo(estado["cultivo"])
                m2 = _superficie_m2()
                if m2 <= 0:
                    raise ValueError("Cargá una superficie mayor a 0.")
                esperado = validate_positive_float(esperado_input.text, "Cosecha esperada")
                cantidad_txt = norm_text(cantidad_input.text).replace(",", ".")
                try:
                    dist = float(norm_text(distancia_input.text).replace(",", ".") or 0)
                except Exception:
                    dist = 0.0
                try:
                    lineas = int(norm_text(lineas_input.text) or 0)
                except Exception:
                    lineas = 0
                fechas = []
                for _row, tf, lbl in filas_fechas:
                    if norm_text(tf.text):
                        fechas.append(validate_fecha(tf.text))
                on_confirm(cultivo, m2, esperado, cantidad_txt, estado["unidad"],
                           estado["tipo"], dist, lineas, _plantas(), fechas)
                if dialog:
                    dialog.dismiss()
            except Exception as e:
                self.safe_error("Error al guardar cultivo", e, "BaseScreen.abrir_dialogo_plan_cultivo()")

        dialog = MDDialog(
            title=titulo,
            type="custom",
            content_cls=content,
            md_bg_color=CARD_BG,
            buttons=[
                MDFlatButton(text="CANCELAR", text_color=SAGE_GREEN,
                             on_release=lambda *_: dialog.dismiss()),
                MDFlatButton(text="GUARDAR", text_color=SAGE_GREEN, on_release=_confirmar),
            ],
        )
        dialog.open()

    def safe_error(self, title: str, exc: BaseException, prefix: str):
        log_exception(prefix, exc)
        # diálogo con mensaje breve
        show_error_dialog(title, f"{exc}\n\nDetalle completo en:\n{LOG_PATH.resolve()}")
        # también a consola para depurar rápido
        print(prefix)
        print(traceback.format_exc())

    def _downloads_dir(self) -> Path:
        """Carpeta donde dejamos los CSV para que los puedas sacar del equipo.

        En Android 10+ la carpeta Descargas está bloqueada por 'scoped storage'
        (escribir ahí falla en silencio). La carpeta externa propia de la app,
        en cambio, se puede escribir sin permisos y se ve por USB y desde el
        explorador de archivos:
            Android/data/com.martintrigo.monagro/files/MonAgric
        """
        if platform == "android":
            try:
                from jnius import autoclass
                PythonActivity = autoclass("org.kivy.android.PythonActivity")
                ctx = PythonActivity.mActivity.getApplicationContext()
                externa = ctx.getExternalFilesDir(None)
                if externa is not None:
                    destino = Path(externa.getAbsolutePath()) / "MonAgric"
                    destino.mkdir(parents=True, exist_ok=True)
                    return destino
            except Exception as e:
                log_exception("No se pudo usar la carpeta externa de la app", e)
        CSV_DIR.mkdir(parents=True, exist_ok=True)
        return CSV_DIR

    def choose_export_dir(self, on_select):
        # En el celular no tiene sentido pedir una ruta a mano: exportamos
        # directo a la carpeta accesible de la app.
        if platform == "android":
            on_select(self._downloads_dir())
            return

        dialog = None

        def _use_downloads(*_):
            try:
                if dialog:
                    dialog.dismiss()
            except Exception:
                pass
            on_select(self._downloads_dir())

        def _custom_path(*_):
            try:
                if dialog:
                    dialog.dismiss()
            except Exception:
                pass
            input_path = MDTextField(
                hint_text="Ruta de carpeta",
                text=str(self._downloads_dir()),
            )
            content = MDBoxLayout(
                orientation="vertical",
                adaptive_height=True,
                spacing=dp(8),
                padding=[dp(8), dp(6), dp(8), dp(0)],
            )
            content.add_widget(input_path)
            custom_dialog = None

            def _confirm(*_args):
                try:
                    out_dir = Path(norm_text(input_path.text))
                    if not norm_text(str(out_dir)):
                        raise ValueError("La ruta no puede estar vacia.")
                    on_select(out_dir)
                    if custom_dialog:
                        custom_dialog.dismiss()
                except Exception as e:
                    self.safe_error("Ruta invalida", e, "BaseScreen.choose_export_dir()")

            custom_dialog = MDDialog(
                title="Otra carpeta",
                type="custom",
                content_cls=content,
                md_bg_color=SAGE_GREEN,
                buttons=[
                    MDFlatButton(
                        text="CANCELAR",
                        text_color=(1, 1, 1, 1),
                        on_release=lambda *_: custom_dialog.dismiss(),
                    ),
                    MDFlatButton(
                        text="GUARDAR AQUI",
                        text_color=(1, 1, 1, 1),
                        on_release=_confirm,
                    ),
                ],
            )
            custom_dialog.open()

        dialog = MDDialog(
            title="Elegir donde guardar",
            text="Selecciona carpeta de destino",
            md_bg_color=SAGE_GREEN,
            buttons=[
                MDFlatButton(
                    text="DESCARGAS",
                    text_color=(1, 1, 1, 1),
                    on_release=_use_downloads,
                ),
                MDFlatButton(
                    text="OTRA CARPETA",
                    text_color=(1, 1, 1, 1),
                    on_release=_custom_path,
                ),
                MDFlatButton(
                    text="CANCELAR",
                    text_color=(1, 1, 1, 1),
                    on_release=lambda *_: dialog.dismiss(),
                ),
            ],
        )
        dialog.open()


class HomeScreen(BaseScreen):
    def on_pre_enter(self):
        self.refresh_backend_status()
        app = MDApp.get_running_app()
        if app and hasattr(app, "actualizar_encabezado"):
            app.actualizar_encabezado()
        self.armar_tablero()

    def _tablero_label(self, texto, style="Body2", color=None, height=22, bold=False):
        lbl = MDLabel(text=texto, font_style=style, size_hint_y=None, height=dp(height), bold=bold)
        if color:
            lbl.theme_text_color = "Custom"
            lbl.text_color = color
        return lbl

    def armar_tablero(self):
        box = self.ids.get("tablero_box")
        if box is None:
            return
        box.clear_widgets()
        try:
            temporada = get_temporada_activa()
        except Exception:
            temporada = None

        if not temporada:
            card = Factory.SectionCard()
            card.add_widget(self._tablero_label(
                "Todavía no hay una temporada configurada.", "Body2", height=24))
            btn = MDRaisedButton(
                text="CONFIGURAR TEMPORADA",
                md_bg_color=WARM_AMBER, text_color=(1, 1, 1, 1),
                size_hint=(1, None), height=dp(44),
            )
            btn.bind(on_release=lambda *_: setattr(self.manager, "current", "setup"))
            card.add_widget(btn)
            box.add_widget(card)
            return

        hoy = date.today().isoformat()
        try:
            plan = list_plan(temporada["id"])
        except Exception:
            plan = []

        # --- Cosechas (global, tocable -> relojes por cultivo) ---
        resumenes = self.resumenes_cosecha_temporada(temporada, hoy)
        real_total = round(sum(r["kg_real"] for r in resumenes), 1)
        esperado_temporada_total = round(sum(r["esperado_total"] for r in resumenes), 1)

        card = Factory.SectionCard()
        fila_titulo = MDBoxLayout(orientation="horizontal", size_hint_y=None, height=dp(26), spacing=dp(6))
        fila_titulo.add_widget(self._tablero_label("Cosechas", "Subtitle1", SAGE_GREEN, height=26, bold=True))
        fila_titulo.add_widget(MDIcon(icon="chevron-right", theme_text_color="Custom",
                                      text_color=SAGE_GREEN, size_hint_x=None,
                                      width=dp(22), font_size="20sp"))
        card.add_widget(fila_titulo)

        # Avance sobre el TOTAL de la temporada (no prorrateado por fecha)
        if not plan:
            card.add_widget(self._tablero_label(
                "La temporada no tiene cultivos en el plan (Configuración de Temporada).",
                "Caption", height=20))
        elif esperado_temporada_total > 0:
            pct_global = round(real_total / esperado_temporada_total * 100.0, 1)
            card.add_widget(self._tablero_label(
                f"{real_total:g} / {esperado_temporada_total:g} kg del objetivo ({pct_global:g}%)",
                "Body2", height=24, bold=True))
            card.add_widget(ProgressBar(max=100, value=min(100.0, pct_global),
                                        size_hint_y=None, height=dp(5)))
            card.add_widget(self._tablero_label(
                "Tocá para ver el avance por cultivo", "Caption", height=16))
        elif real_total > 0:
            card.add_widget(self._tablero_label(
                f"{real_total:g} kg cosechados (sin objetivo en el plan)",
                "Body2", height=22, bold=True))
        else:
            card.add_widget(self._tablero_label(
                "Sin cosechas registradas todavía", "Caption", height=20))
        card.bind(on_release=lambda *_: setattr(self.manager, "current", "cosechas_logro"))
        box.add_widget(card)

        # --- Siembras efectivas (global, tocable -> detalle por cultivo) ---
        try:
            semillas = semillas_por_cultivo(temporada["id"]) if self.modo_local() else {}
        except Exception:
            semillas = {}
        objetivo_total = sum(int(p[6] or 0) for p in plan if len(p) > 6)
        sembradas_total = sum(semillas.values())

        card_siembras = Factory.SectionCard()
        fila_titulo = MDBoxLayout(orientation="horizontal", size_hint_y=None, height=dp(26), spacing=dp(6))
        fila_titulo.add_widget(self._tablero_label("Siembras efectivas", "Subtitle1", SAGE_GREEN, height=26, bold=True))
        fila_titulo.add_widget(MDIcon(icon="chevron-right", theme_text_color="Custom",
                                      text_color=SAGE_GREEN, size_hint_x=None,
                                      width=dp(22), font_size="20sp"))
        card_siembras.add_widget(fila_titulo)
        if objetivo_total > 0:
            pct_siembras = round(sembradas_total / objetivo_total * 100.0, 1)
            card_siembras.add_widget(self._tablero_label(
                f"{sembradas_total} / {objetivo_total} plantas objetivo ({pct_siembras:g}%)",
                "Body2", height=22, bold=True))
            card_siembras.add_widget(ProgressBar(max=100, value=min(100.0, pct_siembras),
                                                 size_hint_y=None, height=dp(5)))
            card_siembras.add_widget(self._tablero_label(
                "Tocá para ver el avance por cultivo", "Caption", height=16))
        else:
            card_siembras.add_widget(self._tablero_label(
                "Definí las plantas objetivo en el plan (editá cada cultivo) para activar esta barra.",
                "Caption", height=32))
        card_siembras.bind(on_release=lambda *_: setattr(self.manager, "current", "siembras_logro"))
        box.add_widget(card_siembras)

        # --- Trasplantes (plantas trasplantadas vs objetivo, tocable -> detalle) ---
        card_tras = Factory.SectionCard()
        fila_tras = MDBoxLayout(orientation="horizontal", size_hint_y=None, height=dp(26), spacing=dp(6))
        fila_tras.add_widget(self._tablero_label("Trasplantes", "Subtitle1", SAGE_GREEN, height=26, bold=True))
        fila_tras.add_widget(MDIcon(icon="chevron-right", theme_text_color="Custom",
                                    text_color=SAGE_GREEN, size_hint_x=None,
                                    width=dp(22), font_size="20sp"))
        card_tras.add_widget(fila_tras)
        try:
            datos_tras = trasplantes_por_cultivo(hoy) if self.modo_local() else {}
        except Exception:
            datos_tras = {}
        # Objetivo: plantas del plan de los cultivos de almácigo
        plan_almacigo = []
        for fila in plan:
            tipo_s = norm_text(fila[3]) if len(fila) > 3 else ""
            if not tipo_s:
                try:
                    tipo_s = get_perfil_cultivo(fila[0])["tipo_siembra"]
                except Exception:
                    tipo_s = ""
            if tipo_s == TIPO_SIEMBRA_ALMACIGO:
                plan_almacigo.append(fila)
        objetivo_tras = sum(int(f[6] or 0) for f in plan_almacigo if len(f) > 6)
        real_tras = sum(d["plantas"] for d in datos_tras.values())
        if objetivo_tras > 0:
            pct_tras = round(real_tras / objetivo_tras * 100.0, 1)
            card_tras.add_widget(self._tablero_label(
                f"{real_tras} / {objetivo_tras} plantas trasplantadas ({pct_tras:g}%)",
                "Body2", height=22, bold=True))
            card_tras.add_widget(ProgressBar(max=100, value=min(100.0, pct_tras),
                                             size_hint_y=None, height=dp(5)))
            card_tras.add_widget(self._tablero_label(
                "Tocá para ver el detalle por cultivo", "Caption", height=16))
        elif real_tras > 0:
            card_tras.add_widget(self._tablero_label(
                f"{real_tras} plantas trasplantadas · definí plantas objetivo en el plan",
                "Caption", height=20))
        else:
            card_tras.add_widget(self._tablero_label(
                "Sin trasplantes registrados. Definí plantas objetivo en el plan para el reloj.",
                "Caption", height=32))
        card_tras.bind(on_release=lambda *_: setattr(self.manager, "current", "trasplantes_logro"))
        box.add_widget(card_tras)

        # --- Riego (calidad ponderada de los últimos 7 días, tocable -> pantalla Riego) ---
        card_riego = Factory.SectionCard()
        fila_riego = MDBoxLayout(orientation="horizontal", size_hint_y=None, height=dp(26), spacing=dp(6))
        fila_riego.add_widget(self._tablero_label("Riego (últimos 7 días)", "Subtitle1",
                                                  SAGE_GREEN, height=26, bold=True))
        fila_riego.add_widget(MDIcon(icon="chevron-right", theme_text_color="Custom",
                                     text_color=SAGE_GREEN, size_hint_x=None,
                                     width=dp(22), font_size="20sp"))
        card_riego.add_widget(fila_riego)
        try:
            puntaje, total_riegos, sin_logro = calidad_riego_ultimos_dias(7)
        except Exception:
            puntaje, total_riegos, sin_logro = None, 0, 0
        if puntaje is not None:
            if puntaje >= RIEGO_VERDE:
                estado_riego = "verde"
            elif puntaje >= RIEGO_AMBAR:
                estado_riego = "ambar"
            else:
                estado_riego = "rojo"
            fila_val = MDBoxLayout(orientation="horizontal", size_hint_y=None, height=dp(24), spacing=dp(6))
            fila_val.add_widget(MDIcon(icon="circle", theme_text_color="Custom",
                                       text_color=SEMAFORO_COLORES[estado_riego],
                                       size_hint_x=None, width=dp(18), font_size="16sp"))
            fila_val.add_widget(self._tablero_label(
                f"Calidad de riego: {round(puntaje * 100):g}% · {total_riegos} riego(s)",
                "Body2", height=24, bold=True))
            card_riego.add_widget(fila_val)
            card_riego.add_widget(ProgressBar(max=100, value=round(puntaje * 100),
                                              size_hint_y=None, height=dp(5)))
            detalle_riego = "CC=100 · aceptable=75 · insuficiente=25"
            if sin_logro:
                detalle_riego += f" · {sin_logro} sin logro cargado"
            card_riego.add_widget(self._tablero_label(detalle_riego, "Caption", height=16))
        elif total_riegos > 0:
            card_riego.add_widget(self._tablero_label(
                f"{total_riegos} riego(s) sin logro cargado: registralo para medir la calidad.",
                "Caption", height=32))
        else:
            card_riego.add_widget(self._tablero_label(
                "Sin riegos registrados en los últimos 7 días.", "Caption", height=20))
        card_riego.bind(on_release=lambda *_: setattr(self.manager, "current", "riego"))
        box.add_widget(card_riego)

        # --- Avisos ---
        avisos = []
        try:
            last = self.repo.list_last_riego_by_sector()
            ahora = datetime.now()
            peor = None
            try:
                sectores_aviso = sectores_de_temporada_activa() or SECTORES
            except Exception:
                sectores_aviso = SECTORES
            for s in sectores_aviso:
                v = norm_text(last.get(s, ""))
                if not v:
                    continue
                try:
                    horas = (ahora - datetime.fromisoformat(v)).total_seconds() / 3600.0
                except Exception:
                    continue
                if peor is None or horas > peor[1]:
                    peor = (s, horas)
            if peor and peor[1] >= 48:
                avisos.append((f"Sector {peor[0]}: {int(peor[1])} h sin riego", "water-alert"))
        except Exception:
            pass
        try:
            tareas = self.repo.list_tareas_pendientes()
            altas = sum(1 for t in tareas if t[3] == "Alta")
            if altas:
                avisos.append((f"{altas} tarea(s) de importancia ALTA pendiente(s)", "alert"))
        except Exception:
            pass
        try:
            kg_hoy = sum(float(k) for _c, k, _s, _b in self.repo.list_cosechas_by_fecha(hoy))
            if kg_hoy > 0:
                avisos.append((f"Cosechado hoy: {round(kg_hoy, 1):g} kg", "basket"))
        except Exception:
            pass

        if avisos:
            aviso_card = Factory.SectionCard()
            aviso_card.add_widget(self._tablero_label("Avisos", "Subtitle1", SAGE_GREEN, height=24, bold=True))
            for texto, icono in avisos:
                fila = MDBoxLayout(orientation="horizontal", size_hint_y=None, height=dp(24), spacing=dp(6))
                fila.add_widget(MDIcon(icon=icono, theme_text_color="Custom",
                                       text_color=WARM_AMBER, size_hint_x=None,
                                       width=dp(20), font_size="18sp"))
                fila.add_widget(self._tablero_label(texto, "Body2", height=24))
                aviso_card.add_widget(fila)
            box.add_widget(aviso_card)

    def ir_riego(self):
        self.manager.current = "riego"

    def ir_cosechas(self):
        self.manager.current = "cosechas"

    def ir_horas(self):
        self.manager.current = "horas"

    def ir_trasplantes(self):
        self.manager.current = "trasplantes"

    def ir_sanidad(self):
        self.manager.current = "sanidad"

    def ir_objetivo(self):
        # La configuración de temporada es del administrador: si hay PIN, se pide
        if not admin_pin_definido():
            self.manager.current = "objetivo"
            return

        def _verificar(valor):
            if verificar_admin_pin(valor):
                self.manager.current = "objetivo"
            else:
                self.safe_snackbar("PIN incorrecto.")

        self.abrir_teclado_numerico("PIN de administrador", _verificar,
                                    permitir_coma=False, oculto=True)

    def ir_stock(self):
        self.manager.current = "stock"

    def ir_siembras(self):
        self.manager.current = "siembras"

    def ir_tareas(self):
        self.manager.current = "tareas"

    def ir_manejo(self):
        self.no_disponible("MANEJO")

    def confirm_exit(self):
        # La sesión de trabajo se guarda sola en on_stop; salir es directo.
        app = MDApp.get_running_app()
        if app:
            app.stop()

    def no_disponible(self, modulo: str):
        # No navegamos, solo avisamos.
        self.safe_snackbar(f"{modulo}: aún no implementado (MVP).")


class SetupScreen(BaseScreen):
    """Configuracion inicial / nueva temporada: productor, chacra y plan de cultivos."""

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        # cultivo -> {"superficie_m2": float, "esperado_kg": float, "detalle": str}
        self._plan = {}
        self._sector_rows = []

    def on_pre_enter(self):
        try:
            perfil = get_perfil_usuario()
            if not norm_text(self.ids.productor.text):
                self.ids.productor.text = perfil["productor"]
            if not norm_text(self.ids.chacra.text):
                self.ids.chacra.text = perfil["chacra"]
            if not norm_text(self.ids.largo_bancal.text):
                self.ids.largo_bancal.text = f"{perfil['largo_bancal_m']:g}"
            if not norm_text(self.ids.ancho_bancal.text):
                self.ids.ancho_bancal.text = f"{perfil['ancho_bancal_m']:g}"
            if not norm_text(self.ids.pasillo.text):
                self.ids.pasillo.text = f"{perfil['pasillo_m']:g}"
            if not norm_text(self.ids.n_bancales.text) and perfil["n_bancales"] > 0:
                self.ids.n_bancales.text = str(perfil["n_bancales"])
            if not norm_text(self.ids.temporada_nombre.text):
                self.ids.temporada_nombre.text = sugerir_nombre_temporada()
            if not norm_text(self.ids.fecha.text):
                self.ids.fecha.text = date.today().isoformat()
            if not self._sector_rows:
                self.agregar_sector_riego()
            self.actualizar_superficie()
            self.refrescar_plan()
        except Exception as e:
            self.safe_error("Error al preparar configuracion", e, "SetupScreen.on_pre_enter()")

    def _leer_float(self, widget_id: str) -> float:
        try:
            return float(norm_text(self.ids[widget_id].text).replace(",", "."))
        except Exception:
            return 0.0

    def actualizar_superficie(self):
        info = self.ids.get("superficie_info")
        if info is None:
            return
        largo = self._leer_float("largo_bancal")
        ancho = self._leer_float("ancho_bancal")
        pasillo = self._leer_float("pasillo")
        try:
            n = int(norm_text(self.ids.n_bancales.text) or 0)
        except Exception:
            n = 0
        if largo > 0 and ancho > 0 and n > 0:
            neta = round(largo * ancho * n, 1)
            ocupada = round(largo * (ancho + pasillo) * n, 1)
            info.text = (f"Bancal: {round(largo * ancho, 1):g} m² · Sup. neta de cultivo: {neta:g} m² · "
                         f"Con pasillos: {ocupada:g} m²")
        elif largo > 0 and ancho > 0:
            info.text = f"Bancal: {round(largo * ancho, 1):g} m² (cargá el n° de bancales para la superficie total)"
        else:
            info.text = ""

    # ---- Sectores de riego
    def _reletra_sectores(self):
        for i, fila in enumerate(self._sector_rows):
            fila.set_letra(chr(ord("A") + i) if i < 26 else f"S{i + 1}")

    def agregar_sector_riego(self, bancales: str = "", tipo: str = ""):
        box = self.ids.get("sectores_box")
        if box is None:
            return
        fila = SectorRiegoRow("?", on_eliminar=self._eliminar_sector_riego,
                              bancales=bancales, tipo=tipo)
        self._sector_rows.append(fila)
        box.add_widget(fila)
        self._reletra_sectores()

    def _eliminar_sector_riego(self, fila):
        box = self.ids.get("sectores_box")
        if fila in self._sector_rows:
            self._sector_rows.remove(fila)
        if box and fila in box.children:
            box.remove_widget(fila)
        self._reletra_sectores()

    def _sectores_datos(self) -> list:
        datos = []
        for fila in self._sector_rows:
            d = fila.get_datos()
            if d:
                datos.append(d)
        return datos

    def ir_home(self):
        self.manager.current = "home"

    def abrir_calendario_fin(self):
        self._abrir_calendario_para(self.ids.fecha_fin)

    def _bancal_m2_actual(self) -> float:
        largo = self._leer_float("largo_bancal")
        ancho = self._leer_float("ancho_bancal")
        if largo > 0 and ancho > 0:
            return round(largo * ancho, 2)
        return DEFAULT_BANCAL_M2

    def _ancho_bancal_actual(self) -> float:
        v = self._leer_float("ancho_bancal")
        return v if v > 0 else DEFAULT_ANCHO_BANCAL_M

    def refrescar_plan(self):
        lista = self.ids.get("plan_lista")
        if lista is None:
            return
        lista.clear_widgets()
        if not self._plan:
            lista.add_widget(OneLineListItem(text="Sin cultivos todavía. Agregá al menos uno."))
            return
        for cultivo in sorted(self._plan.keys()):
            data = self._plan[cultivo]
            item = OneLineListItem(text=data["detalle"])
            item.bind(on_release=lambda _w, c=cultivo, d=data["detalle"]: self.abrir_acciones_registro(
                d, on_eliminar=lambda: self.quitar_del_plan(c)))
            lista.add_widget(item)

    def quitar_del_plan(self, cultivo: str):
        self._plan.pop(cultivo, None)
        self.refrescar_plan()

    def abrir_agregar_cultivo(self):
        def _confirmado(cultivo, m2, esperado, cantidad_txt, unidad, tipo, dist_cm, lineas, plantas, fechas):
            detalle = f"{cultivo} — {cantidad_txt} {unidad} ({m2:g} m²) → {esperado:g} kg"
            if plantas > 0:
                detalle += f" · {plantas} pl"
            if fechas:
                detalle += f" · {len(fechas)} gen."
            self._plan[cultivo] = {
                "superficie_m2": m2, "esperado_kg": esperado, "detalle": detalle,
                "tipo": tipo, "distancia_cm": dist_cm, "lineas": lineas, "plantas": plantas,
                "fechas": fechas,
            }
            self.refrescar_plan()

        self.abrir_dialogo_plan_cultivo(
            self._bancal_m2_actual(), _confirmado,
            ancho_bancal_m=self._ancho_bancal_actual(),
        )

    def comenzar_temporada(self):
        try:
            productor = norm_text(self.ids.productor.text)
            chacra = norm_text(self.ids.chacra.text)
            if not chacra:
                raise ValueError("El nombre de la chacra es obligatorio.")
            largo_bancal = validate_positive_float(self.ids.largo_bancal.text, "Largo de bancal")
            ancho_bancal = validate_positive_float(self.ids.ancho_bancal.text, "Ancho de bancal")
            pasillo = validate_positive_float(self.ids.pasillo.text, "Ancho de pasillo")
            try:
                n_bancales = int(norm_text(self.ids.n_bancales.text))
            except Exception:
                raise ValueError("Cargá el número de bancales disponibles.")
            if n_bancales < 1:
                raise ValueError("El número de bancales debe ser >= 1.")
            nombre = norm_text(self.ids.temporada_nombre.text)
            inicio = validate_fecha(self.ids.fecha.text)
            fin = validate_optional_fecha(self.ids.fecha_fin.text)
            if not self._plan:
                raise ValueError("Agregá al menos un cultivo al plan.")
            sectores = self._sectores_datos()

            set_perfil_usuario(productor, chacra, largo_bancal, ancho_bancal, pasillo, n_bancales)
            temporada_id = create_temporada(nombre, inicio, fin)
            if sectores:
                save_sectores_riego(temporada_id, sectores)
            for cultivo, data in self._plan.items():
                upsert_plan_item(
                    temporada_id, cultivo, data["superficie_m2"], data["esperado_kg"],
                    tipo_siembra=data.get("tipo", ""),
                    distancia_cm=data.get("distancia_cm", 0),
                    lineas=data.get("lineas", 0),
                    plantas=data.get("plantas", 0),
                )
                try:
                    save_fechas_siembra_plan(temporada_id, cultivo, data.get("fechas") or [])
                except Exception:
                    pass
                # Compatibilidad con el indicador viejo de objetivos
                try:
                    upsert_objetivo(cultivo, data["superficie_m2"], data["esperado_kg"])
                except Exception:
                    pass
            try:
                set_temporada_config(inicio, fin)
            except Exception:
                pass

            app = MDApp.get_running_app()
            if app:
                app.actualizar_encabezado()
            self._plan = {}
            self.refrescar_plan()
            self.manager.current = "home"
            self.safe_snackbar(f"Temporada {nombre} iniciada. ¡Buen año productivo!")
        except Exception as e:
            self.safe_error("Error al comenzar Temporada", e, "SetupScreen.comenzar_temporada()")


class TareasScreen(BaseScreen):
    _importancia_opciones = ["Alta", "Media", "Baja"]
    _importancia_sel = StringProperty("Media")
    _fecha_sel = StringProperty("")
    _menu_importancia = None

    def on_pre_enter(self):
        self.refrescar()

    def ir_home(self):
        self.manager.current = "home"

    def refrescar(self):
        lista = self.ids.get("lista_tareas")
        if lista is None:
            return
        lista.clear_widgets()
        try:
            tareas = self.repo.list_tareas_pendientes()
        except Exception as e:
            self.safe_snackbar(f"Error al cargar tareas: {e}")
            return
        if not tareas:
            lbl = MDLabel(
                text="Sin tareas pendientes",
                halign="center",
                theme_text_color="Secondary",
                size_hint_y=None,
                height=dp(40),
            )
            lista.add_widget(lbl)
            return
        color_imp = {"Alta": (0.75, 0.18, 0.18, 1), "Media": WARM_AMBER, "Baja": SAGE_GREEN}
        for row in tareas:
            tarea_id, tarea_txt, fecha, importancia, n_personas = row
            item = TwoLineAvatarIconListItem(
                text=tarea_txt,
                secondary_text=f"{fecha}  ·  {importancia}  ·  {n_personas} pers.",
            )
            dot = IconLeftWidget(
                icon="circle",
                theme_text_color="Custom",
                text_color=color_imp.get(importancia, SAGE_GREEN),
            )
            check = IconRightWidget(
                icon="checkbox-blank-circle-outline",
                theme_text_color="Custom",
                text_color=SAGE_GREEN,
            )
            check.bind(on_release=lambda _btn, tid=tarea_id: self.marcar_realizada(tid))
            item.add_widget(dot)
            item.add_widget(check)
            lista.add_widget(item)

    def abrir_dialogo_nueva_tarea(self):
        self._fecha_sel = ""
        self._importancia_sel = "Media"

        campo_tarea = MDTextField(
            hint_text="Descripción",
            size_hint_y=None,
            height=dp(48),
        )
        campo_fecha = MDTextField(
            hint_text="Fecha (toca para seleccionar)",
            size_hint_y=None,
            height=dp(48),
            readonly=True,
        )
        campo_fecha.bind(on_touch_down=lambda w, t: self._abrir_datepicker(campo_fecha) if w.collide_point(*t.pos) else None)

        btn_imp = MDRaisedButton(
            text=f"Importancia: {self._importancia_sel}",
            md_bg_color=WARM_AMBER,
            text_color=(1, 1, 1, 1),
            size_hint_y=None,
            height=dp(42),
        )

        campo_personas = MDTextField(
            hint_text="N° operadores",
            input_filter="int",
            text="1",
            size_hint_y=None,
            height=dp(48),
        )

        def _abrir_menu_imp(*_):
            items = [
                {"text": op, "viewclass": "OneLineListItem",
                 "on_release": lambda x=op: _sel_imp(x)}
                for op in self._importancia_opciones
            ]
            self._menu_importancia = MDDropdownMenu(
                caller=btn_imp,
                items=items,
                width_mult=3,
            )
            self._menu_importancia.open()

        def _sel_imp(val):
            self._importancia_sel = val
            btn_imp.text = f"Importancia: {val}"
            if self._menu_importancia:
                self._menu_importancia.dismiss()

        btn_imp.bind(on_release=_abrir_menu_imp)

        content = MDBoxLayout(
            orientation="vertical",
            adaptive_height=True,
            spacing=dp(10),
            padding=[dp(8), dp(4), dp(8), dp(0)],
        )
        content.add_widget(campo_tarea)
        content.add_widget(campo_fecha)
        content.add_widget(btn_imp)
        content.add_widget(campo_personas)

        dialog = None

        def _guardar(*_):
            tarea_txt = campo_tarea.text.strip()
            fecha = campo_fecha.text.strip()
            try:
                n_pers = max(1, int(campo_personas.text.strip() or "1"))
            except ValueError:
                n_pers = 1
            if not tarea_txt:
                self.safe_snackbar("Escribí la descripción de la tarea.")
                return
            if not fecha:
                self.safe_snackbar("Seleccioná una fecha.")
                return
            try:
                self.repo.insert_tarea(tarea_txt, fecha, self._importancia_sel, n_pers)
            except Exception as e:
                self.safe_snackbar(f"Error al guardar: {e}")
                return
            if dialog:
                dialog.dismiss()
            self.refrescar()
            self.safe_snackbar("Tarea guardada.")

        dialog = MDDialog(
            title="Nueva tarea",
            type="custom",
            content_cls=content,
            md_bg_color=CARD_BG,
            buttons=[
                MDFlatButton(text="CANCELAR", text_color=SAGE_GREEN, on_release=lambda *_: dialog.dismiss()),
                MDFlatButton(text="GUARDAR", text_color=SAGE_GREEN, on_release=_guardar),
            ],
        )
        dialog.open()

    def _abrir_datepicker(self, campo):
        picker = _date_picker()
        picker.bind(on_save=lambda inst, val, rng: self._on_fecha_sel(campo, val))
        picker.open()

    def _on_fecha_sel(self, campo, val):
        self._fecha_sel = val.strftime("%Y-%m-%d")
        campo.text = self._fecha_sel

    def marcar_realizada(self, tarea_id: int):
        try:
            self.repo.marcar_tarea_realizada(tarea_id)
            self.refrescar()
            self.safe_snackbar("Tarea completada.")
        except Exception as e:
            self.safe_snackbar(f"Error: {e}")

    def abrir_resumen(self):
        try:
            pendientes = self.repo.list_tareas_pendientes()
        except Exception as e:
            self.safe_snackbar(f"Error: {e}")
            return
        total = len(pendientes)
        por_imp = {"Alta": 0, "Media": 0, "Baja": 0}
        for row in pendientes:
            por_imp[row[3]] = por_imp.get(row[3], 0) + 1
        texto = (
            f"Total pendientes: {total}\n"
            f"Alta: {por_imp['Alta']}  ·  Media: {por_imp['Media']}  ·  Baja: {por_imp['Baja']}"
        )
        dialog = MDDialog(
            title="Resumen de Tareas",
            text=texto,
            md_bg_color=SAGE_GREEN,
            buttons=[MDFlatButton(text="OK", text_color=(1, 1, 1, 1), on_release=lambda *_: dialog.dismiss())],
        )
        dialog.open()

    def abrir_descarga(self):
        import csv as _csv
        try:
            from datetime import date
            hoy = date.today().isoformat()
            inicio = hoy[:7] + "-01"
            filas = self.repo.list_tareas_realizadas_between(inicio, hoy)
        except Exception as e:
            self.safe_snackbar(f"Error: {e}")
            return
        dest = self._downloads_dir() / "tareas_realizadas.csv"
        try:
            with open(dest, "w", newline="", encoding="utf-8") as f:
                w = _csv.writer(f)
                w.writerow(["id", "tarea", "fecha", "importancia", "n_personas", "fecha_realizada"])
                w.writerows(filas)
            self.safe_snackbar(f"Guardado en {dest}")
        except Exception as e:
            self.safe_snackbar(f"Error al guardar CSV: {e}")


class RiegoScreen(BaseScreen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self._sector_menu = None
        self._sector_value = "A"
        self._filtro_menu = None
        self._filtro_value = FILTRO_TODOS
        self._logro_menu = None
        self._logro_value = "Riego aceptable"
        self._horas_val = 1
        self._ver_todos = True

    def on_pre_enter(self):
        self.refresh_backend_status()
        self.preparar_menu_operador()
        # Sectores definidos en la temporada activa; A..K solo como respaldo
        sectores = []
        try:
            if self.modo_local():
                sectores = sectores_de_temporada_activa()
        except Exception:
            sectores = []
        self._sectores = sectores or SECTORES
        for menu in (self._sector_menu, self._filtro_menu):
            if menu:
                try:
                    menu.dismiss()
                except Exception:
                    pass
        self._sector_menu = MDDropdownMenu(
            caller=self.ids.sector_item,
            items=[{
                "text": s,
                "on_release": (lambda x=s: self.set_sector(x))
            } for s in self._sectores],
            width_mult=2
        )
        self._filtro_menu = MDDropdownMenu(
            caller=self.ids.filtro_item,
            items=[{
                "text": s,
                "on_release": (lambda x=s: self.set_filtro(x))
            } for s in ([FILTRO_TODOS] + self._sectores)],
            width_mult=3
        )
        if self._sector_value not in self._sectores:
            self._sector_value = self._sectores[0]
            self.ids.sector_item.text = self._sector_value
        if self._filtro_value != FILTRO_TODOS and self._filtro_value not in self._sectores:
            self._filtro_value = FILTRO_TODOS
            self.ids.filtro_item.text = FILTRO_TODOS
        if not self._logro_menu:
            self._logro_menu = MDDropdownMenu(
                caller=self.ids.logro_item,
                items=[{
                    "text": lg,
                    "on_release": (lambda x=lg: self.set_logro(x))
                } for lg in LOGROS_RIEGO],
                width_mult=5
            )
        self.ids.horas.text = str(self._horas_val)
        self.ids.ver_btn.text = "VER HOY" if self._ver_todos else "VER REGISTROS"
        self.ids.modo_label.text = "Mostrando: Todos" if self._ver_todos else "Mostrando: Hoy"
        self.refrescar()
        self.actualizar_prioridad_riego()

    def hoy(self):
        return date.today().isoformat()

    def ir_home(self):
        self.manager.current = "home"

    def open_sector_menu(self):
        if self._sector_menu:
            abrir_menu(self._sector_menu)

    def open_filtro_menu(self):
        if self._filtro_menu:
            abrir_menu(self._filtro_menu)

    def set_sector(self, sector: str):
        self._sector_value = sector
        self.ids.sector_item.text = sector
        self.ids.sector_item.set_item(sector)
        if self._sector_menu:
            self._sector_menu.dismiss()

    def set_filtro(self, filtro: str):
        self._filtro_value = filtro
        self.ids.filtro_item.text = filtro
        self.ids.filtro_item.set_item(filtro)
        if self._filtro_menu:
            self._filtro_menu.dismiss()
        self.refrescar()

    def open_logro_menu(self):
        if self._logro_menu:
            abrir_menu(self._logro_menu)

    def set_logro(self, logro: str):
        self._logro_value = logro
        self.ids.logro_item.text = logro
        self.ids.logro_item.set_item(logro)
        if self._logro_menu:
            self._logro_menu.dismiss()

    def toggle_ver_todos(self):
        self._ver_todos = not self._ver_todos
        self.ids.ver_btn.text = "VER HOY" if self._ver_todos else "VER REGISTROS"
        self.ids.modo_label.text = "Mostrando: Todos" if self._ver_todos else "Mostrando: Hoy"
        self.refrescar()

    def inc_horas(self):
        self._horas_val += 1
        self.ids.horas.text = str(self._horas_val)

    def dec_horas(self):
        if self._horas_val > 1:
            self._horas_val -= 1
            self.ids.horas.text = str(self._horas_val)

    def registrar(self):
        try:
            self.ensure_write_allowed()
            fecha = validate_fecha(self.ids.fecha.text)
            horas = self._horas_val
            operador = norm_text(self._operador_value)
            sector = validate_sector(self._sector_value or self.ids.sector_item.text)
            logro = validate_logro(self._logro_value or self.ids.logro_item.text)

            if not operador:
                raise ValueError("Operador es obligatorio.")

            self.repo.insert_riego(fecha, horas, operador, sector, logro)
            self.safe_snackbar(f"Riego registrado: {horas} h | Sector {sector} | {LOGRO_RIEGO_ABREV.get(logro, logro)}")
            self.refrescar()
            self.actualizar_prioridad_riego()

        except Exception as e:
            self.safe_error("Error al registrar Riego", e, "RiegoScreen.registrar()")

    def resumen_ultimas_48h(self):
        try:
            rows = self.repo.list_riego_last_48h()
            if not rows:
                raise ValueError("No hay riegos en las últimas 48 hs.")

            content = MDBoxLayout(orientation="vertical", spacing=dp(6), size_hint_y=None)
            content.bind(minimum_height=content.setter("height"))

            header = MDBoxLayout(orientation="horizontal", size_hint_y=None, height=dp(28))
            header.add_widget(MDLabel(text="Fecha", size_hint_x=0.35, bold=True))
            header.add_widget(MDLabel(text="Sector", size_hint_x=0.15, bold=True))
            header.add_widget(MDLabel(text="Hs", size_hint_x=0.15, halign="right", bold=True))
            header.add_widget(MDLabel(text="Operador", size_hint_x=0.35, bold=True))
            content.add_widget(header)

            for fecha, sector, horas, operador in rows:
                row = MDBoxLayout(orientation="horizontal", size_hint_y=None, height=dp(26))
                row.add_widget(MDLabel(text=str(fecha), size_hint_x=0.35))
                row.add_widget(MDLabel(text=str(sector), size_hint_x=0.15))
                row.add_widget(MDLabel(text=str(horas), size_hint_x=0.15, halign="right"))
                row.add_widget(MDLabel(text=str(operador), size_hint_x=0.35))
                content.add_widget(row)

            dialog = MDDialog(
                title="Resumen últimas 48 hs",
                type="custom",
                content_cls=content,
                md_bg_color=SAGE_GREEN,
                buttons=[MDFlatButton(
                    text="OK",
                    text_color=(1, 1, 1, 1),
                    on_release=lambda *_: dialog.dismiss()
                )],
            )
            dialog.open()
        except Exception as e:
            self.safe_error("Error al generar Resumen", e, "RiegoScreen.resumen_ultimas_48h()")

    def actualizar_alerta_48h(self):
        self.actualizar_prioridad_riego()

    def actualizar_prioridad_riego(self):
        try:
            self.ids.prioridad_lista.clear_widgets()
            last_by_sector = self.repo.list_last_riego_by_sector()
            now = datetime.now()

            sectores = getattr(self, "_sectores", None) or SECTORES
            ranking = []
            for sector in sectores:
                last = norm_text(last_by_sector.get(sector, ""))
                if not last:
                    ranking.append((float("inf"), sector, "Sin registros previos"))
                    continue
                try:
                    dt = datetime.fromisoformat(last)
                    hours_ago = max(0.0, (now - dt).total_seconds() / 3600.0)
                    ranking.append((hours_ago, sector, f"Ultimo riego hace {round(hours_ago, 1)} h"))
                except Exception:
                    ranking.append((float("inf"), sector, "Fecha invalida en historial"))

            ranking.sort(key=lambda x: (-x[0], x[1]))
            for idx, (_, sector, detalle) in enumerate(ranking[:10], start=1):
                self.ids.prioridad_lista.add_widget(
                    OneLineListItem(text=f"{idx}. Sector {sector} - {detalle}")
                )
        except Exception as e:
            self.safe_error("Error calculando prioridad de riego", e, "RiegoScreen.actualizar_prioridad_riego()")

    def abrir_descarga(self):
        fecha_desde = MDTextField(hint_text="Desde: AAAA-MM-DD", text=self.hoy(), size_hint_y=None, height=dp(48))
        fecha_hasta = MDTextField(hint_text="Hasta: AAAA-MM-DD", text=self.hoy(), size_hint_y=None, height=dp(48))
        content = MDBoxLayout(
            orientation="vertical",
            adaptive_height=True,
            spacing=dp(8),
            padding=[dp(8), dp(6), dp(8), dp(0)],
        )
        content.add_widget(fecha_desde)
        content.add_widget(fecha_hasta)
        dialog = None

        def _descargar(*_):
            try:
                desde = validate_fecha(fecha_desde.text)
                hasta = validate_fecha(fecha_hasta.text)
                if desde > hasta:
                    raise ValueError("La fecha 'desde' no puede ser mayor que 'hasta'.")
                self.choose_export_dir(lambda out_dir: self.descargar_riego_csv(desde, hasta, out_dir))
                if dialog:
                    dialog.dismiss()
            except Exception as e:
                self.safe_error("Error al descargar Riegos", e, "RiegoScreen.abrir_descarga()")

        dialog = MDDialog(
            title="Descargar resumen de riegos",
            type="custom",
            content_cls=content,
            md_bg_color=SAGE_GREEN,
            buttons=[
                MDFlatButton(
                    text="CANCELAR",
                    text_color=(1, 1, 1, 1),
                    on_release=lambda *_: dialog.dismiss(),
                ),
                MDFlatButton(
                    text="DESCARGAR",
                    text_color=(1, 1, 1, 1),
                    on_release=_descargar,
                ),
            ],
        )
        dialog.open()

    def descargar_riego_csv(self, fecha_desde: str, fecha_hasta: str, output_dir: Path | None = None):
        rows = self.repo.list_riego_between(fecha_desde, fecha_hasta)
        if not rows:
            raise ValueError("No hay riegos en ese intervalo.")

        out_dir = Path(output_dir) if output_dir else self._downloads_dir()
        out_dir.mkdir(parents=True, exist_ok=True)
        path = out_dir / f"riego_resumen_{fecha_desde.replace('-', '')}_{fecha_hasta.replace('-', '')}.csv"
        with path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["fecha", "sector", "horas_riego", "operador"])
            for fecha, sector, horas, operador in rows:
                writer.writerow([fecha, sector, horas, operador])
        self.safe_snackbar(f"Descarga generada: {path}")

    def refrescar(self):
        try:
            self.ids.lista.clear_widgets()
            if self.modo_local():
                if self._ver_todos:
                    rows = self.repo.list_riego_con_id(limit=200)
                else:
                    fecha = norm_text(self.ids.fecha.text) or self.hoy()
                    rows = self.repo.list_riego_con_id(fecha=fecha)
                if self._filtro_value != FILTRO_TODOS:
                    rows = [r for r in rows if r[3] == self._filtro_value]
                if not rows:
                    self.ids.lista.add_widget(OneLineListItem(text="Sin registros en esta temporada"))
                    return
                for idx, fila in enumerate(rows, start=1):
                    rid, f, h, sec, op = fila[:5]
                    logro = norm_text(fila[5]) if len(fila) > 5 and fila[5] else ""
                    sufijo = f" | {LOGRO_RIEGO_ABREV.get(logro, logro)}" if logro else ""
                    texto = (f"{idx}. {f} | Sector {sec} | {h} h | {op}{sufijo}"
                             if self._ver_todos else f"{idx}. Sector {sec} | {h} h | {op}{sufijo}")
                    item = OneLineListItem(text=texto)
                    item.bind(on_release=lambda _w, r=fila: self.abrir_acciones_riego(r))
                    self.ids.lista.add_widget(item)
                return

            if self._ver_todos:
                rows = self.repo.list_riego_all(limit=200)
            else:
                fecha = norm_text(self.ids.fecha.text) or self.hoy()
                rows = self.repo.list_riego_by_fecha(fecha)
            if self._filtro_value != FILTRO_TODOS:
                if self._ver_todos:
                    rows = [r for r in rows if r[2] == self._filtro_value]
                else:
                    rows = [r for r in rows if r[1] == self._filtro_value]
            if not rows:
                self.ids.lista.add_widget(OneLineListItem(text="Sin registros"))
                return
            if self._ver_todos:
                for idx, (f, h, sec, op) in enumerate(rows, start=1):
                    self.ids.lista.add_widget(
                        OneLineListItem(text=f"{idx}. {f} | Sector {sec} | {h} h | {op}")
                    )
            else:
                for idx, (h, sec, op) in enumerate(rows, start=1):
                    self.ids.lista.add_widget(
                        OneLineListItem(text=f"{idx}. Sector {sec} | {h} h | {op}")
                    )
        except Exception as e:
            self.safe_error("Error listando Riegos", e, "RiegoScreen.refrescar()")

    def abrir_acciones_riego(self, registro):
        rid, f, h, sec, op = registro[:5]
        logro = norm_text(registro[5]) if len(registro) > 5 and registro[5] else ""
        desc = f"{f} | Sector {sec} | {h} h | {op}" + (f" | {logro}" if logro else "")
        self.abrir_acciones_registro(
            desc,
            on_editar=lambda: self.abrir_editar_riego(registro),
            on_eliminar=lambda: self.eliminar_riego(rid),
        )

    def eliminar_riego(self, riego_id: int):
        try:
            self.repo.delete_riego(riego_id)
            self.safe_snackbar("Riego eliminado.")
            self.refrescar()
            self.actualizar_prioridad_riego()
        except Exception as e:
            self.safe_error("Error al eliminar Riego", e, "RiegoScreen.eliminar_riego()")

    def abrir_editar_riego(self, registro):
        rid, f, h, sec, op = registro[:5]
        logro_actual = norm_text(registro[5]) if len(registro) > 5 and registro[5] else "Riego aceptable"
        if logro_actual not in LOGROS_RIEGO:
            logro_actual = "Riego aceptable"
        estado = {"logro": logro_actual}
        fecha_input = MDTextField(hint_text="Fecha (AAAA-MM-DD)", text=str(f),
                                  size_hint_y=None, height=dp(48))
        horas_input = MDTextField(hint_text="Horas de riego", text=str(h), input_filter="float",
                                  size_hint_y=None, height=dp(48))
        operador_input = MDTextField(hint_text="Operador", text=str(op),
                                     size_hint_y=None, height=dp(48))
        logro_btn = MDRaisedButton(
            text=f"Logro: {estado['logro']}",
            md_bg_color=SAGE_GREEN, text_color=(1, 1, 1, 1),
            size_hint=(1, None), height=dp(40),
        )

        def _sel_logro(lg):
            estado["logro"] = lg
            logro_btn.text = f"Logro: {lg}"
            logro_menu.dismiss()

        logro_menu = MDDropdownMenu(
            caller=logro_btn,
            items=[{"text": lg, "on_release": (lambda x=lg: _sel_logro(x))}
                   for lg in LOGROS_RIEGO],
            width_mult=5,
        )
        logro_btn.bind(on_release=lambda *_: abrir_menu(logro_menu))

        content = MDBoxLayout(orientation="vertical", spacing=dp(12),
                              padding=[dp(8), dp(12), dp(8), 0],
                              size_hint_y=None, height=dp(272))
        content.add_widget(MDLabel(
            text=f"Sector {sec}",
            size_hint_y=None, height=dp(24), font_style="Caption",
        ))
        content.add_widget(fecha_input)
        content.add_widget(horas_input)
        content.add_widget(operador_input)
        content.add_widget(logro_btn)
        dialog = None

        def _guardar(*_):
            try:
                nueva_fecha = validate_fecha(fecha_input.text)
                nuevas_horas = validate_positive_float(horas_input.text, "Horas de riego")
                nuevo_operador = norm_text(operador_input.text)
                if not nuevo_operador:
                    raise ValueError("Operador es obligatorio.")
                self.repo.update_riego(rid, nueva_fecha, nuevas_horas, nuevo_operador, estado["logro"])
                if dialog:
                    dialog.dismiss()
                self.safe_snackbar("Riego actualizado.")
                self.refrescar()
                self.actualizar_prioridad_riego()
            except Exception as e:
                self.safe_error("Error al editar Riego", e, "RiegoScreen.abrir_editar_riego()")

        dialog = MDDialog(
            title="Editar riego",
            type="custom",
            content_cls=content,
            md_bg_color=CARD_BG,
            buttons=[
                MDFlatButton(text="CANCELAR", text_color=SAGE_GREEN,
                             on_release=lambda *_: dialog.dismiss()),
                MDFlatButton(text="GUARDAR", text_color=SAGE_GREEN, on_release=_guardar),
            ],
        )
        dialog.open()


class CosechasScreen(BaseScreen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self._cultivo_menu = None
        self._cultivo_value = "Lechuga"
        self._ver_todos = True

    def on_pre_enter(self):
        self.refresh_backend_status()
        # Solo se cosechan cultivos del plan de la temporada activa;
        # sin plan, se ofrece el catalogo completo para no bloquear.
        opciones = []
        try:
            temporada = get_temporada_activa()
            if temporada and self.modo_local():
                opciones = [f[0] for f in list_plan(temporada["id"])]
        except Exception:
            opciones = []
        if not opciones:
            opciones = get_cultivos()
        if self._cultivo_menu:
            try:
                self._cultivo_menu.dismiss()
            except Exception:
                pass
        self._cultivo_menu = MDDropdownMenu(
            caller=self.ids.cultivo_item,
            items=[{
                "text": c,
                "on_release": (lambda x=c: self.set_cultivo(x))
            } for c in opciones],
            width_mult=6
        )
        if self._cultivo_value not in opciones:
            self._cultivo_value = opciones[0]
            self.ids.cultivo_item.text = self._cultivo_value
        self.ids.ver_btn.text = "VER HOY" if self._ver_todos else "VER REGISTROS"
        self.ids.modo_label.text = "Mostrando: Todos" if self._ver_todos else "Mostrando: Hoy"
        self.refrescar()
        self.update_logro_indicator()

    def hoy(self):
        return date.today().isoformat()

    def ir_home(self):
        self.manager.current = "home"

    def open_cultivo_menu(self):
        if self._cultivo_menu:
            abrir_menu(self._cultivo_menu)

    def set_cultivo(self, cultivo: str):
        self._cultivo_value = cultivo
        self.ids.cultivo_item.text = cultivo
        self.ids.cultivo_item.set_item(cultivo)
        if self._cultivo_menu:
            self._cultivo_menu.dismiss()
        self.update_logro_indicator()

    def toggle_ver_todos(self):
        self._ver_todos = not self._ver_todos
        self.ids.ver_btn.text = "VER HOY" if self._ver_todos else "VER REGISTROS"
        self.ids.modo_label.text = "Mostrando: Todos" if self._ver_todos else "Mostrando: Hoy"
        self.refrescar()

    def guardar(self):
        try:
            self.ensure_write_allowed()
            fecha = validate_fecha(self.ids.fecha.text)
            cultivo = validate_cultivo(self._cultivo_value or self.ids.cultivo_item.text)
            kg = validate_kg(self.ids.kg.text)

            # La cosecha se registra por cultivo (los cosecheros levantan de
            # varios bancales); sector/bancal quedan como marcador neutro.
            self.repo.insert_cosecha(fecha, cultivo, kg, "-", 0)
            self.safe_snackbar(f"Cosecha guardada: {cultivo} ({kg} kg)")
            self.refrescar()
            self.update_logro_indicator()

        except Exception as e:
            self.safe_error("Error al guardar Cosecha", e, "CosechasScreen.guardar()")

    def update_logro_indicator(self):
        try:
            cultivo = validate_cultivo(self._cultivo_value or self.ids.cultivo_item.text)
            temporada = get_temporada_activa()
            if not temporada:
                self.ids.logro_label.text = "Logro: configurá la temporada desde la pantalla de inicio"
                self.ids.logro_bar.value = 0
                return

            plan = {c: (float(s), float(e)) for c, s, e, *_marco in list_plan(temporada["id"])}
            if cultivo not in plan:
                self.ids.logro_label.text = f"Logro ({cultivo}): no está en el plan de esta temporada"
                self.ids.logro_bar.value = 0
                return
            superficie, esperado = plan[cultivo]

            fecha_actual = validate_fecha(norm_text(self.ids.fecha.text) or self.hoy())
            fecha_corte = fecha_actual
            fin = norm_text(temporada.get("fin", ""))
            if fin and fin < fecha_corte:
                fecha_corte = fin
            if fecha_corte < temporada["inicio"]:
                self.ids.logro_label.text = f"Logro ({cultivo}): fuera de temporada"
                self.ids.logro_bar.value = 0
                return

            cosechado = 0.0
            primera_cosecha = None
            if self.modo_local():
                filas = list_cosechas_de_temporada(temporada["id"])
            else:
                filas = [(f, c, kg) for f, c, kg, _s, _b in
                         self.repo.list_cosechas_between(temporada["inicio"], fecha_corte)]
            for f_row, row_cultivo, kg in filas:
                if row_cultivo != cultivo or str(f_row) > fecha_corte:
                    continue
                cosechado += float(kg)
                if primera_cosecha is None:
                    primera_cosecha = f_row

            primera_siembra = None
            try:
                if self.modo_local():
                    filas_s = list_siembras_de_temporada(temporada["id"])
                else:
                    filas_s = [(f, c) for f, c, *_resto in
                               self.repo.list_siembras_between(temporada["inicio"], fecha_corte)]
                for f, c in filas_s:
                    if c == cultivo and str(f) <= fecha_corte:
                        primera_siembra = f
                        break
            except Exception:
                pass

            try:
                fechas_plan = get_fechas_siembra_plan(temporada["id"], cultivo)
            except Exception:
                fechas_plan = []
            r = resumen_cultivo_a_fecha(
                cultivo, fecha_corte, temporada, superficie, esperado, cosechado,
                primera_siembra, primera_cosecha, fechas_plan,
            )
            # Avance sobre el TOTAL del plan (no prorrateado por fecha)
            pct = (round(r["kg_real"] / r["esperado_total"] * 100.0, 1)
                   if r["esperado_total"] > 0 else 0.0)
            self.ids.logro_bar.value = max(0.0, min(100.0, pct))
            self.ids.logro_label.theme_text_color = "Custom"
            self.ids.logro_label.text_color = SAGE_GREEN
            self.ids.logro_label.text = (
                f"{cultivo}: {r['kg_real']:g} / {r['esperado_total']:g} kg ({pct:g}%) · "
                f"{r['kg_m2_real']:g} kg/m² (obj {r['kg_m2_obj']:g})"
            )
        except Exception:
            self.ids.logro_label.text = "Logro: sin datos"
            self.ids.logro_bar.value = 0

    def resumen_del_dia(self):
        try:
            fecha = validate_fecha(self.ids.fecha.text)
            rows = self.repo.list_cosechas_by_fecha(fecha)
            if not rows:
                raise ValueError("No hay cosechas para esa fecha.")

            resumen = {}
            for cultivo, kg, _sec, _b in rows:
                resumen[cultivo] = resumen.get(cultivo, 0.0) + float(kg)

            content = MDBoxLayout(orientation="vertical", spacing=dp(6), size_hint_y=None)
            content.bind(minimum_height=content.setter("height"))
            header = MDBoxLayout(orientation="horizontal", size_hint_y=None, height=dp(28))
            header.add_widget(MDLabel(text="Cultivo", size_hint_x=0.65, bold=True))
            header.add_widget(MDLabel(text="Kg", size_hint_x=0.35, halign="right", bold=True))
            content.add_widget(header)

            total = 0.0
            for cultivo in sorted(resumen.keys()):
                kg = round(resumen[cultivo], 3)
                total += kg
                row = MDBoxLayout(orientation="horizontal", size_hint_y=None, height=dp(26))
                row.add_widget(MDLabel(text=str(cultivo), size_hint_x=0.65))
                row.add_widget(MDLabel(text=str(kg), size_hint_x=0.35, halign="right"))
                content.add_widget(row)

            total_row = MDBoxLayout(orientation="horizontal", size_hint_y=None, height=dp(28))
            total_row.add_widget(MDLabel(text="TOTAL", size_hint_x=0.65, bold=True))
            total_row.add_widget(MDLabel(text=str(round(total, 3)), size_hint_x=0.35, halign="right", bold=True))
            content.add_widget(total_row)

            dialog = MDDialog(
                title=f"Resumen de Cosechas ({fecha})",
                type="custom",
                content_cls=content,
                md_bg_color=SAGE_GREEN,
                buttons=[MDFlatButton(
                    text="OK",
                    text_color=(1, 1, 1, 1),
                    on_release=lambda *_: dialog.dismiss()
                )],
            )
            dialog.open()
        except Exception as e:
            self.safe_error("Error al generar Resumen de Cosechas", e, "CosechasScreen.resumen_del_dia()")

    def abrir_descarga(self):
        fecha_desde = MDTextField(hint_text="Desde: AAAA-MM-DD", text=self.hoy(), size_hint_y=None, height=dp(48))
        fecha_hasta = MDTextField(hint_text="Hasta: AAAA-MM-DD", text=self.hoy(), size_hint_y=None, height=dp(48))
        formato_item = MDFlatButton(text="csv", size_hint_y=None, height=dp(44))
        formato_menu = MDDropdownMenu(
            caller=formato_item,
            items=[
                {"text": "csv", "on_release": lambda *_: _set_formato("csv")},
                {"text": "odt", "on_release": lambda *_: _set_formato("odt")},
            ],
            width_mult=3,
        )

        def _set_formato(fmt: str):
            formato_item.text = fmt
            formato_menu.dismiss()

        formato_item.bind(on_release=lambda *_: abrir_menu(formato_menu))

        content = MDBoxLayout(
            orientation="vertical",
            adaptive_height=True,
            spacing=dp(8),
            padding=[dp(8), dp(6), dp(8), dp(0)],
        )
        content.add_widget(fecha_desde)
        content.add_widget(fecha_hasta)
        content.add_widget(formato_item)
        dialog = None

        def _descargar(*_):
            try:
                desde = validate_fecha(fecha_desde.text)
                hasta = validate_fecha(fecha_hasta.text)
                if desde > hasta:
                    raise ValueError("La fecha 'desde' no puede ser mayor que 'hasta'.")
                formato = norm_text(formato_item.text).lower() or "csv"
                self.choose_export_dir(lambda out_dir: self.descargar_cosechas(desde, hasta, formato, out_dir))
                if dialog:
                    dialog.dismiss()
            except Exception as e:
                self.safe_error("Error al descargar Cosechas", e, "CosechasScreen.abrir_descarga()")

        dialog = MDDialog(
            title="Descargar cosechas por intervalo",
            type="custom",
            content_cls=content,
            md_bg_color=SAGE_GREEN,
            buttons=[
                MDFlatButton(
                    text="CANCELAR",
                    text_color=(1, 1, 1, 1),
                    on_release=lambda *_: dialog.dismiss(),
                ),
                MDFlatButton(
                    text="DESCARGAR",
                    text_color=(1, 1, 1, 1),
                    on_release=_descargar,
                ),
            ],
        )
        dialog.open()

    def descargar_cosechas(self, fecha_desde: str, fecha_hasta: str, formato: str, output_dir: Path | None = None):
        rows = self.repo.list_cosechas_between(fecha_desde, fecha_hasta)
        if not rows:
            raise ValueError("No hay cosechas en ese intervalo.")

        out_dir = Path(output_dir) if output_dir else self._downloads_dir()
        out_dir.mkdir(parents=True, exist_ok=True)
        base = f"cosechas_{fecha_desde.replace('-', '')}_{fecha_hasta.replace('-', '')}"

        if formato == "csv":
            path = out_dir / f"{base}.csv"
            with path.open("w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["fecha", "cultivo", "kg", "sector", "bancal"])
                writer.writerows(rows)
            self.safe_snackbar(f"Descarga generada: {path}")
            return

        if formato == "odt":
            try:
                from odf.opendocument import OpenDocumentText
                from odf.text import P
            except Exception:
                raise RuntimeError("Para exportar a ODT instala 'odfpy' en el entorno.")

            path = out_dir / f"{base}.odt"
            doc = OpenDocumentText()
            doc.text.addElement(P(text=f"Cosechas {fecha_desde} a {fecha_hasta}"))
            for fecha, cultivo, kg, sector, bancal in rows:
                doc.text.addElement(P(text=f"{fecha} | {cultivo} | {kg} kg | {sector}-{bancal}"))
            doc.save(str(path))
            self.safe_snackbar(f"Descarga generada: {path}")
            return

        raise ValueError("Formato invalido. Usa csv u odt.")

    @staticmethod
    def _ubicacion(sec, b) -> str:
        sec = norm_text(str(sec or ""))
        return f" | {sec}-{b}" if sec and sec != "-" else ""

    def refrescar(self):
        try:
            self.ids.lista.clear_widgets()
            if self.modo_local():
                if self._ver_todos:
                    rows = self.repo.list_cosechas_con_id(limit=200)
                else:
                    fecha = norm_text(self.ids.fecha.text) or self.hoy()
                    rows = self.repo.list_cosechas_con_id(fecha=fecha)
                if not rows:
                    self.ids.lista.add_widget(OneLineListItem(text="Sin registros en esta temporada"))
                    return
                total = 0.0
                for rid, f, cultivo, kg, sec, b in rows:
                    total += float(kg)
                    ubic = self._ubicacion(sec, b)
                    texto = f"{f} | {cultivo}: {kg} kg{ubic}" if self._ver_todos else f"{cultivo}: {kg} kg{ubic}"
                    item = OneLineListItem(text=texto)
                    item.bind(on_release=lambda _w, r=(rid, f, cultivo, kg, sec, b): self.abrir_acciones_cosecha(r))
                    self.ids.lista.add_widget(item)
            else:
                if self._ver_todos:
                    rows = self.repo.list_cosechas_all(limit=200)
                    rows = [(f, c, kg, s, b) for f, c, kg, s, b in rows]
                else:
                    fecha = norm_text(self.ids.fecha.text) or self.hoy()
                    rows = [(None, c, kg, s, b) for c, kg, s, b in self.repo.list_cosechas_by_fecha(fecha)]
                if not rows:
                    self.ids.lista.add_widget(OneLineListItem(text="Sin registros"))
                    return
                total = 0.0
                for f, cultivo, kg, sec, b in rows:
                    total += float(kg)
                    prefijo = f"{f} | " if f else ""
                    self.ids.lista.add_widget(OneLineListItem(text=f"{prefijo}{cultivo}: {kg} kg | {sec}-{b}"))

            self.ids.lista.add_widget(
                OneLineListItem(text=f"TOTAL: {round(total, 2)} kg")
            )
            self.update_logro_indicator()
        except Exception as e:
            self.safe_error("Error listando Cosechas", e, "CosechasScreen.refrescar()")

    def abrir_acciones_cosecha(self, registro):
        rid, f, cultivo, kg, sec, b = registro
        desc = f"{f} | {cultivo}: {kg} kg{self._ubicacion(sec, b)}"
        self.abrir_acciones_registro(
            desc,
            on_editar=lambda: self.abrir_editar_cosecha(registro),
            on_eliminar=lambda: self.eliminar_cosecha(rid),
        )

    def eliminar_cosecha(self, cosecha_id: int):
        try:
            self.repo.delete_cosecha(cosecha_id)
            self.safe_snackbar("Cosecha eliminada.")
            self.refrescar()
        except Exception as e:
            self.safe_error("Error al eliminar Cosecha", e, "CosechasScreen.eliminar_cosecha()")

    def abrir_editar_cosecha(self, registro):
        rid, f, cultivo, kg, sec, b = registro
        fecha_input = MDTextField(hint_text="Fecha (AAAA-MM-DD)", text=str(f),
                                  size_hint_y=None, height=dp(48))
        kg_input = MDTextField(hint_text="Kg", text=str(kg), input_filter="float",
                               size_hint_y=None, height=dp(48))
        content = MDBoxLayout(orientation="vertical", spacing=dp(12),
                              padding=[dp(8), dp(12), dp(8), 0],
                              size_hint_y=None, height=dp(160))
        content.add_widget(MDLabel(
            text=f"{cultivo}{self._ubicacion(sec, b)}",
            size_hint_y=None, height=dp(24), font_style="Caption",
        ))
        content.add_widget(fecha_input)
        content.add_widget(kg_input)
        dialog = None

        def _guardar(*_):
            try:
                nueva_fecha = validate_fecha(fecha_input.text)
                nuevo_kg = validate_kg(kg_input.text)
                self.repo.update_cosecha(rid, nueva_fecha, nuevo_kg)
                if dialog:
                    dialog.dismiss()
                self.safe_snackbar("Cosecha actualizada.")
                self.refrescar()
            except Exception as e:
                self.safe_error("Error al editar Cosecha", e, "CosechasScreen.abrir_editar_cosecha()")

        dialog = MDDialog(
            title="Editar cosecha",
            type="custom",
            content_cls=content,
            md_bg_color=CARD_BG,
            buttons=[
                MDFlatButton(text="CANCELAR", text_color=SAGE_GREEN,
                             on_release=lambda *_: dialog.dismiss()),
                MDFlatButton(text="GUARDAR", text_color=SAGE_GREEN, on_release=_guardar),
            ],
        )
        dialog.open()


class ObjetivoScreen(BaseScreen):
    """Configuración de Temporada: ver/editar/activar temporadas y su plan."""

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self._temporada_vista = None  # temporada seleccionada en el desplegable
        self._temporada_menu = None

    def on_pre_enter(self):
        self.refresh_backend_status()
        # Por defecto se muestra la temporada activa
        if self._temporada_vista is None:
            self._temporada_vista = self._temporada_activa()
        else:
            # Refresca datos por si la editaron/activaron desde otra pantalla
            actual = next(
                (t for t in self._listar_temporadas() if t["id"] == self._temporada_vista["id"]),
                None,
            )
            self._temporada_vista = actual or self._temporada_activa()
        self.refrescar()

    def ir_home(self):
        self.manager.current = "home"

    def ir_setup(self):
        self.manager.current = "setup"

    def _temporada_activa(self):
        try:
            return get_temporada_activa()
        except Exception:
            return None

    def _listar_temporadas(self):
        try:
            return list_temporadas()
        except Exception:
            return []

    def open_temporada_menu(self):
        temporadas = self._listar_temporadas()
        if not temporadas:
            self.safe_snackbar("No hay temporadas: creá una con NUEVA.")
            return
        activa_id = (self._temporada_activa() or {}).get("id")

        def _etiqueta(t):
            marca = " (activa)" if t["id"] == activa_id else ""
            return f"{t['nombre']} — {t['inicio']}{marca}"

        if self._temporada_menu:
            try:
                self._temporada_menu.dismiss()
            except Exception:
                pass
        self._temporada_menu = MDDropdownMenu(
            caller=self.ids.temporada_item,
            items=[{"text": _etiqueta(t), "on_release": (lambda x=t: self.set_temporada_vista(x))}
                   for t in temporadas],
            width_mult=6,
        )
        abrir_menu(self._temporada_menu)

    def set_temporada_vista(self, temporada: dict):
        self._temporada_vista = temporada
        if self._temporada_menu:
            self._temporada_menu.dismiss()
        self.refrescar()

    def activar_temporada_vista(self):
        try:
            t = self._temporada_vista
            if not t:
                raise ValueError("Elegí una temporada primero.")
            activa = self._temporada_activa()
            if activa and activa["id"] == t["id"]:
                self.safe_snackbar(f"La temporada {t['nombre']} ya está activa.")
                return
            activar_temporada(t["id"])
            app = MDApp.get_running_app()
            if app:
                app.actualizar_encabezado()
            self.refrescar()
            self.safe_snackbar(f"Temporada {t['nombre']} activada.")
        except Exception as e:
            self.safe_error("Error activando Temporada", e, "ObjetivoScreen.activar_temporada_vista()")

    def abrir_editar_temporada(self):
        try:
            t = self._temporada_vista
            if not t:
                raise ValueError("Elegí una temporada primero.")
            nombre_input = MDTextField(hint_text="Nombre (ej: 2026-27)", text=t["nombre"],
                                       size_hint_y=None, height=dp(48))
            inicio_input = MDTextField(hint_text="Inicio (AAAA-MM-DD)", text=t["inicio"],
                                       size_hint_y=None, height=dp(48))
            fin_input = MDTextField(hint_text="Fin (opcional)", text=norm_text(t.get("fin", "")),
                                    size_hint_y=None, height=dp(48))
            content = MDBoxLayout(orientation="vertical", spacing=dp(12),
                                  padding=[dp(8), dp(12), dp(8), 0],
                                  size_hint_y=None, height=dp(190))
            content.add_widget(nombre_input)
            content.add_widget(inicio_input)
            content.add_widget(fin_input)
            dialog = None

            def _guardar(*_):
                try:
                    update_temporada(t["id"], nombre_input.text, inicio_input.text, fin_input.text)
                    self._temporada_vista = next(
                        (x for x in self._listar_temporadas() if x["id"] == t["id"]), None)
                    app = MDApp.get_running_app()
                    if app:
                        app.actualizar_encabezado()
                    self.refrescar()
                    if dialog:
                        dialog.dismiss()
                    self.safe_snackbar("Temporada actualizada.")
                except Exception as e:
                    self.safe_error("Error editando Temporada", e, "ObjetivoScreen.abrir_editar_temporada()")

            dialog = MDDialog(
                title="Editar temporada",
                type="custom",
                content_cls=content,
                md_bg_color=CARD_BG,
                buttons=[
                    MDFlatButton(text="CANCELAR", text_color=SAGE_GREEN,
                                 on_release=lambda *_: dialog.dismiss()),
                    MDFlatButton(text="GUARDAR", text_color=SAGE_GREEN, on_release=_guardar),
                ],
            )
            dialog.open()
        except Exception as e:
            self.safe_error("Error abriendo edicion de Temporada", e, "ObjetivoScreen.abrir_editar_temporada()")

    def refrescar(self):
        try:
            t = self._temporada_vista
            if t:
                activa = self._temporada_activa()
                es_activa = bool(activa and activa["id"] == t["id"])
                marca = "ACTIVA" if es_activa else "archivada — tocá ACTIVAR para retomarla"
                self.ids.temporada_item.text = t["nombre"]
                fin = norm_text(t.get("fin", "")) or "abierta"
                self.ids.temporada_label.text = f"{t['inicio']} a {fin} · {marca}"

                self.ids.objetivos_lista.clear_widgets()
                rows = list_plan(t["id"])
                if not rows:
                    self.ids.objetivos_lista.add_widget(
                        OneLineListItem(text="Sin cultivos: usá + AGREGAR CULTIVO"))
                    return
                for fila in rows:
                    cultivo, superficie_m2, esperado_kg = fila[0], float(fila[1]), float(fila[2])
                    plantas = int(fila[6]) if len(fila) > 6 else 0
                    kg_m2 = round(esperado_kg / superficie_m2, 2) if superficie_m2 > 0 else 0
                    texto = (
                        f"{cultivo} | {round(superficie_m2, 1)} m2 | "
                        f"{round(esperado_kg, 1)} kg (obj {kg_m2} kg/m2)"
                    )
                    if plantas > 0:
                        texto += f" | {plantas} pl"
                    item = OneLineListItem(text=texto)
                    item.bind(on_release=lambda _w, f=fila: self.abrir_acciones_plan(f))
                    self.ids.objetivos_lista.add_widget(item)
                return

            # Sin temporadas creadas: comportamiento legado
            self.ids.temporada_item.text = "Seleccionar"
            legacy = self.repo.get_temporada() or {}
            inicio = norm_text(legacy.get("inicio", ""))
            fin = norm_text(legacy.get("fin", ""))
            self.ids.temporada_label.text = f"Sin temporadas. {inicio or ''} {fin or ''}".strip()
            self.ids.objetivos_lista.clear_widgets()
            rows = self.repo.list_objetivos()
            if not rows:
                self.ids.objetivos_lista.add_widget(OneLineListItem(text="Sin objetivos cargados"))
                return
            for row in rows:
                cultivo = row[0]
                superficie_m2 = row[1] if len(row) > 1 else 0
                cosecha_esperada_kg = row[2] if len(row) > 2 else 0
                self.ids.objetivos_lista.add_widget(
                    OneLineListItem(
                        text=(
                            f"{cultivo} | {round(float(superficie_m2), 2)} m2 | "
                            f"Esperado: {round(float(cosecha_esperada_kg), 2)} kg"
                        )
                    )
                )
        except Exception as e:
            self.safe_error("Error listando Objetivos", e, "ObjetivoScreen.refrescar()")

    def abrir_perfil_cultivo(self):
        """Editor del perfil por cultivo: curva de cosecha y marco de plantación."""
        try:
            estado = {"cultivo": "", "tipo": TIPO_SIEMBRA_DIRECTA, "cosecha": COSECHA_ESCALONADA}
            usuario = get_perfil_usuario()
            bancal_m2 = usuario["bancal_m2"]
            ancho_bancal = usuario["ancho_bancal_m"]

            cultivo_btn = MDRaisedButton(
                text="SELECCIONAR CULTIVO",
                md_bg_color=SAGE_GREEN, text_color=(1, 1, 1, 1),
                size_hint=(1, None), height=dp(44),
            )
            tipo_btn = MDRaisedButton(
                text=f"Siembra: {estado['tipo']}",
                md_bg_color=WARM_AMBER, text_color=(1, 1, 1, 1),
                size_hint=(1, None), height=dp(40),
            )
            cosecha_btn = MDRaisedButton(
                text=f"Cosecha: {estado['cosecha']}",
                md_bg_color=SAGE_GREEN, text_color=(1, 1, 1, 1),
                size_hint=(1, None), height=dp(40),
            )
            alm_input = MDTextField(hint_text="Siembra a trasplante (días)",
                                    input_filter="int", size_hint_y=None, height=dp(48))
            tc_input = MDTextField(hint_text="Trasplante a cosecha (días)",
                                   input_filter="int", size_hint_y=None, height=dp(48))
            total_input = MDTextField(hint_text="Siembra a cosecha (días)",
                                      input_filter="int", size_hint_y=None, height=dp(48))
            ventana_input = MDTextField(hint_text="Ventana (días)",
                                        input_filter="int", size_hint_y=None, height=dp(48))
            rinde_input = MDTextField(hint_text="Rinde (kg/m²)",
                                      input_filter="float", size_hint_y=None, height=dp(48))
            distancia_input = MDTextField(hint_text="Dist. plantas (cm)",
                                          input_filter="float", size_hint_y=None, height=dp(48))
            lineas_input = MDTextField(hint_text="Líneas/bancal",
                                       input_filter="int", size_hint_y=None, height=dp(48))
            plantas_label = MDLabel(
                text="", font_style="Caption", bold=True,
                theme_text_color="Custom", text_color=SAGE_GREEN,
                size_hint_y=None, height=dp(20),
            )
            ayuda = MDLabel(
                text=("Concentrada: se cosecha junta (lechuga, zapallo). Escalonada: varias "
                      "recolecciones (tomate). Continua: hasta el fin de temporada (acelga, kale)."),
                font_style="Caption", theme_text_color="Secondary",
                size_hint_y=None, height=dp(44),
            )

            cosecha_menu = MDDropdownMenu(
                caller=cosecha_btn, width_mult=3,
                items=[{"text": t, "on_release": (lambda x=t: _sel_cosecha(x))}
                       for t in TIPOS_COSECHA],
            )

            def _sel_cosecha(t):
                estado["cosecha"] = t
                cosecha_btn.text = f"Cosecha: {t}"
                cosecha_menu.dismiss()

            cosecha_btn.bind(on_release=lambda *_: abrir_menu(cosecha_menu))

            def _aplicar_tipo():
                es_almacigo = estado["tipo"] == TIPO_SIEMBRA_ALMACIGO
                # En directa los campos de almácigo quedan deshabilitados y el
                # total es editable; en almácigo el total se calcula solo.
                alm_input.disabled = not es_almacigo
                tc_input.disabled = not es_almacigo
                total_input.disabled = es_almacigo
                if es_almacigo:
                    _recalc_total()

            def _recalc_total(*_):
                if estado["tipo"] != TIPO_SIEMBRA_ALMACIGO:
                    return
                try:
                    alm = int(norm_text(alm_input.text) or 0)
                except Exception:
                    alm = 0
                try:
                    tc = int(norm_text(tc_input.text) or 0)
                except Exception:
                    tc = 0
                total_input.text = str(alm + tc)

            def _recalc_plantas(*_):
                try:
                    dist = float(norm_text(distancia_input.text).replace(",", ".") or 0)
                except Exception:
                    dist = 0.0
                try:
                    lineas = int(norm_text(lineas_input.text) or 0)
                except Exception:
                    lineas = 0
                n = calcular_plantas(bancal_m2, dist, lineas, ancho_bancal)
                plantas_label.text = (
                    f"≈ {n} plantas por bancal ({bancal_m2:g} m²)" if n > 0 else ""
                )

            def _toggle_tipo(*_):
                estado["tipo"] = (
                    TIPO_SIEMBRA_ALMACIGO if estado["tipo"] == TIPO_SIEMBRA_DIRECTA
                    else TIPO_SIEMBRA_DIRECTA
                )
                tipo_btn.text = f"Siembra: {estado['tipo']}"
                _aplicar_tipo()

            def _sel_cultivo(c):
                estado["cultivo"] = c
                cultivo_btn.text = c
                p = get_perfil_cultivo(c)
                estado["tipo"] = p["tipo_siembra"]
                tipo_btn.text = f"Siembra: {estado['tipo']}"
                alm_input.text = str(p["dias_almacigo"])
                tc_input.text = str(p["dias_trasplante_cosecha"])
                total_input.text = str(p["dias_a_cosecha"])
                ventana_input.text = str(p["ventana"])
                rinde_input.text = f"{p['rinde_ref']:g}"
                distancia_input.text = f"{p['distancia_cm']:g}"
                lineas_input.text = str(p["lineas_bancal"])
                estado["cosecha"] = p.get("tipo_cosecha", COSECHA_ESCALONADA)
                cosecha_btn.text = f"Cosecha: {estado['cosecha']}"
                _aplicar_tipo()
                _recalc_plantas()

            cultivo_btn.bind(on_release=lambda *_: self.abrir_selector_cultivo(_sel_cultivo))
            tipo_btn.bind(on_release=_toggle_tipo)
            alm_input.bind(text=_recalc_total)
            tc_input.bind(text=_recalc_total)
            distancia_input.bind(text=_recalc_plantas)
            lineas_input.bind(text=_recalc_plantas)

            fila_curva = MDBoxLayout(orientation="horizontal", spacing=dp(8),
                                     size_hint_y=None, height=dp(52))
            fila_curva.add_widget(ventana_input)
            fila_curva.add_widget(rinde_input)

            fila_marco = MDBoxLayout(orientation="horizontal", spacing=dp(8),
                                     size_hint_y=None, height=dp(52))
            fila_marco.add_widget(distancia_input)
            fila_marco.add_widget(lineas_input)

            content = MDBoxLayout(orientation="vertical", spacing=dp(8),
                                  padding=[dp(8), dp(12), dp(8), 0],
                                  size_hint_y=None, height=dp(536))
            content.add_widget(cultivo_btn)
            content.add_widget(tipo_btn)
            content.add_widget(alm_input)
            content.add_widget(tc_input)
            content.add_widget(total_input)
            content.add_widget(cosecha_btn)
            content.add_widget(fila_curva)
            content.add_widget(fila_marco)
            content.add_widget(plantas_label)
            content.add_widget(ayuda)
            dialog = None
            _aplicar_tipo()

            def _guardar(*_):
                try:
                    cultivo = validate_cultivo(estado["cultivo"] or cultivo_btn.text)
                    tipo = estado["tipo"]
                    if tipo == TIPO_SIEMBRA_ALMACIGO:
                        alm = int(norm_text(alm_input.text) or 0)
                        tc = int(norm_text(tc_input.text) or 0)
                        if alm < 0 or tc < 1:
                            raise ValueError("Días de almácigo >= 0 y de trasplante a cosecha >= 1.")
                    else:
                        alm = 0
                        tc = int(norm_text(total_input.text) or 0)
                        if tc < 1:
                            raise ValueError("Días de siembra a cosecha >= 1.")
                    ventana = int(norm_text(ventana_input.text) or 0)
                    if ventana < 1:
                        raise ValueError("La ventana de cosecha debe ser >= 1 día.")
                    rinde = float(norm_text(rinde_input.text).replace(",", ".") or 0)
                    dist = float(norm_text(distancia_input.text).replace(",", ".") or 0)
                    lineas = int(norm_text(lineas_input.text) or 0)
                    upsert_perfil_cultivo(cultivo, tipo, alm, tc, ventana, rinde, dist, lineas,
                                          tipo_cosecha=estado["cosecha"])
                    if dialog:
                        dialog.dismiss()
                    self.safe_snackbar(f"Perfil actualizado: {cultivo}")
                except Exception as e:
                    self.safe_error("Error guardando Perfil", e, "ObjetivoScreen.abrir_perfil_cultivo()")

            dialog = MDDialog(
                title="Perfil de cultivo (curva y marco de plantación)",
                type="custom",
                content_cls=content,
                md_bg_color=CARD_BG,
                buttons=[
                    MDFlatButton(text="CANCELAR", text_color=SAGE_GREEN,
                                 on_release=lambda *_: dialog.dismiss()),
                    MDFlatButton(text="GUARDAR", text_color=SAGE_GREEN, on_release=_guardar),
                ],
            )
            dialog.open()
        except Exception as e:
            self.safe_error("Error abriendo Perfil de cultivo", e, "ObjetivoScreen.abrir_perfil_cultivo()")

    def abrir_integrantes(self):
        """Fichas de los integrantes del proyecto: ver, agregar, editar y quitar."""
        try:
            filas_box = MDBoxLayout(orientation="vertical", spacing=dp(6), size_hint_y=None)
            filas_box.bind(minimum_height=filas_box.setter("height"))
            scroll = ScrollView(size_hint=(1, None), height=dp(260))
            scroll.add_widget(filas_box)
            dialog = None

            def _cerrar(*_):
                try:
                    dialog.dismiss()
                except Exception:
                    pass

            def _reabrir():
                _cerrar()
                self.abrir_integrantes()

            def _ficha(i):
                titulo = MDLabel(text=i["nombre"], bold=True, valign="middle",
                                 theme_text_color="Custom", text_color=SAGE_GREEN)
                fila = MDBoxLayout(orientation="horizontal", size_hint_y=None,
                                   height=dp(48), spacing=dp(4))
                fila.add_widget(titulo)
                editar = MDIconButton(icon="pencil-outline", theme_text_color="Custom",
                                      text_color=SAGE_GREEN, pos_hint={"center_y": 0.5})
                editar.bind(on_release=lambda *_, x=i: (_cerrar(), self.abrir_form_integrante(x, _reabrir_tras_form)))
                quitar = MDIconButton(icon="delete-outline", theme_text_color="Custom",
                                      text_color=(0.7, 0.25, 0.25, 1), pos_hint={"center_y": 0.5})
                quitar.bind(on_release=lambda *_, x=i: (_cerrar(), self.confirmar_eliminar(
                    f"Integrante: {x['nombre']}", lambda: (_quitar(x)))))
                fila.add_widget(editar)
                fila.add_widget(quitar)
                return fila

            def _reabrir_tras_form():
                self.abrir_integrantes()

            def _quitar(i):
                try:
                    delete_integrante(i["id"])
                    self.safe_snackbar(f"Integrante quitado: {i['nombre']}")
                except Exception as e:
                    self.safe_error("Error quitando Integrante", e, "ObjetivoScreen.abrir_integrantes()")
                self.abrir_integrantes()

            integrantes = list_integrantes()
            if not integrantes:
                filas_box.add_widget(MDLabel(
                    text="Sin integrantes cargados todavía.",
                    font_style="Caption", theme_text_color="Secondary",
                    size_hint_y=None, height=dp(30)))
            for i in integrantes:
                filas_box.add_widget(_ficha(i))

            nuevo_btn = MDRaisedButton(
                text="+ NUEVO INTEGRANTE",
                md_bg_color=WARM_AMBER, text_color=(1, 1, 1, 1),
                size_hint=(1, None), height=dp(44),
            )
            nuevo_btn.bind(on_release=lambda *_: (_cerrar(), self.abrir_form_integrante(None, _reabrir_tras_form)))

            content = MDBoxLayout(orientation="vertical", spacing=dp(8),
                                  padding=[dp(4), dp(8), dp(4), 0],
                                  size_hint_y=None, height=dp(320))
            content.add_widget(scroll)
            content.add_widget(nuevo_btn)

            dialog = MDDialog(
                title="Integrantes del proyecto",
                type="custom",
                content_cls=content,
                md_bg_color=CARD_BG,
                buttons=[MDFlatButton(text="CERRAR", text_color=SAGE_GREEN, on_release=_cerrar)],
            )
            dialog.open()
        except Exception as e:
            self.safe_error("Error abriendo Integrantes", e, "ObjetivoScreen.abrir_integrantes()")

    def abrir_form_integrante(self, integrante=None, al_terminar=None):
        """Alta/edición de un integrante: SOLO el nombre.

        Estos nombres son los que después se eligen (lista desplegable) en Riego,
        Siembras, Trasplantes y Horas de Trabajo."""
        nombre_input = MDTextField(hint_text="Nombre",
                                   text=(integrante or {}).get("nombre", ""),
                                   size_hint_y=None, height=dp(48))
        content = MDBoxLayout(orientation="vertical", spacing=dp(8),
                              padding=[dp(8), dp(16), dp(8), 0],
                              size_hint_y=None, height=dp(96))
        content.add_widget(nombre_input)
        content.add_widget(MDLabel(
            text="Se usa en Riego, Siembras, Trasplantes y Horas.",
            font_style="Caption", theme_text_color="Secondary",
            size_hint_y=None, height=dp(20)))
        dialog = None

        def _guardar(*_):
            try:
                nombre = norm_text(nombre_input.text)
                if not nombre:
                    raise ValueError("El nombre es obligatorio.")
                if integrante:
                    update_integrante(integrante["id"], nombre, "", "", "", 0.0)
                else:
                    insert_integrante(nombre, "", "", "", 0.0)
                if dialog:
                    dialog.dismiss()
                self.safe_snackbar(f"Integrante guardado: {nombre}")
                if al_terminar:
                    al_terminar()
            except Exception as e:
                self.safe_error("Error guardando Integrante", e, "ObjetivoScreen.abrir_form_integrante()")

        dialog = MDDialog(
            title=("Editar integrante" if integrante else "Nuevo integrante"),
            type="custom",
            content_cls=content,
            md_bg_color=CARD_BG,
            buttons=[
                MDFlatButton(text="CANCELAR", text_color=SAGE_GREEN,
                             on_release=lambda *_: dialog.dismiss()),
                MDFlatButton(text="GUARDAR", text_color=SAGE_GREEN, on_release=_guardar),
            ],
        )
        dialog.open()

    def configurar_pin_admin(self):
        """Define, cambia o quita el PIN que protege esta pantalla."""
        accion = "cambiar" if admin_pin_definido() else "definir"

        def _ok(valor):
            valor = norm_text(valor)
            if not valor:
                if admin_pin_definido():
                    set_admin_pin("")
                    self.safe_snackbar("PIN quitado: la configuración queda sin protección.")
                return
            if len(valor) < 4:
                self.safe_snackbar("El PIN debe tener al menos 4 dígitos.")
                return
            set_admin_pin(valor)
            self.safe_snackbar("PIN de administrador guardado.")

        self.abrir_teclado_numerico(
            f"PIN de administrador ({accion}; OK vacío lo quita)",
            _ok, permitir_coma=False, oculto=True,
        )

    def abrir_sectores_riego(self):
        """Editar/borrar/reiniciar los sectores de riego de la temporada vista."""
        try:
            t = self._temporada_en_uso()
            if not t:
                self.safe_snackbar("Primero creá una temporada con NUEVA.")
                return

            filas = []
            box = MDBoxLayout(orientation="vertical", spacing=dp(4), size_hint_y=None)
            box.bind(minimum_height=box.setter("height"))
            scroll = ScrollView(size_hint=(1, None), height=dp(240))
            scroll.add_widget(box)

            def _reletra():
                for i, fila in enumerate(filas):
                    fila.set_letra(chr(ord("A") + i) if i < 26 else f"S{i + 1}")

            def _quitar(fila):
                if fila in filas:
                    filas.remove(fila)
                if fila in box.children:
                    box.remove_widget(fila)
                _reletra()

            def _agregar(bancales="", tipo=""):
                fila = SectorRiegoRow("?", on_eliminar=_quitar,
                                      bancales=bancales, tipo=tipo)
                filas.append(fila)
                box.add_widget(fila)
                _reletra()

            for _sector, bancales, tipo in list_sectores_riego(t["id"]):
                _agregar(str(bancales), tipo)
            if not filas:
                _agregar()

            agregar_btn = MDRaisedButton(
                text="+ AGREGAR SECTOR",
                md_bg_color=SAGE_GREEN, text_color=(1, 1, 1, 1),
                size_hint=(1, None), height=dp(40),
            )
            agregar_btn.bind(on_release=lambda *_: _agregar())

            content = MDBoxLayout(orientation="vertical", spacing=dp(8),
                                  padding=[dp(4), dp(8), dp(4), 0],
                                  size_hint_y=None, height=dp(300))
            content.add_widget(scroll)
            content.add_widget(agregar_btn)
            dialog = None

            def _reiniciar(*_):
                for fila in list(filas):
                    _quitar(fila)
                self.safe_snackbar("Sectores vaciados: GUARDAR confirma el reinicio.")

            def _guardar(*_):
                try:
                    datos = []
                    for fila in filas:
                        d = fila.get_datos()
                        if d:
                            datos.append(d)
                    save_sectores_riego(t["id"], datos)
                    if dialog:
                        dialog.dismiss()
                    self.safe_snackbar(f"Sectores de riego guardados: {len(datos)}.")
                except Exception as e:
                    self.safe_error("Error guardando Sectores", e, "ObjetivoScreen.abrir_sectores_riego()")

            dialog = MDDialog(
                title=f"Sectores de riego — {t['nombre']}",
                type="custom",
                content_cls=content,
                md_bg_color=CARD_BG,
                buttons=[
                    MDFlatButton(text="REINICIAR", text_color=(0.75, 0.18, 0.18, 1),
                                 on_release=_reiniciar),
                    MDFlatButton(text="CANCELAR", text_color=SAGE_GREEN,
                                 on_release=lambda *_: dialog.dismiss()),
                    MDFlatButton(text="GUARDAR", text_color=SAGE_GREEN, on_release=_guardar),
                ],
            )
            dialog.open()
        except Exception as e:
            self.safe_error("Error abriendo Sectores de riego", e, "ObjetivoScreen.abrir_sectores_riego()")

    def _bancal_m2(self) -> float:
        try:
            return float(get_perfil_usuario()["bancal_m2"])
        except Exception:
            return DEFAULT_BANCAL_M2

    def _temporada_en_uso(self):
        return self._temporada_vista or self._temporada_activa()

    def _ancho_bancal(self) -> float:
        try:
            return float(get_perfil_usuario()["ancho_bancal_m"])
        except Exception:
            return DEFAULT_ANCHO_BANCAL_M

    def abrir_agregar_cultivo(self):
        t = self._temporada_en_uso()
        if not t:
            self.safe_snackbar("Primero creá una temporada con NUEVA.")
            return

        def _confirmado(cultivo, m2, esperado, _cant, _uni, tipo, dist_cm, lineas, plantas, fechas):
            upsert_plan_item(t["id"], cultivo, m2, esperado,
                             tipo_siembra=tipo, distancia_cm=dist_cm, lineas=lineas, plantas=plantas)
            save_fechas_siembra_plan(t["id"], cultivo, fechas)
            try:
                self.repo.upsert_objetivo(cultivo, m2, esperado)
            except Exception:
                pass
            self.refrescar()
            self.safe_snackbar(f"{cultivo} agregado al plan de {t['nombre']}.")

        self.abrir_dialogo_plan_cultivo(
            self._bancal_m2(), _confirmado,
            titulo=f"Agregar cultivo — {t['nombre']}",
            ancho_bancal_m=self._ancho_bancal(),
        )

    def abrir_acciones_plan(self, fila):
        t = self._temporada_en_uso()
        if not t:
            return
        cultivo, superficie_m2, esperado_kg = fila[0], float(fila[1]), float(fila[2])
        desc = f"{cultivo} | {superficie_m2:g} m² | {esperado_kg:g} kg esperados"
        self.abrir_acciones_registro(
            desc,
            on_editar=lambda: self.abrir_editar_plan(t, fila),
            on_eliminar=lambda: self.eliminar_plan_item(t, cultivo),
        )

    def abrir_editar_plan(self, t: dict, fila):
        cultivo, superficie_m2, esperado_kg = fila[0], float(fila[1]), float(fila[2])
        tipo = norm_text(fila[3]) if len(fila) > 3 else ""
        dist = float(fila[4]) if len(fila) > 4 and fila[4] else 0.0
        lineas = int(fila[5]) if len(fila) > 5 and fila[5] else 0
        perfil = get_perfil_cultivo(cultivo)
        if not tipo:
            tipo = perfil["tipo_siembra"]
        if dist <= 0:
            dist = perfil["distancia_cm"]
        if lineas <= 0:
            lineas = perfil["lineas_bancal"]

        def _confirmado(c, m2, esperado, _cant, _uni, tipo_s, dist_cm, lin, plantas, fechas):
            upsert_plan_item(t["id"], c, m2, esperado,
                             tipo_siembra=tipo_s, distancia_cm=dist_cm, lineas=lin, plantas=plantas)
            save_fechas_siembra_plan(t["id"], c, fechas)
            try:
                self.repo.upsert_objetivo(c, m2, esperado)
            except Exception:
                pass
            self.refrescar()
            self.safe_snackbar(f"{c} actualizado.")

        self.abrir_dialogo_plan_cultivo(
            self._bancal_m2(), _confirmado,
            titulo=f"Editar {cultivo}",
            cultivo_fijo=cultivo,
            cantidad_inicial=f"{superficie_m2:g}",
            unidad_inicial="m2",
            esperado_inicial=f"{esperado_kg:g}",
            tipo_inicial=tipo,
            distancia_inicial=f"{dist:g}",
            lineas_inicial=str(lineas),
            ancho_bancal_m=self._ancho_bancal(),
            fechas_iniciales=get_fechas_siembra_plan(t["id"], cultivo),
        )

    def eliminar_plan_item(self, t: dict, cultivo: str):
        try:
            delete_plan_item(t["id"], cultivo)
            try:
                self.repo.delete_objetivo(cultivo)
            except Exception:
                pass
            self.refrescar()
            self.safe_snackbar(f"{cultivo} eliminado del plan.")
        except Exception as e:
            self.safe_error("Error eliminando cultivo del plan", e, "ObjetivoScreen.eliminar_plan_item()")

    def eliminar_temporada_vista(self):
        try:
            t = self._temporada_vista
            if not t:
                self.safe_snackbar("Elegí una temporada primero.")
                return
            desc = (
                f"Temporada {t['nombre']} ({t['inicio']} a {t.get('fin') or 'abierta'}).\n"
                "Se eliminará la planificación y TODOS sus registros: "
                "cosechas, riegos, siembras, tareas y stock."
            )

            def _ok():
                delete_temporada(t["id"])
                self._temporada_vista = self._temporada_activa()
                app = MDApp.get_running_app()
                if app:
                    app.actualizar_encabezado()
                self.refrescar()
                self.safe_snackbar(f"Temporada {t['nombre']} eliminada.")

            self.confirmar_eliminar(desc, _ok)
        except Exception as e:
            self.safe_error("Error eliminando Temporada", e, "ObjetivoScreen.eliminar_temporada_vista()")

    def confirmar_reiniciar_objetivos(self):
        dialog = None

        def _confirmar(*_):
            try:
                temporada = self._temporada_vista or self._temporada_activa()
                if temporada:
                    for cultivo, *_resto in list_plan(temporada["id"]):
                        delete_plan_item(temporada["id"], cultivo)
                self.repo.clear_objetivos()
                self.refrescar()
                self.safe_snackbar("Objetivos reiniciados.")
                if dialog:
                    dialog.dismiss()
            except Exception as e:
                self.safe_error("Error reiniciando Objetivos", e, "ObjetivoScreen.confirmar_reiniciar_objetivos()")

        dialog = MDDialog(
            title="Reiniciar objetivos",
            text="Se eliminaran todos los objetivos cargados. Esta accion no se puede deshacer.",
            md_bg_color=SAGE_GREEN,
            buttons=[
                MDFlatButton(
                    text="CANCELAR",
                    text_color=(1, 1, 1, 1),
                    on_release=lambda *_: dialog.dismiss(),
                ),
                MDFlatButton(
                    text="REINICIAR",
                    text_color=(1, 1, 1, 1),
                    on_release=_confirmar,
                ),
            ],
        )
        dialog.open()


class StockRow(MDBoxLayout):
    def __init__(self, bancal: int, on_cultivo_change=None, cultivos=None, **kwargs):
        super().__init__(orientation="horizontal", spacing=dp(4), size_hint_y=None, height=dp(40), **kwargs)
        self.bancal = bancal
        self._on_cultivo_change = on_cultivo_change

        self.bancal_label = MDLabel(text=str(bancal), size_hint_x=None, width=dp(34), valign="middle")
        # Evitamos MDDropDownItem en filas creadas en Python para prevenir errores de ids en KivyMD
        self.cultivo_item = MDFlatButton(text="Cultivo", size_hint_x=None, width=dp(120))
        self.cantidad_input = MDTextField(hint_text="Cantidad", size_hint_x=None, width=dp(80), input_filter="int")
        self.peso_input = MDTextField(hint_text="Kg/u", size_hint_x=None, width=dp(70), input_filter="float")
        self.kg_totales_label = MDLabel(
            text="0",
            size_hint_x=None,
            width=dp(70),
            halign="right",
            valign="middle",
            theme_text_color="Custom",
            text_color=SAGE_GREEN,
        )

        self.add_widget(self.bancal_label)
        self.add_widget(self.cultivo_item)
        self.add_widget(self.cantidad_input)
        self.add_widget(self.peso_input)
        self.add_widget(self.kg_totales_label)

        self._cultivo_value = ""
        opciones = cultivos or CULTIVOS
        self._cultivo_menu = MDDropdownMenu(
            caller=self.cultivo_item,
            items=[{
                "text": c,
                "on_release": (lambda x=c: self.set_cultivo(x))
            } for c in opciones],
            width_mult=6
        )
        self.cultivo_item.bind(on_release=lambda *_: self.open_cultivo_menu())
        self.cantidad_input.bind(text=lambda *_: self.update_kg_totales())
        self.peso_input.bind(text=lambda *_: self.update_kg_totales())
        self.bind(width=lambda *_: self.apply_responsive_layout())

        self.cantidad_input.line_color_focus = SAGE_GREEN
        self.peso_input.line_color_focus = SAGE_GREEN
        self.apply_responsive_layout()

    def apply_responsive_layout(self):
        row_width = self.width if self.width > dp(220) else max(dp(260), Window.width - dp(80))
        spacing_total = self.spacing * 4
        available = max(dp(240), row_width - spacing_total)

        mins = {
            "bancal": dp(26),
            "cultivo": dp(82),
            "cantidad": dp(58),
            "peso": dp(52),
            "kg": dp(54),
        }
        maxs = {
            "bancal": dp(36),
            "cultivo": dp(150),
            "cantidad": dp(95),
            "peso": dp(85),
            "kg": dp(80),
        }
        weights = {
            "bancal": 0.4,
            "cultivo": 4.0,
            "cantidad": 2.0,
            "peso": 1.6,
            "kg": 1.4,
        }
        total_min = sum(mins.values())
        extra = max(0.0, available - total_min)
        total_weight = sum(weights.values())

        widths = {}
        for key in mins:
            candidate = mins[key] + extra * (weights[key] / total_weight)
            widths[key] = min(maxs[key], candidate)

        self.bancal_label.width = widths["bancal"]
        self.cultivo_item.width = widths["cultivo"]
        self.cantidad_input.width = widths["cantidad"]
        self.peso_input.width = widths["peso"]
        self.kg_totales_label.width = widths["kg"]

    def open_cultivo_menu(self):
        if self._cultivo_menu:
            abrir_menu(self._cultivo_menu)

    def set_cultivo(self, cultivo: str, notify: bool = True):
        self._cultivo_value = cultivo
        self.cultivo_item.text = cultivo
        if hasattr(self.cultivo_item, "set_item"):
            self.cultivo_item.set_item(cultivo)
        if self._cultivo_menu:
            self._cultivo_menu.dismiss()
        if notify and self._on_cultivo_change:
            self._on_cultivo_change(self.bancal, cultivo)

    def update_kg_totales(self):
        try:
            cantidad = float(norm_text(self.cantidad_input.text).replace(",", ".") or 0)
            peso = float(norm_text(self.peso_input.text).replace(",", ".") or 0)
            kg_totales = round(cantidad * peso, 3) if peso > 0 else 0
            self.kg_totales_label.text = str(kg_totales)
        except Exception:
            self.kg_totales_label.text = "0"

    def set_raw(self, cultivo: str, cantidad_txt: str, peso_txt: str):
        if cultivo:
            self.set_cultivo(cultivo, notify=False)
        self.cantidad_input.text = cantidad_txt or ""
        self.peso_input.text = peso_txt or ""
        self.update_kg_totales()

    def get_raw(self):
        return {
            "bancal": self.bancal,
            "cultivo": self._cultivo_value,
            "cantidad_txt": norm_text(self.cantidad_input.text),
            "peso_txt": norm_text(self.peso_input.text),
        }

    def get_data(self):
        cultivo = self._cultivo_value
        cantidad_txt = norm_text(self.cantidad_input.text)
        peso_txt = norm_text(self.peso_input.text)
        if not cultivo and not cantidad_txt and not peso_txt:
            return None
        cultivo = validate_cultivo(cultivo)
        peso = validate_peso_unit(peso_txt)
        unidades = validate_unidades(cantidad_txt)
        kg = round(unidades * peso, 3)
        return {
            "bancal": self.bancal,
            "cultivo": cultivo,
            "kg": kg,
            "peso": peso,
            "unidades": unidades,
        }


TIPOS_BANDEJA = [72, 98, 128, 162]


class SiembrasScreen(BaseScreen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self._plan_map = {}
        self._semillas_map = {}
        self._cultivo_menu = None
        self._cultivo_value = ""
        self._sector_menu = None
        self._sector_value = ""
        self._bancal_menu = None
        self._bancal_value = ""
        self._bancales_por_sector = {}
        self._tipo_menu = None
        self._tipo_value = TIPOS_SIEMBRA[0]
        self._tipo_bandeja_menu = None
        self._tipo_bandeja_val = 72
        self._generacion_val = 1
        self._bandejas_val = 1
        self._ver_todos = False

    def on_pre_enter(self):
        self.refresh_backend_status()
        self.preparar_menu_operador()
        self._cargar_plan()
        # El menú de cultivos se arma con el plan de la temporada activa;
        # si todavía no hay plan, se ofrece la lista completa.
        cultivos_menu = sorted(self._plan_map.keys()) if self._plan_map else get_cultivos()
        if self._cultivo_menu:
            try:
                self._cultivo_menu.dismiss()
            except Exception:
                pass
        self._cultivo_menu = MDDropdownMenu(
            caller=self.ids.cultivo_item,
            items=[{"text": c, "on_release": (lambda x=c: self._set_cultivo(x))}
                   for c in cultivos_menu],
            width_mult=5,
        )
        self._actualizar_info_plan()
        # Sectores de la temporada para siembra directa
        self._bancales_por_sector = {}
        sectores = []
        try:
            temporada = get_temporada_activa()
            if temporada and self.modo_local():
                for sec, bancales, _t in list_sectores_riego(temporada["id"]):
                    sectores.append(sec)
                    self._bancales_por_sector[sec] = int(bancales)
        except Exception:
            pass
        if not sectores:
            sectores = SECTORES
        self._sector_menu = MDDropdownMenu(
            caller=self.ids.sector_item,
            items=[{"text": s, "on_release": (lambda x=s: self._set_sector(x))}
                   for s in sectores],
            width_mult=2,
        )
        if not self._tipo_menu:
            self._tipo_menu = MDDropdownMenu(
                caller=self.ids.tipo_item,
                items=[{"text": t, "on_release": (lambda x=t: self._set_tipo(x))}
                       for t in TIPOS_SIEMBRA],
                width_mult=5,
            )
        if not self._tipo_bandeja_menu:
            self._tipo_bandeja_menu = MDDropdownMenu(
                caller=self.ids.tipo_bandeja_item,
                items=[{"text": str(v), "on_release": (lambda x=v: self._set_tipo_bandeja(x))}
                       for v in TIPOS_BANDEJA],
                width_mult=3,
            )
        self.ids.generacion.text = str(self._generacion_val)
        self.ids.bandejas.text = str(self._bandejas_val)
        self._update_semillas()
        self.refrescar()

    def hoy(self):
        return date.today().isoformat()

    def ir_home(self):
        self.manager.current = "home"

    def open_cultivo_menu(self):
        if self._cultivo_menu:
            abrir_menu(self._cultivo_menu)

    def open_tipo_menu(self):
        if self._tipo_menu:
            abrir_menu(self._tipo_menu)

    def open_sector_menu(self):
        if self._sector_menu:
            abrir_menu(self._sector_menu)

    def open_bancal_menu(self):
        if not self._sector_value:
            self.safe_snackbar("Elegí primero el sector.")
            return
        if self._bancal_menu:
            abrir_menu(self._bancal_menu)

    def _set_sector(self, sector: str):
        self._sector_value = sector
        self.ids.sector_item.text = sector
        self.ids.sector_item.set_item(sector)
        if self._sector_menu:
            self._sector_menu.dismiss()
        maximo = self._bancales_por_sector.get(sector, BANCAL_MAX)
        opciones = [str(i) for i in range(1, max(1, maximo) + 1)]
        self._bancal_menu = MDDropdownMenu(
            caller=self.ids.bancal_item,
            items=[{"text": b, "on_release": (lambda x=b: self._set_bancal(x))}
                   for b in opciones],
            width_mult=2,
        )
        if self._bancal_value not in opciones:
            self._set_bancal(opciones[0])

    def _set_bancal(self, bancal: str):
        self._bancal_value = bancal
        self.ids.bancal_item.text = bancal
        self.ids.bancal_item.set_item(bancal)
        if self._bancal_menu:
            self._bancal_menu.dismiss()

    def _cargar_plan(self):
        """Plan de la temporada activa: tipo de siembra y plantas objetivo por cultivo."""
        self._plan_map = {}
        self._semillas_map = {}
        try:
            temporada = get_temporada_activa()
            if not temporada or not self.modo_local():
                return
            for fila in list_plan(temporada["id"]):
                cultivo = fila[0]
                tipo = norm_text(fila[3]) if len(fila) > 3 else ""
                plantas = int(fila[6]) if len(fila) > 6 and fila[6] else 0
                if not tipo:
                    tipo = get_perfil_cultivo(cultivo)["tipo_siembra"]
                self._plan_map[cultivo] = {"tipo": tipo, "plantas": plantas}
            self._semillas_map = semillas_por_cultivo(temporada["id"])
        except Exception:
            pass

    def _actualizar_info_plan(self):
        label = self.ids.get("plan_info")
        if label is None:
            return
        cultivo = self._cultivo_value
        datos = self._plan_map.get(cultivo)
        if not datos:
            label.text = ""
            return
        objetivo = datos["plantas"]
        sembradas = self._semillas_map.get(cultivo, 0)
        if objetivo > 0:
            pct = round(sembradas / objetivo * 100.0, 1)
            faltan = max(0, objetivo - sembradas)
            label.text = f"Plan: {datos['tipo']} · objetivo {objetivo} pl · sembradas {sembradas} ({pct:g}%) · faltan {faltan}"
        else:
            label.text = f"Plan: {datos['tipo']} · sin plantas objetivo (editá el cultivo en el plan)"

    def _set_cultivo(self, cultivo: str):
        self._cultivo_value = cultivo
        self.ids.cultivo_item.text = cultivo
        self.ids.cultivo_item.set_item(cultivo)
        if self._cultivo_menu:
            self._cultivo_menu.dismiss()
        # Tipo de siembra automático según el plan/perfil (el usuario puede cambiarlo)
        datos = self._plan_map.get(cultivo)
        tipo_plan = datos["tipo"] if datos else get_perfil_cultivo(cultivo)["tipo_siembra"]
        self._set_tipo("Siembra almácigo" if tipo_plan == TIPO_SIEMBRA_ALMACIGO else "Siembra directa")
        self._actualizar_info_plan()

    def _set_tipo(self, tipo: str):
        self._tipo_value = tipo
        self.ids.tipo_item.text = tipo
        self.ids.tipo_item.set_item(tipo)
        if self._tipo_menu:
            self._tipo_menu.dismiss()

    def open_tipo_bandeja_menu(self):
        if self._tipo_bandeja_menu:
            abrir_menu(self._tipo_bandeja_menu)

    def _set_tipo_bandeja(self, valor: int):
        self._tipo_bandeja_val = valor
        self.ids.tipo_bandeja_item.text = str(valor)
        self.ids.tipo_bandeja_item.set_item(str(valor))
        if self._tipo_bandeja_menu:
            self._tipo_bandeja_menu.dismiss()
        self._update_semillas()

    def _update_semillas(self):
        total = self._bandejas_val * self._tipo_bandeja_val
        self.ids.semillas_label.text = f"= {total} semillas"

    def inc_generacion(self):
        self._generacion_val += 1
        self.ids.generacion.text = str(self._generacion_val)

    def dec_generacion(self):
        if self._generacion_val > 1:
            self._generacion_val -= 1
            self.ids.generacion.text = str(self._generacion_val)

    def inc_bandejas(self):
        self._bandejas_val += 1
        self.ids.bandejas.text = str(self._bandejas_val)
        self._update_semillas()

    def dec_bandejas(self):
        if self._bandejas_val > 1:
            self._bandejas_val -= 1
            self.ids.bandejas.text = str(self._bandejas_val)
            self._update_semillas()

    def guardar(self):
        try:
            fecha = validate_fecha(self.ids.fecha.text)
            cultivo = validate_cultivo(self._cultivo_value or self.ids.cultivo_item.text)
            variedad = norm_text(self.ids.variedad.text)
            tipo = validate_tipo_siembra(self._tipo_value or self.ids.tipo_item.text)
            generacion = self._generacion_val
            bandejas = self._bandejas_val
            tipo_bandeja = self._tipo_bandeja_val
            operador = norm_text(self._operador_value)
            observaciones = norm_text(self.ids.observaciones.text)

            # En siembra directa la ubicación es obligatoria (va directo al bancal)
            sector, bancal = "", 0
            if tipo == "Siembra directa":
                if not self._sector_value or not self._bancal_value:
                    raise ValueError("Siembra directa: elegí sector y bancal.")
                sector = validate_sector(self._sector_value)
                bancal = int(self._bancal_value)
            elif self._sector_value and self._bancal_value:
                sector = validate_sector(self._sector_value)
                bancal = int(self._bancal_value)

            self.repo.insert_siembra(fecha, cultivo, variedad, tipo, generacion, bandejas,
                                     tipo_bandeja, observaciones, operador, sector, bancal)
            semillas = bandejas * tipo_bandeja
            self.safe_snackbar(f"Guardado: {cultivo} | {tipo} | {bandejas}×{tipo_bandeja}={semillas} sem.")
            self.ids.variedad.text = ""
            self.ids.observaciones.text = ""
            self._cargar_plan()
            self._actualizar_info_plan()
            self.refrescar()
        except Exception as e:
            self.safe_error("Error al guardar Siembra", e, "SiembrasScreen.guardar()")

    def resumen_del_dia(self):
        try:
            fecha = norm_text(self.ids.fecha.text) or self.hoy()
            rows = self.repo.list_siembras_by_fecha(fecha)
            if not rows:
                raise ValueError(f"No hay registros para {fecha}.")
            lines = [f"Resumen {fecha}:", ""]
            for cultivo, variedad, tipo, gen, bandejas, obs, operador in rows:
                desc = f"• {cultivo}"
                if variedad:
                    desc += f" ({variedad})"
                desc += f" | {tipo} | Gen {gen} | {bandejas} bandeja(s)"
                if operador:
                    desc += f" | Op: {operador}"
                lines.append(desc)
            dialog = MDDialog(
                title="Resumen del día",
                text="\n".join(lines),
                md_bg_color=SAGE_GREEN,
                buttons=[MDFlatButton(text="OK", text_color=(1, 1, 1, 1),
                                      on_release=lambda *_: dialog.dismiss())],
            )
            dialog.open()
        except Exception as e:
            self.safe_error("Error al generar Resumen", e, "SiembrasScreen.resumen_del_dia()")

    def abrir_descarga(self):
        fecha_desde = MDTextField(hint_text="Desde: AAAA-MM-DD", text=self.hoy(), size_hint_y=None, height=dp(48))
        fecha_hasta = MDTextField(hint_text="Hasta: AAAA-MM-DD", text=self.hoy(), size_hint_y=None, height=dp(48))
        content = MDBoxLayout(orientation="vertical", adaptive_height=True,
                              spacing=dp(8), padding=[dp(8), dp(6), dp(8), dp(0)])
        content.add_widget(fecha_desde)
        content.add_widget(fecha_hasta)
        dialog = None

        def _descargar(*_):
            try:
                desde = validate_fecha(fecha_desde.text)
                hasta = validate_fecha(fecha_hasta.text)
                if desde > hasta:
                    raise ValueError("La fecha 'desde' no puede ser mayor que 'hasta'.")
                self.choose_export_dir(lambda out_dir: self._descargar_csv(desde, hasta, out_dir))
                if dialog:
                    dialog.dismiss()
            except Exception as e:
                self.safe_error("Error al descargar", e, "SiembrasScreen.abrir_descarga()")

        dialog = MDDialog(
            title="Descargar siembras por intervalo",
            type="custom",
            content_cls=content,
            md_bg_color=SAGE_GREEN,
            buttons=[
                MDFlatButton(text="CANCELAR", text_color=(1, 1, 1, 1),
                             on_release=lambda *_: dialog.dismiss()),
                MDFlatButton(text="DESCARGAR", text_color=(1, 1, 1, 1),
                             on_release=_descargar),
            ],
        )
        dialog.open()

    def _descargar_csv(self, fecha_desde: str, fecha_hasta: str, out_dir: Path):
        try:
            rows = self.repo.list_siembras_between(fecha_desde, fecha_hasta)
            if not rows:
                raise ValueError("No hay registros en ese intervalo.")
            path = out_dir / f"siembras_{fecha_desde}_{fecha_hasta}.csv"
            with path.open("w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["fecha", "cultivo", "variedad", "tipo", "generacion",
                                  "bandejas", "observaciones", "operador"])
                writer.writerows(rows)
            self.safe_snackbar(f"Descarga generada: {path}")
        except Exception as e:
            self.safe_error("Error al generar descarga", e, "SiembrasScreen._descargar_csv()")

    def toggle_ver_todos(self):
        self._ver_todos = not self._ver_todos
        self.ids.ver_btn.text = "VER HOY" if self._ver_todos else "VER REGISTROS"
        self.ids.modo_label.text = "Mostrando: Todos" if self._ver_todos else "Mostrando: Hoy"
        self.refrescar()

    def refrescar(self):
        try:
            self.ids.lista.clear_widgets()
            if self.modo_local():
                if self._ver_todos:
                    rows = self.repo.list_siembras_con_id(limit=200)
                else:
                    fecha = norm_text(self.ids.fecha.text) or self.hoy()
                    rows = self.repo.list_siembras_con_id(fecha=fecha)
                if not rows:
                    self.ids.lista.add_widget(OneLineListItem(text="Sin registros en esta temporada"))
                    return
                for fila in rows:
                    rid, f, cultivo, variedad, tipo, gen, bandejas = fila[:7]
                    tipo_bandeja = int(fila[7]) if len(fila) > 7 and fila[7] else 72
                    texto = f"{f} | {cultivo}" if self._ver_todos else f"{cultivo}"
                    if variedad:
                        texto += f" ({variedad})"
                    texto += f" | {tipo} | Gen {gen} | {bandejas}×{tipo_bandeja}={bandejas * tipo_bandeja} sem."
                    sec = norm_text(str(fila[10])) if len(fila) > 10 and fila[10] else ""
                    if sec:
                        texto += f" | {sec}-{fila[11]}"
                    item = OneLineListItem(text=texto)
                    item.bind(on_release=lambda _w, r=fila, t=texto: self.abrir_acciones_registro(
                        t,
                        on_editar=lambda: self.abrir_editar_siembra(r),
                        on_eliminar=lambda: self.eliminar_siembra(r[0])))
                    self.ids.lista.add_widget(item)
                return

            if self._ver_todos:
                rows = self.repo.list_siembras_all(limit=200)
                if not rows:
                    self.ids.lista.add_widget(OneLineListItem(text="Sin registros"))
                    return
                for fecha, cultivo, variedad, tipo, gen, bandejas, _obs, _op in rows:
                    texto = f"{fecha} | {cultivo}"
                    if variedad:
                        texto += f" ({variedad})"
                    texto += f" | {tipo} | Gen {gen} | {bandejas}b"
                    self.ids.lista.add_widget(OneLineListItem(text=texto))
            else:
                fecha = norm_text(self.ids.fecha.text) or self.hoy()
                rows = self.repo.list_siembras_by_fecha(fecha)
                if not rows:
                    self.ids.lista.add_widget(OneLineListItem(text="Sin registros"))
                    return
                for cultivo, variedad, tipo, gen, bandejas, _obs, _op in rows:
                    texto = f"{cultivo}"
                    if variedad:
                        texto += f" ({variedad})"
                    texto += f" | {tipo} | Gen {gen} | {bandejas}b"
                    self.ids.lista.add_widget(OneLineListItem(text=texto))
        except Exception as e:
            self.safe_error("Error listando Siembras", e, "SiembrasScreen.refrescar()")

    def eliminar_siembra(self, siembra_id: int):
        try:
            self.repo.delete_siembra(siembra_id)
            self.safe_snackbar("Siembra eliminada.")
            self._cargar_plan()
            self._actualizar_info_plan()
            self.refrescar()
        except Exception as e:
            self.safe_error("Error al eliminar Siembra", e, "SiembrasScreen.eliminar_siembra()")

    def abrir_editar_siembra(self, fila):
        rid, f, cultivo, variedad, tipo, gen, bandejas = fila[:7]
        tipo_bandeja = int(fila[7]) if len(fila) > 7 and fila[7] else 72
        observaciones = str(fila[8]) if len(fila) > 8 and fila[8] else ""
        operador = str(fila[9]) if len(fila) > 9 and fila[9] else ""

        estado = {"tipo_bandeja": tipo_bandeja}
        fecha_input = MDTextField(hint_text="Fecha (AAAA-MM-DD)", text=str(f),
                                  size_hint_y=None, height=dp(48))
        variedad_input = MDTextField(hint_text="Variedad", text=str(variedad or ""),
                                     size_hint_y=None, height=dp(48))
        gen_input = MDTextField(hint_text="Generación", text=str(gen), input_filter="int",
                                size_hint_y=None, height=dp(48))
        bandejas_input = MDTextField(hint_text="Bandejas", text=str(bandejas), input_filter="int",
                                     size_hint_y=None, height=dp(48))
        bandeja_btn = MDRaisedButton(
            text=f"Tipo bandeja: {tipo_bandeja}",
            md_bg_color=WARM_AMBER, text_color=(1, 1, 1, 1),
            size_hint=(1, None), height=dp(40),
        )
        semillas_label = MDLabel(text="", font_style="Caption", bold=True,
                                 theme_text_color="Custom", text_color=SAGE_GREEN,
                                 size_hint_y=None, height=dp(18))
        operador_input = MDTextField(hint_text="Operador", text=operador,
                                     size_hint_y=None, height=dp(48))
        obs_input = MDTextField(hint_text="Observaciones", text=observaciones,
                                size_hint_y=None, height=dp(48))

        def _recalc_semillas(*_):
            try:
                b = int(norm_text(bandejas_input.text) or 0)
            except Exception:
                b = 0
            semillas_label.text = f"= {b * estado['tipo_bandeja']} semillas" if b > 0 else ""

        def _sel_bandeja(v):
            estado["tipo_bandeja"] = v
            bandeja_btn.text = f"Tipo bandeja: {v}"
            bandeja_menu.dismiss()
            _recalc_semillas()

        bandeja_menu = MDDropdownMenu(
            caller=bandeja_btn,
            items=[{"text": str(v), "on_release": (lambda x=v: _sel_bandeja(x))}
                   for v in TIPOS_BANDEJA],
            width_mult=3,
        )
        bandeja_btn.bind(on_release=lambda *_: abrir_menu(bandeja_menu))
        bandejas_input.bind(text=_recalc_semillas)
        _recalc_semillas()

        fila_gen = MDBoxLayout(orientation="horizontal", spacing=dp(8),
                               size_hint_y=None, height=dp(52))
        fila_gen.add_widget(gen_input)
        fila_gen.add_widget(bandejas_input)

        content = MDBoxLayout(orientation="vertical", spacing=dp(8),
                              padding=[dp(8), dp(12), dp(8), 0],
                              size_hint_y=None, height=dp(360))
        content.add_widget(MDLabel(text=f"{cultivo} | {tipo}", bold=True, halign="center",
                                   theme_text_color="Custom", text_color=SAGE_GREEN,
                                   size_hint_y=None, height=dp(24)))
        content.add_widget(fecha_input)
        content.add_widget(variedad_input)
        content.add_widget(fila_gen)
        content.add_widget(bandeja_btn)
        content.add_widget(semillas_label)
        content.add_widget(operador_input)
        content.add_widget(obs_input)
        dialog = None

        def _guardar(*_):
            try:
                nueva_fecha = validate_fecha(fecha_input.text)
                nueva_gen = validate_generacion(gen_input.text)
                nuevas_bandejas = validate_bandejas(bandejas_input.text)
                self.repo.update_siembra(
                    rid, nueva_fecha, norm_text(variedad_input.text), nueva_gen,
                    nuevas_bandejas, estado["tipo_bandeja"],
                    norm_text(obs_input.text), norm_text(operador_input.text),
                )
                if dialog:
                    dialog.dismiss()
                self.safe_snackbar("Siembra actualizada.")
                self._cargar_plan()
                self._actualizar_info_plan()
                self.refrescar()
            except Exception as e:
                self.safe_error("Error al editar Siembra", e, "SiembrasScreen.abrir_editar_siembra()")

        dialog = MDDialog(
            title="Editar siembra",
            type="custom",
            content_cls=content,
            md_bg_color=CARD_BG,
            buttons=[
                MDFlatButton(text="CANCELAR", text_color=SAGE_GREEN,
                             on_release=lambda *_: dialog.dismiss()),
                MDFlatButton(text="GUARDAR", text_color=SAGE_GREEN, on_release=_guardar),
            ],
        )
        dialog.open()


class CosechasLogroScreen(BaseScreen):
    """Relojes de cosecha por cultivo: real vs esperado a la fecha, con semáforo."""

    def on_pre_enter(self):
        self.armar()

    def ir_home(self):
        self.manager.current = "home"

    def _fila_label(self, texto, style="Body2", color=None, height=22, bold=False):
        lbl = MDLabel(text=texto, font_style=style, size_hint_y=None, height=dp(height), bold=bold)
        if color:
            lbl.theme_text_color = "Custom"
            lbl.text_color = color
        return lbl

    def armar(self):
        box = self.ids.get("logro_box")
        if box is None:
            return
        box.clear_widgets()
        try:
            temporada = get_temporada_activa()
        except Exception:
            temporada = None
        if not temporada:
            card = Factory.SectionCard()
            card.add_widget(self._fila_label("No hay temporada activa.", height=24))
            box.add_widget(card)
            return

        hoy = date.today().isoformat()
        resumenes = self.resumenes_cosecha_temporada(temporada, hoy)

        if not resumenes:
            card = Factory.SectionCard()
            card.add_widget(self._fila_label("El plan no tiene cultivos todavía.", "Caption", height=20))
            box.add_widget(card)
            return

        # Una tarjeta por cultivo, con su gráfico de desarrollo + cosecha
        for r in resumenes:
            card = Factory.SectionCard()
            color = SEMAFORO_COLORES.get(r["estado"], SAGE_GREEN)
            card.add_widget(self._fila_label(
                r["cultivo"], "Subtitle2", SAGE_GREEN, height=24, bold=True))

            # Avance sobre el TOTAL del plan (sin prorratear por fecha)
            pct_temporada = (round(r["kg_real"] / r["esperado_total"] * 100.0, 1)
                             if r["esperado_total"] > 0 else 0.0)
            card.add_widget(self._fila_label(
                f"{r['kg_real']:g} / {r['esperado_total']:g} kg ({pct_temporada:g}%)",
                "Body2", height=22, bold=True))
            card.add_widget(ProgressBar(max=100, value=min(100.0, pct_temporada),
                                        size_hint_y=None, height=dp(5)))
            card.add_widget(self._fila_label(
                f"{r['kg_m2_real']:g} kg/m² (objetivo {r['kg_m2_obj']:g})",
                "Caption", height=18))
            box.add_widget(card)


class TrasplantesScreen(BaseScreen):
    """Registro de trasplantes: un registro por bancal trasplantado."""

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self._nombre_menu = None
        self._nombre_value = ""
        self._cultivo_menu = None
        self._cultivo_value = ""
        self._sector_menu = None
        self._sector_value = ""
        self._bancal_menu = None
        self._bancal_value = "1"
        self._generacion_val = 1
        self._bancales_por_sector = {}

    def hoy(self):
        return date.today().isoformat()

    def ir_home(self):
        self.manager.current = "home"

    def on_pre_enter(self):
        try:
            temporada = get_temporada_activa()

            nombres = [i["nombre"] for i in list_integrantes()]
            self._nombre_menu = MDDropdownMenu(
                caller=self.ids.nombre_item,
                items=[{"text": n, "on_release": (lambda x=n: self._set_nombre(x))}
                       for n in nombres],
                width_mult=5,
            )
            if self._nombre_value not in nombres:
                self._nombre_value = nombres[0] if nombres else ""
                self.ids.nombre_item.text = self._nombre_value or "Seleccionar"

            # Cultivos de almácigo del plan (los que se trasplantan)
            cultivos = []
            if temporada and self.modo_local():
                for fila in list_plan(temporada["id"]):
                    tipo = norm_text(fila[3]) if len(fila) > 3 else ""
                    if not tipo:
                        tipo = get_perfil_cultivo(fila[0])["tipo_siembra"]
                    if tipo == TIPO_SIEMBRA_ALMACIGO:
                        cultivos.append(fila[0])
            if not cultivos:
                cultivos = get_cultivos()
            self._cultivo_menu = MDDropdownMenu(
                caller=self.ids.cultivo_item,
                items=[{"text": c, "on_release": (lambda x=c: self._set_cultivo(x))}
                       for c in cultivos],
                width_mult=5,
            )
            if self._cultivo_value not in cultivos:
                self._cultivo_value = cultivos[0]
                self.ids.cultivo_item.text = self._cultivo_value

            # Sectores de la temporada y bancales por sector
            self._bancales_por_sector = {}
            sectores = []
            if temporada and self.modo_local():
                for sector, bancales, _tipo in list_sectores_riego(temporada["id"]):
                    sectores.append(sector)
                    self._bancales_por_sector[sector] = int(bancales)
            if not sectores:
                sectores = SECTORES
            self._sector_menu = MDDropdownMenu(
                caller=self.ids.sector_item,
                items=[{"text": s, "on_release": (lambda x=s: self._set_sector(x))}
                       for s in sectores],
                width_mult=2,
            )
            if self._sector_value not in sectores:
                self._set_sector(sectores[0])
            else:
                self._rebuild_bancal_menu()

            self.ids.generacion.text = str(self._generacion_val)
            self.refrescar()
        except Exception as e:
            self.safe_error("Error al preparar Trasplantes", e, "TrasplantesScreen.on_pre_enter()")

    def open_nombre_menu(self):
        if not list_integrantes():
            self.safe_snackbar("Cargá los integrantes en Configuración de Temporada → INTEGRANTES.")
            return
        if self._nombre_menu:
            abrir_menu(self._nombre_menu)

    def open_cultivo_menu(self):
        if self._cultivo_menu:
            abrir_menu(self._cultivo_menu)

    def open_sector_menu(self):
        if self._sector_menu:
            abrir_menu(self._sector_menu)

    def open_bancal_menu(self):
        if self._bancal_menu:
            abrir_menu(self._bancal_menu)

    def _set_nombre(self, nombre):
        self._nombre_value = nombre
        self.ids.nombre_item.text = nombre
        self.ids.nombre_item.set_item(nombre)
        if self._nombre_menu:
            self._nombre_menu.dismiss()

    def _set_cultivo(self, cultivo):
        self._cultivo_value = cultivo
        self.ids.cultivo_item.text = cultivo
        self.ids.cultivo_item.set_item(cultivo)
        if self._cultivo_menu:
            self._cultivo_menu.dismiss()

    def _set_sector(self, sector):
        self._sector_value = sector
        self.ids.sector_item.text = sector
        self.ids.sector_item.set_item(sector)
        if self._sector_menu:
            self._sector_menu.dismiss()
        self._rebuild_bancal_menu()

    def _rebuild_bancal_menu(self):
        maximo = self._bancales_por_sector.get(self._sector_value, BANCAL_MAX)
        opciones = [str(i) for i in range(1, max(1, maximo) + 1)]
        self._bancal_menu = MDDropdownMenu(
            caller=self.ids.bancal_item,
            items=[{"text": b, "on_release": (lambda x=b: self._set_bancal(x))}
                   for b in opciones],
            width_mult=2,
        )
        if self._bancal_value not in opciones:
            self._set_bancal(opciones[0])

    def _set_bancal(self, bancal):
        self._bancal_value = bancal
        self.ids.bancal_item.text = bancal
        self.ids.bancal_item.set_item(bancal)
        if self._bancal_menu:
            self._bancal_menu.dismiss()

    def inc_generacion(self):
        self._generacion_val += 1
        self.ids.generacion.text = str(self._generacion_val)

    def dec_generacion(self):
        if self._generacion_val > 1:
            self._generacion_val -= 1
            self.ids.generacion.text = str(self._generacion_val)

    def guardar(self):
        try:
            fecha = validate_fecha(self.ids.fecha.text)
            if not self._nombre_value:
                raise ValueError("Elegí el nombre del integrante.")
            cultivo = validate_cultivo(self._cultivo_value)
            sector = validate_sector(self._sector_value)
            bancal = int(self._bancal_value)
            insert_trasplante(fecha, self._nombre_value, cultivo, self._generacion_val, sector, bancal)
            plantas = plantas_por_bancal_cultivo(cultivo)
            extra = f" (≈{plantas} plantas)" if plantas > 0 else ""
            self.safe_snackbar(f"Trasplante: {cultivo} Gen {self._generacion_val} en {sector}-{bancal}{extra}")
            self.refrescar()
        except Exception as e:
            self.safe_error("Error al guardar Trasplante", e, "TrasplantesScreen.guardar()")

    def refrescar(self):
        try:
            self.ids.lista.clear_widgets()
            rows = list_trasplantes_con_id(limit=200)
            if not rows:
                self.ids.lista.add_widget(OneLineListItem(text="Sin trasplantes en esta temporada"))
                return
            for fila in rows:
                rid, f, integrante, cultivo, gen, sector, bancal = fila
                texto = f"{f} | {cultivo} Gen {gen} | {sector}-{bancal} | {integrante}"
                item = OneLineListItem(text=texto)
                item.bind(on_release=lambda _w, r=fila, t=texto: self.abrir_acciones_registro(
                    t,
                    on_editar=lambda: self.abrir_editar(r),
                    on_eliminar=lambda: self.eliminar(r[0])))
                self.ids.lista.add_widget(item)
        except Exception as e:
            self.safe_error("Error listando Trasplantes", e, "TrasplantesScreen.refrescar()")

    def eliminar(self, registro_id: int):
        try:
            delete_trasplante(registro_id)
            self.safe_snackbar("Trasplante eliminado.")
            self.refrescar()
        except Exception as e:
            self.safe_error("Error al eliminar Trasplante", e, "TrasplantesScreen.eliminar()")

    def abrir_editar(self, fila):
        rid, f, integrante, cultivo, gen, sector, bancal = fila
        fecha_input = MDTextField(hint_text="Fecha (AAAA-MM-DD)", text=str(f),
                                  size_hint_y=None, height=dp(48))
        gen_input = MDTextField(hint_text="Generación", text=str(gen), input_filter="int",
                                size_hint_y=None, height=dp(48))
        sector_input = MDTextField(hint_text="Sector", text=str(sector),
                                   size_hint_y=None, height=dp(48))
        bancal_input = MDTextField(hint_text="Bancal", text=str(bancal), input_filter="int",
                                   size_hint_y=None, height=dp(48))
        fila_ub = MDBoxLayout(orientation="horizontal", spacing=dp(8),
                              size_hint_y=None, height=dp(52))
        fila_ub.add_widget(sector_input)
        fila_ub.add_widget(bancal_input)

        content = MDBoxLayout(orientation="vertical", spacing=dp(8),
                              padding=[dp(8), dp(16), dp(8), 0],
                              size_hint_y=None, height=dp(210))
        content.add_widget(MDLabel(text=f"{cultivo} — {integrante}", bold=True, halign="center",
                                   theme_text_color="Custom", text_color=SAGE_GREEN,
                                   size_hint_y=None, height=dp(24)))
        content.add_widget(fecha_input)
        content.add_widget(gen_input)
        content.add_widget(fila_ub)
        dialog = None

        def _guardar(*_):
            try:
                nueva_fecha = validate_fecha(fecha_input.text)
                nueva_gen = validate_generacion(gen_input.text)
                nuevo_sector = validate_sector(sector_input.text)
                nuevo_bancal = int(norm_text(bancal_input.text) or 0)
                if nuevo_bancal < 1:
                    raise ValueError("Bancal debe ser >= 1.")
                update_trasplante(rid, nueva_fecha, nueva_gen, nuevo_sector, nuevo_bancal)
                if dialog:
                    dialog.dismiss()
                self.safe_snackbar("Trasplante actualizado.")
                self.refrescar()
            except Exception as e:
                self.safe_error("Error al editar Trasplante", e, "TrasplantesScreen.abrir_editar()")

        dialog = MDDialog(
            title="Editar trasplante",
            type="custom",
            content_cls=content,
            md_bg_color=CARD_BG,
            buttons=[
                MDFlatButton(text="CANCELAR", text_color=SAGE_GREEN,
                             on_release=lambda *_: dialog.dismiss()),
                MDFlatButton(text="GUARDAR", text_color=SAGE_GREEN, on_release=_guardar),
            ],
        )
        dialog.open()


class TrasplantesLogroScreen(BaseScreen):
    """Avance de trasplantes por cultivo y tabla completa de registros."""

    def on_pre_enter(self):
        self.armar()

    def ir_home(self):
        self.manager.current = "home"

    def _fila_label(self, texto, style="Body2", color=None, height=22, bold=False):
        lbl = MDLabel(text=texto, font_style=style, size_hint_y=None, height=dp(height), bold=bold)
        if color:
            lbl.theme_text_color = "Custom"
            lbl.text_color = color
        return lbl

    def armar(self):
        box = self.ids.get("logro_box")
        if box is None:
            return
        box.clear_widgets()
        try:
            temporada = get_temporada_activa()
        except Exception:
            temporada = None
        if not temporada:
            card = Factory.SectionCard()
            card.add_widget(self._fila_label("No hay temporada activa.", height=24))
            box.add_widget(card)
            return

        hoy = date.today().isoformat()
        datos = trasplantes_por_cultivo(hoy)

        card = Factory.SectionCard()
        card.add_widget(self._fila_label(
            f"Temporada {temporada['nombre']} — plantas trasplantadas vs objetivo",
            "Subtitle1", SAGE_GREEN, height=26, bold=True))
        hay_almacigo = False
        for fila in list_plan(temporada["id"]):
            cultivo = fila[0]
            tipo = norm_text(fila[3]) if len(fila) > 3 else ""
            if not tipo:
                tipo = get_perfil_cultivo(cultivo)["tipo_siembra"]
            if tipo != TIPO_SIEMBRA_ALMACIGO:
                continue
            hay_almacigo = True
            objetivo = int(fila[6]) if len(fila) > 6 and fila[6] else 0
            d = datos.get(cultivo, {"registros": 0, "plantas": 0})
            if objetivo > 0:
                pct = round(d["plantas"] / objetivo * 100.0, 1)
                card.add_widget(self._fila_label(
                    f"{cultivo}: {d['plantas']} / {objetivo} pl ({pct:g}%) · {d['registros']} bancal(es)",
                    "Body2", height=24, bold=True))
                card.add_widget(ProgressBar(max=100, value=min(100.0, pct),
                                            size_hint_y=None, height=dp(5)))
            else:
                card.add_widget(self._fila_label(
                    f"{cultivo}: {d['plantas']} pl en {d['registros']} bancal(es) · sin objetivo (editá el plan)",
                    "Caption", height=20))
        if not hay_almacigo:
            card.add_widget(self._fila_label(
                "El plan no tiene cultivos de almácigo.", "Caption", height=20))
        box.add_widget(card)

        # Tabla completa: qué se trasplantó, dónde y cuándo
        tabla = Factory.SectionCard()
        tabla.add_widget(self._fila_label("Registros (qué, dónde y cuándo)",
                                          "Subtitle1", SAGE_GREEN, height=24, bold=True))
        rows = list_trasplantes_con_id(limit=300)
        if not rows:
            tabla.add_widget(self._fila_label("Sin trasplantes registrados.", "Caption", height=20))
        for _rid, f, integrante, cultivo, gen, sector, bancal in rows:
            plantas = plantas_por_bancal_cultivo(cultivo)
            extra = f" · ≈{plantas} pl" if plantas > 0 else ""
            tabla.add_widget(self._fila_label(
                f"{f} · {cultivo} Gen {gen} · {sector}-{bancal} · {integrante}{extra}",
                "Caption", height=18))
        box.add_widget(tabla)


class HorasScreen(BaseScreen):
    """Horas de trabajo por integrante, con actividades y montos."""

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self._nombre_menu = None
        self._nombre_value = ""
        self._horas_txt = "0"
        self._checks = {}

    def hoy(self):
        return date.today().isoformat()

    def ir_home(self):
        self.manager.current = "home"

    def on_pre_enter(self):
        try:
            nombres = [i["nombre"] for i in list_integrantes()]
            if self._nombre_menu:
                try:
                    self._nombre_menu.dismiss()
                except Exception:
                    pass
            self._nombre_menu = MDDropdownMenu(
                caller=self.ids.nombre_item,
                items=[{"text": n, "on_release": (lambda x=n: self._set_nombre(x))}
                       for n in nombres],
                width_mult=5,
            )
            if self._nombre_value not in nombres:
                self._nombre_value = nombres[0] if nombres else ""
                self.ids.nombre_item.text = self._nombre_value or "Seleccionar"
            self._armar_checks()
            self.refrescar()
        except Exception as e:
            self.safe_error("Error al preparar Horas", e, "HorasScreen.on_pre_enter()")

    def open_nombre_menu(self):
        if not list_integrantes():
            self.safe_snackbar("Cargá los integrantes en Configuración de Temporada → INTEGRANTES.")
            return
        if self._nombre_menu:
            abrir_menu(self._nombre_menu)

    def _set_nombre(self, nombre: str):
        self._nombre_value = nombre
        self.ids.nombre_item.text = nombre
        self.ids.nombre_item.set_item(nombre)
        if self._nombre_menu:
            self._nombre_menu.dismiss()

    def _armar_checks(self):
        box = self.ids.get("actividades_box")
        if box is None or self._checks:
            return
        for act in ACTIVIDADES_TRABAJO:
            fila = MDBoxLayout(orientation="horizontal", size_hint_y=None,
                               height=dp(38), spacing=dp(4))
            chk = MDCheckbox(size_hint=(None, None), size=(dp(40), dp(38)))
            fila.add_widget(chk)
            fila.add_widget(MDLabel(text=act, valign="middle"))
            box.add_widget(fila)
            self._checks[act] = chk

    def _actividades_marcadas(self) -> list:
        return [act for act, chk in self._checks.items() if chk.active]

    @staticmethod
    def _parse_horas(texto: str) -> float:
        try:
            v = float(norm_text(texto).replace(",", "."))
        except Exception:
            raise ValueError("Cargá las horas con el teclado numérico.")
        if v <= 0 or v > 24:
            raise ValueError("Las horas deben estar entre 0 y 24.")
        return round(v, 2)

    def abrir_teclado_horas(self):
        def _ok(valor):
            self._horas_txt = valor or "0"
            self.ids.horas_btn.text = self._horas_txt or "0"

        self.abrir_teclado_numerico("Horas trabajadas", _ok, inicial="", permitir_coma=True)

    def guardar(self):
        try:
            fecha = validate_fecha(self.ids.fecha.text)
            if not self._nombre_value:
                raise ValueError("Elegí el nombre del integrante.")
            horas = self._parse_horas(self._horas_txt)
            actividades = self._actividades_marcadas()
            if not actividades:
                raise ValueError("Marcá al menos una tarea realizada.")
            insert_horas_trabajo(fecha, self._nombre_value, horas, actividades)
            self.safe_snackbar(f"{self._nombre_value}: {self._horas_txt} h registradas.")
            self._horas_txt = "0"
            self.ids.horas_btn.text = "0"
            for chk in self._checks.values():
                chk.active = False
            self.refrescar()
        except Exception as e:
            self.safe_error("Error al guardar Horas", e, "HorasScreen.guardar()")

    def refrescar(self):
        try:
            self.ids.lista.clear_widgets()
            rows = list_horas_con_id(limit=200)
            if not rows:
                self.ids.lista.add_widget(OneLineListItem(text="Sin registros en esta temporada"))
                return
            for fila in rows:
                rid, f, integrante, horas, actividades = fila
                horas_txt = f"{horas:g}".replace(".", ",")
                texto = f"{f} | {integrante} | {horas_txt} h | {actividades}"
                item = OneLineListItem(text=texto)
                item.bind(on_release=lambda _w, r=fila, t=texto: self.abrir_acciones_registro(
                    t,
                    on_editar=lambda: self.abrir_editar(r),
                    on_eliminar=lambda: self.eliminar(r[0])))
                self.ids.lista.add_widget(item)
        except Exception as e:
            self.safe_error("Error listando Horas", e, "HorasScreen.refrescar()")

    def eliminar(self, registro_id: int):
        try:
            delete_horas_trabajo(registro_id)
            self.safe_snackbar("Registro eliminado.")
            self.refrescar()
        except Exception as e:
            self.safe_error("Error al eliminar Horas", e, "HorasScreen.eliminar()")

    def abrir_editar(self, fila):
        rid, f, integrante, horas, actividades = fila
        marcadas = {norm_text(a) for a in str(actividades).split(",")}
        estado = {"horas": f"{horas:g}".replace(".", ",")}

        fecha_input = MDTextField(hint_text="Fecha (AAAA-MM-DD)", text=str(f),
                                  size_hint_y=None, height=dp(48))
        horas_btn = MDRaisedButton(
            text=f"Horas: {estado['horas']}",
            md_bg_color=SAGE_GREEN, text_color=(1, 1, 1, 1),
            size_hint=(1, None), height=dp(40),
        )

        def _teclado(*_):
            def _ok(valor):
                if valor:
                    estado["horas"] = valor
                    horas_btn.text = f"Horas: {valor}"
            self.abrir_teclado_numerico("Horas trabajadas", _ok)

        horas_btn.bind(on_release=_teclado)

        checks = {}
        checks_box = MDBoxLayout(orientation="vertical", size_hint_y=None,
                                 height=dp(38 * len(ACTIVIDADES_TRABAJO)))
        for act in ACTIVIDADES_TRABAJO:
            fila_chk = MDBoxLayout(orientation="horizontal", size_hint_y=None,
                                   height=dp(38), spacing=dp(4))
            chk = MDCheckbox(size_hint=(None, None), size=(dp(40), dp(38)),
                             active=(act in marcadas))
            fila_chk.add_widget(chk)
            fila_chk.add_widget(MDLabel(text=act, valign="middle"))
            checks_box.add_widget(fila_chk)
            checks[act] = chk

        content = MDBoxLayout(orientation="vertical", spacing=dp(8),
                              padding=[dp(8), dp(12), dp(8), 0],
                              size_hint_y=None,
                              height=dp(24 + 48 + 40 + 38 * len(ACTIVIDADES_TRABAJO) + 40))
        content.add_widget(MDLabel(text=integrante, bold=True, halign="center",
                                   theme_text_color="Custom", text_color=SAGE_GREEN,
                                   size_hint_y=None, height=dp(24)))
        content.add_widget(fecha_input)
        content.add_widget(horas_btn)
        content.add_widget(checks_box)
        dialog = None

        def _guardar(*_):
            try:
                nueva_fecha = validate_fecha(fecha_input.text)
                nuevas_horas = self._parse_horas(estado["horas"])
                actividades_sel = [a for a, c in checks.items() if c.active]
                if not actividades_sel:
                    raise ValueError("Marcá al menos una tarea realizada.")
                update_horas_trabajo(rid, nueva_fecha, nuevas_horas, actividades_sel)
                if dialog:
                    dialog.dismiss()
                self.safe_snackbar("Registro actualizado.")
                self.refrescar()
            except Exception as e:
                self.safe_error("Error al editar Horas", e, "HorasScreen.abrir_editar()")

        dialog = MDDialog(
            title="Editar horas de trabajo",
            type="custom",
            content_cls=content,
            md_bg_color=CARD_BG,
            buttons=[
                MDFlatButton(text="CANCELAR", text_color=SAGE_GREEN,
                             on_release=lambda *_: dialog.dismiss()),
                MDFlatButton(text="GUARDAR", text_color=SAGE_GREEN, on_release=_guardar),
            ],
        )
        dialog.open()

    def abrir_resumen(self):
        try:
            por_persona, por_actividad = resumen_horas_trabajo()
            if not por_persona:
                raise ValueError("No hay horas registradas en esta temporada.")
            lineas = ["POR INTEGRANTE", ""]
            for nombre in sorted(por_persona):
                lineas.append(f"• {nombre}: {por_persona[nombre]['horas']:g} h")
            lineas += ["", "POR ACTIVIDAD (horas repartidas)", ""]
            for act in sorted(por_actividad):
                lineas.append(f"• {act}: {round(por_actividad[act]['horas'], 1):g} h")
            dialog = MDDialog(
                title="Resumen de la temporada",
                text="\n".join(lineas),
                md_bg_color=SAGE_GREEN,
                buttons=[MDFlatButton(text="OK", text_color=(1, 1, 1, 1),
                                      on_release=lambda *_: dialog.dismiss())],
            )
            dialog.open()
        except Exception as e:
            self.safe_error("Error en Resumen de Horas", e, "HorasScreen.abrir_resumen()")

    def abrir_descarga(self):
        try:
            rows = list_horas_con_id()
            if not rows:
                raise ValueError("No hay horas registradas para descargar.")
            def _exportar(out_dir):
                out_dir = Path(out_dir)
                out_dir.mkdir(parents=True, exist_ok=True)
                path = out_dir / f"horas_trabajo_{date.today().isoformat().replace('-', '')}.csv"
                with path.open("w", newline="", encoding="utf-8") as fh:
                    w = csv.writer(fh)
                    w.writerow(["fecha", "integrante", "horas", "actividades"])
                    for _rid, f, integrante, horas, actividades in rows:
                        w.writerow([f, integrante, horas, actividades])
                self.safe_snackbar(f"Descarga generada: {path}")

            self.choose_export_dir(_exportar)
        except Exception as e:
            self.safe_error("Error al descargar Horas", e, "HorasScreen.abrir_descarga()")


class SiembrasLogroScreen(BaseScreen):
    """Avance de siembras por cultivo: semillas sembradas vs plantas objetivo."""

    def on_pre_enter(self):
        self.armar()

    def ir_home(self):
        self.manager.current = "home"

    def _fila_label(self, texto, style="Body2", color=None, height=22, bold=False):
        lbl = MDLabel(text=texto, font_style=style, size_hint_y=None, height=dp(height), bold=bold)
        if color:
            lbl.theme_text_color = "Custom"
            lbl.text_color = color
        return lbl

    def armar(self):
        box = self.ids.get("logro_box")
        if box is None:
            return
        box.clear_widgets()
        try:
            temporada = get_temporada_activa()
        except Exception:
            temporada = None
        if not temporada:
            card = Factory.SectionCard()
            card.add_widget(self._fila_label("No hay temporada activa.", height=24))
            box.add_widget(card)
            return

        try:
            plan = list_plan(temporada["id"])
        except Exception:
            plan = []
        try:
            semillas = semillas_por_cultivo(temporada["id"])
        except Exception:
            semillas = {}

        card = Factory.SectionCard()
        card.add_widget(self._fila_label(
            f"Temporada {temporada['nombre']} — semillas vs plantas objetivo",
            "Subtitle1", SAGE_GREEN, height=26, bold=True))

        if not plan:
            card.add_widget(self._fila_label("El plan no tiene cultivos todavía.", "Caption", height=20))
        for fila in plan:
            cultivo = fila[0]
            objetivo = int(fila[6]) if len(fila) > 6 and fila[6] else 0
            sembradas = semillas.get(cultivo, 0)
            if objetivo > 0:
                pct = round(sembradas / objetivo * 100.0, 1)
                faltan = max(0, objetivo - sembradas)
                card.add_widget(self._fila_label(
                    f"{cultivo}: {sembradas} / {objetivo} pl ({pct:g}%)",
                    "Body2", height=24, bold=True))
                card.add_widget(ProgressBar(max=100, value=min(100.0, pct),
                                            size_hint_y=None, height=dp(5)))
                detalle = f"Faltan {faltan} plantas" if faltan > 0 else "Objetivo cumplido"
                card.add_widget(self._fila_label(detalle, "Caption", height=18))
            else:
                card.add_widget(self._fila_label(
                    f"{cultivo}: {sembradas} sembradas · sin objetivo (editá el plan)",
                    "Caption", height=22))
        box.add_widget(card)

        # Siembras de cultivos que no están en el plan (para no perderlas de vista)
        fuera = {c: n for c, n in semillas.items() if c not in {p[0] for p in plan}}
        if fuera:
            extra = Factory.SectionCard()
            extra.add_widget(self._fila_label("Sembrado fuera del plan", "Subtitle1",
                                              WARM_AMBER, height=24, bold=True))
            for cultivo, n in sorted(fuera.items()):
                extra.add_widget(self._fila_label(f"{cultivo}: {n} semillas", "Caption", height=20))
            box.add_widget(extra)


class SanidadScreen(BaseScreen):
    """Monitoreo fitosanitario: aplicaciones de rutina, detección de enfermedad
    o plaga con diagnóstico automático, planificación de tratamientos y una
    planilla única con TODAS las aplicaciones."""

    _SEV_SEL = SAGE_GREEN
    _SEV_OFF = (0.82, 0.85, 0.82, 1)

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self._r_producto_menu = None
        self._r_producto_value = ""
        self._r_cultivo_menu = None
        self._r_cultivo_value = ""
        self._r_sector_menu = None
        self._r_sector_value = ""
        self._r_bancal_menu = None
        self._r_bancal_value = "1"
        self._bancales_por_sector = {}
        self._sectores = []
        self._cultivos = []
        self._deteccion_abierta = False
        self._det = {}
        self._trat = {}
        self._menus_vivos = []
        self._ver_hoy = False

    def hoy(self):
        return date.today().isoformat()

    def ir_home(self):
        self.manager.current = "home"

    # ------------------------------------------------------------------ carga
    def on_pre_enter(self):
        try:
            self._cargar_opciones()
            self._construir_menus_rutina()
            self._deteccion_abierta = False
            self._det = {}
            self._trat = {}
            self.ids.deteccion_box.clear_widgets()
            self.ids.btn_deteccion.text = "DETECCIÓN DE ENFERMEDAD O PLAGA"
            self.refrescar_tratamientos()
            self.refrescar_planilla()
        except Exception as e:
            self.safe_error("Error al preparar Sanidad", e, "SanidadScreen.on_pre_enter()")

    def _cargar_opciones(self):
        temporada = get_temporada_activa()
        cultivos = []
        if temporada and self.modo_local():
            for fila in list_plan(temporada["id"]):
                cultivos.append(fila[0])
        self._cultivos = cultivos or get_cultivos()
        self._bancales_por_sector = {}
        sectores = []
        if temporada and self.modo_local():
            for sector, bancales, _t in list_sectores_riego(temporada["id"]):
                sectores.append(sector)
                self._bancales_por_sector[sector] = int(bancales)
        self._sectores = sectores or SECTORES

    def _bancales_de(self, sector):
        maximo = self._bancales_por_sector.get(sector, BANCAL_MAX)
        return [str(i) for i in range(1, max(1, maximo) + 1)]

    # -------------------------------------------------- helpers de widgets
    def _make_dropdown(self, opciones, on_set, inicial=None, width_mult=4):
        """MDDropDownItem + su menú ya cableados. Devuelve (item, menu)."""
        item = MDDropDownItem()
        menu_holder = {}

        def _choose(v):
            item.text = v
            try:
                item.set_item(v)
            except Exception:
                pass
            m = menu_holder.get("m")
            if m:
                m.dismiss()
            on_set(v)

        menu = MDDropdownMenu(
            caller=item, width_mult=width_mult,
            items=[{"text": o, "on_release": (lambda x=o: _choose(x))} for o in opciones],
        )
        menu_holder["m"] = menu
        item.bind(on_release=lambda *_: abrir_menu(menu))
        val = inicial if (inicial in opciones) else (opciones[0] if opciones else "")
        if val:
            item.text = val
            try:
                item.set_item(val)
            except Exception:
                pass
        self._menus_vivos.append(menu)
        return item, menu

    def _fila_label_item(self, label_text, item, ancho=90):
        row = MDBoxLayout(adaptive_height=True, spacing=dp(12))
        row.add_widget(MDLabel(text=label_text, size_hint_x=None, width=dp(ancho),
                               valign="middle"))
        row.add_widget(item)
        return row

    def _pedir_texto(self, titulo, on_ok, inicial=""):
        campo = MDTextField(text=inicial, size_hint_y=None, height=dp(48))
        content = MDBoxLayout(orientation="vertical", spacing=dp(8),
                              padding=[dp(8), dp(16), dp(8), 0],
                              size_hint_y=None, height=dp(70))
        content.add_widget(campo)
        dlg = None

        def _ok(*_):
            valor = norm_text(campo.text)
            if dlg:
                dlg.dismiss()
            if valor:
                on_ok(valor)

        dlg = MDDialog(title=titulo, type="custom", content_cls=content, md_bg_color=CARD_BG,
                       buttons=[MDFlatButton(text="CANCELAR", text_color=SAGE_GREEN,
                                             on_release=lambda *_: dlg.dismiss()),
                                MDFlatButton(text="OK", text_color=SAGE_GREEN, on_release=_ok)])
        dlg.open()

    # ------------------------------------------------- formulario de RUTINA
    def _construir_menus_rutina(self):
        self._r_producto_menu = MDDropdownMenu(
            caller=self.ids.r_producto_item,
            items=[{"text": p, "on_release": (lambda x=p: self._set_r_producto(x))}
                   for p in PRODUCTOS_SANIDAD_LISTA],
            width_mult=5)
        if self._r_producto_value not in PRODUCTOS_SANIDAD_LISTA:
            self._set_r_producto(PRODUCTOS_SANIDAD_LISTA[0])
        self._r_cultivo_menu = MDDropdownMenu(
            caller=self.ids.r_cultivo_item,
            items=[{"text": c, "on_release": (lambda x=c: self._set_r_cultivo(x))}
                   for c in self._cultivos],
            width_mult=5)
        if self._r_cultivo_value not in self._cultivos:
            self._set_r_cultivo(self._cultivos[0])
        self._r_sector_menu = MDDropdownMenu(
            caller=self.ids.r_sector_item,
            items=[{"text": s, "on_release": (lambda x=s: self._set_r_sector(x))}
                   for s in self._sectores],
            width_mult=2)
        if self._r_sector_value not in self._sectores:
            self._set_r_sector(self._sectores[0])
        else:
            self._rebuild_r_bancal()

    def open_r_producto_menu(self):
        if self._r_producto_menu:
            abrir_menu(self._r_producto_menu)

    def open_r_cultivo_menu(self):
        if self._r_cultivo_menu:
            abrir_menu(self._r_cultivo_menu)

    def open_r_sector_menu(self):
        if self._r_sector_menu:
            abrir_menu(self._r_sector_menu)

    def open_r_bancal_menu(self):
        if self._r_bancal_menu:
            abrir_menu(self._r_bancal_menu)

    def _set_r_producto(self, p):
        if self._r_producto_menu:
            self._r_producto_menu.dismiss()
        if p == "Otro":
            self._pedir_texto("Nombre del producto", self._aplicar_r_producto_otro)
            return
        self._aplicar_r_producto(p)

    def _aplicar_r_producto_otro(self, nombre):
        self._aplicar_r_producto(nombre)

    def _aplicar_r_producto(self, p):
        self._r_producto_value = p
        self.ids.r_producto_item.text = p
        try:
            self.ids.r_producto_item.set_item(p)
        except Exception:
            pass
        sug = PRODUCTOS_SANIDAD.get(p, {}).get("dosis", "")
        if sug and not norm_text(self.ids.r_dosis.text):
            self.ids.r_dosis.text = sug

    def _set_r_cultivo(self, c):
        self._r_cultivo_value = c
        self.ids.r_cultivo_item.text = c
        try:
            self.ids.r_cultivo_item.set_item(c)
        except Exception:
            pass
        if self._r_cultivo_menu:
            self._r_cultivo_menu.dismiss()

    def _set_r_sector(self, s):
        self._r_sector_value = s
        self.ids.r_sector_item.text = s
        try:
            self.ids.r_sector_item.set_item(s)
        except Exception:
            pass
        if self._r_sector_menu:
            self._r_sector_menu.dismiss()
        self._rebuild_r_bancal()

    def _rebuild_r_bancal(self):
        opciones = self._bancales_de(self._r_sector_value)
        self._r_bancal_menu = MDDropdownMenu(
            caller=self.ids.r_bancal_item,
            items=[{"text": b, "on_release": (lambda x=b: self._set_r_bancal(x))}
                   for b in opciones],
            width_mult=2)
        if self._r_bancal_value not in opciones:
            self._set_r_bancal(opciones[0])

    def _set_r_bancal(self, b):
        self._r_bancal_value = b
        self.ids.r_bancal_item.text = b
        try:
            self.ids.r_bancal_item.set_item(b)
        except Exception:
            pass
        if self._r_bancal_menu:
            self._r_bancal_menu.dismiss()

    def guardar_rutina(self):
        try:
            fecha = validate_fecha(self.ids.r_fecha.text)
            producto = norm_text(self._r_producto_value)
            if not producto:
                raise ValueError("Elegí el producto aplicado.")
            dosis = norm_text(self.ids.r_dosis.text)
            cultivo = norm_text(self._r_cultivo_value)
            sector = validate_sector(self._r_sector_value)
            bancal = int(self._r_bancal_value)
            insert_aplicacion_sanidad(fecha, producto, dosis, cultivo, sector, bancal,
                                      tipo="rutina")
            self.safe_snackbar(f"Aplicación registrada: {producto} en {sector}-{bancal}")
            self.refrescar_planilla()
        except Exception as e:
            self.safe_error("Error al registrar aplicación", e, "SanidadScreen.guardar_rutina()")

    # -------------------------------------------- DETECCIÓN de enfermedad/plaga
    def toggle_deteccion(self):
        if self._deteccion_abierta:
            self.ids.deteccion_box.clear_widgets()
            self._deteccion_abierta = False
            self._det = {}
            self._trat = {}
            self.ids.btn_deteccion.text = "DETECCIÓN DE ENFERMEDAD O PLAGA"
        else:
            self._construir_form_deteccion()
            self._deteccion_abierta = True
            self.ids.btn_deteccion.text = "OCULTAR DETECCIÓN"

    def _construir_form_deteccion(self):
        box = self.ids.deteccion_box
        box.clear_widgets()
        self._det = {"sintoma": "Ninguno", "signo": "Ninguno", "severidad": 0,
                     "deteccion_id": None, "diagnostico": "", "productos_sugeridos": []}

        card = Factory.SectionCard()
        card.add_widget(MDLabel(text="Detección de enfermedad o plaga", font_style="Subtitle1",
                                bold=True, theme_text_color="Custom", text_color=SAGE_GREEN,
                                size_hint_y=None, height=dp(26)))

        # fecha
        fecha_tf = MDTextField(hint_text="AAAA-MM-DD", text=self.hoy())
        cal = MDIconButton(icon="calendar", size_hint=(None, None), size=(dp(48), dp(48)),
                           pos_hint={"center_y": 0.5})
        cal.bind(on_release=lambda *_: self._abrir_calendario_para(fecha_tf))
        fila_fecha = MDBoxLayout(adaptive_height=True, spacing=dp(4))
        fila_fecha.add_widget(fecha_tf)
        fila_fecha.add_widget(cal)
        card.add_widget(fila_fecha)
        self._det["fecha_tf"] = fecha_tf

        # cultivo (de la planificación)
        cult_item, _ = self._make_dropdown(
            self._cultivos, lambda v: self._det.__setitem__("cultivo", v),
            inicial=self._cultivos[0] if self._cultivos else None, width_mult=5)
        self._det["cultivo"] = self._cultivos[0] if self._cultivos else ""
        card.add_widget(self._fila_label_item("Cultivo", cult_item, ancho=70))

        # ubicación
        sec_item, _ = self._make_dropdown(
            self._sectores, self._on_det_sector,
            inicial=self._sectores[0] if self._sectores else None, width_mult=2)
        self._det["sector"] = self._sectores[0] if self._sectores else ""
        ban_opts = self._bancales_de(self._det["sector"])
        ban_item, _ = self._make_dropdown(
            ban_opts, lambda v: self._det.__setitem__("bancal", v),
            inicial="1", width_mult=2)
        self._det["bancal"] = "1"
        self._det["ban_item"] = ban_item
        fila_ub = MDBoxLayout(adaptive_height=True, spacing=dp(12))
        fila_ub.add_widget(MDLabel(text="Sector", size_hint_x=None, width=dp(70), valign="middle"))
        fila_ub.add_widget(sec_item)
        fila_ub.add_widget(MDLabel(text="Bancal", size_hint_x=None, width=dp(60), valign="middle"))
        fila_ub.add_widget(ban_item)
        card.add_widget(fila_ub)

        # síntoma
        sint_item, _ = self._make_dropdown(SINTOMAS_FITO, self._on_sintoma,
                                           inicial="Ninguno", width_mult=5)
        self._det["sint_item"] = sint_item
        card.add_widget(self._fila_label_item("Síntoma", sint_item, ancho=70))

        # signo
        sig_item, _ = self._make_dropdown(SIGNOS_FITO, self._on_signo,
                                          inicial="Ninguno", width_mult=5)
        self._det["sig_item"] = sig_item
        card.add_widget(self._fila_label_item("Signo", sig_item, ancho=70))

        # diagnóstico automático
        diag_lbl = MDLabel(text="Posible causa: elegí síntoma y/o signo.", font_style="Caption",
                           theme_text_color="Custom", text_color=WARM_AMBER,
                           size_hint_y=None, height=dp(44))
        self._det["diag_lbl"] = diag_lbl
        card.add_widget(MDLabel(text="Posible enfermedad o plaga", font_style="Overline",
                                size_hint_y=None, height=dp(18)))
        card.add_widget(diag_lbl)

        # severidad
        card.add_widget(MDLabel(text="Severidad (% de planta/órgano afectado)", font_style="Overline",
                                size_hint_y=None, height=dp(18)))
        fila_sev = MDBoxLayout(adaptive_height=True, spacing=dp(6))
        self._det["sev_btns"] = {}
        for v in SEVERIDADES_SANIDAD:
            b = MDRaisedButton(text=f"{v}%", size_hint=(1, None), height=dp(38),
                               md_bg_color=self._SEV_OFF, text_color=(0.2, 0.2, 0.2, 1))
            b.bind(on_release=lambda _w, val=v: self._set_severidad(val))
            self._det["sev_btns"][v] = b
            fila_sev.add_widget(b)
        card.add_widget(fila_sev)
        self._set_severidad(0)

        # acciones
        fila_btn = MDBoxLayout(adaptive_height=True, spacing=dp(8))
        b_guardar = Factory.MobilePrimaryButton(text="GUARDAR DETECCIÓN")
        b_guardar.bind(on_release=lambda *_: self.guardar_deteccion())
        b_plan = Factory.MobileActionButton(text="PLANIFICAR TRATAMIENTO")
        b_plan.bind(on_release=lambda *_: self.mostrar_form_tratamiento())
        fila_btn.add_widget(b_guardar)
        fila_btn.add_widget(b_plan)
        card.add_widget(fila_btn)

        box.add_widget(card)
        # contenedor para el formulario de tratamiento (debajo de la detección)
        self._det["trat_holder"] = MDBoxLayout(orientation="vertical", spacing=dp(10),
                                               size_hint_y=None, height=dp(0))
        self._det["trat_holder"].bind(minimum_height=self._det["trat_holder"].setter("height"))
        box.add_widget(self._det["trat_holder"])

    def _on_det_sector(self, sector):
        self._det["sector"] = sector
        # reconstruir opciones de bancal del sector elegido
        opciones = self._bancales_de(sector)
        item = self._det.get("ban_item")
        if item is not None:
            nuevo, _ = self._make_dropdown(opciones,
                                           lambda v: self._det.__setitem__("bancal", v),
                                           inicial="1", width_mult=2)
            # reemplazar el item viejo en su fila
            parent = item.parent
            if parent is not None:
                idx = parent.children.index(item)
                parent.remove_widget(item)
                parent.add_widget(nuevo, index=idx)
            self._det["ban_item"] = nuevo
            self._det["bancal"] = "1"

    def _on_sintoma(self, v):
        if v == "Otro":
            self._pedir_texto("Describí el síntoma",
                              lambda t: self._fijar_det_texto("sintoma", "sint_item", t))
            return
        self._det["sintoma"] = v
        self._actualizar_diagnostico()

    def _on_signo(self, v):
        if v == "Otro":
            self._pedir_texto("Describí el signo",
                              lambda t: self._fijar_det_texto("signo", "sig_item", t))
            return
        self._det["signo"] = v
        self._actualizar_diagnostico()

    def _fijar_det_texto(self, clave, item_key, texto):
        self._det[clave] = texto
        item = self._det.get(item_key)
        if item is not None:
            item.text = texto
            try:
                item.set_item(texto)
            except Exception:
                pass
        self._actualizar_diagnostico()

    def _actualizar_diagnostico(self):
        matches = diagnosticar_sanidad(self._det.get("sintoma", ""), self._det.get("signo", ""))
        lbl = self._det.get("diag_lbl")
        if not matches:
            self._det["diagnostico"] = ""
            self._det["productos_sugeridos"] = []
            if lbl:
                lbl.text = "Sin coincidencias claras — revisá manualmente."
            return
        self._det["diagnostico"] = matches[0][0]
        self._det["productos_sugeridos"] = matches[0][3]
        texto = " · ".join(f"{n} ({t})" for n, t, _s, _p in matches)
        if lbl:
            lbl.text = texto

    def _set_severidad(self, v):
        self._det["severidad"] = v
        for val, btn in self._det.get("sev_btns", {}).items():
            if val == v:
                btn.md_bg_color = self._SEV_SEL
                btn.text_color = (1, 1, 1, 1)
            else:
                btn.md_bg_color = self._SEV_OFF
                btn.text_color = (0.2, 0.2, 0.2, 1)

    def guardar_deteccion(self, silencioso=False):
        """Guarda la detección; devuelve el id (o None). Reusa el ya guardado."""
        if self._det.get("deteccion_id"):
            return self._det["deteccion_id"]
        try:
            fecha = validate_fecha(self._det["fecha_tf"].text)
            cultivo = norm_text(self._det.get("cultivo", ""))
            sector = norm_text(self._det.get("sector", ""))
            bancal = int(self._det.get("bancal", "0") or 0)
            sintoma = norm_text(self._det.get("sintoma", "Ninguno"))
            signo = norm_text(self._det.get("signo", "Ninguno"))
            diagnostico = norm_text(self._det.get("diagnostico", ""))
            severidad = int(self._det.get("severidad", 0))
            det_id = insert_deteccion(fecha, cultivo, sector, bancal, sintoma, signo,
                                      diagnostico, severidad)
            self._det["deteccion_id"] = det_id
            if not silencioso:
                dx = diagnostico or "sin diagnóstico"
                self.safe_snackbar(f"Detección guardada: {cultivo} — {dx} ({severidad}%)")
            return det_id
        except Exception as e:
            self.safe_error("Error al guardar detección", e, "SanidadScreen.guardar_deteccion()")
            return None

    # ------------------------------------------------ PLANIFICAR tratamiento
    def mostrar_form_tratamiento(self):
        holder = self._det.get("trat_holder")
        if holder is None:
            return
        holder.clear_widgets()
        sugeridos = self._det.get("productos_sugeridos", [])
        prod_inicial = sugeridos[0] if sugeridos else PRODUCTOS_SANIDAD_LISTA[0]
        self._trat = {"producto": prod_inicial, "n": 3, "frecuencia": 7}

        card = Factory.SectionCard()
        card.add_widget(MDLabel(text="Planificar tratamiento", font_style="Subtitle1", bold=True,
                                theme_text_color="Custom", text_color=SAGE_GREEN,
                                size_hint_y=None, height=dp(26)))

        # producto
        prod_item, _ = self._make_dropdown(PRODUCTOS_SANIDAD_LISTA, self._on_trat_producto,
                                           inicial=prod_inicial, width_mult=5)
        self._trat["prod_item"] = prod_item
        card.add_widget(self._fila_label_item("Producto", prod_item, ancho=80))

        # dosis recomendada (editable)
        dosis_tf = MDTextField(hint_text="Dosis recomendada",
                               text=PRODUCTOS_SANIDAD.get(prod_inicial, {}).get("dosis", ""))
        self._trat["dosis_tf"] = dosis_tf
        card.add_widget(dosis_tf)

        # número de aplicaciones (stepper)
        fila_n = MDBoxLayout(size_hint_y=None, height=dp(48), spacing=dp(6))
        fila_n.add_widget(MDLabel(text="N° aplicaciones", size_hint_x=None, width=dp(120),
                                  valign="middle"))
        btn_menos = MDIconButton(icon="minus-circle-outline", theme_text_color="Custom",
                                 text_color=SAGE_GREEN, pos_hint={"center_y": 0.5})
        btn_menos.bind(on_release=lambda *_: self._paso_n(-1))
        n_lbl = MDLabel(text="3", halign="center", valign="middle", size_hint_x=None,
                        width=dp(44), font_style="H6", theme_text_color="Custom",
                        text_color=SAGE_GREEN)
        self._trat["n_lbl"] = n_lbl
        btn_mas = MDIconButton(icon="plus-circle-outline", theme_text_color="Custom",
                               text_color=SAGE_GREEN, pos_hint={"center_y": 0.5})
        btn_mas.bind(on_release=lambda *_: self._paso_n(1))
        fila_n.add_widget(btn_menos)
        fila_n.add_widget(n_lbl)
        fila_n.add_widget(btn_mas)
        card.add_widget(fila_n)

        # frecuencia
        freq_item, _ = self._make_dropdown(
            [f"cada {d} días" for d in FRECUENCIAS_APLICACION], self._on_trat_frecuencia,
            inicial="cada 7 días", width_mult=3)
        card.add_widget(self._fila_label_item("Frecuencia", freq_item, ancho=90))

        # fecha de inicio
        fini_tf = MDTextField(hint_text="AAAA-MM-DD", text=self.hoy())
        cal = MDIconButton(icon="calendar", size_hint=(None, None), size=(dp(48), dp(48)),
                           pos_hint={"center_y": 0.5})
        cal.bind(on_release=lambda *_: self._abrir_calendario_para(fini_tf))
        fila_fini = MDBoxLayout(adaptive_height=True, spacing=dp(4))
        fila_fini.add_widget(MDLabel(text="Inicio", size_hint_x=None, width=dp(70), valign="middle"))
        fila_fini.add_widget(fini_tf)
        fila_fini.add_widget(cal)
        self._trat["fini_tf"] = fini_tf
        card.add_widget(fila_fini)

        b_crear = Factory.MobilePrimaryButton(text="CREAR TRATAMIENTO")
        b_crear.bind(on_release=lambda *_: self.crear_tratamiento())
        card.add_widget(b_crear)

        holder.add_widget(card)

    def _on_trat_producto(self, p):
        if p == "Otro":
            self._pedir_texto("Nombre del producto", self._aplicar_trat_producto_otro)
            return
        self._aplicar_trat_producto(p)

    def _aplicar_trat_producto_otro(self, nombre):
        item = self._trat.get("prod_item")
        if item is not None:
            item.text = nombre
            try:
                item.set_item(nombre)
            except Exception:
                pass
        self._aplicar_trat_producto(nombre)

    def _aplicar_trat_producto(self, p):
        self._trat["producto"] = p
        sug = PRODUCTOS_SANIDAD.get(p, {}).get("dosis", "")
        tf = self._trat.get("dosis_tf")
        if tf is not None and sug:
            tf.text = sug

    def _paso_n(self, delta):
        n = max(1, self._trat.get("n", 1) + delta)
        self._trat["n"] = n
        if self._trat.get("n_lbl") is not None:
            self._trat["n_lbl"].text = str(n)

    def _on_trat_frecuencia(self, etiqueta):
        try:
            self._trat["frecuencia"] = int(norm_text(etiqueta.replace("cada", "").replace("días", "")))
        except Exception:
            self._trat["frecuencia"] = 7

    def crear_tratamiento(self):
        try:
            det_id = self.guardar_deteccion(silencioso=True)
            producto = norm_text(self._trat.get("producto", ""))
            if not producto:
                raise ValueError("Elegí el producto a aplicar.")
            dosis = norm_text(self._trat["dosis_tf"].text)
            n = int(self._trat.get("n", 1))
            freq = int(self._trat.get("frecuencia", 7))
            fini = validate_fecha(self._trat["fini_tf"].text)
            crear_tratamiento(det_id, norm_text(self._det.get("cultivo", "")),
                              norm_text(self._det.get("sector", "")),
                              int(self._det.get("bancal", "0") or 0),
                              producto, dosis, n, freq, fini,
                              diagnostico=norm_text(self._det.get("diagnostico", "")))
            self.safe_snackbar(f"Tratamiento creado: {producto} × {n} (cada {freq} días)")
            # cerrar el formulario de detección y refrescar
            self.ids.deteccion_box.clear_widgets()
            self._deteccion_abierta = False
            self._det = {}
            self._trat = {}
            self.ids.btn_deteccion.text = "DETECCIÓN DE ENFERMEDAD O PLAGA"
            self.refrescar_tratamientos()
            self.refrescar_planilla()
        except Exception as e:
            self.safe_error("Error al crear tratamiento", e, "SanidadScreen.crear_tratamiento()")

    # ------------------------------------------ tabla de TRATAMIENTOS activos
    def refrescar_tratamientos(self):
        box = self.ids.tratamientos_box
        box.clear_widgets()
        try:
            activos = list_tratamientos_activos()
        except Exception:
            activos = []
        if not activos:
            return
        titulo = MDLabel(text="Tratamientos en curso", font_style="Subtitle1", bold=True,
                         theme_text_color="Custom", text_color=SAGE_GREEN,
                         size_hint_y=None, height=dp(26))
        box.add_widget(titulo)
        for (tid, cultivo, sector, bancal, producto, dosis, n_apl, freq, fini, diag) in activos:
            card = Factory.SectionCard()
            enc = MDBoxLayout(adaptive_height=True, spacing=dp(6))
            enc.add_widget(MDLabel(
                text=f"{producto}  ·  {cultivo}  ·  {sector}-{bancal}",
                font_style="Subtitle2", bold=True, theme_text_color="Custom",
                text_color=SAGE_GREEN))
            cancelar = MDIconButton(icon="delete-outline", theme_text_color="Custom",
                                    text_color=(0.7, 0.25, 0.25, 1), pos_hint={"center_y": 0.5})
            cancelar.bind(on_release=lambda _w, t=tid: self.cancelar_tratamiento(t))
            enc.add_widget(cancelar)
            card.add_widget(enc)
            sub = f"Dosis {dosis or '—'} · {n_apl} aplicaciones cada {freq} días"
            if diag:
                sub += f" · {diag}"
            card.add_widget(MDLabel(text=sub, font_style="Caption", size_hint_y=None, height=dp(18)))
            for (pid, fprog, realizada) in list_plan_aplicaciones(tid):
                fila = MDBoxLayout(size_hint_y=None, height=dp(36), spacing=dp(6))
                chk = MDCheckbox(size_hint=(None, None), size=(dp(32), dp(32)),
                                 pos_hint={"center_y": 0.5})
                if realizada:
                    chk.active = True
                    chk.disabled = True
                else:
                    chk.bind(active=lambda _w, val, p=pid: self._on_check(p, val))
                fila.add_widget(chk)
                estado = "✓ realizada" if realizada else "pendiente"
                fila.add_widget(MDLabel(text=f"{fprog}   ·   {estado}",
                                        font_style="Body2", valign="middle"))
                card.add_widget(fila)
            box.add_widget(card)

    def _on_check(self, plan_id, activo):
        if not activo:
            return
        try:
            marcar_aplicacion_plan(plan_id)
            self.safe_snackbar("Aplicación registrada en la planilla.")
            self.refrescar_tratamientos()
            self.refrescar_planilla()
        except Exception as e:
            self.safe_error("Error al marcar aplicación", e, "SanidadScreen._on_check()")

    def cancelar_tratamiento(self, tratamiento_id):
        def _ok():
            delete_tratamiento(tratamiento_id)
            self.safe_snackbar("Tratamiento cancelado.")
            self.refrescar_tratamientos()
        self.confirmar_accion("¿Cancelar este tratamiento y su calendario pendiente?", _ok)

    def confirmar_accion(self, texto, on_ok):
        dlg = None

        def _si(*_):
            if dlg:
                dlg.dismiss()
            on_ok()

        dlg = MDDialog(title="Confirmar", text=texto, md_bg_color=CARD_BG,
                       buttons=[MDFlatButton(text="NO", text_color=SAGE_GREEN,
                                             on_release=lambda *_: dlg.dismiss()),
                                MDFlatButton(text="SÍ", text_color=(0.7, 0.25, 0.25, 1),
                                             on_release=_si)])
        dlg.open()

    # -------------------------------------------- PLANILLA general (últimas 10)
    def refrescar_planilla(self):
        box = self.ids.planilla_box
        box.clear_widgets()
        self.ids.planilla_modo.text = "Hoy" if self._ver_hoy else "Últimas 10"
        try:
            filas = list_aplicaciones_sanidad(limit=None if self._ver_hoy else 10,
                                              solo_hoy=self._ver_hoy)
        except Exception:
            filas = []
        if not filas:
            box.add_widget(MDLabel(text="Sin aplicaciones registradas.", font_style="Caption",
                                   size_hint_y=None, height=dp(22)))
            return
        for (_id, fecha, producto, dosis, cultivo, sector, bancal, tipo, diag) in filas:
            ubic = f"{sector}-{bancal}" if sector else "—"
            etiqueta = "rutina" if tipo == "rutina" else "tratam."
            texto = f"{fecha} · {producto}"
            if dosis:
                texto += f" ({dosis})"
            texto += f" · {cultivo or '—'} · {ubic} · {etiqueta}"
            box.add_widget(MDLabel(text=texto, font_style="Caption", size_hint_y=None, height=dp(20)))

    def toggle_ver_hoy(self):
        self._ver_hoy = not self._ver_hoy
        self.refrescar_planilla()

    def abrir_resumen(self):
        try:
            r = resumen_aplicaciones_sanidad()
            n_det = len(list_detecciones())
            n_trat = len(list_tratamientos_activos())
        except Exception as e:
            self.safe_snackbar(f"Error: {e}")
            return
        top = sorted(r["por_producto"].items(), key=lambda kv: kv[1], reverse=True)[:5]
        detalle = "\n".join(f"  · {p}: {n}" for p, n in top) or "  (sin datos)"
        texto = (f"Aplicaciones totales: {r['total']}\n"
                 f"  rutina: {r['por_tipo'].get('rutina', 0)}  ·  "
                 f"tratamiento: {r['por_tipo'].get('tratamiento', 0)}\n"
                 f"Detecciones registradas: {n_det}\n"
                 f"Tratamientos activos: {n_trat}\n\n"
                 f"Productos más usados:\n{detalle}")
        dlg = MDDialog(title="Resumen de sanidad", text=texto, md_bg_color=SAGE_GREEN,
                       buttons=[MDFlatButton(text="OK", text_color=(1, 1, 1, 1),
                                             on_release=lambda *_: dlg.dismiss())])
        dlg.open()

    def abrir_descarga(self):
        import csv as _csv
        try:
            filas = list_aplicaciones_sanidad()
        except Exception as e:
            self.safe_snackbar(f"Error: {e}")
            return
        dest = self._downloads_dir() / "sanidad_aplicaciones.csv"
        try:
            with open(dest, "w", newline="", encoding="utf-8") as f:
                w = _csv.writer(f)
                w.writerow(["id", "fecha", "producto", "dosis", "cultivo", "sector",
                            "bancal", "tipo", "diagnostico"])
                w.writerows(filas)
            self.safe_snackbar(f"Guardado en {dest}")
        except Exception as e:
            self.safe_snackbar(f"Error al guardar CSV: {e}")


class StockScreen(BaseScreen):
    """Stock: carga rápida de kg por bancal, lista editable, resumen y exportación.

    Un registro = kg de un cultivo, en un bancal, en una fecha. Los sectores y
    los cultivos salen de la planificación de la temporada activa.
    """

    PASO_LISTA = 5   # cuántos registros se ven, y de a cuántos crece "VER MÁS"

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self._sector_menu = None
        self._sector_value = ""
        self._bancal_menu = None
        self._bancal_value = "1"
        self._cultivo_menu = None
        self._cultivo_value = ""
        self._sectores = []
        self._cultivos = []
        self._bancales_por_sector = {}
        self._visibles = self.PASO_LISTA

    def hoy(self):
        return date.today().isoformat()

    def ir_home(self):
        self.manager.current = "home"

    # ------------------------------------------------------------------ carga
    def on_pre_enter(self):
        try:
            self._cargar_opciones()
            self._armar_menus()
            self.ids.fecha.text = self.hoy()
            self._visibles = self.PASO_LISTA
            self.refrescar()
        except Exception as e:
            self.safe_error("Error al preparar Stock", e, "StockScreen.on_pre_enter()")

    def _cargar_opciones(self):
        """Sectores y cultivos SIEMPRE de la planificación de la temporada."""
        temporada = get_temporada_activa()
        self._sectores, self._bancales_por_sector, self._cultivos = [], {}, []
        if temporada:
            try:
                for sector, bancales, _tipo in list_sectores_riego(temporada["id"]):
                    self._sectores.append(sector)
                    self._bancales_por_sector[sector] = int(bancales)
            except Exception:
                pass
            try:
                self._cultivos = [fila[0] for fila in list_plan(temporada["id"])]
            except Exception:
                pass
        if not self._sectores:
            self._sectores = SECTORES
        if not self._cultivos:
            self._cultivos = get_cultivos()

    def _bancales_de(self, sector):
        maximo = self._bancales_por_sector.get(sector, BANCAL_MAX)
        return [str(i) for i in range(1, max(1, maximo) + 1)]

    def _armar_menus(self):
        self._sector_menu = MDDropdownMenu(
            caller=self.ids.sector_item, width_mult=2,
            items=[{"text": s, "on_release": (lambda x=s: self._set_sector(x))}
                   for s in self._sectores])
        if self._sector_value not in self._sectores:
            self._set_sector(self._sectores[0])
        else:
            self._rebuild_bancal()

        self._cultivo_menu = MDDropdownMenu(
            caller=self.ids.cultivo_item, width_mult=5,
            items=[{"text": c, "on_release": (lambda x=c: self._set_cultivo(x))}
                   for c in self._cultivos])
        if self._cultivo_value not in self._cultivos:
            self._set_cultivo(self._cultivos[0])

    def open_sector_menu(self):
        abrir_menu(self._sector_menu)

    def open_bancal_menu(self):
        abrir_menu(self._bancal_menu)

    def open_cultivo_menu(self):
        abrir_menu(self._cultivo_menu)

    def _set_sector(self, sector):
        self._sector_value = sector
        self._fijar_item("sector_item", sector)
        if self._sector_menu:
            self._sector_menu.dismiss()
        self._rebuild_bancal()

    def _rebuild_bancal(self):
        opciones = self._bancales_de(self._sector_value)
        self._bancal_menu = MDDropdownMenu(
            caller=self.ids.bancal_item, width_mult=2,
            items=[{"text": b, "on_release": (lambda x=b: self._set_bancal(x))}
                   for b in opciones])
        if self._bancal_value not in opciones:
            self._set_bancal(opciones[0])

    def _set_bancal(self, bancal):
        self._bancal_value = bancal
        self._fijar_item("bancal_item", bancal)
        if self._bancal_menu:
            self._bancal_menu.dismiss()

    def _set_cultivo(self, cultivo):
        self._cultivo_value = cultivo
        self._fijar_item("cultivo_item", cultivo)
        if self._cultivo_menu:
            self._cultivo_menu.dismiss()

    def _fijar_item(self, id_item, valor):
        item = self.ids.get(id_item)
        if item is None:
            return
        item.text = valor
        try:
            item.set_item(valor)
        except Exception:
            pass

    # ----------------------------------------------------------------- guardar
    def guardar(self):
        try:
            fecha = validate_fecha(self.ids.fecha.text)
            sector = validate_sector(self._sector_value)
            bancal = int(self._bancal_value)
            cultivo = validate_cultivo(self._cultivo_value)
            texto_kg = norm_text(self.ids.kg.text).replace(",", ".")
            if not texto_kg:
                raise ValueError("Cargá los kg.")
            kg = float(texto_kg)
            if kg <= 0:
                raise ValueError("Los kg deben ser mayores a 0.")
            insert_stock_registro(fecha, sector, bancal, cultivo, kg)
            self.ids.kg.text = ""
            self.safe_snackbar(f"{cultivo}: {kg:g} kg en {sector}-{bancal}")
            self.refrescar()
        except Exception as e:
            self.safe_error("Error al ingresar el dato", e, "StockScreen.guardar()")

    # ------------------------------------------------------------------- lista
    def refrescar(self):
        box = self.ids.lista_box
        box.clear_widgets()
        try:
            filas = list_stock_registros(limit=self._visibles)
            total = len(list_stock_registros())
        except Exception:
            filas, total = [], 0
        self.ids.modo_label.text = f"{len(filas)} de {total}"
        self.ids.ver_mas_btn.disabled = len(filas) >= total
        if not filas:
            box.add_widget(MDLabel(text="Sin registros de stock todavía.",
                                   font_style="Caption", size_hint_y=None, height=dp(22)))
            return
        for fila in filas:
            rid, fecha, sector, bancal, cultivo, kg = fila
            texto = f"{fecha} · {cultivo} · {sector}-{bancal} · {float(kg):g} kg"
            item = OneLineListItem(text=texto)
            item.bind(on_release=lambda _w, r=fila, t=texto: self.abrir_acciones_registro(
                t,
                on_editar=lambda: self.abrir_editar(r),
                on_eliminar=lambda: self.eliminar(r[0])))
            box.add_widget(item)

    def ver_mas(self):
        self._visibles += self.PASO_LISTA
        self.refrescar()

    def eliminar(self, registro_id):
        try:
            delete_stock_registro(registro_id)
            self.safe_snackbar("Registro eliminado.")
            self.refrescar()
        except Exception as e:
            self.safe_error("Error al eliminar", e, "StockScreen.eliminar()")

    def abrir_editar(self, fila):
        rid, fecha, sector, bancal, cultivo, kg = fila
        fecha_tf = MDTextField(hint_text="Fecha (AAAA-MM-DD)", text=str(fecha),
                               size_hint_y=None, height=dp(48))
        sector_tf = MDTextField(hint_text="Sector", text=str(sector),
                                size_hint_y=None, height=dp(48))
        bancal_tf = MDTextField(hint_text="Bancal", text=str(bancal), input_filter="int",
                                size_hint_y=None, height=dp(48))
        kg_tf = MDTextField(hint_text="Cant. (kg)", text=f"{float(kg):g}", input_filter="float",
                            size_hint_y=None, height=dp(48))
        fila_ub = MDBoxLayout(orientation="horizontal", spacing=dp(8),
                              size_hint_y=None, height=dp(52))
        fila_ub.add_widget(sector_tf)
        fila_ub.add_widget(bancal_tf)

        content = MDBoxLayout(orientation="vertical", spacing=dp(8),
                              padding=[dp(8), dp(16), dp(8), 0],
                              size_hint_y=None, height=dp(212))
        content.add_widget(MDLabel(text=cultivo, bold=True, halign="center",
                                   theme_text_color="Custom", text_color=SAGE_GREEN,
                                   size_hint_y=None, height=dp(24)))
        content.add_widget(fecha_tf)
        content.add_widget(fila_ub)
        content.add_widget(kg_tf)
        dialog = None

        def _guardar(*_):
            try:
                nueva_fecha = validate_fecha(fecha_tf.text)
                nuevo_sector = validate_sector(sector_tf.text)
                nuevo_bancal = int(norm_text(bancal_tf.text) or 0)
                if nuevo_bancal < 1:
                    raise ValueError("Bancal debe ser >= 1.")
                nuevos_kg = float(norm_text(kg_tf.text).replace(",", ".") or 0)
                if nuevos_kg <= 0:
                    raise ValueError("Los kg deben ser mayores a 0.")
                update_stock_registro(rid, nueva_fecha, nuevo_sector, nuevo_bancal,
                                      cultivo, nuevos_kg)
                if dialog:
                    dialog.dismiss()
                self.safe_snackbar("Registro actualizado.")
                self.refrescar()
            except Exception as e:
                self.safe_error("Error al editar", e, "StockScreen.abrir_editar()")

        dialog = MDDialog(
            title="Editar registro de stock",
            type="custom", content_cls=content, md_bg_color=CARD_BG,
            buttons=[
                MDFlatButton(text="CANCELAR", text_color=SAGE_GREEN,
                             on_release=lambda *_: dialog.dismiss()),
                MDFlatButton(text="GUARDAR", text_color=SAGE_GREEN, on_release=_guardar),
            ],
        )
        dialog.open()

    # ----------------------------------------------------------------- resumen
    def abrir_resumen(self):
        """kg totales por cultivo en la fecha del formulario (suma varios bancales)."""
        try:
            fecha = validate_fecha(self.ids.fecha.text)
            totales = resumen_stock_fecha(fecha)
        except Exception as e:
            self.safe_snackbar(f"Error: {e}")
            return
        if not totales:
            texto = f"No hay stock registrado el {fecha}."
        else:
            lineas = [f"• {c}: {kg:g} kg" for c, kg in sorted(totales.items())]
            total = round(sum(totales.values()), 2)
            texto = "\n".join(lineas) + f"\n\nTOTAL: {total:g} kg"
        dialog = MDDialog(
            title=f"Stock del {fecha}",
            text=texto, md_bg_color=SAGE_GREEN,
            buttons=[MDFlatButton(text="OK", text_color=(1, 1, 1, 1),
                                  on_release=lambda *_: dialog.dismiss())],
        )
        dialog.open()

    # -------------------------------------------------- descargar / compartir
    def _dialogo_rango(self, titulo, on_ok):
        """Pide dos fechas y si el detalle va por cultivo o por bancal."""
        hoy = date.today()
        desde_tf = MDTextField(hint_text="Desde (AAAA-MM-DD)",
                               text=hoy.replace(day=1).isoformat(),
                               size_hint_y=None, height=dp(48))
        hasta_tf = MDTextField(hint_text="Hasta (AAAA-MM-DD)", text=hoy.isoformat(),
                               size_hint_y=None, height=dp(48))
        estado = {"detalle": "cultivo"}
        btn_cultivo = MDRaisedButton(text="Totales por cultivo", size_hint=(1, None),
                                     height=dp(40), md_bg_color=SAGE_GREEN,
                                     text_color=(1, 1, 1, 1))
        btn_bancal = MDRaisedButton(text="Detalle por bancal", size_hint=(1, None),
                                    height=dp(40), md_bg_color=(0.82, 0.85, 0.82, 1),
                                    text_color=(0.2, 0.2, 0.2, 1))

        def _modo(cual):
            estado["detalle"] = cual
            eleg, otro = (btn_cultivo, btn_bancal) if cual == "cultivo" else (btn_bancal, btn_cultivo)
            eleg.md_bg_color, eleg.text_color = SAGE_GREEN, (1, 1, 1, 1)
            otro.md_bg_color, otro.text_color = (0.82, 0.85, 0.82, 1), (0.2, 0.2, 0.2, 1)

        btn_cultivo.bind(on_release=lambda *_: _modo("cultivo"))
        btn_bancal.bind(on_release=lambda *_: _modo("bancal"))

        content = MDBoxLayout(orientation="vertical", spacing=dp(8),
                              padding=[dp(8), dp(16), dp(8), 0],
                              size_hint_y=None, height=dp(212))
        content.add_widget(desde_tf)
        content.add_widget(hasta_tf)
        content.add_widget(btn_cultivo)
        content.add_widget(btn_bancal)
        dialog = None

        def _ok(*_):
            try:
                desde = validate_fecha(desde_tf.text)
                hasta = validate_fecha(hasta_tf.text)
                if desde > hasta:
                    raise ValueError("La fecha 'desde' es posterior a 'hasta'.")
                if dialog:
                    dialog.dismiss()
                on_ok(desde, hasta, estado["detalle"])
            except Exception as e:
                self.safe_error("Rango inválido", e, "StockScreen._dialogo_rango()")

        dialog = MDDialog(
            title=titulo, type="custom", content_cls=content, md_bg_color=CARD_BG,
            buttons=[
                MDFlatButton(text="CANCELAR", text_color=SAGE_GREEN,
                             on_release=lambda *_: dialog.dismiss()),
                MDFlatButton(text="ACEPTAR", text_color=SAGE_GREEN, on_release=_ok),
            ],
        )
        dialog.open()

    def _armar_planilla(self, desde, hasta, detalle):
        """Genera el .xlsx y devuelve su ruta."""
        if detalle == "bancal":
            encabezados = ["Fecha", "Sector", "Bancal", "Cultivo", "Kg"]
            filas = [[f, s, b, c, float(kg)]
                     for f, s, b, c, kg in stock_entre_fechas(desde, hasta)]
            hoja = "Stock por bancal"
        else:
            encabezados = ["Cultivo", "Kg totales"]
            filas = [[c, kg] for c, kg in sorted(stock_totales_por_cultivo(desde, hasta).items())]
            hoja = "Stock por cultivo"
        if not filas:
            raise ValueError("No hay stock registrado en ese rango de fechas.")
        nombre = f"stock_{desde.replace('-', '')}_{hasta.replace('-', '')}_{detalle}.xlsx"
        return exportar_xlsx(self._downloads_dir() / nombre, hoja, encabezados, filas)

    def abrir_descarga(self):
        def _hacer(desde, hasta, detalle):
            try:
                archivo = self._armar_planilla(desde, hasta, detalle)
                uri = guardar_en_descargas(archivo)
                destino = "Descargas/MonAgric" if uri else str(archivo)
                self.safe_snackbar(f"Planilla guardada en {destino}")
            except Exception as e:
                self.safe_error("Error al descargar", e, "StockScreen.abrir_descarga()")

        self._dialogo_rango("Descargar planilla de stock", _hacer)

    def compartir(self):
        def _hacer(desde, hasta, detalle):
            try:
                archivo = self._armar_planilla(desde, hasta, detalle)
                if compartir_archivo(archivo, MIME_XLSX, f"Stock MonAgric {desde} a {hasta}"):
                    return
                self.safe_snackbar(f"Planilla lista: {archivo}")
            except Exception as e:
                self.safe_error("Error al compartir", e, "StockScreen.compartir()")

        self._dialogo_rango("Compartir stock", _hacer)



# ==========================================================
# APP
# ==========================================================

class MonAgricSM(ScreenManager):
    """ScreenManager con carga perezosa de pantallas.

    Al arrancar solo existen 'home' y 'setup'. Las demás se instancian la
    primera vez que se navega a ellas, y se precargan en segundo plano un rato
    después del primer frame para que la navegación sea instantánea. Así el
    arranque no paga el costo de construir las ~20 pantallas de una."""

    # nombre de pantalla -> clase registrada en Factory
    LAZY = {
        "setup": "SetupScreen",
        "tareas": "TareasScreen",
        "riego": "RiegoScreen",
        "cosechas": "CosechasScreen",
        "objetivo": "ObjetivoScreen",
        "siembras": "SiembrasScreen",
        "siembras_logro": "SiembrasLogroScreen",
        "cosechas_logro": "CosechasLogroScreen",
        "horas": "HorasScreen",
        "trasplantes": "TrasplantesScreen",
        "trasplantes_logro": "TrasplantesLogroScreen",
        "sanidad": "SanidadScreen",
        "stock": "StockScreen",
    }

    def _asegurar_pantalla(self, name):
        if not name or self.has_screen(name):
            return
        clsname = self.LAZY.get(name)
        if not clsname:
            return
        try:
            self.add_widget(Factory.get(clsname)())
        except Exception as e:
            log_exception(f"No se pudo crear la pantalla {name}", e)

    def on_current(self, instance, value):
        self._asegurar_pantalla(value)
        return super().on_current(instance, value)

    def precargar_resto(self):
        """Instancia las pantallas que falten (en segundo plano)."""
        for name in self.LAZY:
            self._asegurar_pantalla(name)


class MonAgricApp(MDApp):
    backend_mode_name = StringProperty("local")
    backend_status_text = StringProperty("Datos: LOCAL")
    backend_online = BooleanProperty(True)
    temporada_header_text = StringProperty("")

    def actualizar_encabezado(self):
        try:
            perfil = get_perfil_usuario()
            temporada = get_temporada_activa()
            partes = []
            if perfil["chacra"]:
                partes.append(perfil["chacra"])
            if temporada:
                partes.append(f"Temporada {temporada['nombre']}")
            self.temporada_header_text = " · ".join(partes)
        except Exception:
            self.temporada_header_text = ""

    def _apply_desktop_preview_size(self):
        profile = norm_text(os.getenv("MONAGRIC_PREVIEW_DEVICE", DEFAULT_PREVIEW_DEVICE)).lower()
        if profile not in PREVIEW_SIZES:
            profile = DEFAULT_PREVIEW_DEVICE
        width, height = PREVIEW_SIZES[profile]
        Window.size = (width, height)
        Window.minimum_width, Window.minimum_height = (320, 560)

    def build(self):
        self.title = APP_NAME
        if platform not in ("android", "ios"):
            self._apply_desktop_preview_size()
        else:
            # Que el teclado empuje el campo activo en vez de taparlo
            Window.softinput_mode = "below_target"
        self.theme_cls.primary_palette = "Green"
        self.theme_cls.primary_hue = "700"
        self.theme_cls.accent_palette = "Amber"
        self.theme_cls.accent_hue = "700"
        self.theme_cls.theme_style = "Light"
        self.card_bg      = CARD_BG
        self.card_outline = CARD_BORDER
        self.header_color = SAGE_GREEN
        self.accent_color = WARM_AMBER
        self.surface_bg   = SURFACE_BG
        init_db()
        try:
            # Datos previos a la version con temporadas: crea la temporada activa
            migrar_objetivos_a_temporada()
            migrar_registros_a_temporadas()
        except Exception as e:
            log_exception("Fallo migracion a temporadas", e)
        self.backend_mode_name = norm_text(os.getenv("MONAGRIC_DATA_BACKEND", "local")).lower() or "local"
        self.api_base_url = norm_text(os.getenv("MONAGRIC_API_BASE_URL", DEFAULT_API_BASE_URL))
        self.repo = create_repository_from_env(self.backend_mode_name, self.api_base_url)
        self.refresh_backend_status()
        self.actualizar_encabezado()
        root = Builder.load_string(KV)
        # Primer uso: sin temporada configurada arrancamos en la pantalla de setup
        try:
            if not get_temporada_activa():
                root.current = "setup"
        except Exception:
            pass
        return root

    def refresh_backend_status(self):
        mode = self.backend_mode_name
        if mode == "local":
            self.backend_online = True
            self.backend_status_text = "Datos: LOCAL (este dispositivo)"
            return

        online = api_healthcheck(self.api_base_url)
        self.backend_online = online
        if mode == "remote":
            status = "CONECTADO" if online else "SIN CONEXION"
            self.backend_status_text = f"Datos: REMOTO ({status}) {self.api_base_url}"
            return
        if mode == "auto":
            status = "API OK" if online else "FALLBACK LOCAL"
            self.backend_status_text = f"Datos: AUTO ({status}) {self.api_base_url}"
            return
        self.backend_status_text = f"Datos: {mode.upper()}"

    def on_start(self):
        # En Android/iOS, filtrar los eventos de mouse sintéticos que SDL2
        # genera por cada toque (causaban el doble-toque: menús que se abrían y
        # cerraban solos, teclado que no aparecía, scroll errático, duplicados).
        _instalar_filtro_mouse_sintetico()
        # En el arranque frio el on_pre_enter de la pantalla inicial corre
        # antes del primer layout y el tablero queda vacio: lo construimos
        # apenas arranca el loop grafico.
        def _tablero_inicial(*_):
            try:
                if self.root and self.root.current == "home":
                    self.root.get_screen("home").armar_tablero()
            except Exception as e:
                log_exception("Fallo tablero inicial", e)
        Clock.schedule_once(_tablero_inicial, 0)

        # Precargar las pantallas restantes de a una, ya con la UI a la vista,
        # para que abrir cada sección sea instantáneo sin frenar el arranque.
        if isinstance(self.root, MonAgricSM):
            pendientes = list(self.root.LAZY)

            def _precargar_una(*_):
                if not pendientes:
                    return
                self.root._asegurar_pantalla(pendientes.pop(0))
                if pendientes:
                    Clock.schedule_once(_precargar_una, 0.3)

            Clock.schedule_once(_precargar_una, 1.5)

    def on_stop(self):
        # El Stock ahora guarda cada registro en la base al momento de cargarlo,
        # así que no queda nada pendiente por persistir al cerrar (antes había
        # cachés en JSON de sesión, chacras y cultivos por bancal).
        pass


if __name__ == "__main__":
    # Si hay errores fuera de Kivy, los registramos también
    try:
        MonAgricApp().run()
    except Exception as e:
        log_exception("Fallo fatal al ejecutar la app", e)
        print(traceback.format_exc())
        raise

            